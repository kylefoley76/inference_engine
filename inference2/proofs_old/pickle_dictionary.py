from settings import *
from openpyxl import load_workbook
import pickle
import copy
import json
import sys
from analyze_definition import process_sentences
from classes import get_dictionary, row_class


def update_synonyms(definition):
    str6 = definition[definition.find("=") + 1:-1]
    str6 = str6.strip()
    str7 = definition[1:definition.find("=")]
    str7 = str7.strip()
    dictionary.synonyms.update({str7: str6})
    dictionary.definitions.update({str7: definition})


def cut_word(word):
    try:
        if "(" in word:
            cc = word.index("(")
            word = word[:cc - 1]
            word = word.strip()
        else:
            word = word.strip()
    except:
        word = ""

    return word


def is_a_definition(str1):
    if str1 == None:
        return False
    if "e.g." in str1:
        return False
    formal_symbols = ["&", "=", conditional, iff, xorr]
    for n in formal_symbols:
        if n in str1:
            return True
    return False


def get_prepositional_relations():
    str1 = third_sheet.cell(row=2, column=5).value
    str1a = third_sheet.cell(row=3, column=5).value
    str2 = third_sheet.cell(row=6, column=5).value
    str2a = third_sheet.cell(row=7, column=5).value
    str3 = third_sheet.cell(row=10, column=5).value
    str3a = third_sheet.cell(row=11, column=5).value
    prepositional = str1.split() + str1a.split()
    non_spatio_temporal = str2.split() + str2a.split()
    spatio_temporal = str3.split() + str3a.split()
    dictionary.prepositional_relations = prepositional
    dictionary.spatio_temporal_relations = spatio_temporal
    dictionary.non_spatio_temporal_relations = non_spatio_temporal


# ex!fill_row,build_dictionary
def fill_row(i, rw, worksheet):
    rw.row_num = worksheet.cell(row=i, column=1).value
    rw.pos = worksheet.cell(row=i, column=3).value
    rw.word = worksheet.cell(row=i, column=4).value
    rw.next_word = worksheet.cell(row=i + 1, column=4).value
    rw.abbrev_relat = worksheet.cell(row=i, column=5).value
    rw.defin = worksheet.cell(row=i, column=6).value
    rw.next_defin = worksheet.cell(row=i + 1, column=6).value
    rw.edisj = worksheet.cell(row=i, column=13).value
    rw.embed = worksheet.cell(row=i, column=10).value
    rw.easy_embed = worksheet.cell(row=i, column=11).value

    return rw


def build_dictionary(kind2):
    global ws, wsar, third_sheet, dictionary
    dictionary = get_dictionary()
    if kind2 == "":
        get_prepositional_relations()
        worksheet = ws

    i = 25
    while i < 3000:

        i += 1
        rw = row_class()
        if kind2 == "":
            rw = fill_row(i, rw, worksheet)
        elif kind2 == 'mysql':
            rw = fill_mysql_row(i, rw, worksheet)

        if rw.word == 'spiesx on' or rw.abbrev_relat == 'INM':
            bb = 8

        if not not_blank(rw.word) and not not_blank(rw.next_word) and i > 300:
            break

        if not_blank(rw.pos):
            if rw.easy_embed in [1, 6]: rw.embed = False
            if rw.embed == 'done': rw.embed = False
            if not isinstance(rw.pos, int): rw.pos = rw.pos.strip()
            if isinstance(rw.word, int): rw.word = str(rw.word)
            if rw.word == "true*": rw.word = "true"
            if rw.word == "false*": rw.word = "false"
            rw.word = cut_word(rw.word)
            rw.next_word = cut_word(rw.next_word)
            # print (word)

            fir_let, sec_let = put_in_categories(rw)

            # universals are not defined, synonyms and determinative nouns are already done
            if sec_let not in ['a', 'b', 's', 'd'] and not rw.embed and not_blank(rw.defin):

                rw.defin = rw.defin.strip()

                if rw.next_word == rw.word and "e.g." not in rw.next_defin:
                    while is_a_definition(worksheet.cell(row=i + 1, column=6).value) \
                            and rw.next_word == rw.word:
                        i += 1
                        rw.defin += "| " + worksheet.cell(row=i, column=6).value.strip()
                        rw.next_word = worksheet.cell(row=i + 1, column=4).value
                        rw.next_word = cut_word(rw.next_word)

                if fir_let == "r":
                    dictionary.definitions.update({rw.abbrev_relat: rw.defin})
                    if rw.easy_embed in [1, 6] and fir_let == 'd':
                        dictionary.easy_embed.append(rw.abbrev_relat)
                        dictionary.embed_type.update({rw.abbrev_relat: rw.easy_embed})

                else:
                    if rw.easy_embed in [1, 6] and fir_let == 'd':
                        dictionary.easy_embed.append(rw.word)
                    dictionary.definitions.update({rw.word: rw.defin})

    return reduce_definitions(dictionary)


def put_in_categories(rw):
    abbrev_relat = rw.abbrev_relat
    pos = rw.pos
    word = rw.word
    defin = rw.defin
    row_num = rw.row_num

    if len(pos) == 1: pos += "z"
    dictionary.pos.update({word: pos})
    fir_let = pos[0]
    sec_let = pos[1]
    thir_let = pos[2] if len(pos) > 2 else ""
    four_let = pos[3] if len(pos) > 3 else ""
    fif_let = pos[4] if len(pos) > 4 else ""
    place_in_decision_procedure2(word, fir_let, sec_let)
    if " " in word: make_doubles(word)
    dictionary.words_to_row.update({word: row_num})
    if pos[1] in ['s', "d"]: update_synonyms(defin)
    if not_blank(abbrev_relat):
        place_in_decision_procedure2(abbrev_relat, fir_let, sec_let)
        if abbrev_relat not in dictionary.words_to_row.keys():
            dictionary.words_to_row.update({abbrev_relat: row_num})
        if fir_let == 'r': dictionary.pos.update({abbrev_relat: pos})
        if fir_let == 'r': dictionary.rel_abbrev.update({word: abbrev_relat})
        if len(abbrev_relat) == 4 and abbrev_relat[-1] == "P": dictionary.past_participles.append(abbrev_relat)
    if fir_let == 'r' and sec_let not in ['s', 'a'] and not not_blank(abbrev_relat):
        raise Exception(f"you forgot to give {word} an abbreviation")

    return fir_let, sec_let


def make_doubles(word):
    m = word.count(" ")
    if m == 1:
        word1 = copy.copy(word)
        y = word1.find(" ")
        word1 = word1[:y]
        dictionary.doubles.setdefault(word1, []).append(word)
    if m == 2:
        word1 = copy.copy(word)
        y = word1.find(" ")
        word1 = word1[:y]
        dictionary.triples.setdefault(word1, []).append(word)


def place_in_decision_procedure2(word, part_of_speech, sub_part):
    category = None
    if sub_part == 'd':  # determinative nouns
        category = 19
    elif word in ['distinct from', 'different']:  # special synonyms, distinct from
        category = 20
    elif sub_part == 's':  # synonyms
        category = 18
    elif word == "i":
        category = .5
    elif part_of_speech == 'd' and sub_part not in ['b', 'i', 'e', 'd']:
        category = 1  # determinative, possessive pronouns
    elif part_of_speech == 'p':
        category = 2
    elif word == 'and' + uc:
        category = 5  # and
    elif part_of_speech == 'u':  # relative pronouns
        category = 9
    elif word in ['AS', 'as']:  # AS
        category = 11
    elif word in dictionary.spatio_temporal_relations:
        category = 13
    elif word in dictionary.non_spatio_temporal_relations:
        category = 13.5
    elif word == 'there':  # there
        category = 14
    elif part_of_speech == 'd' and sub_part == 'b':  # universals
        category = 15
    elif part_of_speech == 'd' and sub_part == 'e':  # many
        category = 16
    elif part_of_speech == 'm':  # negative determiners
        category = .4

    if category != None: dictionary.decision_procedure.update({word: category})


def unembed(def_info, definiendum, red_kind):
    abbrev = []
    ogreek = def_info[5]
    type3 = False
    if red_kind == 6:
        type3 = True

    for sent, greek_sent in zip(def_info[3], def_info[6]):
        if mini_e in sent and len(greek_sent) == 1:
            abbrev.append(greek_sent)

    new_conjunct = " & ".join(abbrev)
    new_conjunct = " & " + new_conjunct

    e = 0
    special = False
    for lst1, lst2, lst3, lst4, lst6 in zip(def_info[1], def_info[2], def_info[3], def_info[4], def_info[6]):
        e += 1

        if e > 1:
            parent = ".".join(lst1[:-1])
            parent_pos = def_info[2].index(parent)
            parent_type = def_info[4][parent_pos][1]

            if parent_type == iff and parent in ['1', '1.1', '1.2']:
                if lst1[-1] == "2":

                    ogreek_definiens = lst6
                    if "&" not in ogreek_definiens:
                        special = True

                    break

    for gsent in abbrev: ogreek = ogreek.replace(" & " + gsent, "")

    if special:
        ngreek_definiens = "(" + ogreek_definiens + new_conjunct + ")"

    elif type3:

        ngreek_definiens = ogreek_definiens[:-2] + new_conjunct + "))"
    else:
        ngreek_definiens = ogreek_definiens[:-1] + new_conjunct + ")"
    ogreek = ogreek.replace(ogreek_definiens, ngreek_definiens)

    for greek_sent, english in zip(def_info[6], def_info[3]):
        ogreek = ogreek.replace(greek_sent, english)

    ogreek = ogreek.strip()
    ogreek = remove_outer_paren(ogreek)

    return find_sentences(ogreek, definiendum), ogreek


def print_new_definitions():
    for word in dictionary.easy_embed:
        definition = dictionary.definitions.get(word)
        row_num = dictionary.words_to_row.get(word)
        if "|" in definition:
            list1 = definition.split("|")
            for i, part in enumerate(list1):
                ws.cell(row=row_num + i, column=6).value = part
        else:
            ws.cell(row=row_num, column=6).value = definition
        ws.cell(row=row_num, column=8).value = "done"
    return


def reduce_definitions(dictionary):
    for definiendum, definition in dictionary.definitions.items():

        pos = dictionary.pos.get(definiendum)

        if definiendum != "0":
            bb = 8

            # print (definiendum)

            if pos[1] not in ["s", "d", "b"]:
                definition = dictionary.definitions.get(definiendum)

                definition = definition.replace("|", "")

                dictionary = process_sentences(definition, definiendum, dictionary)

    return dictionary


def save_dictionary(dictionary):
    for k, v in dictionary.categorized_sent.items():
        lst2 = [x.sentences for x in v]
        with open("json_dict/" + k + ".json", "w") as fp:
            json.dump(lst2, fp)
        for x in v: x.sentences = None


def start_pickle(user="", kind2=""):
    global dictionary, wb4, ws, third_sheet, kind

    if user == "":

        if kind2 != "":
            kind = kind2

        arguments = sys.argv
        if len(arguments) > 1:
            kind = int(arguments[1])
        else:
            kind = 6

        wb4 = load_workbook('/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/dictionary5.xlsx')
        ws = wb4.worksheets[0]
        third_sheet = wb4.worksheets[2]

    if kind == 6:

        print ('dictionary built and reduced')
        dictionary = build_dictionary(kind2)

    elif kind == 7:
        wb4 = load_workbook('/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/dictionary5.xlsx')
        ws = wb4.worksheets[0]
        dictionary = build_dictionary()
        for word in dictionary.easy_embed:
            definition = dictionary[1].get(word)
            definition = adjust_definition(word, definition, dictionary)
            dictionary[1][word] = definition

        print_new_definitions()
        wb4.save('/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/dictionary5.xlsx')

    if user == "":
        save_dictionary(dictionary)
        dictionary2 = open(user + 'z_dict_words.pkl', 'wb')
        pickle.dump(dictionary, dictionary2)
        dictionary2.close()
    else:
        return dictionary


start_pickle()
