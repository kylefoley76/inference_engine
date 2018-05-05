
from collections import defaultdict
import itertools
from itertools import combinations
import copy
import json, operator, collections
from openpyxl import load_workbook
from classes import implications
import time
import random
# from random import *

try:
    from settings import *
    from general_functions import *
    from classes import get_output
    from search_for_instantiation import loop_through_gsent, try_instantiation
    from prepare_for_print import rearrange
except:
    from .settings import *
    from .general_functions import *
    from .classes import get_output
    from .search_for_instantiation import loop_through_gsent, try_instantiation
    from .prepare_for_print import rearrange



def determine_class(user, dictionary2=""):
    global word, dictionary, greek_var2, words_used
    global constituents
    begin_time = time.time()
    if dictionary2 == "":
        pkl_file = open(user + 'z_dict_words.pkl', 'rb')
        dictionary = pickle.load(pkl_file)
        pkl_file.close()
    else:
        dictionary = dictionary2

    dictionary.level = {}
    constituents = {}
    dictionary.necessary_predicates = {}
    word_list = list(dictionary.categorized_sent.keys())

    by_basicity = order_list(word_list)

    list_of_exc(by_basicity)

    # redo_groups()

    determine_entailments(by_basicity)

    num_lemmas = quick_pairings(by_basicity)

    time_stats(num_lemmas, begin_time)

    if dictionary2 == "":
        result = open('z_dict_words.pkl', 'wb')
        pickle.dump(dictionary, result)
        result.close()
    else:
        return dictionary


def time_stats(num_lemmas, begin_time):
    e = len(dictionary.groups)
    end_time = time.time() - begin_time
    b = end_time / num_lemmas
    b = 1 / b
    print (str(e) + " unique predicates")
    print (decimal_numbers2(end_time))
    print (str(int(b)) + " lemmas per second")


def ranking(parts, by_basicity):
    list1 = []
    for part in parts:
        level = by_basicity.get(part)
        english, _ = translate_to_english(part)
        str1 = english + " " + str(level)
        list1.append(str1)

    return ", ".join(list1)


def decompose_hword(word):
    nword = word[1:]
    if nword[0] in ["a", "e", "i", "o", "u"]:
        article = "an"
    else:
        article = "a"
    return "has " + article + " " + nword


def translate_to_english(word):
    is_hword = False
    if hprop(word):
        english = decompose_hword(word)
        is_hword = True
    elif word[0].isupper():
        english = get_key(dictionary.rel_abbrev, word)
    else:
        english = word
    return english, is_hword


def list_of_exc(by_basicity, kind=0):
    kind = 0
    if kind == 1:
        by_basicity = dict(by_basicity)
        wb4 = load_workbook('/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/dictionary5.xlsx')
        ws = wb4.worksheets[4]
        num = 1
        for word, level in by_basicity.items():
            if "," not in word:
                num += 1
                english, is_hword = translate_to_english(word)
                if not is_hword:
                    parts = constituents.get(word)
                    if parts != None:
                        parts_rank = ranking(parts, by_basicity)
                else:
                    parts_rank = ""

                ws.cell(row=num, column=1).value = english
                ws.cell(row=num, column=2).value = level
                ws.cell(row=num, column=3).value = parts_rank

        wb4.save('/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/dictionary5.xlsx')


def redo_groups():
    for key in dictionary.groups.keys():
        if key not in dictionary.ontology[1].keys():
            dictionary.predicates.append(key)

    for k, v in dictionary.groups.items():
        if isinstance(v, set):
            if dictionary.ontology[2].get(k) == None:

                dictionary.possible.setdefault(k, set()).union(v)

                for concept in v:
                    dictionary.possible.setdefault(k, set()).add(concept)

        elif dictionary.ontology[2].get(k) == None:

            dictionary.necessary.update({k: v})

    return



def neg_property(sent, abbreviations):
    if sent[13] == "J" and sent[14] in abbreviations.keys():
        prop = abbreviations.get(sent[14])
        return "~", prop
    return "", ""


def get_var_match(sentences, sent_list, abbreviations, dword, svars, cls = "", output=""):
    premise = True if output != "" else False
    if not premise:
        perfect_disjunct = cls.def_stats.perfect_disjunct
        isdisjunctive = cls.def_stats.isdisjunctive
    else:
        perfect_disjunct = False
        isdisjunctive = []
    vars = {}

    for idx in sent_list:
        if dword == 'dead':
            bb = 8
        if idx in isdisjunctive or perfect_disjunct:
            disjunctive = True
        else:
            disjunctive = False

        sent = sentences[idx]
        if sent[10] == 'c':
            bb = 8
        for num in sent[42]:
            if num == 8:
                vars.setdefault(sent[num], []).append("relationship")
                svars.setdefault(sent[num], set()).add("relationship")

            elif sent[num] in abbreviations.values():
                pass
            elif num == 14 and sent[13] in ["I", "J", "V"] and \
                    sent[14] in abbreviations.keys():
                pass
            else:
                tilde = ""
                const = ""
                if sent[3] == "~":
                    if sent[7] == 'cf':
                        tilde, const = neg_property(sent, abbreviations)

                    elif "q" in sent[7]:
                        tilde, const = get_antecedent(sent, sentences, dword, svars)

                str2 = sent[58] + const
                str1 = get_constant(sent, dword, premise)
                if hprop(str2):
                    if num == 10:
                        if disjunctive:
                            disjunctive_abbreviation(vars, svars, sent, num, str1, str2)
                        else:
                            vars.setdefault(sent[num], []).append(str1)
                            svars.setdefault(sent[num], set()).add(sent[58] + tilde)
                else:
                    if isrelat(str1[0]):
                        str1 += sent_pos_name.get(num)
                    if isrelat(str2[0]):
                        str2 += sent_pos_name.get(num)
                    if disjunctive:
                        disjunctive_abbreviation(vars, svars, sent, num, str1, str2 + tilde)
                    else:
                        vars.setdefault(sent[num], []).append(str1)
                        svars.setdefault(sent[num], set()).add(str2 + tilde)

    make_vars_different(vars)
    return vars

def disjunctive_abbreviation(vars, svars, sent, num, str1, str2):
    if sent[7][:2] == 'cx':
        parent = sent[6].split(".")
        parent = ".".join(parent[:-1])
    else:
        parent = sent[6]
    var = sent[num] + "," + parent
    vars.setdefault(var, []).append(str1)
    svars.setdefault(var, set()).add(str2)


def make_vars_different(vars):
    for var, lst in vars.items():
        if len(set(lst)) != len(lst):
            ct = collections.Counter(lst)
            for k, v in dict(ct).items():
                if v > 1:
                    e = 0
                    for f, prop in enumerate(lst):
                        if prop == k:
                            e += 1
                            prop = prop + str(e)
                            lst[f] = prop

    for var, lst in vars.items():
        vars[var] = set(lst)

    return


def get_conditionals(cls, cond_sent, sbicond, word, arity, svars):
    if cls.def_stats.connection_type == 'e' and len(cls.def_stats.flat_con_index) == 1:
        if len(arity) > 1:
            bb = 8
        else:
            for k, v in svars.items():
                if word in v:
                    for prop in v:
                        if prop != word:
                            str1 = prop + "." + word
                            sbicond.update({str1: "necessary"})



def get_rel_status(predicate, cls):
    if "," in predicate or predicate.startswith("qu"):
        num = cls.def_stats.flat_ant_index[0]
        if cls.sentences[num][13] in ["I", "J", "V"]:
            dictionary.kind.update({predicate: "z"})
            return False
        else:
            dictionary.kind.update({predicate: "r"})
            return True
    else:
        if dictionary.pos.get(predicate)[0] != "r":
            return False
        else:
            return True


def fill_primitives(dictionary, level, word):
    if dictionary.pos.get(word)[0] == 'r':
        dictionary.variable_entail.update({word + "s": [set(), set(), level]})
        dictionary.variable_entail.update({word + "o": [set(), set(), level]})
        if word in dictionary.disjunctive.keys():
            dictionary.disjunctive.update({word + "s": True})
            dictionary.disjunctive.update({word + "o": False})
    else:
        dictionary.variable_entail.update({word: [set(), set(), level]})

    if word == 'personhood':
        bb = 8

    if word in dictionary.categorized_sent.keys():
        reduced_def = get_word_info(dictionary, word, "")
        for cls in reduced_def:
            necessary_predicates(cls)
    return


def determine_entailments(by_basicity, dictionary2=[], reduced_def="", output=""):
    global dictionary
    premise = False
    cond_sent = []
    sbicond = {}
    if dictionary2 != []:
        dictionary = dictionary2
        premise = True
    else:
        predicates = {}
    z = -1
    for word, level in by_basicity:
        z += 1
        if z == 39:
            bb = 8
        if word != '0':

            if "," not in word:
                if level == 0 or word in dictionary.disjunctive.keys():
                    fill_primitives(dictionary, level, word)
                else:

                    if word == 'sexed':
                        bb = 8

                    var_dict = []
                    var_cat = {}
                    oword = word
                    if not premise:
                        if "," not in word:
                            reduced_def = get_word_info(dictionary, word, "")
                        else:
                            reduced_def, oword = get_embed_rd(word)
                        abbreviations = dictionary.def_constants.get(oword, {})
                    else:
                        abbreviations = output.abbreviations

                    for f, cls in enumerate(reduced_def):
                        svars = {}
                        if cls.def_stats.isdisjunctive != []:
                            bb = 8
                        is_relation = get_rel_status(word, cls)
                        list1 = cls.def_stats.flat_ant_index
                        ant_var = get_var_match(cls.sentences, list1, abbreviations, word, svars, cls, output)
                        list2 = cls.def_stats.flat_con_index
                        if not word.startswith('qu') and not word.startswith('ax'):
                            arity = cls.def_stats.arity
                            fill_var_dict(arity, cls.sentences, list2, var_dict, level)
                        con_var = get_var_match(cls.sentences, list2, abbreviations, word, svars, cls, output)
                        loop_through_index(cls, ant_var, con_var, word + str(f), premise)
                        check_category_errors(svars, var_cat, word, cls)
                        get_conditionals(cls, cond_sent, sbicond, word, arity, svars)
                        necessary_predicates(cls)

                    if word.startswith("qu"):
                        output.gsent.update({word: cls.def_stats.entailments})
                        output.gstats.update({word: cls})

                    elif not word.startswith("ax"):
                        pass
                        # fill_dict(dictionary.variable_entail, word, var_dict, is_relation)
                        # fill_var_cat(var_cat, arity, word)



def fill_var_cat(var_cat, arity, word, is_relation, predicates):
    main_var = [x[0] for x in arity]

    for var, categories in var_cat.items():
        if var in main_var:
            if is_relation:
                str2 = sent_pos_name(var)
                str1 = word + str2
                predicates.update({str1: categories})


def fill_dict(dict1, word, var_dict, is_relation):
    if is_relation:
        for idx in var_dict:
            dict1.update({word + idx[0]: idx[1]})
    else:
        dict1.update({word: var_dict[0][1]})


def fill_var_dict(arity, sentences, idx, var_dict, level):
    fill_set = lambda p, n, s, c: p.add(c) if s[3] == "" else n.add(c)
    for var, pos in arity:
        pst = set()
        nst = set()
        for num in idx:
            for sent_idx in sentences[num][42]:
                if sentences[num][sent_idx] == var:
                    posp = dictionary.pos.get(sentences[num][58])
                    if sent_idx == 8:
                        pst.add("relationship")

                    if posp == None or posp[0] != 'r':
                        fill_set(pst, nst, sentences[num], sentences[num][58])
                    else:
                        fill_set(pst, nst, sentences[num], sentences[num][58] + sent_pos_name.get(sent_idx))

        if pst != set() or nst != set():
            var_dict.append([sent_pos_name.get(pos), [pst, nst, level]])


def loop_through_index(cls, ant_var, con_var, word, premise=False):
    sentences = cls.sentences
    ant_list = cls.def_stats.flat_ant_index
    con_list = cls.def_stats.flat_con_index
    implied = implications()
    ant_neg = implied.ant_neg
    ant_pos = implied.ant_pos
    con_neg = implied.con_neg
    con_pos = implied.con_pos
    if word == 'particle0':
        bb = 8

    implied.ant_var = ant_var
    implied.con_var = con_var
    bb = 8

    for i in range(2):
        list1 = ant_list if i == 0 else con_list
        for num in list1:
            sent = sentences[num]
            is_prop_constant(i, ant_var, con_var, ant_pos, con_pos, sent)
            constant = get_constant(sent, word, premise)
            if i == 0:
                if sent[3] == "~":
                    ant_neg.add(constant)
                else:
                    ant_pos.add(constant)
            elif i == 1:
                if sent[3] == "~":
                    con_neg.add(constant)
                    for_contradiction(sentences, num, implied, cls, word)
                else:
                    con_pos.add(constant)

    cls.def_stats.entailments = implied

    return




def is_prop_constant(i, ant_var, con_var, ant_pos, con_pos, sent):
    if sent[8] != None:
        prop_var = sent[8]
        st = con_var.get(prop_var)
        if st != None:
            st.add("relationship")
            con_pos.add("relationship")
        else:
            ant_var.get(prop_var).add("relationship")
            ant_pos.add("relationship")


def for_contradiction(sentences, num, implied, cls, word):
    main_var = [x[0] for x in cls.def_stats.arity]
    if 'q' in sentences[num][7] \
            and any(sentences[num][x] in main_var for x in sentences[num][42]):
        implied.con_neg_4c.add(sentences[num][58])
    else:
        constant = get_constant(sentences[num], word)
        implied.con_neg_4c.add(constant)



def get_constant(sent, word, premise=False):
    if premise:
        conn_type = 7
        hnum = 6
    elif "," not in word:
        conn_type = 56
        hnum = 48
    else:
        hnum = 48
        conn_type = 50
    if sent[conn_type] == 'c':
        return sent[58]
    else:
        return sent[conn_type] + sent[hnum] + "," + sent[58]




def quick_pairings(by_basicity):
    lemmata2 = necessary_inferences()
    # concepts, properties, disjunctive_concepts = get_concepts(by_basicity)
    temp_test(lemmata2)
    # new_lemmas = disjunctive_lemmas(disjunctive_concepts, lemmata2)

    # lemmata2 = {**new_lemmas, **lemmata2}
    # lemmata2 = {**dictionary.ontology[2], **lemmata2}
    # for lemma, value in dictionary.necessary_predicates.items():
    #     lemmata2.update({lemma: value})

    # pkl_file = open('lemmata.pkl', 'rb')
    # lemmata = pickle.load(pkl_file)
    # pkl_file.close()

    # for k, v in lemmata2.items():
    #     tv = lemmata.get(k)
    #     if tv != None:
    #         if (v == 'impossible' and tv != 'impossible') or \
    #             tv == 'impossible' and v != 'impossible':
    #
    #             print (k + " " + v)



    # print_lemmas(lemmata2, concepts)

    print (str(len(lemmata2)) + " lemmas")



    pkl_file = open('lemmata.pkl', 'wb')
    pickle.dump(lemmata2, pkl_file)
    pkl_file.close()

    return len(lemmata2)



def print_lemmas(lemmata2, concepts):
    kind = 2
    if kind == 1:
        lemmata3 = []

        while True:
            b = random.choice(lemmata2.items())



        for k, v in lemmata2.items():
            list1 = k.split(".")
            if list1[0] in concepts or list1[1] in concepts:
                lemmata3.append([list1[0], list1[1], v])

        for lst in lemmata3:
            if lst[2] == 'possible':
                print (lst[0] + " " + lst[1] + " " + lst[2])
        print ("")
        for lst in lemmata3:
            if lst[2] == 'impossible':
                print (lst[0] + " " + lst[1] + " " + lst[2])

        print ("")
        for lst in lemmata3:
            if lst[2] == 'necessary':
                print (lst[0] + " " + lst[1] + " " + lst[2])




def disjunctive_lemmas(disjunctive_concepts, lemmata):
    new_lemmas = {}
    for k, v in disjunctive_concepts.items():
        for concept in v:
            for lemma, modality in lemmata.items():
                if concept in lemma:

                    new_lemma = lemma.replace(concept, k)
                    list1 = new_lemma.split(".")
                    new_lemma = pair(list1[0], list1[1], ".")
                    if new_lemma == 'AGs.symbol':
                        bb = 8

                    if new_lemmas.get(new_lemma) == None:
                        new_lemmas.update({new_lemma: modality})

                    elif new_lemmas.get(new_lemma) == 'impossible' and modality in ['possible', 'necesary']:
                        new_lemmas.update({new_lemma: "possible"})

                    elif new_lemmas.get(new_lemma) == 'necessary' and modality in ['possible', 'necesary']:
                        new_lemmas.update({new_lemma: "possible"})

    return new_lemmas




def get_antecedent(sent, sentences, dword, svars):
    if dword == 'non-whole':
        bb = 8
    kind = dictionary.kind.get(dword)
    if kind == 'r':
        return "", ""
    else:
        for var, prop in svars.items():
            if dword in prop:
                key_var = var
                break

    if sent[7].startswith("q"):
        parent = ".".join(sent[6].split(".")[:-1])

    elif sent[7][1] == 'q':
        parent = ".".join(sent[6].split(".")[:-2])

    const = ""
    for e, sent2 in enumerate(sentences):
        if sent2[7].startswith("a"):
            tparent = ".".join(sent2[6].split(".")[:-1])
            if tparent == parent:
                const = sent2[58]
                break

        elif len(sent2[7]) > 1 and sent2[7][1] == 'a':
            tparent = ".".join(sent2[6].split(".")[:-2])
            if tparent == parent:
                const = sent2[58]
                break

    if const == 'thing':
        for num in sent[42]:
            if sent[num] == key_var:
                pos = sent_pos_name.get(num)
                str1 = sent[58] + pos
                dictionary.impossible.update({dword: str1})

                break
        else:
            print (f"failed to find thing ant in {dword}")
    elif sent[13] in ["H", "W"]:
        str1 = sent[13] + const
        dictionary.impossible.update({dword: str1})

    if const == "":
        return "", ""
    else:
        if const == 'thing':
            return "~", ""
        else:
            # print (f"the word {const} is negated in definition {dword}")
            return "~", const

def necessary_predicates(cls):
    bicond = cls.def_stats.connection_type
    if cls.def_stats.def_word == 'personhood':
        bb = 8

    if bicond == 'e' and len(cls.def_stats.flat_con_index) == 1:
        ant_const = cls.def_stats.def_word
        if len(cls.def_stats.flat_ant_index) > 1:
            pass
            # print (f"{cls.def_stats.def_word} does not have a single antecedent")

        con_const = cls.sentences[cls.def_stats.flat_con_index[0]][58]
        str1 = ant_const + "." + con_const
        str2 = con_const + "." + ant_const

        if str2 == "essence.particular" + un:
            bb = 8
        if str1 == "essence.particular" + un:
            bb = 8

        # print (str1)
        dictionary.necessary_predicates.update({str1: "necessary"})
        dictionary.necessary_predicates.update({str2: "necessary"})

    return


def necessary_inferences():
    new_possible = {}
    new_necessary = {}
    new_impossible = {}
    new_impossible2 = []
    lemmata2 = {}
    # for k, v in dictionary.necessary.items():
    #     parents = dictionary.ontology[1].get(v)
    #     if parents != None:
    #         new_possible.update({k: parents})
    #
    #     children = dictionary.ontology[0].get(v)
    #     if children != None:
    #         new_necessary.update({k: children})

    for k, v in dictionary.impossible.items():
        children = dictionary.ontology[0].get(k)
        if children != None:
            for child in children:
                new_impossible.setdefault(child, []).append(v)

    dictionary.impossible = {}

    for k, v in new_impossible.items():
        for prop in v:
            pair1 = build_lemma(k, prop)
            pair2 = build_lemma(prop, k)
            lemmata2.update({pair1: "impossible"})
            lemmata2.update({pair2: "impossible"})
            dictionary.impossible.update({pair1: "impossible"})
            dictionary.impossible.update({pair2: "impossible"})
            new_impossible2.append(pair1)

    # for k, v in dictionary.possible.items():
    #     for st in v:
    #         children = dictionary.ontology[0].get(st)
    #         if children != None:
    #             new_possible.update({st: children})


    return lemmata2


def distribute(tot_var, cls):
    if cls.def_stats.isdisjunctive != [] and \
        not cls.def_stats.perfect_disjunct:
        disj_var = set()
        for key in tot_var.keys():
            if "," in key:
                list1 = key.split(",")
                disj_var.add(list1[0])

        for key, var in tot_var.items():
            if "," not in key and key in disj_var:
                for key2, var2 in tot_var.items():
                    key3 = key2[:key2.find(",")]
                    if key == key3:
                        var2 = var2 | var
                        tot_var.update({key2: var2})

        for var in disj_var:
            if var in tot_var.keys():
                del tot_var[var]

    return


def check_category_errors(tot_var, var_cat, word, cls):
    if word == 'INC':
        bb = 8
    distribute(tot_var, cls)
    for var, properties in tot_var.items():
        categories = set()
        sets_of_cat = []
        tvalues = {}
        for property in properties:
            neg = ""
            if "~" in property:
                property = property[:-1]
                neg = "~"
            if neg == "":
                category = dictionary.groups.get(property)

                # assert category != None, f"{property} does not have a category in {word}"
                if category == None:
                    # print (f"{property} does not have a category in {word}")
                    pass
                elif isinstance(category, set):
                    assert neg == ""
                    for st in category: tvalues.update({st: neg})
                    sets_of_cat.append(category)
                else:
                    tvalues.update({category: neg})
                    categories.add(category)

        categories = list(categories)
        for st in sets_of_cat: categories.append(st)
        var_cat.update({var: categories})
        if len(categories) > 2:
            tpls = combinations(categories, 2)
            for tpl in tpls:
                if tpl[0] != tpl[1]:
                    outcome, pair1 = check_category_errors3(tpl, tvalues, word)
                    if not outcome:
                        print (f"{var} in {word} contains the contradiction {pair1} ")
                        # raise Exception

    return


def check_category_errors3(tpl, tvalues, word):
    b = 3
    if isinstance(tpl[0], set) and isinstance(tpl[1], str):
        b = 0
        c = 1
    elif isinstance(tpl[0], str) and isinstance(tpl[1], set):
        b = 1
        c = 0
    elif isinstance(tpl[0], set) and isinstance(tpl[1], set):

        cart_product = [i for i in itertools.product(*[tpl[0], tpl[1]])]

        for tpl in cart_product:
            if tpl[0] == tpl[1]:
                return True, ""
            else:
                result, pair1 = check_category_errors2(tpl[0], tpl[1], tvalues)
                if result:
                    return True, ""
        else:
            return False

    if b == 3:
        result, pair1 = check_category_errors2(tpl[0], tpl[1], tvalues)
        if not result:
            return False, pair1

    else:
        for st in tpl[b]:
            result, pair1 = check_category_errors2(st, tpl[c], tvalues)
            if result:
                return True, ""
        return False, pair1

    return True, ""


def check_category_errors2(tpl1, tpl2, tvalues):
    if tpl1 == tpl2: return True, ""
    if tpl1 == 'thing' or tpl2 == 'thing': return True, ""
    pair1 = pair(tpl1, tpl2, ".")
    tv1 = tvalues.get(tpl1)
    tv2 = tvalues.get(tpl2)
    if tv1 == "~" and tv2 == "~":
        pass
    else:
        tvalue2 = False if tv1 == "~" or tv2 == "~" else True
        tvalue = dictionary.ontology[2].get(pair1)
        if tvalue in ["possible", "necessary"]:
            tvalue = True
        else:
            tvalue = False

        return tvalue == tvalue2, pair1



def get_concepts(by_basicity):
    disjunctive_concepts = {}
    concepts = set()
    for concept, _ in by_basicity:

        if concept[2] == 'n':
            if concept[0] in dictionary.groups.keys():
                category = dictionary.groups.get(concept[0])
                assert category != None
                if isinstance(category, set):
                    disjunctive_concepts.update({concept[0]: category})
                    for noun in category:
                        concepts.add(noun)
                else:
                    concepts.add(concept[0])
            else:
                concepts.add(concept[0])

    concepts = list(concepts)
    concepts.sort()
    concepts.remove('thing')

    properties = []
    for property in dictionary.popular:
        if property[2] == "a" and property[0] not in dictionary.ontology[1].keys():
            properties.append(property[0])
        elif property[2] == 'a' and property[0] in dictionary.ontology[1].keys():
            concepts.append(property[0])

    list1 = ["s", "o", "b", "c", "d"]
    for relation in dictionary.popular:
        if relation[2] == 'r':
            for i in range(relation[3]):
                properties.append(relation[1] + list1[i])

    return concepts, properties, disjunctive_concepts

def axiom_of_entity(lemmata2):
    if word2 in dictionary.biconditional_words:
        lemmata2.update({"thing" + "." + word2: "necessary"})
    else:
        lemmata2.update({"thing" + "." + word2: "possible"})


def get_word_kind(word):
    if hprop(word):
        return "h"
    elif isrelat(word):
        return 'r'
    else:
        return dictionary.kind.get(word)



def temp_test(lemmata2):
    global okind1, okind2, word1, word2
    total = dictionary.groups.keys()
    combos = combinations(total, 2)

    for combo in combos:
        word1 = combo[0]
        word2 = combo[1]
        # if "male" in combo and "female" in combo:
        category1 = dictionary.groups.get(word1)
        category2 = dictionary.groups.get(word2)
        okind1 = get_word_kind(word1)
        okind2 = get_word_kind(word2)

        if okind1 == 'i' and okind2 == 'i':
            test_modality2(category1, category2, combo, lemmata2, "impossible")
        elif word1 == 'thing':
            axiom_of_entity(lemmata2)

        elif isinstance(category1, set) and isinstance(category2, set):
            two_sets(category1, category2, combo, lemmata2)

        elif isinstance(category1, set):
            set_modality(category1, category2, combo, lemmata2)

        elif isinstance(category2, set):
            set_modality(category2, category1, combo, lemmata2)

        else:
            test_modality2(category1, category2, combo, lemmata2)

    return



def set_modality(set1, category2, combo, lemmata2):
    opair1 = build_lemma(combo[0], combo[1])
    opair2 = build_lemma(combo[1], combo[0])

    last_tv = ''
    last_tv2 = ''

    for category1 in set1:
        pair1 = build_lemma(category1, category2)
        pair2 = build_lemma(category2, category1)
        last_tv, last_tv2 = set_modality3(last_tv, last_tv2, pair1, pair2)


    lemmata2.update({opair1: last_tv})
    lemmata2.update({opair2: last_tv2})


def set_modality3(last_tv, last_tv2, pair1, pair2):
    modality = dictionary.ontology[2].get(pair1)
    modality = degrade_necessity(modality)
    last_tv = set_modality2(modality, last_tv)
    modality2 = dictionary.ontology[2].get(pair2)
    modality2 = degrade_necessity(modality2)
    last_tv2 = set_modality2(modality2, last_tv2)
    return last_tv, last_tv2


def set_modality2(modality, last_tv):
    if modality == 'possible':
        last_tv = "possible"
    elif modality == 'necessary' and last_tv not in ['impossible', 'possible']:
        last_tv = 'necessary'
    elif modality == 'impossible' and last_tv in ["", "impossible"]:
        last_tv = "impossible"
    else:
        last_tv = 'possible'

    return last_tv


def degrade_necessity(modality):
    if modality in ['impossible', 'possible']: return modality
    if okind2 in ['p', "i"]:
        return 'possible'
    else:
        return modality

def test_modality2(category1, category2, combo, lemmata2, modality = ""):
    opair1 = build_lemma(combo[0], combo[1])
    pair1 = build_lemma(category1, category2)
    opair2 = build_lemma(combo[1], combo[0])
    pair2 = build_lemma(category2, category1)


    if modality != "":
        lemmata2.update({opair1: modality})
        lemmata2.update({opair2: modality})
    else:
        modality = dictionary.ontology[2].get(opair1)
        modality2 = dictionary.ontology[2].get(opair2)
        if modality != None:
            lemmata2.update({opair1: modality})
            lemmata2.update({opair2: modality2})

        elif category2 == category1:
            modality = degrade_necessity("necessary")
            lemmata2.update({opair1: modality})
            lemmata2.update({opair2: modality})
        else:
            modality = dictionary.ontology[2].get(pair1)
            modality2 = dictionary.ontology[2].get(pair2)
            modality = degrade_necessity(modality)
            modality2 = degrade_necessity(modality2)
            lemmata2.update({opair1: modality})
            lemmata2.update({opair2: modality2})





def two_sets(category1, category2, combo, lemmata2):
    opair1 = build_lemma(combo[0], combo[1])
    opair2 = build_lemma(combo[1], combo[0])
    last_tv = ""
    last_tv2 = ""
    cart_product = [i for i in itertools.product(*[category1, category2])]
    for tpl in cart_product:
        pair1 = build_lemma(tpl[0], tpl[1])
        pair2 = build_lemma(tpl[1], tpl[0])
        last_tv, last_tv2 = set_modality3(last_tv, last_tv2, pair1, pair2)

    lemmata2.update({opair1: last_tv})
    lemmata2.update({opair2: last_tv2})


def order_list(word_list):
    j = 0
    while j < len(word_list):
        p = 0
        recent_words = []
        word = word_list[j]
        recent_words.append("start")
        recent_words.append(word)

        if word == 'Henergy':
            bb = 8
        if "," not in word:
            pos = dictionary.pos.get(word)
            if pos == None:
                print (f"you mispelled {word} during ordering")
                raise Exception
            pos = 'y' if word == 'there' else pos
            if pos == 'nh':
                dictionary.level.update({word: 0})
            elif (word in dictionary.categorized_sent.keys()
                  and pos[0] in ['n', 'r', 'a'] and pos[1] != "h") or "," in word:

                level = dictionary.level.get(word)
                if level == None:
                    level, p = order_list2(word, 0, p, recent_words, word_list)
                dictionary.level.update({word: level})

            elif pos[0] in ['n', 'r', 'a']:
                dictionary.level.update({word: 0})

        j += 1

    return sorted(dictionary.level.items(), key=operator.itemgetter(1))


def get_embed_rd(word):
    e, idx, parent = word.split(",")
    e = int(e)
    parent_info = dictionary.categorized_sent.get(parent)
    return [parent_info[e].embeds.get(idx)], parent


def order_list2(word, level, p, recent_words, word_list):
    p += 1
    if word in ['sentient being']:
        bb = 8
    if "," not in word:
        reduced_def = get_word_info(dictionary, word, "")
    else:
        reduced_def, _ = get_embed_rd(word)

    for e, cls in enumerate(reduced_def):
        if word not in dictionary.disjunctive.keys():
            for k, v in cls.embeds.items():
                word_list.append(str(e) + "," + k + "," + word)

    for cls in reduced_def:
        for num in cls.def_stats.flat_con_index:
            level, p = order_list3(cls, num, level, word_list, p, recent_words, word)

    return level, p


def order_list3(cls, num, level, word_list, p, recent_words, oword):
    p += 1
    if p > 300:
        print ('you have defined a word circulalarly here is a list of your recent definitions')
        for word in recent_words:
            print (word)
        raise Exception

    new_word = cls.sentences[num][58]
    recent_words.append(new_word)
    constituents.setdefault(oword, set()).add(new_word)
    if new_word in ["Hbody" + uc]:
        bb = 8

    if dictionary.pos.get(new_word) == 'nh':
        return level, p

    res = dictionary.level.get(new_word)
    if res == None:
        if new_word not in dictionary.categorized_sent.keys():
            word_list.append(new_word)
            dictionary.level.update({new_word: 0})
            if level == 0: level = 1
        else:
            res, p = order_list2(new_word, 0, p, recent_words, word_list)
            dictionary.level.update({new_word: res})
            if res >= level: level = res + 1
    else:
        if res >= level: level = res + 1

    return level, p


def convert(str1):
    list3 = str1.split("|")
    return list3[0] + get_key(superscript_dict, list3[1])


def print_some_lemmas(arg1, arg2):
    pkl_file = open('lemmata.pkl', 'rb')
    lemmata = pickle.load(pkl_file)
    pkl_file.close()
    # arg1 = 'po'
    # arg2 = 'man'

    if arg1 == "":
        lemma = input("input lemma: ")
        if "|" in lemma:
            list2 = lemma.split(".")
            if "|" in list2[0]:
                str1 = convert(list2[0])
            else:
                str1 = list2[0]
            if "|" in list2[1]:
                str2 = convert(list2[1])
            else:
                str2 = list2[1]
            lemma = str1 + "." + str2

        print (lemmata.get(lemma))

    elif arg1 in ['po', 'po1']:
        list1 = []
        if arg1 == 'po':
            word = arg2
        user = ""
        pkl_file = open(user + 'z_dict_words.pkl', 'rb')
        dictionary = pickle.load(pkl_file)
        pkl_file.close()
        popular = [x[1] if x[2] == 'r' else x[0] for x in dictionary.popular]
        b = 0
        # c = random.sample(lemmata.items(), 10)

        for lemma, value in lemmata.items():
            assert value in ['necessary', 'possible', 'impossible']
            b += 1
            if b % 2000 == 0:
                print (b)
            word1 = lemma[:lemma.find(".")]
            word2 = lemma[lemma.find(".") + 1:]
            if word1 == word:
                if word2 in popular:
                    list1.append([lemma, value])
            elif word2 == word:
                if word1 in popular:
                    list1.append([lemma, value])

        list1 = sorted(list1, key=operator.itemgetter(1))
        for lst in list1:
            print (lst[0] + " " + lst[1])




    else:
        for lemma, value in lemmata.items():
            if arg1 in lemma:
                print (lemma + " " + value)



