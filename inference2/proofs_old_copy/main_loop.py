import time
import sys
from openpyxl import load_workbook

#
# from natural_language import step_one
# from general_functions import parameters
# from classes import ErrorWithCode
# from settings import *
#
# try:
from natural_language import step_one
from general_functions import parameters
from classes import ErrorWithCode
from settings import *
# except:
#     from .natural_language import step_one
#     from .general_functions import parameters
#     from .classes import ErrorWithCode
#     from .settings import *


def calculate_time_statistics(num_proved, total_time):
    total_time = time.time() - total_time
    print("")
    print("average " + str("{0:.4f}".format(total_time / num_proved)))
    print("total " + str("{0:.3f}".format(total_time)))
    print("")
    print(num_proved)


def determine_words_used(words_used):
    wb5 = load_workbook('/Users/kylefoley/Desktop/inference_engine/dictionary5.xlsx')
    w5 = wb5.worksheets[0]
    for word in words_used:
        j = dictionary.words_to_row.get(word, 28)
        if j == 28:
            print(word)
        elif j == None:
            print(word)
        else:
            try:
                w5.cell(row=j, column=2).value = 1
            except:
                pass
    wb5.save('/Users/kylefoley/Desktop/inference_engine/dictionary5.xlsx')
    words = open('words_used.pkl', 'wb')
    pickle.dump(words_used, words)
    words.close()


def print_on_error(order, dictionary, test_sent, print_type, lemmata, user):
    global words_used
    j = -1
    num_proved = 0
    while j < len(order) - 1:
        j += 1
        k = order[j]

        if test_sent[k][0] != 'pass':
            num_proved += 1
            st1 = time.time()

            if k == 182:
                bb = 7
            # print (k)
            try:

                _ = step_one(dictionary, user, test_sent, lemmata, k)
                consistent, total_sent, twords_used = _
                if total_sent != 'skip':

                    test_sent[k] = json.loads(json.dumps(total_sent))
                    words_used = words_used | twords_used

                else:
                    if k in order:
                        order.remove(k)
                        j -= 1

                reaction(consistent, k, print_type, st1)

            except:
                print("bug")

                if k in order:
                    order.remove(k)
                    j -= 1


        elif print_type[0] in ["0", "4"] and k % 50 == 0:
            print(k)

    return num_proved


def reaction(consistent, k, print_type, st1):
    if print_type[0] in ['1', "2"]:
        if consistent:
            print('RIGHT')
        else:
            print('WRONG')

    elif print_type[0] != "4":
        if not consistent:
            if print_type[0] == "3":
                print(str(k) + " - " + str("{0:.3f}".format(time.time() - st1) + " False"))
            elif print_type[0] == "0":
                print(str(k) + " - False")
                sys.exit()


        elif print_type[0] == "3":
            print(str(k) + " - " + str("{0:.3f}".format(time.time() - st1)))
        elif print_type[0] == "1":
            print(str(k) + " - True")
    elif print_type[0] == "4" and not consistent:
        print(str(k) + " - False")


def stop_if_error(order, dictionary, test_sent, print_type, lemmata, user):
    global words_used
    j = -1
    num_proved = 0

    while j < len(order) - 1:
        j += 1
        k = order[j]

        if test_sent[k][0] != 'pass':
            num_proved += 1
            st1 = time.time()

            if k == 182:
                bb = 7
            # print (k)
            _ = step_one(dictionary, user, test_sent, lemmata, k)
            consistent, total_sent, twords_used = _

            if total_sent != 'skip':
                test_sent[k] = json.loads(json.dumps(total_sent))
                words_used = words_used | twords_used

            reaction(consistent, k, print_type, st1)


        elif print_type[0] in ["0", "4"] and k % 50 == 0:
            print(k)

    return num_proved


def get_result(one_sent, user="", print_type="40", order=[0], get_words_used=0):
    global words_used
    # import os
    # from django.conf import settings

    total_time = time.time()

    if one_sent == 'a':
        proof_type, print_type, get_words_used, order = parameters()
        pkl_file = open(user + 'zz_claims.pkl', 'rb')
        test_sent = pickle.load(pkl_file)
        pkl_file.close()
    elif one_sent != "":
        test_sent = one_sent
    else:
        # file = os.path.join(settings.BASE_DIR, 'inference2/proofs_old/' + user + 'zz_claims.pkl')
        file = user + 'zz_claims.pkl'
        pkl_file = open(file, 'rb')
        test_sent = pickle.load(pkl_file)
        pkl_file.close()

    # pkl_file = open(user + 'z_dict_words.pkl', 'rb')
    # dictionary =pickle.load(pkl_file)
    # pkl_file.close()
    # import pdb;pdb.set_trace()

    # file = os.path.join(settings.BASE_DIR, 'inference2/proofs_old_copy/' + user + 'data.json')
    file = 'data.json'
    # pkl_file = open(file, 'rb')
    # lemmata = pickle.load(pkl_file)
    # pkl_file.close()

    with open(file, 'r') as fp:
        lemmata = json.load(fp)
    # pkl_file = open('data.json', 'r')
    # lemmata = json.loads(pkl_file)
    # lemmata = pickle.load(pkl_file)

    # import json
    # with open('data.json', 'w') as fp:
    #     json.dump(lemmata, fp)

    # pkl_file.close()

    # file = os.path.join(settings.BASE_DIR, 'inference2/proofs_old/' + user + 'z_dict_words.pkl')
    file = user + 'z_dict_words.pkl'
    pkl_file = open(file, 'rb')
    # from classes import get_dictionary
    dictionary = pickle.load(pkl_file)
    pkl_file.close()

    words_used = set()
    print_type = "31"

    if print_type[1] == "1":
        num_proved = print_on_error(order, dictionary, test_sent, print_type, lemmata, user)
    else:
        num_proved = stop_if_error(order, dictionary, test_sent, print_type, lemmata, user)

    if print_type[1] in ["0", "2", "3", "4"]:
        calculate_time_statistics(num_proved, total_time)

    if get_words_used == 1:
        determine_words_used(words_used)

    with open("test3.txt", "wb") as fp:  # Pickling
        pickle.dump(test_sent, fp)

    # with open('test_file.txt', 'w') as f:
    #     f.write('score = %s' % test_sent)
