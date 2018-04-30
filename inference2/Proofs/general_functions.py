from openpyxl import load_workbook
from itertools import chain, combinations


try:
    from settings import *
except:
    from .settings import *

# from settings import *



def powerset(list1):
    return chain.from_iterable(combinations(list1, r) for r in range(len(list1) + 1))


def get_last_row(worksheet, col_num):
    row_num = 15000
    str3 = worksheet.cell(row=row_num, column=col_num).value
    while str3 == None:
        str3 = worksheet.cell(row=row_num, column=col_num).value
        row_num -= 1

    return row_num + 1


################ group: naming sentences

def build_sent_pos(sent):
    list1 = list(filter(lambda x: sent[x] != "~" and sent[x] != "", sent[54]))
    return nbuild_sent2(sent, sent, list1)


def name_and_build(output, list1):
    if list1[46] == 'do not rename': return
    list1[1] = build_sent_pos(list1)
    list1[2] = name_sent(list1[1], output.prop_name)
    output.oprop_name[list1[2]] = list1[1]
    list1[0] = list1[3] + list1[1]


def direct_equivalence(output, ant_sent, ant_sentp, sent, rule, anc1=""):
    name_and_build(output, sent)
    sent1 = build_connection(ant_sent, iff, sent[0])
    sent1p = build_connection(ant_sentp, iff, sent[3] + sent[2])
    add_to_tsent(output, sent1, sent1p, "", rule, anc1)
    sent[44] = output.tindex

def name_build_pre_cat(output, list1, ant_sent, ant_sentp, rule, anc1):
    sent = "(" + " ".join(list1) + ")"
    sent_ns = "(" + "".join(list1) + ")"
    sentp = name_sent(sent_ns, output.prop_name)
    output.oprop_name[sentp] = sent
    sent1 = build_connection(ant_sent, iff, sent)
    sent1p = build_connection(ant_sentp, iff, sentp)
    add_to_tsent(output, sent1, sent1p, "", rule, anc1)
    output.inferences.append([sent, sentp, "", "EF", anc1,
                              output.tindex, "not standard"])


def svo_sent(output, subject, relation, object, tvalue=""):
    sent1 = [None] * 60
    sent1[3] = tvalue
    sent1[10], sent1[13], sent1[14] = subject, relation, object
    sent1[54] = [10, 13, 14]
    sent1[42] = [10, 14]
    sent1[45] = []
    name_and_build(output, sent1)
    sent1[46] = "do not rename"

    return sent1


def svop_sent(output, prop, subject, relation, object, tvalue=""):
    sent1 = [None] * 60
    sent1[3] = tvalue
    _ = prop, mini_e, subject, relation, object
    sent1[8], sent1[9], sent1[10], sent1[13], sent1[14] = _
    sent1[54] = [8, 9, 10, 13, 14]
    sent1[42] = [8, 10, 14]
    sent1[45] = []
    name_and_build(output, sent1)
    sent1[46] = "do not rename"

    return sent1


#####################group: index functions

def find_counterpart_inlist(str1, list1, i, j):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the matching second element

    for d in range(len(list1)):
        if str1 == list1[d][i]:
            str2 = list1[d][j]
            return str2
    return None


def find_2posinlist(str1, str2, list1, p, q):
    for i, lst in enumerate(list1):
        if str1 == lst[p] and str2 == lst[q]:
            return i

    return -1


def findposinmd(str1, list1, p, start=0, no_error=False):
    # this determines the position of an element in a multidimensional list

    for i in range(start, len(list1)):
        try:
            if list1[i][p] == str1:
                return i
        except:
            pass
    if no_error: raise Exception

    return -1


def findposinmdlistint(i, list1, p):
    for j in range(len(list1)):
        if list1[j][p] == i:
            return j
    else:
        return -1


def find_sibling_int(i, list1, p, q):
    for j in range(len(list1)):
        if list1[j][p] == i:
            return list1[j][q]
    else:
        return -1


############### the following functions are related to adding to the total sent
###############  and finding a contradiction


def check_consistency(output):
    new_sent_abbr = output.total_sent[-1][2]
    tvalue = output.total_sent[-1][3]
    for i in range(len(output.total_sent) - 2, -1, -1):
        if output.total_sent[i][2] == new_sent_abbr and output.total_sent[i][3] != tvalue:
            build_contradiction(output, i)
            return False

    for lst in output.negated_conjunction:
        for sent in lst:
            if sent[0] == new_sent_abbr and sent[1] == tvalue:
                del sent[0]
                del sent[1]
                lst[3].append(output.total_sent[-1][0])
                if lst == []:
                    raise Exception
                    build_contradictory_conjunction(output, lst[3])
                    return False

    if output.disj_elim != []:
        if len(new_sent_abbr) < 3:
            return neg_excl_disj_elim(new_sent_abbr, tvalue, output, len(output.total_sent) - 1, -1, output.tindex)

    return True


def neg_excl_disj_elim(det_abb, det_abb_tv, output, tot_sent_num, ex0, anc1):
    while ex0 < len(output.disj_elim) - 1:
        ex0 += 1
        excl_disj = output.disj_elim[ex0]
        ex1 = -1
        while ex1 < len(excl_disj) - 1:
            ex1 += 1
            ex2 = -1
            while ex2 < len(excl_disj[ex1][2]) - 1:
                ex2 += 1
                dsent = excl_disj[ex1][2][ex2]
                if det_abb == dsent[1] and det_abb_tv != dsent[2]:
                    consistent = neg_excl_disj_elim2(ex0, ex1, tot_sent_num, output, anc1)

                    ex1 -= 1
                    return consistent
    return True


def neg_excl_disj_elim2(ex0, delex, tot_sent_num, output, anc3):
    excl_disj = output.disj_elim[ex0]
    anc1 = output.total_sent[tot_sent_num][0]
    connector = " " + xorr + " "
    if len(excl_disj) > 2:
        list1 = []
        list1a = []
        anc2 = excl_disj[0][3]
        for d, lst in enumerate(excl_disj):
            if d != delex:
                list1.append(lst[0])
                list1a.append(lst[1])
        nat_sent = connector.join(list1)
        abb_sent = connector.join(list1a)
        del output.disj_elim[ex0][delex]
        add_to_tsent(output, nat_sent, abb_sent, "", xorr + "E", anc1, anc2)
    else:
        b = 1 if delex == 0 else 0
        anc1 = tot_sent_num
        anc2 = excl_disj[0][3]
        if len(excl_disj[b][2]) == 1:
            add_to_tsent(output, excl_disj[b][2][0][0], excl_disj[b][2][0][1], excl_disj[b][2][0][2],
                         xorr + "E", anc1, anc2)
            consistent = check_consistency(output)
            if not consistent: return
        else:
            add_to_tsent(output, excl_disj[b][0], excl_disj[b][1], "", xorr + "E", anc1, anc2)
            qn = anc3
            for lst in excl_disj[b][2]:
                add_to_tsent(output, lst[0], lst[1], lst[2], "&E", qn)
                consistent = check_consistency(output)
                if not consistent:
                    return False
        del output.disj_elim[ex0]

    return True


def build_contradictory_conjunction(output, list1):
    list3 = tuple(str(output.total_sent[k][0]) for k in list1)
    anc1 = ",".join(list3)
    list2 = [output.total_sent[j][3] + output.total_sent[j][1] for j in list1]
    list2p = [output.total_sent[j][3] + output.total_sent[j][2] for j in list1]
    conjunction = "(" + " & ".join(list2) + ")"
    conjunctionp = "(" + " & ".join(list2p) + ")"
    add_to_tsent(output, conjunction, conjunctionp, "", "&I", anc1)
    build_contradiction(output, -2)


def build_contradiction(output, i):
    str1 = output.total_sent[-1][1] + " & ~" + output.total_sent[i][1]
    str2 = output.total_sent[-1][2] + " & ~" + output.total_sent[i][2]
    add_to_tsent(output, str1, str2, "", "&I", output.total_sent[-1][0], output.total_sent[i][0])
    add_to_tsent(output, bottom, bottom, "", bottom + "I", 0)


def add_to_tsent(output, str1, str2="", tvalue="", rule="", anc1="", anc2="", anc3="", anc4=""):
    if anc1 == 0: anc1 = output.tindex
    if anc2 == 0: anc2 = output.tindex
    list2 = [""] * 9
    output.tindex += 1
    list2[0] = output.tindex
    list2[1] = str1
    list2[2] = str2
    list2[3] = tvalue
    list2[4] = rule
    list2[5] = anc1
    list2[6] = anc2
    list2[7] = anc3
    list2[8] = anc4
    output.total_sent.append(list2)
    if rule == xorr + "E":
        bb = 8


######################group: miscellanious functions



def rename_sentences(old_word, new_word, output):
    for e, sent in enumerate(output.all_sent):
        if old_word in sent[0]:
            output.all_sent[e][0] = output.all_sent[e][0].replace(old_word, new_word)
            for j, slot in enumerate(sent):
                if slot == old_word:
                    sent[j] = new_word
                    break

    for var, prop in output.oprop_name.items():
        if old_word in prop:
            prop = prop.replace(old_word, new_word)
            output.oprop_name[var] = prop
    for prop, var in output.prop_name.items():
        if old_word in prop:
            new_prop = prop
            del output.prop_name[prop]
            new_prop = new_prop.replace(old_word, new_word)
            output.prop_name.update({new_prop: var})
            break
    for sent in output.total_sent:
        if old_word in sent[1]:
            sent[1] = sent[1].replace(old_word, new_word)

    output.words_used.remove(old_word)
    output.words_used.add(new_word)
    return



def determine_if_compound_word(list1, dictionary):
    i = -1
    while i < len(list1) - 1:
        i += 1
        if i + 3 < len(list1):
            third_word_later = list1[i + 3]
        else:
            third_word_later = "0"

        if i + 2 < len(list1):
            after_next_word = list1[i + 2]
        else:
            after_next_word = "0"

        if i + 1 < len(list1):
            next_word = list1[i + 1]
        else:
            next_word = "0"

        quadruple_word = list1[i] + " " + next_word + " " + after_next_word + " " + third_word_later
        triple_word = list1[i] + " " + next_word + " " + after_next_word
        double_word = list1[i] + " " + next_word

        if quadruple_word in dictionary.quadruples:

            list1[i] = quadruple_word
            del list1[i + 1]
            del list1[i + 1]
            del list1[i + 1]

        elif triple_word in dictionary.triples:

            list1[i] = triple_word
            del list1[i + 1]
            del list1[i + 1]

        elif double_word in dictionary.doubles:

            list1[i] = double_word
            del list1[i + 1]

    return list1



def eliminate_blanks(list1):
    i = 0
    while i < len(list1):
        if not_blank(list1[i]):
            i += 1
        else:
            del list1[i]

    return list1


def get_part_of_speech(word, dictionary, abbreviations, output=[]):
    neg_found = False
    if word[0] == neg:
        word = word[1:]
        neg_found = True

    while True:
        pos = dictionary.pos.get(word)
        if isvariable(word):
            reference = abbreviations.get(word)
            pos = dictionary.pos.get(reference)
            pos = 'ny' if pos == None or pos[0] != 'a' else 'ay'
            break
        elif word[-2:] == "'s":
            pos = 's' if dictionary.kind.get(word[:-2]) == 'i' else 'o'
            break
        elif pos == None:
            if output == []:
                print('you mispelled ' + word)
                raise Exception('you mispelled ' + word)
            else:
                old_word = word
                print('you mispelled ' + old_word)
                new_word = input("new word: ")
                while True:
                    if new_word in dictionary.pos.keys():
                        pos = dictionary.pos.get(new_word)
                        break
                    else:
                        print(f'{new_word} is also a mispelling')
                        new_word = input("new word: ")

                rename_sentences(old_word, new_word, output)
                word = new_word
                break
        else:
            break

    if neg_found: word = neg + word

    return pos, word



def mainconn(str1):
    if one_sentence(str1):
        return ["", 0]
    num = 0
    if str1.startswith("((c I b"):
        bb = 8

    next_conn = False
    candidate_conn = False
    potential_error = False
    possibility = []
    for idx, letter in enumerate(str1):
        if letter == "(":
            num += 1
            if candidate_conn:
                candidate_conn = False

        elif letter == ")":
            num -= 1
            if candidate_conn:
                candidate_conn = False
        if next_conn and letter in all_connectives:
            return letter, idx

        if num == 0 and idx > 0:
            next_conn = True

        if candidate_conn and letter in all_connectives:
            if possibility == []:
                possibility = [letter, idx]
            elif letter in ['&', xorr, idisj] and possibility[0] == letter:
                pass
            else:
                potential_error = True

        elif num == 1 and idx + 1 < len(str1) - 1 and not candidate_conn:
            candidate_conn = True

    if potential_error or possibility == []:
        raise Exception

    return possibility[0], possibility[1]


def insert_new_word(new_word_pos, old_word_pos):
    list1 = []
    for idx in old_word_pos:
        if idx > 164:
            list1.append(list1[-1] + .01)
        else:
            list1.append(slot_order.index(idx))
    new_word_index = slot_order.index(new_word_pos)
    for i, num in enumerate(list1):
        if new_word_index < num:
            old_word_pos.insert(i, new_word_pos)
            return old_word_pos


def remove_duplicates(list1, i):
    list2 = []
    j = -1
    while j < len(list1) - 1:
        j += 1
        if list1[j][i] in list2:
            del list1[j]
            j -= 1
        else:
            list2.append(list1[j][i])

    return list1


def isvariable(str3, kind=""):
    bool2 = True
    if str3 == None or str3 == "":
        return False

    if str3 == 'a':
        return False
    elif str3 == 'i' and kind == "":
        return False
    elif str3 == "i":
        return True

    if str3 != "":
        str3 = str3.replace(l1, "")
        str3 = str3.replace(l2, "")
        str3 = str3.replace(l3, "")
        str3 = str3.replace(neg, "")
        if len(str3) == 1 and str3.islower():
            bool2 = True
        else:
            bool2 = False

    return bool2


def is_standard(sent):
    try:
        if sent[13][0].islower():
            return "not standard"
    except:
        pass

    for num in sent[54]:
        if num not in standard_nouns + relational_positions + negative_positions:
            return "not standard"
    if sent[20] in ["AS", 'OFX'] or sent[10] == 'there':
        return "not standard"

    return "standard"


def get_variables():
    variables = [chr(122 - t) for t in range(25)]
    variables.remove("i")
    variables.remove("l")
    variables3 = [chr(122 - t) + l1 for t in range(26)]
    variables3.remove("l" + l1)
    variables4 = [chr(122 - t) + l2 for t in range(26)]
    variables4.remove("l" + l2)
    variables5 = [chr(122 - t) + l3 for t in range(26)]
    variables5.remove("l" + l3)
    variables2 = variables + variables3 + variables4 + variables5

    return variables2


def get_prop_var():
    prop_var = [chr(122 - t) for t in range(26)]
    prop_var2 = [chr(122 - t) + "\u2081" for t in range(26)]
    prop_var3 = [chr(122 - t) + "\u2082" for t in range(26)]
    prop_var5 = [chr(122 - t) + "\u2083" for t in range(26)]
    prop_var6 = [chr(122 - t) + "\u2084" for t in range(26)]
    prop_var7 = [chr(122 - t) + "\u2085" for t in range(26)]
    prop_var8 = [chr(122 - t) + "\u2086" for t in range(26)]
    prop_var9 = [chr(122 - t) + "\u2087" for t in range(26)]
    prop_var2 = prop_var9 + prop_var8 + prop_var7 + prop_var6 + prop_var5 + prop_var3 + prop_var2 + prop_var

    return prop_var2


def get_word_info(dictionary, word, user=""):
    if word not in dictionary.categorized_sent.keys():
        return []
    with open(user + "json_dict/" + word + ".json", "r") as fp:
        lst_sent = json.load(fp)
    item1 = dictionary.categorized_sent.get(word)
    for x, y in zip(lst_sent, item1):
        y.sentences = x
        for cls in item1:
            try:
                for k, v in cls.embeds.items():
                    v.sentences = x
            except:
                pass

    return item1


def make_groups_definite(sentences, output):
    for sentence in sentences:
        if sentence[13] == 'W':
            try:
                if sentence[7][0] in ["b", "a"] or \
                        sentence[7][:2] in ['cb', 'ca']:
                    output.constants.add(sentence[10])
            except:
                pass


def add_to_gsent(item1, output, proof_kind="", asent_idx=[]):
    for e, cls in enumerate(item1):
        def_stats = cls.def_stats
        if proof_kind == "":
            sentences = cls.sentences
        else:
            sentences = output.all_sent
        def_stats.already_instantiated = False
        kind = def_stats.connection_type
        definiendum = def_stats.def_word_num
        if kind != "x":
            first_ant = def_stats.ant_index[0]
            sent_constant = sentences[first_ant][58]
            if sent_constant == 'thing' and len(def_stats.ant_index) > 1:
                first_ant = def_stats.ant_index[1]
                idx = def_stats.ant_index[0]
                del def_stats.ant_index[0]
                def_stats.ant_index.append(idx)
                sent_constant = sentences[first_ant][58]
            output.gsent.append([sent_constant, 0, definiendum, asent_idx])
            if kind == 'e' and cls.disjuncts == []:
                first_con = def_stats.con_index[0]
                if isinstance(first_con, int):
                    sent_constant = sentences[first_con][58]
                else:
                    sent_constant = cls.def_stats.con_comp_const[0]
                output.gsent.append([sent_constant, 1, definiendum, asent_idx])

            elif kind == 'e':
                add_disjunct_to_gsent(cls, sentences, definiendum, output)

        else:
            add_disjunct_to_gsent(cls, sentences, definiendum, output)

        make_groups_definite(sentences, output)

    return


def add_disjunct_to_gsent(cls, sentences, definiendum, output):
    for j, disjunction in enumerate(cls.disjuncts):
        first_con = disjunction.index1[0]
        if isinstance(first_con, int):
            sent_constant = sentences[first_con][58]
            output.gsent.append([sent_constant, j + 2, definiendum, []])
        else:
            raise Exception


def rebuild_hconstant(sentences, abbreviations, word=""):
    for sentence in sentences:
        if sentence[13] == "H":
            var = sentence[14]
            for sentence2 in sentences:
                if sentence2[13] == "I" and sentence2[10] == var:
                    str1 = abbreviations.get(sentence2[14])
                    if str1 != None:
                        sentence[58] = "H" + str1
                        break


def determine_constants(dict1, sentence):
    if sentence[13] == "=" and sentence[14] in dict1.values():
        return sentence[14]
    else:
        str1 = is_concept(sentence[13], sentence[14], dict1)
        if str1 == None:
            return sentence[13]
        return str1


def universal_negations(cls, output):
    definiendum = ""
    done = "not done"
    if cls.def_stats.def_word == '':
        for e, num in enumerate(cls.def_stats.ant_index):
            if isinstance(num, int):
                sentence = cls.sentences[num]
                sent_constant = sentence[58]
                if sent_constant == 'thing':
                    done = "done"
                    definiendum = 'thing'
                    cls.def_stats.def_word = 'thing'
                    disjuncts = cls.disjuncts
                    new_ant_index = []
                    for e, num in enumerate(cls.def_stats.ant_index):
                        sentence = cls.sentences[num]
                        sent_constant = sentence[58]
                        if sent_constant != 'thing':
                            new_ant_index.append(num)

                    for e, num in enumerate(cls.def_stats.con_index):
                        sentence = cls.sentences[num]
                        sent_constant = sentence[58]
                        new_con_index = []
                        new_con_index.append(num)
                        new_con_index += new_ant_index
                        disjuncts.append(json.loads(json.dumps(new_con_index)))

                        if "~" not in sent_constant:
                            tvalue = "~ "
                        else:
                            sent_constant = sent_constant.replace("~ ", "")
                            tvalue = ""
                        sentence[58] = tvalue + sent_constant
                        output.gsent.append([tvalue + sent_constant, 2 + e, "thing0", []])

    return definiendum, done


def get_key(dict1, val):
    for k, v in dict1.items():
        if v == val:
            return k
    else:
        return None


def change_constants(already_used_map, output, from_definitions):
    if from_definitions == None: return
    for k, v in from_definitions.items():
        new_var = get_key(output.abbreviations, v)
        if new_var != None:
            already_used_map.update({k: new_var})
        else:
            old_value = v
            if k not in output.variables:
                v = output.variables[0]
                del output.variables[0]
                already_used_map.update({k: v})
                output.abbreviations.update({v: old_value})
            else:
                output.variables.remove(k)
                already_used_map.update({k: k})
                output.abbreviations.update({k: old_value})
    return


def parameters():
    proof_type = 0
    print_type = 4
    get_words_used = 0
    start = 200
    stop = 201
    order = [x for x in range(start, stop)]

    return proof_type, print_type, get_words_used, order


def remove_extra_paren(sentence, embed=False):
    if embed:
        return sentence[1:-1]

    num = 0
    has_extra_paren = True
    for idx, letter in enumerate(sentence):
        if letter == "(":
            num += 1
        elif letter == ")":
            num -= 1
        if num == 0 and idx != 0 and idx + 1 != len(sentence):
            has_extra_paren = False

    if has_extra_paren:
        return sentence[1:-1]

    return sentence


def flatten_list(lst):
    list2 = []
    for x in lst:
        if isinstance(x, list):
            for y in x: list2.append(y)
        else:
            list2.append(x)

    return list2


def get_super(str1):
    if str1 == "a":
        return "\u1d43"
    elif str1 == "b":
        return "\u1d47"
    elif str1 == "c":
        return "\u1d9c"
    elif str1 == "d":
        return "\u1d48"
    elif str1 == "e":
        return "\u1d49"
    elif str1 == "f":
        return "\u1da0"
    elif str1 == "g":
        return "\u1d4d"
    elif str1 == "h":
        return "\u02b0"
    elif str1 == "i":
        return "\u2071"
    elif str1 == "j":
        return "\u02B2"
    elif str1 == "k":
        return "\u1d4f"
    elif str1 == "l":
        return "\u02E1"
    elif str1 == "m":
        return "\u1d50"
    elif str1 == "n":
        return "\u207f"
    elif str1 == "o":
        return "\u1d52"
    elif str1 == "p":
        return "\u1d56"
    elif str1 == "r":
        return "\u02b3"
    elif str1 == "s":
        return "\u02e2"
    elif str1 == "t":
        return "\u1d57"
    elif str1 == "u":
        return "\u1d58"
    elif str1 == "v":
        return "\u1d5b"
    elif str1 == "w":
        return "\u02b7"
    elif str1 == "y":
        return "\u02b8"


def tran_str(str1, has_sentence_connectives=False):
    if str1 == "":
        return str1
    if "|" in str1:
        for i in range(len(str1)):
            if str1[i:i + 1] == "|":
                str3 = str1[i + 1:i + 2]
                str4 = get_super(str3)
                str1 = str1[:i] + str4 + str1[i + 2:]

    if has_sentence_connectives:

        if "t^" in str1:
            str1 = str1.replace("t^", conditional)
        if "nt+" in str1:
            str1 = str1.replace("nt+", neg)
        if "x^" in str1:
            str1 = str1.replace("x^", iff)
        if "b^" in str1:
            str1 = str1.replace("b^", mini_e)
        if "c^" in str1:
            str1 = str1.replace("c^", mini_c)
        if "ed^" in str1:
            str1 = str1.replace("ed^", xorr)
        if "v+" in str1:
            str1 = str1.replace("v+", idisj)

    return str1



############################
######## group: print sent
##############


def closest_to_max_size(word_str, max_size, sent_type):
    if sent_type == 'sent id' and len(word_str) < max_size:
        for idx in range(max_size, 0, -1):
            if word_str[idx] == ")":
                return idx
    elif one_sentence(word_str[:max_size]):
        for idx in range(max_size, 0, -1):
            if word_str[idx] in " ":
                return idx
    elif word_str.startswith("RELEVANT"):
        return 24

    else:
        for idx in range(max_size, 30, -1):
            if word_str[idx] in all_connectives:
                return idx
        else:
            if word_str[30:max_size].count(")") < 2:
                for idx in range(max_size, 0, -1):
                    if word_str[idx] in " ":
                        return idx
            else:
                raise Exception


def space_sentences(num_str, word_str, largest_rule, space1, rule, name_sent, sent_type):
    if name_sent:
        print (word_str)
    else:
        third_column = largest_rule + 4
        max_size = 75 - third_column
        j = 0
        k = 0
        if " ," in word_str: word_str = word_str.replace(" ,", ",")
        five_spaces = " " * 5
        # the word_str always starts at position 5, there is always
        # 3 spaces between the rule and the word_str

        while 5 + third_column + len(word_str) > 75:
            k += 1
            if k > 10: raise Exception("printer caught in infinite loop")
            location = closest_to_max_size(word_str, max_size - 5, sent_type)
            remainder = word_str[location:]
            word_str = word_str[:location]

            if j == 0:
                space2 = 75 - (4 + len(word_str) + largest_rule)
                space2 = " " * space2
                print(num_str + space1 + word_str + space2 + rule)
            else:
                print(five_spaces + word_str)
            j += 1
            word_str = remainder
        else:
            if all(x in ["", " "] for x in word_str) and j > 0:
                pass
            elif j > 0:
                print(five_spaces + word_str)
            else:
                space2 = 75 - (4 + len(word_str) + largest_rule)
                space2 = " " * space2
                print(num_str + space1 + word_str + space2 + rule)

    return


def print_sent(test_sent, order, print_type):
    if print_type[0] not in ["1", "2"]: return
    if print_type[0] == "1":
        wb4 = load_workbook('/Users/kylefoley/Desktop/inference_engine/temp_proof.xlsx')
        w4 = wb4.worksheets[0]
    row_number = 1
    largest_rule = 1

    for i in order:
        name_sent_now = False
        if test_sent[i][0] != "pass":
            for j in range(len(test_sent[i])):
                rule = ""
                if test_sent[i][j][5] == "id":
                    pass
                elif test_sent[i][j][4] != "":
                    rule = test_sent[i][j][4] + " "
                    if test_sent[i][j][5] != "":
                        rule += str(test_sent[i][j][5])

                        if test_sent[i][j][6] != "":
                            rule += "," + str(test_sent[i][j][6])

                    rule_size = len(rule)
                    if rule_size > largest_rule: largest_rule = rule_size

                    test_sent[i][j][4] = rule

            for j in range(len(test_sent[i])):
                if test_sent[i][j][1].startswith('name sent'):
                    print ("")
                    if test_sent[i][j][1] == 'name sent start':
                        name_sent_now = True
                    elif test_sent[i][j][1] == 'name sent end':
                        name_sent_now = False

                elif print_type[0] == "2":

                    size_num = 5 - len(str(test_sent[i][j][0]))
                    space1 = " " * size_num
                    sent_type = ""
                    if test_sent[i][j][5] == 'id':
                        sent_type = 'sent id'
                    elif test_sent[i][j][5] == 'natural':
                        sent_type = 'natural'
                    word_str = test_sent[i][j][3] + test_sent[i][j][1]
                    if test_sent[i][j][1].startswith("(a" + mini_e):
                        bb = 8

                    space_sentences(str(test_sent[i][j][0]), word_str, largest_rule, space1, test_sent[i][j][4],
                                    name_sent_now, sent_type)


                elif print_type[0] == "1":
                    w4.cell(row=row_number, column=2).value = test_sent[i][j][0]
                    w4.cell(row=row_number, column=3).value = test_sent[i][j][3] + test_sent[i][j][1]
                    w4.cell(row=row_number, column=4).value = test_sent[i][j][4]
                    if test_sent[i][j][1] == bottom:
                        w4.cell(row=row_number, column=5).value = 1

                    if len(test_sent[i][j]) > 8:
                        if test_sent[i][j][8] == "*":
                            xls_cell = w4.cell(row=row_number, column=3)
                            xls_cell.font = xls_cell.font.copy(color='FFFF0000')

                row_number += 1

        row_number += 2
        if print_type[0] == '2':
            print ("")
            print ("")


    if print_type[0] == "1":
        wb4.save('/Users/kylefoley/Desktop/inference_engine/temp_proof.xlsx')

    return
