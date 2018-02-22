from settings import *
from general_functions import *
from classes import get_output
from search_for_instantiation import loop_through_gsent, try_instantiation
from use_lemmas import get_class
from collections import defaultdict
import json, operator
from openpyxl import load_workbook
from prepare_for_print import rearrange

pkl_file = open('words_used.pkl', 'rb')
words_used = pickle.load(pkl_file)
words_used = sorted(words_used)
pkl_file.close()


def modify_abbreviations(youtput, xoutput, xvar, abbrev_dict):
    xabbrev = xoutput.abbreviations
    for k, v in youtput.abbreviations.items():
        if v in xabbrev.values():
            new_var = get_key(xabbrev, v)
            abbrev_dict.update({k: new_var})
        elif k not in xvar:
            new_var = xvar[0]
            abbrev_dict.update({k: new_var})
            del xvar[0]


def modify_variables(youtput, xoutput):
    xvars = xoutput.variables
    yvars = youtput.variables
    xprop_name = xoutput.prop_name
    xoprop_name = xoutput.oprop_name

    tvars = list(set(xvars) | set(yvars))
    abbrev_dict = {}
    modify_abbreviations(youtput, xoutput, xvars, abbrev_dict)

    for sentence in youtput.all_sent:
        changed = False
        for pos in sentence[42]:
            yvar = sentence[pos]
            value = abbrev_dict.get(yvar)
            if value != None:
                changed = True
                sentence[pos] = value
            elif yvar in youtput.main_var:
                pass
            elif yvar not in xvars:
                changed = True
                new_var = tvars[0]
                sentence[pos] = new_var
                abbrev_dict.update({yvar: new_var})
                del tvars[0]
        if changed:
            name_and_build(xoutput, sentence)
        else:
            sentence[2] = name_sent(sentence[1], xprop_name)
            xoprop_name[sentence[2]] = sentence[1]

    return


def get_basic_sent(tword):
    with open("basic/" + tword + ".json", "r") as fp:
        return json.load(fp)


def reload_sentences(output, anc):
    for k, v in output.trans_def.items():
        list2 = []
        list1 = output.lemma_embed.get(k)
        v.def_stats.ant_index = list1[0]
        v.def_stats.con_index = list1[1]
        v.def_stats.tot_sent_idx = anc

        for e, idx in enumerate(list1[0] + list1[1]):
            list2.append(output.all_sent[idx])
        v.sentences = json.loads(json.dumps(list2))

    return


def quick_contradiction(xoutput, youtput):
    xlist = list(filter(lambda x: x[7] == 'c', xoutput.all_sent))
    ylist = list(filter(lambda x: x[7] == 'c', youtput.all_sent))
    xvar = set(map(lambda x: x[2], xlist))
    yvar = set(map(lambda x: x[2], ylist))
    for var in xvar & yvar:
        xtvalue = find_counterpart_inlist(var, xoutput.all_sent, 2, 3)
        ytvalue = find_counterpart_inlist(var, youtput.all_sent, 2, 3)
        if xtvalue != ytvalue:
            return False
    return True


def name_xsent(xoutput):
    for sentence in xoutput.all_sent:
        sentence[2] = name_sent(sentence[1], xoutput.prop_name)
        xoutput.oprop_name.update({sentence[2]: sentence[1]})


def adjust_index2(output, new_sent, list1):
    sentences = output.all_sent
    new_index = []
    for e, num in enumerate(list1):
        if isinstance(num, int):
            new_sent.append(sentences[num])
            new_index.append(len(new_sent) - 1)
        else:
            list2 = []
            for f, cnum in enumerate(num):
                new_sent.append(sentences[cnum])
                list2.append(len(new_sent) - 1)
            new_index.append(list2)

    return new_index


def adjust_index(output):
    for k, v in output.trans_def.items():
        new_sent = []
        new_index = adjust_index2(output, new_sent, v.def_stats.ant_index)
        v.def_stats.ant_index = new_index
        new_index = adjust_index2(output, new_sent, v.def_stats.con_index)
        v.def_stats.con_index = new_index
        v.sentences = json.loads(json.dumps(new_sent))


def do_not_instantiate(xoutput, youtput, len_asent):
    output3 = xoutput if youtput == [] else youtput
    for k, v in output3.trans_def.items():
        for e, sent in enumerate(xoutput.all_sent[len_asent:]):
            xoutput.disj_elim.append([e, v.def_stats.def_word_num])


def make_matrix2(user):
    global dictionary
    pkl_file = open(user + 'z_dict_words.pkl', 'rb')
    dictionary = pickle.load(pkl_file)
    pkl_file.close()
    xword = "whole"
    yword = "moment"
    youtput = dictionary.basic_output.get(yword)
    youtput.all_sent = get_basic_sent(yword)
    reload_sentences(youtput, 2)
    xoutput = dictionary.basic_output.get(xword)
    xoutput.all_sent = get_basic_sent(xword)
    xoutput.prop_var = get_prop_var()
    xoutput.prop_name = defaultdict(lambda: xoutput.prop_var.pop(), {})
    name_xsent(xoutput)
    reload_sentences(xoutput, 1)
    modify_variables(youtput, xoutput)
    consistent = quick_contradiction(xoutput, youtput)
    if not consistent:
        pass
    adjust_index(xoutput)
    adjust_index(youtput)
    do_not_instantiate(xoutput, [], 0)
    len_asent = len(xoutput.all_sent)
    xoutput.all_sent = xoutput.all_sent + youtput.all_sent
    do_not_instantiate(xoutput, youtput, len_asent)
    xoutput.trans_def = {**xoutput.trans_def, **youtput.trans_def}
    for k, v in xoutput.trans_def.items(): add_to_gsent([v], xoutput)
    if xoutput.gsent != []:
        fill_tsent(xoutput, xword, yword, len_asent)
        loop_through_gsent(xoutput, "lemmas2", dictionary)
        consistent = True if xoutput.total_sent[-1][1] == consist else False
        rearrange("last", xoutput, consistent, "lemmas2", xoutput.main_var)

    return


def fill_tsent(xoutput, xword, yword, len_asent):
    xoutput.tindex = 0
    basic_defintion = dictionary.basic_definitions.get(xword)
    add_to_tsent(xoutput, basic_defintion)
    basic_defintion = dictionary.basic_definitions.get(yword)
    add_to_tsent(xoutput, basic_defintion)
    for sent in xoutput.all_sent[:len_asent]:
        if sent[7] == 'c':
            add_to_tsent(xoutput, sent[1], "", sent[3], "&E", 1)
    for sent in xoutput.all_sent[len_asent:]:
        if sent[7] == 'c':
            add_to_tsent(xoutput, sent[1], "", sent[3], "&E", 2)
    for k, v in xoutput.trans_def.items():
        anc1 = v.def_stats.tot_sent_idx
        add_to_tsent(xoutput, v.def_stats.natural_sent, "", "", "&E", anc1)


def adjust_prop_name(output):
    nprop_name = {}
    for k, v in output.prop_name.items():
        s = k.split()
        s = "".join(s)
        nprop_name.update({s: v})

    output.prop_name = nprop_name


def save_json_sent(obj, location):
    with open(location + ".json", "w") as fp:
        json.dump(obj, fp)


def save_output(output):
    adjust_prop_name(output)
    save_json_sent(output.all_sent, "basic/" + word)
    output.all_sent = []
    output.lsent_dict = {}
    output.total_sent = []
    output.prop_var = []
    output.prop_name = {}
    output.oprop_name = {}
    output.lsent_list = []
    output.gsent = []
    output.substitutions = {}

    while True:
        for k, v in output.trans_def.items():
            v.def_stats.already_instantiated = False
            if "," not in k:
                del output.trans_def[k]
                break
        else:
            break

    dictionary.basic_output.update({word: output})
    return


def print_to_excel(proof_type):
    if proof_type != 5:
        wb5 = load_workbook('/Users/kylefoley/Desktop/inference_engine/dictionary5.xlsx')
        w5 = wb5.worksheets[4]
        row_number = 1
        for k, v in dictionary.basic_definitions.items():
            w5.cell(row=row_number, column=1).value = k
            w5.cell(row=row_number, column=2).value = v
            row_number += 1

        wb5.save('/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/dictionary5.xlsx')


def get_from_all_sent(num, output, sentences, list2):
    idx = findposinmd(sentences[num][0], output.all_sent, 0)
    if idx == -1:
        output.all_sent.append(sentences[num])
        list2.append(len(output.all_sent) - 1)
    else:
        list2.append(idx)


def adjust_sides(list1, output, sentences):
    list2 = []
    for e, num in enumerate(list1):
        if isinstance(num, int):
            get_from_all_sent(num, output, sentences, list2)
        else:
            list3 = []
            for cnum in num:
                get_from_all_sent(cnum, output, sentences, list3)
            list2.append(json.loads(json.dumps(list3)))

    return list2


def build_basic_definition(output, reduced, word):
    if reduced:
        list1 = [x[0] for x in output.all_sent]
        for k, defin in output.trans_def.items():
            if "," in k:
                sentences = defin.sentences
                conn_sent = defin.def_stats.tot_greek_sent
                nant = adjust_sides(defin.def_stats.ant_index, output, sentences)
                ncon = adjust_sides(defin.def_stats.con_index, output, sentences)

                for e, sentence in enumerate(sentences):
                    conn_sent = conn_sent.replace(sentence[5], sentence[0])

                output.lemma_embed.update({k: json.loads(json.dumps([nant, ncon]))})
                list1.append(conn_sent)

        basic_definition = " & ".join(list1)
        if output.abbreviations != {}:
            str1 = " & ".join([build_connection("(" + k, "=", v + ")") for k, v in output.abbreviations.items()])
            basic_definition = build_connection(basic_definition, " & ", str1)

        dictionary.basic_definitions.update({word: basic_definition})


def make_matrix(user):
    global dictionary
    pkl_file = open(user + 'z_dict_words.pkl', 'rb')
    dictionary = pickle.load(pkl_file)
    pkl_file.close()
    word_list = []
    for xword in words_used:
        for yword in words_used:
            if xword != yword:
                list1 = ([xword, yword])
                list1.sort()
                if list1 not in word_list:
                    xpos = dictionary.pos.get(xword)
                    ypos = dictionary.pos.get(yword)
                    if xpos[0] in ['n', 'a', 'r'] and ypos[0] in ['n', 'a', 'r']:
                        pass


def determine_class2(output, vars):
    abbrev_to_group = {}
    for var in vars:
        for sent in output.all_sent:
            for pos in sent[42]:
                if sent[pos] == var:
                    group = get_class(sent, pos, "", True)
                    if group != "":
                        abbrev_to_group.update({var: group})

    dictionary.groups.update({word: abbrev_to_group})


def determine_class(user, size = "small"):
    global word, dictionary
    pkl_file = open(user + 'z_dict_words.pkl', 'rb')
    dictionary = pickle.load(pkl_file)
    pkl_file.close()
    kind = 1
    proof_type = 5
    dictionary.basic_definitions = {}
    if size == 'small':
        list1 = words_used
        list1.append('whole')
    elif size == 'large':
        list1 = dictionary.categorized_sent.keys()

    for word in list1:
        pos = dictionary.pos.get(word)

        if word in dictionary.categorized_sent.keys() and \
                pos[0] in ['n', 'r', 'a']:
            # print (word)
            if word == 'imagination':
                if word == 'INM':
                    bb = 8

                output, vars = prepare_output(word)

                if kind == 0:
                    try:
                        output, consistent, reduced = try_instantiation(output, dictionary, "lemmas")
                        if not consistent:
                            raise Exception
                        determine_class2(output, vars)
                        build_basic_definition(output, reduced, word)
                        save_output(output)
                    except:
                        print ("bug: " + word)
                else:
                    output, consistent, reduced = try_instantiation(output, dictionary, "lemmas")
                    if not consistent and proof_type == 5:
                        output = rearrange("last", output, consistent, "", output.main_var)
                        print_sent([output.total_sent], [0], 2)
                        print (word)
                        raise Exception
                    determine_class2(output, vars)
                    build_basic_definition(output, reduced, word)
                    save_output(output)
        print_to_excel(proof_type)

        result = open('z_dict_words.pkl', 'wb')
        pickle.dump(dictionary, result)
        result.close()


def prepare_output(word):
    reduced_def = get_word_info(dictionary, word, "")
    sent = []
    ant_sent = []
    vars = set()
    cvars = set()
    output = get_output()
    for z, cls in enumerate(reduced_def):
        if z == 3:
            bb = 8

        sentences = cls.sentences
        adjust_ant_index(ant_sent, cls, sent, sentences, vars)
        adjust_con_index(cls, cvars, output, sent, sentences)
    output.prop_var = get_prop_var()
    output.main_var = vars
    output.prop_name = defaultdict(lambda: output.prop_var.pop(), {})
    for tsent in sent:
        tsent[2] = name_sent(tsent[1], output.prop_name)
        output.oprop_name[tsent[2]] = tsent[1]
    output.all_sent = sent
    output.abbreviations = json.loads(json.dumps(dictionary.def_constants.get(word, {})))
    output.variables = get_variables()
    output.tindex = 1
    output.user = ""
    for k in output.abbreviations.keys(): output.variables.remove(k)
    for var in vars | cvars:
        if var in output.variables: output.variables.remove(var)
    return output, vars


def adjust_ant_index(ant_sent, cls, sent, sentences, vars):
    for num in cls.def_stats.ant_index:
        if isinstance(num, int):
            ant_sent.append(sentences[num])
            for noun in sentences[num][42]: vars.add(sentences[num][noun])
        else:
            for cnum in num:
                for noun in sentences[cnum][42]: vars.add(sentences[cnum][noun])
                ant_sent.append(sentences[cnum])
                sent.append(sentences[cnum])


def adjust_con_index(cls, cvars, output, sent, sentences):
    for num in cls.def_stats.con_index:
        cons_count = 0
        if isinstance(num, int):
            sentences[num][7] = 'c'
            sent.append(sentences[num])
            for noun in sentences[num][42]: cvars.add(sentences[num][noun])
        else:
            tnum = cls.def_stats.con_hnum[cons_count]
            embed = cls.embeds.get(tnum)
            embed.sentences = sentences
            sub_word = embed.def_stats.def_word_num
            add_to_gsent([embed], output)
            output.trans_def.update({sub_word: embed})
            cons_count += 1
            for cnum in num:
                sent.append(sentences[cnum])
                for noun in sentences[cnum][42]: cvars.add(sentences[cnum][noun])
