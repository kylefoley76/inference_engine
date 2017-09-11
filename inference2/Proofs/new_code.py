from dictionary_new import large_dict
from claims_new import pop_sent
from openpyxl import load_workbook
from collections import Counter
import copy
import time
import operator
import sys
import cProfile
import re
from pprint import pprint
import collections
from start_and_stop import info
import os

# import pdb







# what's up


# hey
# checked coverage up to line 5575

# averaged .076 on 5.22 (proof type 'n'), .039 definitions, .004 statement logic
# averaged .059 on 5.22 proof type 'n', .023 definitions, .004 statement
# but just prior to that the speed was .066
# time spent in instantiation is .029

# on 6/8 time spent in instantiation = .009, .014
# on 6/10 time spent in instantiation = .018, definitions = .031, total .074

# on 6/26 average .026 (up to instantiation) definitions: .019
# trial 2, .020, .025, trial 3: same as 2, total 2.562

# with a lot of globals: 2.831, 3.330, 2.934

# 7/2 average .050, definitions .029, statement .0015, instantiation .010

# 7/19 average .0275 3 trials using old dictionary

# 7/20 average .0226 using new dictionary, .022, 0.216

# 7/26 average .0317, change_var = .0206, reduction .0056

# 8/2 average .0397, change_var = .239, reduction .0059

# 8/2 average .0464, change var .286, instantiation .012

# 9/4 average .0386, change var .0198, instantiation .0167

total_time = time.time()

######### tahir begin

#tahir - the code for the website should always be mysql == 1 for the prove site
# for the test machine site, it needs to be mysql == 2 and input = True
mysql = 0
excel = 0
if mysql == 0:
    proof_type, get_words_used, order = info()
    if proof_type == 1:
        wb4 = load_workbook('/Users/kylefoley/Desktop/inference engine/temp_proof.xlsx')
        w4 = wb4.worksheets[0]
    if get_words_used == 1:
        wb5 = load_workbook('/Users/kylefoley/Desktop/inference engine/dictionary4.xlsx')
        ws = wb5.worksheets[0]
else:
    proof_type = 0
    get_words_used = 0
    order = [0, 0, 1]

if mysql == 1:
    import os

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    print(BASE_DIR)
    sys.path.append(BASE_DIR)
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "inference_engine2.settings")
    import django

    django.setup()
    from inference2 import views
    from inference2.models import Define3, Archives, Input

########### tahir end

total_sent = []
all_sent = []
attach_sent = []
detach_sent = []
prop_name = []
prop_var = []
variables = []
abbreviations = []
words_used = []
dictionary = {}
object_classes = {}
definite_assignments = {}
propositional_constants = {}
object_properties = {}
do_not_instantiate = {}
already_defined = []
variable_type = []

tot_prop_name = []
result_data = {}
time1 = 0
st_log_time = 0
inst_tim = 0
instan_used = 0  # the number of times the instan function is used
instan_time = 0  # measures the time used in instantiation
lemmas_used = 0
time_spent_in_lemma_function = 0
time_spent_reducing = 0
time_spent_defining = 0
build_sent_slots_time = 0
build_sent_slots_counter = 0
start = 0
stop = 0

cond_r = chr(8835)
consist = "\u2102"  # consistency
top = chr(8868)
bottom = chr(8869)
neg = chr(172)
idd = chr(8781)  # translation symbol
iff = chr(8801)
mini_c = chr(8658)
mini_e = chr(8703)
implies = chr(8866)
conditional = chr(8594)
nonseq = chr(8876)
xorr = chr(8891)
idisj = chr(8744)
cj = chr(8896)
aid = chr(8776)
disj = chr(8855)
equi = chr(8660)
ne = "\u2260"  # not equal

l1 = "\u2081"
l2 = "\u2082"
l3 = "\u2083"
l4 = "\u2084"
l5 = "\u2085"
l6 = "\u2086"
l7 = "\u2087"
ua = "\u1d43"
ub = "\u1d47"
uc = "\u1d9c"
ud = "\u1d48"
ue = "\u1d49"
uf = "\u1da0"
ug = "\u1d4d"
ui = "\u2071"
uk = "\u1d4f"
um = "\u1d50"
un = "\u207f"
uo = "\u1d52"
up = "\u1d56"
ut = "\u1d57"
uv = "\u1d5b"
uu = "\u1d58"
uw = "\u02b7"
uy = "\u02b8"
uj = "\u02B2"
ul = "\u02E1"
ur = "\u02b3"
us = "\u02e2"
uh = "\u02b0"

prop_var4 = [chr(97 + t) for t in range(26)]
prop_var2 = [chr(97 + t) + "\u2081" for t in range(26)]
prop_var3 = [chr(97 + t) + "\u2082" for t in range(26)]
prop_var5 = [chr(97 + t) + "\u2083" for t in range(26)]
prop_var6 = [chr(97 + t) + "\u2084" for t in range(26)]
prop_var7 = [chr(97 + t) + "\u2085" for t in range(26)]
prop_var8 = [chr(97 + t) + "\u2086" for t in range(26)]
prop_var9 = [chr(97 + t) + "\u2087" for t in range(26)]
prop_var4 = prop_var4 + prop_var2 + prop_var3 + prop_var5 + prop_var6 + prop_var7 + prop_var8 + prop_var9
variables2 = [chr(122 - t) for t in range(25)]
variables2.remove("i")
variables2.remove("l")
variables3 = [chr(122 - t) + l1 for t in range(25)]
variables4 = [chr(122 - t) + l2 for t in range(25)]
variables2 = variables2 + variables3 + variables4

subscripts = [l1, l2, l3, l4]
alpha = chr(945)
beta = chr(946)
delta = chr(948)


#
# >> 8835
# ta^ 8868
# co^ 8869
# nt+ 172
# x^ 8801
# c^ 8658
# # 8703
# i^ 8866
# t^ 8594
# nf^ 8876
# ed^ 8891
# v+ 8744
# && 8896
# @ 8855
# if^ 8660

# tahir
def tran_str(str1, has_sentence_connectives=False):
    if str1 == "":
        return str1
    if "|" in str1:
        for i in range(len(str1)):
            if str1[i:i + 1] == "|":
                str3 = str1[i + 1:i + 2]
                str4 = get_super(str3)
                str1 = str1[:i] + str4 + str1[i + 2:]

    if has_sentence_connectives:

        if "t^" in str1:
            str1 = str1.replace("t^", conditional)
        if "nt+" in str1:
            str1 = str1.replace("nt+", neg)
        if "zzz" in str1:
            str1 = str1.replace("zzz", ne)
        if "x^" in str1:
            str1 = str1.replace("x^", iff)
        if "b^" in str1:
            str1 = str1.replace("b^", mini_e)
        if "c^" in str1:
            str1 = str1.replace("c^", mini_c)
        if "ed^" in str1:
            str1 = str1.replace("ed^", xorr)
        if "v+" in str1:
            str1 = str1.replace("v+", idisj)

    return str1


def get_super(str1):
    if str1 == "a":
        return "\u1d43"
    elif str1 == "b":
        return "\u1d47"
    elif str1 == "c":
        return "\u1d9c"
    elif str1 == "d":
        return "\u1d48"
    elif str1 == "e":
        return "\u1d49"
    elif str1 == "f":
        return "\u1da0"
    elif str1 == "g":
        return "\u1d4d"
    elif str1 == "h":
        return "\u02b0"
    elif str1 == "i":
        return "\u2071"
    elif str1 == "j":
        return "\u02B2"
    elif str1 == "k":
        return "\u1d4f"
    elif str1 == "l":
        return "\u02E1"
    elif str1 == "m":
        return "\u1d50"
    elif str1 == "n":
        return "\u207f"
    elif str1 == "o":
        return "\u1d52"
    elif str1 == "p":
        return "\u1d56"
    elif str1 == "r":
        return "\u02b3"
    elif str1 == "s":
        return "\u02e2"
    elif str1 == "t":
        return "\u1d57"
    elif str1 == "u":
        return "\u1d58"
    elif str1 == "v":
        return "\u1d5b"
    elif str1 == "w":
        return "\u02b7"
    elif str1 == "y":
        return "\u02b8"


def remove_outer_paren(str1, bool1=False):
    if str1 == "":
        return ""
    elif str1.count(")") == 0:
        if not bool1:
            return str1
        else:
            return False

    j = 0
    # on very rare occasions we will encounter strings of the following form ((p))
    if str1[0] != "(" and str1[-1] != ")":
        if not bool1:
            return str1
        else:
            return True
    if str1[:2] == "((" and str1[-2:] == "))":
        d = 2
    else:
        d = 1

    for k in range(0, d):
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 0 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]
                if bool1:
                    return True
    if not bool1:
        return str1
    else:
        return False


def remove_redundant_paren(str1):
    j = 0
    str2 = str1[:2]
    str3 = str1[2:]
    if str2 == '((' and str3 == '))':

        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 1 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]
    return str1


def mainconn(str1):
    ostring = copy.copy(str1)
    if os(str1):
        return ["", 0]
    if str1.find("&") < -1 and str1.find(idisj) < -1 and str1.find(iff) < -1 and str1.find(conditional) < -1 and \
                    str1.find(implies) < -1 and str1.find(nonseq) < -1 and str1.find(xorr) < -1:
        return ["", 0]

    str3 = str1
    bool1 = False

    if str1[0] == "~":
        str1 = str1[1:]
        bool1 = True

    j = 0
    bool2 = False
    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1

        if j == 0 and i + 1 != len(str1):
            break
        elif j == 0 and i + 1 == len(str1):
            str1 = str1[1:len(str1) - 1]
            bool2 = True

    j = -1
    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == conditional:
            f = -1
        if str2 == idisj or str2 == "&" or str2 == iff or str2 == implies or \
                        str2 == nonseq or str2 == conditional or str2 == xorr:
            if str3 != str2 and str3 != "":
                j = j + 1

            str3 = str2

    if j == -1:
        i = ostring.find(str3)
        return [str3, i]
    k = -1
    j = -1
    while True:
        k += 1
        if k > 150:
            break
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]

            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1

            if j == -1 and (str2 == idisj or str2 == "&" or str2 == iff or str2 == implies
                            or str2 == nonseq or str2 == conditional or str2 == xorr):
                if bool1 and bool2:
                    return [str2, i + 2]
                elif bool2 or bool1:
                    return [str2, i + 1]
                else:
                    return [str2, i]
        else:
            str1 = str1[1:-1]


def isvariable(str3, kind=""):
    bool2 = True
    if str3 == None or str3 == "":
        return False

    if str3 == 'a':
        return False
    elif str3 == 'i' and kind == "":
        return False
    elif str3 == "i":
        return True

    if str3 != "":
        str3 = str3.replace(l1, "")
        str3 = str3.replace(l2, "")
        str3 = str3.replace(l3, "")
        str3 = str3.replace(neg, "")
        if len(str3) == 1 and str3.islower():
            bool2 = True
        else:
            bool2 = False

    return bool2


def os(str1):
    cnx = [xorr, iff, idisj, conditional, implies, nonseq, "&"]
    for i in range(0, len(cnx)):
        if str1.find(cnx[i]) > -1:
            os = False
            return os
    os = True
    return os


def enclose(str1):
    i = -1
    global subscripts
    while i < len(str1) - 1:
        i += 1
        str2 = str1[i:i + 1]
        str3 = str1[i - 1:i]
        str4 = str1[i + 1:i + 2]
        if str2.islower() and str4 in subscripts:
            if str3 == "~":
                str1 = str1[:i - 1] + "(~" + str2 + str4 + ")" + str1[i + 2:]
            else:
                str1 = str1[:i] + "(" + str2 + str4 + ")" + str1[i + 2:]
            i += 4
        elif str2.islower():
            if str3 == "~":
                str1 = str1[:i - 1] + "(~" + str2 + ")" + str1[i + 1:]
            else:
                str1 = str1[:i] + "(" + str2 + ")" + str1[i + 1:]
            i += 3
    return str1


def is_natural_language(sentence, i):
    # this determines if the sentence is natural or an abbreviation
    if sentence[i + 1] == "~":
        if sentence[i + 3] in subscripts and sentence[i + 4] == ")":
            return True
        elif sentence[i + 3] == ")":
            return True
    else:
        if sentence[i + 2] in subscripts and sentence[i + 3] == ")":
            return True
        elif sentence[i + 2] == ")":
            return True
    return False


def find_sentences(sentence):
    global subscripts
    if sentence == None: g = 4 / 0
    if os(sentence): g = 4 / 0
    g = sentence.count('(')
    h = sentence.count(')')
    if g != h:
        print('wrong number of parentheses in sentence:' + sentence)
        g = 4 / 0
    if sentence.startswith("(~g)"):
        bb = 8
    marker = False
    il = -1
    total = -1
    c = -1
    neg_value = []
    str1 = ""
    sent1 = []
    sent_type2 = []
    wneg = []
    output = [None] * 9
    # the skel name list names each single sentence after a greek letter, even if
    # the same sentence appears twice it obtains a different name on the second
    # appearance
    skel_nam = []
    sent_num = []
    if sentence.find("~(") > -1:
        sentence = sentence.replace("~(", "(!")
    if sentence.find(implies) > -1:
        str2 = implies
    elif sentence.find(nonseq) > -1:
        str2 = nonseq
    str3 = mainconn(sentence)
    sentence = sentence.strip()
    str4 = str3[0]
    f = str3[1]
    id_num = []

    id_num.append(["1", str4, f])
    sent_num.append([1, '1', sentence, str4, f])
    str21 = ""
    p = 947
    greek_english_dict = {}
    unenclose_at_end = False
    connectives = ["&", idisj, iff, conditional, nonseq, implies, xorr]
    arr1 = []
    mini_c2 = mini_c + neg
    prt = copy.copy(sentence)
    more_num = [chr(945 + x) for x in range(24)]
    temp_string = mainconn(sentence)
    if sentence.find(implies) > -1:
        str1 = implies
    elif sentence.find(nonseq) > -1:
        str1 = nonseq
    else:
        if temp_string == iff:
            str1 = "bicond"
        elif temp_string == conditional:
            str1 = "cond"
        elif temp_string == "&":
            str1 = "cj"

    sent1.append(sentence)
    neg_value.append("")
    sent_type2.append(str1)
    wneg.append(sentence)
    skel_nam.append(None)

    j = 0
    n = 0
    for i in range(0, len(sentence)):
        str1 = sentence[i:(i + 1)]
        for o in connectives:
            if str1 == o:
                j += 1

    while n < j + 1:

        il += 1
        if il > 15:
            break

        e = 0
        l = len(sentence)
        x = -1
        while x < l - 1:
            x += 1
            temp_string = sentence[x:x + 1]
            if sentence[x:x + 1] == "(":
                if not unenclose_at_end:
                    unenclose_at_end = is_natural_language(sentence, x)

                if marker == False:
                    z = x
                    marker = True

                total += 1
            elif sentence[x: x + 1] == ")":
                total -= 1
                if total == -1:
                    marker = False
                    e += 1
                    c += 1

                    temp_sent = sentence[z: x + 1]
                    if temp_sent == '(bIc)':
                        pp = 7
                    otemp_sent = copy.copy(temp_sent)

                    if (len(sentence) - len(temp_sent)) > 2:
                        if temp_sent in prt and temp_sent in str21 and prt != str21:
                            prtnum = findinlist(str21, sent_num, 2, 1)
                            numb = prtnum + "1"
                        elif temp_sent in prt:
                            prtnum = findinlist(prt, sent_num, 2, 1)
                            numb = prtnum + "1"
                        else:
                            prtnum = ""
                            for bb in range(len(sent_num) - 1, -1, -1):
                                str3 = sent_num[bb][2]
                                if os(temp_sent) and str21 != "":
                                    if str21 == str3 and temp_sent in sent_num[bb][2]:
                                        # this is for those basic molecules for which the same sentence appears
                                        # in the definition several times
                                        prtnum = sent_num[bb][1]
                                        break
                                else:
                                    if temp_sent in sent_num[bb][2]:
                                        prtnum = sent_num[bb][1]
                                        break
                            # if prtnum == "":
                            #     easygui.msgbox('your sentences are not numbered properly')
                            g = len(prtnum) + 1
                            f = 0

                            for bb in range(len(sent_num) - 1, -1, -1):
                                temp_sn = sent_num[bb][1]
                                h = temp_sn[:g - 1]
                                hh = sent_num[bb][0]
                                if g > sent_num[bb][0]:
                                    break
                                if sent_num[bb][0] == g and temp_sn[:g - 1] == prtnum:
                                    f += 1

                            f += 1
                            if f < 10:
                                numb = prtnum + str(f)
                            else:
                                numb = prtnum + more_num[f - 10]

                        prt = temp_sent
                        temp_mc = mainconn(temp_sent)
                        mc = temp_mc[0]
                        str3 = temp_mc[0]
                        g = temp_mc[1]
                        mc_num = temp_mc[1]
                        sent_num.append([len(numb), numb, temp_sent, str3, g])

                        if os(temp_sent):

                            # n counts the number of single sentences
                            n += 1
                            if temp_sent.find("~") > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace("~", "")

                            elif temp_sent.find(mini_c2) > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace(mini_c2, mini_c)

                            elif temp_sent.find(mini_c2) == -1 and temp_sent.find(mini_c) > -1:
                                neg_value.append("")
                            elif temp_sent.find("~") == -1:
                                neg_value.append("")
                            else:
                                break  # stop
                        else:
                            neg_value.append("")

                        sent1.append(temp_sent)
                        wneg.append(otemp_sent)
                        id_num.append([numb, mc, mc_num])

                        if os(otemp_sent):
                            if otemp_sent in greek_english_dict:
                                skel_nam.append(greek_english_dict.get(otemp_sent))
                            else:
                                p += 1
                                greek_english_dict.update({otemp_sent: chr(p)})
                                skel_nam.append(chr(p))
                        else:
                            skel_nam.append(None)
                    else:
                        sentence = sentence[1:len(sentence) - 1]
                        l = len(sentence)
                        x = -1
                        c -= - 1
                        e -= - 1

        total = -1
        marker = False
        w = -1

        if n < j + 1:
            if len(sent1) > w:
                while w + 1 < len(sent1):
                    if w == 13:
                        pp = 7
                    w += 1
                    str21 = sent1[w]
                    if not os(str21) and w != 0:
                        if str21 not in arr1:
                            sentence = str21
                            arr1.append(sentence)
                            break

    for i in range(len(sent1)):
        temp_string = sent1[i]
        if temp_string.find("(!") > -1:
            sent1[i] = sent1[i].replace("(!", "~(")
            wneg[i] = wneg[i].replace("(!", "~(")

    if unenclose_at_end:
        for i in range(len(sent1)):
            sent1[i] = unenclose(sent1[i])
            wneg[i] = unenclose(wneg[i])

    output[0] = sent1
    output[1] = neg_value
    output[2] = sent_type2
    output[3] = wneg
    output[4] = id_num
    output[6] = translate_to_greek(skel_nam, wneg, id_num)
    output[5] = skel_nam[0]

    return output


def translate_to_greek(skel_nam, wneg, id_num):
    for i in range(len(skel_nam)):
        if skel_nam[i] == None:
            to_be_translated = wneg[i]
            for j in range(len(skel_nam) - 1, -1, -1):
                if skel_nam[j] != None:
                    if wneg[j] in to_be_translated and id_num[j][1] == "":
                        to_be_translated = to_be_translated.replace(wneg[j], skel_nam[j])
            skel_nam[i] = to_be_translated

    return skel_nam


def remove_duplicates(list1, i):
    list2 = []
    j = -1
    while j < len(list1) - 1:
        j += 1
        if list1[j][i] in list2:
            del list1[j]
            j -= 1
        else:
            list2.append(list1[j][i])

    return list1


def sort_decisions(list1):
    for lists in list1:
        lists[46].sort()


def define_irregular_terms(list1, type=""):
    # the code for avoiding the circularity of defining 'i' is found in the add to all sent function

    do_not_define_again = []
    sort_decisions(list1)
    i_defined = 0 if type == "" else 2
    m = -1
    while m < len(list1) - 1:
        m += 1
        while list1[m][46] != []:
            category = list1[m][46][0]
            i = list1[m][45][category][0]
            del list1[m][46][0]
            del list1[m][45][category][0]
            if list1[m][i] == 'i': i_defined += 1
            if list1[m][i] == 'a':
                bb = 8

            if not lies_wi_scope_of_univ_quant(list1[m], i) and \
                    list1[m][42] not in do_not_define_again and \
                    (i_defined < 2 or list1[m][i] != 'i'):
                antecedent = copy.deepcopy(list1[m])
                consequent = copy.deepcopy(list1[m])
                consequent, rule = determine_which_function_to_use(consequent, i, category, list1, type)
                prepare_irregular_att_sent(antecedent, consequent, rule, list1, type)
                del list1[m]
                m -= 1
                break

    return list1


def prepare_irregular_att_sent(antecedent, consequent, rule, list1, type=""):
    if rule == None:
        return

    for sent in consequent:
        list1.append(sent)

    if type == "universal":
        return

    if len(consequent) == 1:
        prepare_att_sent_1_sent(antecedent, rule, iff, consequent)
    elif len(consequent) == 2:
        prepare_att_sent_2_sent(antecedent, iff, consequent, rule)
    elif len(consequent) == 3:
        prepare_att_sent_3_sent(antecedent, iff, consequent, rule)
    elif len(consequent) == 4:
        prepare_att_sent_4_sent(antecedent, consequent, iff, rule)


def determine_which_function_to_use(list1, i, j, list3, type=""):
    if j == 1:
        consequent, rule = change_variables(list1, i, list3, type)
    elif j == 3:
        consequent, rule = eliminate_common_name_possessives(list1, i)
    elif j == 4:
        consequent, rule = eliminate_proper_name_possessives3(list1, i)
    elif j == 5:
        consequent, rule = eliminate_and_coordinator(list1, i)
    elif j == 6:
        consequent, rule = eliminate_adjectives(list1, i)
    elif j == 7:
        consequent, rule = eliminate_concept_instance_apposition(list1, i)
    elif j == 8:
        consequent, rule = eliminate_parenthetical_phrase(list1, i)
    elif j == 9:
        consequent, rule = eliminate_relative_pronouns(list1, i)
    elif j == 11:
        consequent, rule = eliminate_as(list1, i)
    elif j == 13:
        consequent, rule = divide_relations(list1, i)
    elif j == 14:
        consequent, rule = eliminate_there(i, list1)
    elif j == 15:
        consequent, rule = eliminate_universals(list1, i)
    elif j == 16:
        consequent, rule = change_variables(list1, i, list3)

    return consequent, rule


def eliminate_proper_name_possessives3(list1, i):
    if i == 69:
        concept_position = 5
    elif i == 70:
        concept_position = 14
    new_possessor = list1[i][:-2]
    possessee_concept = list1[concept_position]
    new_possessee = definite_assignments.get(possessee_concept)
    if new_possessee == None:
        definite_assignments[possessee_concept] = variables[0]
        new_possessee = variables[0]
        del variables[0]
    list1[i] = None
    list1[concept_position] = new_possessee
    con_parts1 = build_sent2(list1)
    con_parts2 = build_sent1(new_possessor, "", "OWN", new_possessee)
    con_parts3 = build_sent1(new_possessee, "", "I", possessee_concept)

    return [con_parts1, con_parts2, con_parts3], "PPE"


def eliminate_common_name_possessives(list1, i):
    if i == 69:
        concept_position = 5
    elif i == 70:
        concept_position = 14
    possessor_concept = list1[i][:-2]
    possessee = list1[concept_position]
    new_possessor = definite_assignments.get(possessor_concept)
    if new_possessor == None:
        definite_assignments[possessor_concept] = variables[0]
        new_possessor = variables[0]
        del variables[0]
    list1[i] = None
    con_parts1 = build_sent2(list1)
    con_parts2 = build_sent1(new_possessor, "", "OWN", possessee)
    con_parts3 = build_sent1(new_possessor, "", "I", possessor_concept)

    return [con_parts1, con_parts2, con_parts3], "CPE"


def eliminate_and_coordinator(list1, i):
    # this seperates a sentence with an 'and' coordinator into two


    list1[66] = None
    list7 = [None] * 80
    list7[5] = list1[67]
    list1[67] = None
    for i in range(6, 20):
        list7[i] = list1[i]
    list7 = restore_original_sent(list7)
    list1 = build_sent2(list1)
    con_parts = copy.deepcopy(list7)
    con_parts2 = copy.deepcopy(list1)
    consequent = [con_parts, con_parts2]

    return consequent, "DE and" + uc


def eliminate_adjectives(list1, i):
    dict1 = {4: 8, 13: 8, 17: 49}
    relat_to_adj = {13: 9, 17: 15, 4: 9}
    neg_pos = dict1.get(i)
    relat_pos = relat_to_adj.get(i)
    if list1[relat_pos] == "I":
        dict2 = {13: 5, 17: 14}
        noun_pos = dict2.get(i)
    else:
        noun_pos = i + 1
    con_parts = copy.deepcopy(build_sent1(list1[noun_pos], list1[neg_pos], "J", list1[i]))
    list1[i] = None
    list1[neg_pos] = None
    list1 = copy.deepcopy(build_sent2(list1))
    consequent = [list1, con_parts]

    return consequent, "ADJ E"


def eliminate_concept_instance_apposition(list1, i):
    dict1 = {35: 5, 36: 14, 37: 18, 38: 22}
    j = dict1.get(i)
    con_parts = copy.deepcopy(build_sent1(list1[i], "", "I", list1[j]))
    list1[j] = list1[i]
    list1[i] = None
    list1 = copy.deepcopy(build_sent2(list1))
    consequent = [list1, con_parts]

    return consequent, "CIA"


def eliminate_parenthetical_phrase(list1, i):
    sub_begin = list1[57][0]
    sub_end = list1[57][1]
    rule = "DE " + list1[i]
    list1[i] = None
    list2 = [None] * 80
    list2[3] = list1[5]
    k = 3
    for j in allowable_slots():
        if list1[j] != None and list1[j] != "":
            absolute_pos = allowable_slots().index(j)
            if absolute_pos > sub_begin and absolute_pos <= sub_end:
                k += 1
                list2[k] = list1[j]
                list1[j] = None
            if absolute_pos > sub_end:
                break

    list1 = restore_original_sent(list1)
    list2 = categorize_words(list2)

    return [list1, list2], rule


# ddd
def eliminate_relative_pronouns(con_parts1, i):
    rule = "DE " + con_parts1[i]
    con_parts2 = [None] * 80
    map_nouns_to_relative_pronouns = {60: 14, 61: 18, 62: 22}
    j = map_nouns_to_relative_pronouns.get(i)
    relation_positions = [15, 19, 23, 27, 31]
    con_parts1[i] = None
    con_parts2[3] = con_parts1[j]
    k = 3
    started = False
    for j in allowable_slots():
        if j == 49:
            bb = 8
        if j == i:
            started = True
        if started and con_parts1[j] != None:
            k += 1
            if j != i:
                con_parts2[k] = con_parts1[j]
                con_parts1[j] = None
    con_parts2 = categorize_words(con_parts2)
    # the negation sign sometimes get put into the wrong slot during this step
    # this is not an ideal solution but it will do for now
    con_parts2 = quick_negation_transfer(con_parts2)
    con_parts1 = restore_original_sent(con_parts1)
    consequent = [con_parts1, con_parts2]
    return consequent, rule


def eliminate_as(list1, i):
    # modify this if 'as' can be placed in a location other than 15

    con_parts1 = build_sent1(list1[5], "", list1[9], list1[14])
    con_parts2 = build_sent1(list1[18], "", list1[9], list1[14])

    return [con_parts1, con_parts2], "DE AS"


def divide_relations(list1, i):
    # b R c S d = b R c & b S d
    # b R c S d T e = b R c & b S d T e
    list2 = [None] * 80
    dict1 = {5: 5, 49: 8, 15: 9, 18: 14, 19: 15, 22: 18}
    for k, v in dict1.items():
        if list1[k] != None:
            list2[v] = list1[k]
            if k != 5:
                list1[k] = None

    list2[46] = []
    if list2[9][-1] == "P": list2[9] = list2[9][:-1]
    list1 = build_sent2(list1)
    list2 = build_sent2(list2)
    consequent = [list1, list2]

    return consequent, "RDA"


def eliminate_there(i, list1):
    if i == 5:
        dict1 = {14: 5, 10: 3, 13: 4}
    assert i == 5
    for k, v in dict1.items():
        list1[v] = list1[k]
        list1[k] = None
    list1[9] = 'EX'
    list1 = restore_original_sent(list1)

    return [list1], "DE there"


# ddd

def eliminate_universals(list1, i):
    fir_antecedent = copy.deepcopy(list1)
    rule = "DE " + list1[i]
    class_sent = get_class_sent(list1, i)

    ppart_type = has_scope_over_past_participle(list1, i)

    if has_scope_over_parenthetical_phrase(list1, i):
        antecedent, consequent = extract_words_from_parenthetical_phrase(list1, i, class_sent[5])
        antecedent = define_irregular_terms([antecedent], "universal")

    elif has_scope_over_subclause(list1, i):
        antecedent, consequent = extract_words_from_subclause(list1, i, class_sent[5])
        antecedent = define_irregular_terms([antecedent], "universal")

    elif ppart_type != "":
        antecedent, consequent = eliminate_past_participle_in_universal(list1, i, class_sent[5], ppart_type)
        antecedent = define_irregular_terms(antecedent, "universal")

    elif has_scope_over_prepositional_relation(list1, i):
        antecedent, consequent = eliminate_prepositional_relation_in_universal(list1, i, class_sent[5])
        antecedent = define_irregular_terms(antecedent, "universal")

    elif has_scope_over_adjective(list1, i):
        antecedent = turn_adj_into_sent(i, m)

    elif determ_lies_wi_scope_of_univ2(list1, i):
        antecedent, consequent = eliminate_no_w_indef_obj(class_sent, list1)

    else:
        consequent = get_the_simple_consequent(list1, i, class_sent[5])
        antecedent = []
    antecedent.append(class_sent)
    for sent in antecedent:
        s = findposinmd(sent[42], all_sent, 42)
        if s == -1:
            sent[54] = 'do not define'
            all_sent.append(sent)
    consequent[54] = 'do not define'
    s = findposinmd(consequent[42], all_sent, 42)
    if s == -1:
        all_sent.append(consequent)

    prepare_att_sent_univ(fir_antecedent, antecedent, consequent, rule)

    return None, None


def has_scope_over_parenthetical_phrase(list1, i):
    if i == 3 and list1[59] != None:
        return True
    else:
        if list1[39] != None:
            dict1 = {10: 60, 16: 61, 20: 62}
            j = dict1.get(i)
            if j == None:
                return False
            elif list1[j] != None:
                return True
    return False


def extract_words_from_parenthetical_phrase(list1, i, new_var):
    concept_pos = 14 if i == 10 else i + 2
    list1[i] = None
    dict1 = {3: 59, 10: 60, 16: 61, 20: 62}
    i = dict1.get(i)
    list1[concept_pos] = new_var
    list2, _ = eliminate_parenthetical_phrase(list1, i)

    return list2[1], list2[0]


def has_scope_over_prepositional_relation(list1, i):
    prepositional_relations = ['IN', "INB", "OF", "ATC"]
    dict1 = {3: 9, 10: 15, 16: 19}
    j = dict1.get(i)
    if list1[j] in prepositional_relations:
        return True
    else:
        return False


def eliminate_prepositional_relation_in_universal(list1, i, new_var):
    prepositional_relations = ['IN', "INB", "OF", "ATC"]
    quant = list1[i]
    list1[i] = None
    var_pos = 14 if i == 10 else i + 2
    list1[var_pos] = new_var

    if i == 3:
        for j in [15, 19, 23, 27, 31]:
            if list1[j] not in prepositional_relations:
                boundary = j
                break
    else:
        boundary = i

    return divide_antecedent_from_consequent(list1, var_pos, quant, i, new_var, boundary)


def divide_antecedent_from_consequent(list1, var_pos, quant, i, new_var, boundary):
    k = 2
    m = 2
    antecedent = [None] * 80
    consequent = [None] * 80
    first_cons_relat_found = False if quant == 'no' else True
    for j in allowable_slots(1):
        if j == var_pos:
            k += 1
            m += 1
            antecedent[k] = list1[j]
            consequent[m] = list1[j]

        elif list1[j] != None and part_of_ante(i, j, boundary):
            k += 1
            antecedent[k] = list1[j]
        elif list1[j] != None:
            m += 1
            if not first_cons_relat_found:
                pos = dictionary[0].get(list1[j])
                if pos[0] == 'r':
                    consequent[m] = "~"
                    first_cons_relat_found = True
                    m += 1
            consequent[m] = list1[j]
    antecedent = categorize_words(antecedent)
    consequent = categorize_words(consequent)
    antecedent = [antecedent]
    if consequent[46] != []:
        original_cons_relat = consequent[9]
        consequent = define_irregular_terms([consequent], "universal")
        consequent = add_to_antecedent(original_cons_relat, consequent, new_var, antecedent)

    return antecedent, consequent


def part_of_ante(i, j, boundary):
    if i == 3 and j < boundary:
        return True
    elif i == 3:
        return False
    elif j > boundary:
        return True
    else:
        return False


def has_scope_over_past_participle(list1, i):
    indefinite_determinatives = ['a', 'a' + ud, 'many' + un, 'any' + un, 'some' + up, 'few']
    type = ""
    if list1[15] == None:
        return ""
    if list1[i] == 'every':
        if i == 3 and list1[9][-1] == 'P':
            type = "ante first"
            # type 1 = every b BRNP c DRK d
        elif i == 10 and list1[15][-1] == 'P':
            type = "cons first"
            # type 2 = i HLP every b BRNP d

    elif list1[i] == 'no':
        if i == 3 and list1[9][-1] == 'P':
            type = "ante first"
            # type 3 = no b BRNP c DRK d
            # boundary 14
        elif i == 3 and list1[15][-1] == 'P' and list1[10] in indefinite_determinatives:
            type = "cons first"
            #  type 5 = no b THK a c BRNP d <> ((e I b) & (f I c) & (f BRNP d)) -> (e ~ THK f)
            # no b THK d WND e <> ((f I b) &
        elif i == 10 and list1[15][-1] == 'P':
            # i HLP no b BLD c <> ((e I b) & (e BLD c)) -> (i ~ HLP e)
            type = "cons first"
            # boundary = 14

    return type


def eliminate_past_participle_in_universal(list1, i, new_var, type):
    quant = list1[i]
    list1[i] = None
    var_pos = 14 if i == 10 else i + 2
    list1[var_pos] = new_var

    return divide_antecedent_from_consequent2(list1, var_pos, quant, 15, new_var, type)


def divide_antecedent_from_consequent2(list1, var_pos, quant, boundary, new_var, type):
    k = 2
    m = 2
    antecedent = [None] * 80
    consequent = [None] * 80
    if type == 'ante first':
        list1[9] = list1[9][:-1]
    else:
        list1[15] = list1[15][:-1]
    for j in allowable_slots(1):
        if type == 'ante first':
            place_in_antecedent = True if j < boundary else False
        else:
            place_in_antecedent = False if j < boundary else True

        if j == var_pos:
            k += 1
            m += 1
            antecedent[k] = list1[j]
            consequent[m] = list1[j]

        elif list1[j] != None and place_in_antecedent:
            k += 1
            antecedent[k] = list1[j]
        elif list1[j] != None:
            m += 1
            if quant == 'no':
                if (type == "ante first" and not place_in_antecedent) or \
                        (type == "cons first" and not place_in_antecedent):
                    pos = dictionary[0].get(list1[j])
                    if pos != None and pos[0] == 'r':
                        consequent[m] = "~"
                        m += 1
            consequent[m] = list1[j]
    antecedent = categorize_words(antecedent)
    consequent = categorize_words(consequent)
    antecedent = [antecedent]
    if consequent[46] != []:
        original_cons_relat = consequent[9]
        consequent = define_irregular_terms([consequent], "universal")
        consequent = add_to_antecedent(original_cons_relat, consequent, new_var, antecedent)

    return antecedent, consequent


def add_to_antecedent(original_cons_relat, consequent, new_var, antecedent):
    i = -1
    while i < len(consequent) - 1:
        i += 1
        if consequent[i][9] != original_cons_relat:
            antecedent.append(consequent[i])
            del consequent[i]
            consequent = consequent[0]
            return consequent


def get_the_simple_consequent(list1, i, new_var):
    j = 14 if i == 10 else i + 2
    list1[j] = new_var
    if list1[i] == 'no':
        list1[8] = "~"
    list1[i] = None

    return build_sent2(list1)


def eliminate_no_w_indef_obj(class_sent, list1):
    list1[3] = None
    consequent = build_sent1(class_sent[5], "~", list1[9], variables[0])
    sec_antecedent = build_sent1(variables[0], "", "I", list1[14], list1[15], list1[18])
    if variables[0] not in variable_type[0]:
        variable_type[0].append(variables[0])
        variable_type[3].append(variables[0])
    del variables[0]

    return [sec_antecedent], consequent


def has_scope_over_adjective(list1, i):
    k = 13 if i == 10 else i + 1
    if i == 3 and list1[k] != None:
        return True
    return False


def prepare_att_sent_univ(fir_antecedent, antecedent, consequent, rule):
    global sn
    list1 = [""] * 60
    list2 = []
    list42 = []
    bare_sent = []
    str2 = ""
    str3 = ""
    greek = ""
    n = 933
    consequent[44] = chr(933)
    bare_sent.append(consequent[1])
    consequent[68] = "12"
    consequent[53] = 'q'
    heir_num = "11"
    for i, sent in enumerate(antecedent):
        list2.append([sent[1], sent[2]])
        list42.append([sent[72], sent[2]])
        bare_sent.append(sent[1])
        if str2 == "":
            str2 = sent[0]
            str3 = sent[42]
            greek = chr(n + i + 1)
            sent[44] = chr(n + i + 1)
        else:
            str2 += " & " + sent[0]
            str3 += " & " + sent[42]
            greek += " & " + chr(n + i + 1)
            sent[44] = chr(n + i + 1)
        sent[68] = heir_num + str(i + 1)
        if len(antecedent) > 1:
            sent[53] = "ca"
        else:
            sent[53] = "a"
    if len(antecedent) > 1:
        str2 = "(" + str2 + ")"
        str3 = "(" + str3 + ")"
        greek = "(" + greek + ")"

    list1[0] = list2
    list1[1] = [[consequent[1], consequent[2]]]
    list1[3] = 'c'
    list1[4] = str3 + " " + conditional + " " + consequent[42]
    list1[7] = [str3, ""]
    list1[8] = [consequent[1], consequent[2]]
    list1[34] = antecedent
    list1[35] = [consequent]
    list1[37] = str2 + " " + conditional + " " + consequent[0]
    list1[38] = bare_sent
    list1[40] = [str2, ""]
    list1[41] = [consequent[72], ""]
    list1[42] = list42
    list1[43] = [[consequent[72], consequent[2]]]
    list1[47] = greek + " " + conditional + " " + chr(933)
    list1[50] = "axiom of definition"
    bare_sent2 = copy.deepcopy(bare_sent)

    sn += 1
    fir_antecedent[44] = chr(932)
    list5 = [""] * 60
    list5[0] = [[fir_antecedent[1], fir_antecedent[2]]]
    list5[1] = [[list1[4], ""]]
    list5[2] = sn
    list5[3] = "e"
    list5[4] = fir_antecedent[42] + " " + iff + " (" + list1[4] + ")"
    list5[7] = [fir_antecedent[1], fir_antecedent[2]]
    list5[8] = [list1[4], ""]
    list5[34] = [fir_antecedent]
    list5[37] = fir_antecedent[0] + " " + iff + " (" + list1[37] + ")"
    bare_sent2.append(fir_antecedent[1])
    list5[38] = bare_sent2
    list5[39] = [list1]
    list5[40] = [fir_antecedent[72], fir_antecedent[2]]
    list5[41] = [list1[37], ""]
    list5[42] = [[fir_antecedent[0], ""]]
    list5[43] = [[list1[37], ""]]
    list5[47] = fir_antecedent[44] + " " + iff + " (" + list1[47] + ")"
    attach_sent.append(list5)
    add_to_total_sent(sn, list5[37], list5[4], "", rule)


def get_class_sent(list1, i):
    general_variables = copy.deepcopy(variable_type[0])
    class_pos = 14 if i == 10 else i + 2
    list2 = build_sent1(variables[0], "", "I", list1[class_pos])
    general_variables.append(variables[0])
    variable_type[0] = general_variables
    del variables[0]

    return list2


#
# def turn_adj_into_sent(i,m):
#
#     pass



def has_scope_over_subclause(list1, i):
    if i == 3 and list1[59] != None:
        bool1 = True
    elif i == 10 and list1[60] != None:
        bool1 = True
    elif i == 16 and list1[61] != None:
        bool1 = True
    elif i == 20 and list1[62] != None:
        bool1 = True
    else:
        bool1 = False
    return bool1


def get_last_relation(list1):
    relational_positions = [31, 27, 23, 19, 15]
    for pos in relational_positions:
        if list1[pos] != None:
            return pos


def extract_words_from_subclause(consequent, i, new_var):
    # this is the sentence that will be inserted into the antecedent in the
    # definition of 'every' or 'no', in the future it should output
    # a set of lists, not just 1


    quantifier = consequent[i]
    antecedent = [None] * 80
    new_var_loc = 14 if i == 10 else i + 2
    consequent[new_var_loc] = new_var
    consequent[i] = None
    if quantifier == 'no':
        consequent[8] = "~"
    for num in [60, 61, 62]:
        if consequent[num] != None:
            consequent[num] = None
            break

    boundary_found = False
    k = 2
    for n in allowable_slots():
        if n == i:
            boundary_found = True
        if boundary_found and n != i and consequent[n] != None:
            if 55 == allowable_slots().index(n):
                break
            k += 1
            if consequent[n] != "~":
                antecedent[k] = consequent[n]
            if n != new_var_loc:
                consequent[n] = None

    antecedent = categorize_words(antecedent)
    consequent = restore_original_sent(consequent)

    return antecedent, consequent


def restore_original_sent(list1):
    k = 2
    list2 = [None] * 80
    for i in allowable_slots():
        if list1[i] != None and list1[i] != "":
            k += 1
            list2[k] = list1[i]
    list2 = categorize_words(list2)

    return list2


def insert_into_dict(_dict, obj, pos):
    insert_in_dict = lambda _dict, obj, pos: {k: v for k, v in
                                              (list(_dict.items())[:pos] + list(obj.items()) + list(_dict.items())[
                                                                                               pos:])}
    _dict = insert_in_dict(_dict, obj, pos)

    return _dict



def get_more_variable_types(consistent = True):
    if consistent == False:
        return
    global variable_type
    put_in_definite = []
    abbrev_4_definite = abbreviations[1].get("definite")
    variable_type[0] = []
    variable_type[1] = []
    variable_type[2] = []
    variable_type[3] = []

    for k in abbreviations[0].keys():
        if isvariable(k):
            if k not in variable_type[2]:
                variable_type[2].append(k)

    for lst in detach_sent:
        if lst[71] != 'irrelevant':
            if lst[9] == "J" and lst[14] == abbrev_4_definite:
                if lst[5] not in variable_type[2]:
                    variable_type[2].append(lst[5])
                    put_in_definite.append(lst[5])
            else:
                for i in noun_slots():
                    if not ex(lst, i):
                        break
                    elif isvariable(lst[i], "i"):
                        if lst[i] == 'i' and "i" not in variable_type[2]:
                            variable_type[2].append("i")
                        elif lst[i] not in variable_type[2] and lst[i] not in variable_type[1]:
                            variable_type[1].append(lst[i])

    for i in put_in_definite:
        if i in variable_type[1]:
            variable_type[1].remove(i)

    uninstantiable = []
    instantiable = []
    for lst in attach_sent:
        for j in [34,35]:
            for sent in lst[j]:
                if sent[71] != 'irrelevant':
                    for i in noun_slots():
                        if not ex(sent, i):
                            break
                        elif isvariable(sent[i], "i"):
                            if sent[i] not in uninstantiable:
                                uninstantiable.append(sent[i])
                            if sent[i] == 'i' and "i" not in variable_type[2]:
                                variable_type[2].append("i")
                            elif sent[i] not in variable_type[2] and sent[i] not in variable_type[1]\
                                    and sent[i] not in variable_type[0]:
                                if lst[3] == 'e' or (lst[3] == 'c' and j == 34):
                                    variable_type[0].append(sent[i])
                                    instantiable.append(sent[i])
                                elif lst[3] == 'c' and j == 35:
                                    variable_type[1].append(sent[i])
                    for var in instantiable:
                        for var2 in uninstantiable:
                            do_not_instantiate.setdefault(var, []).append(var2)
                    instantiable = []
                    uninstantiable = []

    variable_type[3] = variable_type[0] + variable_type[1] + variable_type[2]


def get_relevant_variables(list1):
    put_in_definite = []
    abbrev_4_definite = abbreviations[1].get("definite")
    for k in abbreviations[0].keys():
        if isvariable(k):
            if k not in variable_type[2]:
                variable_type[2].append(k)

    for lst in list1:
        is_detached = isinmdlist(lst[42], detach_sent, 42)

        if lst[9] == "J" and lst[14] == abbrev_4_definite:
            if lst[5] not in variable_type[2]:
                variable_type[2].append(lst[5])
                put_in_definite.append(lst[5])
        for i in noun_slots():
            if i > 17 and lst[i] == None:
                break
            if isvariable(lst[i], "i"):
                if lst[i] not in variable_type[3] and lst[54] != 'do not define':
                    variable_type[3].append(lst[i])
                if lst[i] not in variable_type[2] and lst[i] not in variable_type[1]:
                    if lst[i] == 'i':
                        if "i" not in variable_type[2]:
                            variable_type[2].append("i")
                    elif is_detached:
                        if lst[i] not in variable_type[1]:
                            variable_type[1].append(lst[i])
                    else:
                        if lst[i] not in variable_type[0]:
                            variable_type[0].append(lst[i])
                            if lst[i] not in variable_type[3]:
                                variable_type[3].append(lst[i])

    for i in put_in_definite:
        if i in variable_type[1]:
            variable_type[1].remove(i)
    for lst in list1:
        if not isinmdlist(lst[1], detach_sent, 1):
            lst[54] = 'do not define'


def is_irrel_var(list1):
    if list1[5] not in variable_type[3] and list1[9] == "I":
        return True
    elif list1[5] not in variable_type[3] and (list1[14] == None or
                                                       list1[14] not in variable_type[3]):
        return True
    else:
        return False



def define_regular_terms(list1):
    dictionary[6] = use_rarely_defined_word()
    do_not_define_again = []
    m = -1
    while m < len(list1) - 1:
        m += 1
        if isdefineable(list1[m]) and list1[m][42] not in do_not_define_again \
                and not list1[m][54] == 'do not define' and not is_irrel_var(list1[m]):
            do_not_define_again.append(list1[m][42])
            change_variables(list1[m], 0, list1)


def get_definiendum(list1, i):
    special_relations = ["I", "J", "H"]
    if i != 0:
        definiendum = list1[i]
        defining_abbreviation = [variables[0]]
        del variables[0]
    elif list1[9] == "=":
        definiendum = list1[14]
        defining_abbreviation = [list1[5]]
    elif list1[9] in special_relations:
        definiendum = abbreviations[0].get(list1[14])
        defining_abbreviation = [list1[5]]
    else:
        definiendum = list1[9]
        if list1[14] != None:
            defining_abbreviation = [list1[5], list1[14]]
        else:
            defining_abbreviation = [list1[5]]

    return definiendum, defining_abbreviation


def change_variables(sentence, def_loc, list1, type=""):
    global time_spent_defining
    aa = time.time()

    definiendum, defining_abbreviations = get_definiendum(sentence, def_loc)

    if definiendum == None or definiendum in dictionary[6]:
        return

    if definiendum == 'a':
        bb = 8

    definition = dictionary[1].get(definiendum)

    if definition == None:
        return

    def_info2 = find_sentences(definition)

    def_info = copy.deepcopy(def_info2)

    constant_map, temp_prop_const = get_abbreviations_from_definition(def_info)

    def_info = eliminate_conjuncts_from_definition(def_info)

    _ = get_new_sent(def_info, defining_abbreviations, def_loc, sentence, definiendum)

    def_abbrev_dict, r_sent_loc, new_sentences, defining_abbreviations = _

    total_dict = {**def_abbrev_dict, **constant_map}

    _ = replace_constants(total_dict, temp_prop_const, defining_abbreviations, new_sentences)

    new_sentences, unfill_positions, prop_unfill = _

    _ = replace_indefinite_variables(new_sentences, unfill_positions, defining_abbreviations, total_dict, sentence[1])

    new_sentences, indefinite_dict, rn_type = _

    total_dict = {**total_dict, **indefinite_dict}

    _ = replace_propositional_constants(temp_prop_const, prop_unfill, new_sentences, total_dict)

    new_sentences, old_prop_new_prop = _

    new_sentences = replace_r_sent(total_dict, r_sent_loc, new_sentences, sentence, definiendum, def_loc)

    remove_starred_general_variables()

    new_sentences = [build_sent2(sent) for sent in new_sentences]

    if type == 'universal':
        # the point of the x is that universals need to add the new sentences to list 1
        # but non-universals do not
        delete_is_indefinite(new_sentences)
        del new_sentences[0]
        return new_sentences, "x"

    new_sentences = add_first_sent_to_def_sent(new_sentences, sentence, r_sent_loc)

    rename = build_rename_sent2(constant_map, def_abbrev_dict, old_prop_new_prop, indefinite_dict, definiendum, rn_type)

    rule = get_rule(definiendum, r_sent_loc, def_info[0], rename)

    add_to_attach_sent(def_info, new_sentences, definition, r_sent_loc, rename, definiendum, type, rule)

    new_sentences = do_not_define(new_sentences, definiendum)

    add_def_sent_to_all_sent(definiendum, list1, new_sentences)

    time_spent_defining += (time.time() - aa)

    if r_sent_loc != [] or definiendum == 'i':
        return sentence, None


# bbb
def get_rule(definiendum, r_sent_loc, def_info, rename):
    if r_sent_loc != []:
        rule = "DE "
    else:
        connective = def_info[4][1][1] if def_info[4][0][1] == "&" else def_info[4][0][1]
        assert connective in [conditional, iff]

        if connective == conditional:
            rule = "NE " if rename == "" else "NC "
        elif connective == iff:
            rule = "DE " if rename == "" else "DF "

    rule += definiendum

    return rule


def delete_is_indefinite(new_sentences):
    for i, sent in enumerate(new_sentences):
        if sent[9] == 'J' and abbreviations[0].get(sent[14]) == 'indefinite':
            del new_sentences[i]
            return


def remove_starred_general_variables():
    for i, var in enumerate(variable_type[0]):
        if "*" in var:
            del variable_type[0][i]
            return


def do_not_define(new_sentences, definiendum):
    if definiendum in ['many' + un, 'many' + ud, "few"]:
        for sent in new_sentences:
            if sent[68] in ['122', '121', '221', '222']:
                sent[54] = 'do not define'

    return new_sentences


def replace_propositional_constants(temp_prop_const, prop_unfill, new_sentences, total_dict):
    if prop_unfill == []:
        return new_sentences, {}

    old_prop_to_new_prop = {}
    for k, v in temp_prop_const.items():
        for num in [5, 14, 18]:
            if v[num] != None:
                var = total_dict.get(v[num])
                if var != None:
                    v[num] = var
                else:
                    if v[num] in variables[0]:
                        variables.remove(v[num])
                        total_dict.update({variables[0]: variables[0]})
                    else:
                        total_dict.update({v[num]: variables[0]})
                        v[num] = variables[0]
                        del variables[0]

    for k, v in temp_prop_const.items():
        old_key = k
        v = build_sent2(v)
        abbreviations[0][v[1]] = v[0]
        variable_type[2].append(v[1])
        old_prop_to_new_prop.update({old_key: v[1]})

    for num in prop_unfill:
        i, j = num[0], num[1]
        new_sentences[i][j] = old_prop_to_new_prop.get(new_sentences[i][j])
        assert new_sentences[i][j] != None

    return new_sentences, old_prop_to_new_prop


def add_to_attach_sent(def_info, new_sentences, definition, r_sent_loc, rename, definiendum, type, rule):
    for i in range(len(def_info)):
        list1 = prepare_attach_sent(def_info[i], new_sentences, rule, r_sent_loc)
        if list1[45] == "append to attach_sent list":
            attach_sent.append(list1)
        if i == 0 and type != 'universal':
            add_definitions_to_total_sent(list1, rule, rename, r_sent_loc, definiendum, definition)
        if def_info[i][2] == 'eliminate as conjunct':
            list1[46] = 'eliminate as conjunct'


def add_definitions_to_total_sent(temp_attach_sent, rule, rename, r_sent_loc, definiendum, definition):
    if r_sent_loc != [] or rename == "":
        add_to_total_sent(temp_attach_sent[2], temp_attach_sent[37], temp_attach_sent[4], "", rule)
    else:
        num = temp_attach_sent[2]
        g = findinlist(rule, total_sent, 4, 0)
        if g == None:
            add_to_total_sent(num - 2, definition, "", "", rule)
            add_to_total_sent(num - 1, rename, "", definiendum, "RN")
            add_to_total_sent(num, temp_attach_sent[37], temp_attach_sent[4], "", "SUB", num - 2, num - 1)
        else:
            add_to_total_sent(num - 1, rename, "", definiendum, "RN")
            add_to_total_sent(num, temp_attach_sent[37], temp_attach_sent[4], "", "SUB", g, num - 1)


def build_rename_sent2(constant_map, def_abbrev_dict, old_prop_to_new_prop,
                       indefinite_dict, definiendum, rn_type):
    rename_sent = ""
    j = 0
    for k, v in def_abbrev_dict.items():
        if k != v and definiendum in abbreviations[2]:
            var = abbreviations[2].get(definiendum)
            j += 1
            var = var[0] if j == 1 else var[1]
            rename_sent += "(" + k + idd + var + ") "
            rename_sent += "(" + var + mini_c + v + ") "
        elif k != v:
            rename_sent += "(" + k + mini_c + v + ") "

    for k, v in constant_map.items():
        if k != v:
            rename_sent += "(" + k + idd + v + ") "

    for k, v in indefinite_dict.items():
        if k != v:
            rn = rn_type.get(v)
            rename_sent += "(" + k + idd + v + ")" + rn + " "

    for k, v in old_prop_to_new_prop.items():
        if k != v:
            rename_sent += "(" + k + idd + v + ") "

    return rename_sent


def add_def_sent_to_all_sent(definiendum, list1, new_sentences):
    if definiendum == 'i':
        b = 0
        new_sentences[1][46].remove(1)
        new_sentences[1] = remove_i_from_45(new_sentences[1])
    else:
        b = 1
    spec_relat = ["I", "H"]
    if new_sentences[0][9] in spec_relat:
        add_to_object_properties = False
        defining_variable = new_sentences[0][5]
    else:
        add_to_object_properties = True
        defining_variable = None

    for i in range(b, len(new_sentences)):
        # we cannot add the first_sent to the all sent list since it is already in there
        # if the definiendum is 'i' then we need to add it since it will later be deleted
        if not add_to_object_properties:
            new_sentences[i][73] = defining_variable

        if not isinmdlist(new_sentences[i][1], all_sent, 1) or definiendum == "i":
            if new_sentences[i][74] == 'add to all_sent':
                list1.append(new_sentences[i])


    for i in range(b, len(new_sentences)):
        if not isinmdlist(new_sentences[i][1], all_sent, 1) or definiendum == "i":
            if new_sentences[i][74] == 'add to all_sent':
                all_sent.append(new_sentences[i])






def add_first_sent_to_def_sent(defin_sent, first_sent, r_sent_loc):
    new_sent = copy.deepcopy(first_sent)
    if r_sent_loc == []:
        if new_sent[8] == "~": #negate definiendum
            new_sent[8] = None
            new_sent = build_sent2(new_sent)

    for i in range(len(defin_sent)):
        if defin_sent[i][68][1] == "1":
            greek_name = defin_sent[i][44]
            del defin_sent[i]
            break
    new_sent[44] = greek_name
    defin_sent.insert(0, new_sent)

    return defin_sent


def replace_r_sent(total_dict, r_sent_location, new_sentences, list1, definiendum, def_loc):
    # modify this if you add on new non-word info to the all_sent list
    # right now the only words that have two r sentences in their definition are 'many' and 'only'
    # since the only thing that distinguishes these r sentences is the negation sign we just
    # hard code that the first r sentence is positive and the second is negative
    if r_sent_location == []:
        return new_sentences
    determinative_positions = [3, 10, 16, 20, 24, 28, 32]
    if def_loc in determinative_positions:
        new_var_loc = 14 if def_loc == 10 else def_loc + 2
    else:
        new_var_loc = def_loc

    for j, location in enumerate(r_sent_location):
        r_sent = [None] * 80
        for i in [44, 45, 46, 56, 68]:
            r_sent[i] = new_sentences[location][i]
        for i in [45, 46, 56, 57]:
            r_sent[i] = list1[i]

        if j == 0: new_var = list(total_dict.values())[0]
        # note that the variable of the second R sentence must always be the second member
        # of the total_dict, it is for this reason that items were inserted into the dict
        # in the get new sentences function
        if j == 1: new_var = list(total_dict.values())[1]
        if j == 2: new_var = list(total_dict.values())[2]
        if j == 3: new_var = list(total_dict.values())[3]
        r_sent[new_var_loc] = new_var

        for i in allowable_slots():
            if list1[i] != None:
                if i != new_var_loc and i != def_loc:
                    r_sent[i] = list1[i]
        if definiendum in ['many' + ud, 'many' + un, 'few']:
            r_sent[8] = new_sentences[location][8]
        # elif j == 0:
        #
        #     r_sent[8] = None if list1[2] == "" else list1[2]
        elif j == 1:
            r_sent[8] = "~"

        new_sentences[location] = r_sent

    return new_sentences


def replace_indefinite_variables(new_sentences, unfill_positions, defining_abbreviations,
                                 total_dict, current_sent):
    # modify this if we allow for possessive pronouns with relations other than 'own'

    dict1 = {14: 5, 5: 14}
    instance_of_thing = ""
    indefinite_dict = {}
    rn_type = {}
    k = -1
    while k < len(unfill_positions) - 1:
        k += 1
        if k == 3:
            bb = 8
        i = unfill_positions[k][0]
        j = unfill_positions[k][1]
        m = dict1.get(j)
        new_sentences[i][j]
        general_thing, instance_of_thing = is_a_general_thing(j, new_sentences[i], instance_of_thing)

        for sent in all_sent:
            if sent[1] != current_sent:
                if sent[2] == new_sentences[i][2] and sent[9] == new_sentences[i][9]:
                    if meets_cond_4_indef_replace(new_sentences[i], m, j, defining_abbreviations,
                                                  total_dict, general_thing, sent):
                        n = dict1.get(m)
                        if sent[n] not in indefinite_dict.values():
                            indefinite_dict.update({new_sentences[i][j]: sent[n]})
                            rn_type.update({sent[n]: l1})
                            # print ("indefinite used")
                            new_sentences[i][j] = sent[n]
                            del unfill_positions[k]
                            k -= 1
                            break

    return replace_indefinite_variables2(indefinite_dict, new_sentences, unfill_positions, rn_type)


def meets_cond_4_indef_replace(new_sentence, m, j, defining_abbreviations, total_dict,
                               general_thing, sent):
    if general_thing and sent[9] == "I" and abbreviations[0].get(sent[14]) == 'thing' and \
            (new_sentence[5] == defining_abbreviations[0] or sent[5] == defining_abbreviations[0]):
        return False

    elif general_thing and sent[9] == "I" and abbreviations[0].get(sent[14]) == 'thing' and \
                    new_sentence[5] != defining_abbreviations[0] and sent[5] != defining_abbreviations[0]:
        return False
    elif (sent[m] == new_sentence[m] and sent[j] in defining_abbreviations):
        # the only sentence that uses this is 'the concept cat is itself a cat' and it makes
        # the new sentence y I y
        # print ("type 1")
        return True
    elif defining_abbreviations[0] == sent[m] and not general_thing:
        return True
    else:
        return False


def is_a_general_thing(j, sent, instance_of_thing):
    if j == 5 and sent[9] == "I" and abbreviations[0].get(sent[14]) == 'thing':
        return True, sent[5]
    elif sent[j] == instance_of_thing:
        return True, instance_of_thing
    else:
        return False, instance_of_thing


def replace_indefinite_variables2(indefinite_dict, new_sentences, unfill_positions, rn_type):
    global variable_type
    if unfill_positions == []:
        return new_sentences, indefinite_dict, rn_type

    for pos in unfill_positions:
        i = pos[0]
        j = pos[1]
        new_var = indefinite_dict.get(new_sentences[i][j])
        if new_var != None:
            new_sentences[i][j] = new_var
        else:
            potentially_general = new_sentences[i][j] + "*"
            new_var = variables[0]
            if potentially_general in variable_type[0]:
                var_pos = variable_type[0].index(potentially_general)
                variable_type[0][var_pos] = new_var
            del variables[0]
            variable_type[1].append(new_var)
            indefinite_dict.update({new_sentences[i][j]: new_var})
            rn_type.update({new_var: l2})
            new_sentences[i][j] = new_var

    return new_sentences, indefinite_dict, rn_type


def replace_constants(total_dict, temp_prop_const, defining_abbreviations, new_sentences):
    unfill_positions = []
    prop_unfill = []
    j = -1

    for sent in new_sentences:
        j += 1
        if (sent[68][1:] != "11" or len(sent[68]) != 3) and \
                (sent[68][1:] != "1" or len(sent[68]) != 2) and sent[9] != "R":
            for i in noun_slots():
                if ex(sent, i):
                    new_var = total_dict.get(sent[i])
                    if new_var in defining_abbreviations:
                        sent[71] = 'relevant'
                    if new_var != None:
                        sent[i] = new_var
                        sent[74] = 'add to all_sent'
                    elif sent[i] in temp_prop_const.keys():
                        prop_unfill.append([j, i])
                    else:
                        unfill_positions.append([j, i])

                else:
                    break

    return new_sentences, unfill_positions, prop_unfill


def get_new_sent(def_info, defining_abbreviations, def_loc, list1, definiendum):
    new_sentences = []
    r_sent_location = []
    need_more_definite_abbreviations = False
    pronouns = ['he', 'she', 'i', 'you']
    n = -1
    o = 0
    for j in range(len(def_info[0][3])):

        if os(def_info[0][3][j]):
            sent = copy.deepcopy(def_info[0][3][j])
            sent = categorize_words(space_words(sent))
            n += 1
            new_sentences.append(sent)
            sent[68] = def_info[0][4][j][0]
            sent[44] = def_info[0][6][j]

            if def_info[0][4][j][0][1:] == "1" or def_info[0][4][j][0][1:] == "11":
                def_abbrev_dict, k = map_defining_abbreviations(sent,
                                                                defining_abbreviations,
                                                                def_loc,
                                                                list1,
                                                                definiendum)
            elif def_info[0][4][j][0][1:] == "12":
                need_more_definite_abbreviations = True
            elif sent[9] == "R":
                o += 1
                r_sent_location.append(n)
                # the following key must be first in the dictionary because
                # later we use the first entry in the dictionary for a special purpose

                if definiendum == 'the':
                    concept = list(def_abbrev_dict.values())[0]
                    instance = definite_assignments.get(concept)
                    if instance == "":
                        definite_assignments[concept] = defining_abbreviations[0]
                        temp1 = {sent[k]: defining_abbreviations[0]}
                    else:
                        temp1 = {sent[k]: instance}
                    def_abbrev_dict = dict(temp1, **def_abbrev_dict)
                elif definiendum in pronouns:
                    new_var = abbreviations[1].get(definiendum)
                    if new_var == None:
                        abbreviations[1].update({definiendum: variables[0]})
                        abbreviations[0].update({variables[0]: definiendum})
                        def_abbrev_dict.update({sent[k]: variables[0]})
                        del variables[0]
                    else:
                        def_abbrev_dict.update({sent[k]: new_var})
                elif o == 1:
                    temp1 = {sent[k]: defining_abbreviations[0]}
                    def_abbrev_dict = dict(temp1, **def_abbrev_dict)
                elif o == 2:
                    # for those definitions which have two r sentences
                    # the second r sentence needs another variable
                    # not that the following must be the third variable
                    def_abbrev_dict = insert_into_dict(def_abbrev_dict, {sent[k]: variables[0]}, 1)
                    del variables[0]
                elif o == 3:
                    def_abbrev_dict = insert_into_dict(def_abbrev_dict, {sent[k]: variables[0]}, 2)
                    del variables[0]
                elif o == 4:
                    def_abbrev_dict = insert_into_dict(def_abbrev_dict, {sent[k]: variables[0]}, 3)
                    del variables[0]

    if need_more_definite_abbreviations:
        def_abbrev_dict = map_double_definienda(new_sentences, definiendum, def_abbrev_dict)

    return def_abbrev_dict, r_sent_location, new_sentences, defining_abbreviations


def map_double_definienda(new_sentences, definiendum, def_abbrev_dict):
    # if a definiendum has the form where the subject of the first conjunct matches the
    # object of the second conjunct, then usually it is the subject of the second conjunct
    # which appears in definiens

    antecedent_variables = set()
    consequent_variables = set()

    for j, sent in enumerate(new_sentences):
        for i in [5, 14, 18, 22]:
            if isvariable(sent[i], "i") and sent[68][1] == "1":
                antecedent_variables.add(sent[i])
                if sent[68][1:] == "12":
                    double_definienda_loc = j

            elif isvariable(sent[i], "i") and sent[68][1] == "2":
                consequent_variables.add(sent[i])
    set1 = antecedent_variables.intersection(consequent_variables)
    abbrev_set = set(abbreviations[0].keys())
    def_abbrev_set = set(def_abbrev_dict.keys())
    new_var = list(set1 - abbrev_set.union(def_abbrev_set))
    assert len(new_var) == 1
    for i in [5, 14, 18, 22]:
        if new_sentences[double_definienda_loc][i] == new_var[0]:
            new_var_loc = i
            break
    dict1 = {5: 14, 14: 5}
    other_var_pos = dict1.get(new_var_loc)
    # right now we simply assume that the variable to be searched is the first value in the def_abbrev_dict
    other_var = list(def_abbrev_dict.values())[0]
    double_def_relat = new_sentences[double_definienda_loc][9]
    for sent in all_sent:
        if sent[9] == double_def_relat and sent[other_var_pos] == other_var:
            rn_var = abbreviations[2].get(definiendum)
            if len(rn_var) == 1:
                rn_var.append(variables[0])
                del variables[0]
                abbreviations[2][definiendum] = rn_var
            corresponding_new_var = sent[new_var_loc]
            def_abbrev_dict = insert_into_dict(def_abbrev_dict, {new_var[0]: corresponding_new_var}, 1)
            break
    else:
        # what this means is that if a corresponding variable is not found for this in the
        # replace_r_sent function then the variable that it is replaced with will be a general variable
        variable_type[0].append(new_var[0] + "*")

    return def_abbrev_dict


def map_defining_abbreviations(sent, defining_abbreviations, def_loc, list1, definiendum):
    special_relations = ["I", "J", "=", "H"]
    # when the pronoun is its|a in the object position the subject does the possessing later on
    special_pronoun = "its" + ua
    def_abbrev_dict = {}
    var_loc = 0
    determinative_positions = [3, 10, 16, 20, 24, 28, 32]
    if definiendum == 'part' + up:
        bb = 8

    concept_loc = 14 if def_loc == 10 else def_loc + 2
    if def_loc != 0 and def_loc not in determinative_positions:
        var_loc = get_non_variable_location(sent, definiendum)
    elif def_loc in determinative_positions:
        var_loc = get_non_variable_det_loc(sent, definiendum)
        # this is not the real def abbreviation as in the subject position
        # later it will be changed in the get new sent function
        if definiendum == 'the':
            if list1[concept_loc] not in definite_assignments:
                definite_assignments.update({list1[concept_loc]: ""})
        elif definiendum == special_pronoun:
            def_abbrev_dict.update({sent[5]: list1[5]})
        def_abbrev_dict.update({sent[var_loc]: list1[concept_loc]})
    elif sent[9] in special_relations or sent[14] == None:
        if definiendum not in abbreviations[2]:
            abbreviations[2][definiendum] = [variables[0]]
            del variables[0]
        def_abbrev_dict.update({sent[5]: defining_abbreviations[0]})
    else:
        if definiendum not in abbreviations[2]:
            abbreviations[2][definiendum] = [variables[0], variables[1]]
            del variables[0]
            del variables[1]
        def_abbrev_dict.update({sent[5]: defining_abbreviations[0],
                                sent[14]: defining_abbreviations[1]})

    return def_abbrev_dict, var_loc


def get_non_variable_location(list1, definiendum):
    # this find out what the definining abbreviation is in pronoun definitions
    # for example if the definition is (bR he) <> (bRc) & etc
    # then the defining abbreviations is c
    for i in list1[56]:
        if list1[i] == definiendum:
            return i
    print('you failed to find location of a the definining variable')
    g = 4 / 0


def get_non_variable_det_loc(list1, definiendum):
    determinative_positions = {3: 5, 10: 14}
    for i in list1[56]:
        if list1[i] == definiendum:
            break
    else:
        print('you failed to find location of a the definining variable')
        g = 4 / 0
    j = determinative_positions.get(i)
    return j


def get_abbreviations_from_definition(def_info):
    # this function picks out that variables in the id sentences of the
    # definition

    constants = {}
    constant_map = {}
    temp_propositional_constants = {}
    list3 = []
    for i in range(len(def_info[0])):
        if os(def_info[0][i]) and mini_e in def_info[0][i]:
            list3.append(def_info[0][i])
        elif os(def_info[0][i]) and "=" in def_info[0][i]:
            str1 = def_info[0][i]
            g = str1.find("=")
            var = str1[1:g]
            word = str1[g + 1:-1]
            if isvariable(var):
                if not isvariable(word):
                    constants.update({var: word})

    for k, v in constants.items():
        if v in abbreviations[1]:
            new_var = abbreviations[1].get(v)
            constant_map.update({k: new_var})
        else:
            if k in variables:
                abbreviations[0].update({k: v})
                abbreviations[1].update({v: k})
                constant_map.update({k: k})
                variables.remove(k)
            else:
                new_var = variables[0]
                del variables[0]
                abbreviations[0].update({new_var: v})
                abbreviations[1].update({v: new_var})
                constant_map.update({k: new_var})

    constant_map.update({"i": "i"})
    temp_propositional_constants = get_propositional_constants(list3, temp_propositional_constants)

    return constant_map, temp_propositional_constants


def get_propositional_constants(list3, temp_propositional_constants):
    if list3 == []:
        return {}

    for i in range(len(list3)):
        prop_con = list3[i][1]
        str2 = list3[i].replace(" ", "")
        str2 = str2[3:-1]
        list8 = space_words(str2)
        list8 = categorize_words(list8)
        temp_propositional_constants.update({prop_con: list8})

    return temp_propositional_constants


def add_necessary_conditions_for_concept(loop_number):
    # if we're talking about concepts in our proof then we need to add their necesssary
    # conditiona to our proof
    if loop_number != 1:
        return
    global sn
    list2 = []
    con_sent_parts = [None] * 80
    for abbrev, word in abbreviations[0].items():
        if word == 'concept' + un or word == 'concept' + ua:
            str1 = abbrev

            for j in range(len(all_sent)):
                if all_sent[j][9] == "I" and all_sent[j][14] == str1:
                    str2 = all_sent[j][5]
                    concept = abbreviations[0].get(str2)
                    if concept != None:
                        pos = dictionary[0].get(concept)
                        pos = pos[0]

                        if concept == "dog":
                            bb = 8
                        if pos == 'a':
                            str4 = "J"
                        elif pos == 'n':
                            str4 = "I"
                        b = 0
                        for k in range(len(all_sent)):
                            if all_sent[k][9] == str4 and all_sent[k][14] == str2 and \
                                            str1 != str2 and str2 not in list2 and all_sent[k][8] == None:
                                str6 = all_sent[k][5]
                                list2.append(str2)
                                b += 1
                        if b > 1:
                            print('you have not coded for multiple concepts')
                        olda = "(" + "b" + ' = ' + concept + ")"
                        oldc = "(" + "c " + str4 + " b" + ")"
                        if str2 != "b":
                            rn1 = "(" + "b" + idd + str2 + ") (" + "c" + idd + str6 + ")" + l1
                        else:
                            rn1 = "(" + "c" + idd + str6 + ")" + l1 + " "
                        nat_sent_b4_sub = olda + " " + conditional + " " + oldc
                        sn += 1
                        add_to_total_sent(sn, nat_sent_b4_sub, "", "", "NC concept " + concept)
                        sn += 1
                        add_to_total_sent(sn, rn1, "", "concept " + concept, "RN")
                        ant_sent_parts = build_sent1(str2, "", "=", concept)
                        con_sent_parts[5] = str6
                        con_sent_parts[9] = str4
                        con_sent_parts[14] = str2
                        prepare_att_sent_1_sent(ant_sent_parts, "SUB", conditional,
                                                [build_sent2(con_sent_parts)], sn, sn - 1)
                        break



def build_sent1(subj, tvalue, relat, obj, relat2="", obj2=""):
    list1 = [None] * 80
    tvalue = " ~ " if tvalue == "~" else " "

    if get_words_used == 1:
        for word in [subj, relat, obj, relat2, obj2]:
            if word not in words_used and not isvariable(word):
                words_used.append(word)

    list1[5] = subj
    list1[9] = relat
    list1[14] = obj
    if relat2 == "" or relat2 == None:
        sent = "(" + subj + tvalue + relat + " " + obj + ")"
        sent_abs = "(" + subj + " " + relat + " " + obj + ")"
    else:
        sent = "(" + subj + tvalue + relat + " " + obj + " " + relat2 + " " + obj2 + ")"
        sent_abs = "(" + subj + " " + relat + " " + obj + " " + relat2 + " " + obj2 + ")"

    abbrev_sent = name_sent(sent_abs)
    tvalue = "~" if tvalue == " ~ " else ""
    list1[8] = tvalue
    list1[0] = sent
    list1[72] = sent_abs
    list1[1] = abbrev_sent
    list1[2] = tvalue
    list1[42] = tvalue + abbrev_sent
    list1[46] = []

    return list1


def build_sent2(list1, type=0):
    # if you revise this list then then you must also revise it in
    # the eliminate_univ_quant_subclause, extract_words_from_subclause, as well as the function 'that', as well as new_categories
    # g=1 means that it is a sentence that identifies a propositional constant, in some cases
    # the proposition itself need not be named
    # also fix list in word sub and isatomic

    str1 = "("
    for i in allowable_slots(type):
        if list1[i] != None and list1[i] != "" and list1[i] != " ":
            if get_words_used == 1:
                if list1[i] not in words_used and not isvariable(list1[i]):
                    words_used.append(list1[i])
            if str1 == "(":
                str1 += list1[i]
            else:
                str1 += " " + list1[i]

    str1 += ")"
    str1p = name_sent(str1)
    list1[0] = str1
    list1[2] = "~" if "~" in str1p else ""
    list1[1] = str1p.replace("~", "") if "~" in str1p else str1p
    list1[72] = str1.replace("~", "") if "~" in str1 else str1
    list1[72] = list1[72].replace("  ", " ")
    list1[42] = str1p

    return list1


def build_temp_sent(list1):
    # The only difference between this and build_sent2 is that it does not abbreviate the
    # sentence with a single letter

    str1 = "("

    for i in allowable_slots():
        if list1[i] != None and list1[i] != "" and list1[i] != " ":
            if get_words_used == 1:
                if list1[i] not in words_used and not isvariable(list1[i]):
                    words_used.append(list1[i])

            if str1 == "(":
                str1 += list1[i]
            else:
                str1 += " " + list1[i]

    str1 += ")"
    list1[0] = str1
    return list1


def build_sent3(list1):
    str1 = "("

    for i in allowable_slots():
        temp_str = list1[i]
        if temp_str != None and temp_str != "":
            if get_words_used == 1:
                if list1[i] not in words_used and not isvariable(list1[i]):
                    words_used.append(list1[i])

            if str1 == "(":
                str1 += temp_str
            else:
                str1 += " " + temp_str

    str1 += ")"

    return str1


def standard_slots():
    return [5, 9, 14, 15, 18, 19, 22, 23, 26, 27, 30, 31]


def standard_slots_w_neg():
    return [5, 8, 9, 14, 49, 15, 18, 50, 19, 22, 51, 23, 26, 52, 27, 30, 31]


def build_sent_standard(list1):
    list2 = []
    for i in standard_slots():
        if not ex(list1, i):
            break
        elif list1[i] != None:
            list2.append(list1[i])

    return "(" + "".join(list2) + ')'


def build_sent_slots_known(list1):
    str1 = "("
    str2 = "("
    tvalue = ""
    for i in list1[56]:
        if get_words_used == 1:
            if list1[i] not in words_used and not isvariable(list1[i]):
                words_used.append(list1[i])

        if str1 == "(":
            str1 += list1[i]

        else:
            str1 += " " + list1[i]
        if str2 == "(":
            str2 += list1[i]
        elif list1[i] != "~":
            str2 += " " + list1[i]
        else:
            tvalue = "~"
    str1 += ")"
    str2 += ")"
    list1[0] = str1
    list1[2] = tvalue
    list1[72] = str2
    list1[1] = name_sent(str2)
    list1[42] = tvalue + list1[1]

    return list1


def build_uncategorized_sent(list1):
    for i in range(3, len(list1)):
        if list1[i] == None or list1[i] == "":
            break
        if i == 3:
            str1 = list1[i]
        else:
            str1 += " " + list1[i]

    str1 = "(" + str1 + ")"
    str1p = name_sent(str1)
    list1[0] = str1
    list1[2] = ""
    list1[1] = str1p
    list1[72] = str1
    list1[42] = str1p

    return list1


def build_sent_list(list1):
    # this list builder does not have the ~ separated from the sentence

    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i]
        else:
            str2 = str2 + ' & ' + list1[i]

    return str2


def build_sent_list2(list1, j):
    # this list builder does not have the ~ separated from the sentence

    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i][j]
        else:
            str2 = str2 + ' & ' + list1[i][j]

    return str2


def use_rarely_defined_word():
    rarely_defined = copy.deepcopy(dictionary[6])
    for word in abbreviations[1].keys():
        if word in rarely_defined:
            rarely_defined.remove(word)

    return rarely_defined


def name_sent(str1, bool2=False, str4=""):
    no_space = copy.copy(str1)
    if str1.find('~') > -1:
        no_space = str1.replace("~", "")
        str1 = str1.replace("~", "")
        ng = '~'
    else:
        ng = ''

    if "  " in str1:
        str1 = str1.replace("  ", " ")

    no_space = remove_outer_paren(no_space)
    no_space = no_space.replace(" ", "")

    if bool2:
        if str4 == 'something':
            no_space = no_space.replace("something", "some thing")
        elif str4 == 'anything':
            no_space = no_space.replace("anything", "any thing")
        elif str4 == 'everything':
            no_space = no_space.replace("everything", "every thing")
        elif str4 == 'anything' + ua:
            no_space = no_space.replace("anything" + ua, "a" + ua + " thing")
        elif str4 == 'anything' + un:
            no_space = no_space.replace("anything" + un, "any" + un + " thing")

    h = findinlist(no_space, prop_name, 1, 0)
    if h != None:
        return ng + h
    else:
        prop_name.append([prop_var[0], no_space, str1])
        str2 = prop_var[0]
        del prop_var[0]
        return ng + str2


def insert_space(str, integer):
    return str[0:integer] + ' ' + str[integer:]


def space_words(str1):
    # this function place a space between variables and relations
    # so as to make it easier to categorize words in a sentence

    str1 = str1.replace("(", "")
    str1 = str1.replace(")", "")
    str1 = str1.replace("~", " ~ ")
    str1 = str1.replace("=", " = ")
    str1 = str1.replace(neg, " " + neg + " ")
    str1 = str1.replace(mini_e, " " + mini_e + " ")
    str1 = str1.replace(ne, " " + ne + " ")
    i = -1
    while i + 1 < len(str1):
        i += 1
        temp_str = str1[i:(i + 1)]
        nxt_str = str1[(i + 1):(i + 2)]
        if nxt_str.isupper() == True and temp_str.islower() == True:
            str1 = insert_space(str1, i + 1)
        elif nxt_str.islower() == True and temp_str.isupper() == True:
            str1 = insert_space(str1, i + 1)

    list1 = prepare_categorize_words(str1)
    list1[79] = 'is in definition'

    return list1


def isdefineable(list1):
    must_be_blank = [3, 4, 6, 7, 10, 11, 13, 16, 17, 18, 20, 21, 23, 24, 25, 27, 28, 29, 31, 32, 33,
                     35, 36, 49, 50, 51, 52, 55]
    must_be_variable = [5, 14, 18, 22]

    if list1[9] == "=": must_be_variable.remove(14)

    for i in must_be_blank:
        if list1[i] != None and list1[i] != '':
            return False
    for i in must_be_variable:
        if list1[i] != None:
            if not isvariable(list1[i], "i"):
                return False
    return True


def is_standard(list1):
    must_be_blank = [3, 4, 6, 7, 10, 11, 13, 16, 17, 20, 21, 23, 24, 25, 27, 28, 29, 31, 32, 33,
                     35, 36, 49, 50, 51, 52, 55, 59, 60, 66, 67, 69, 70]

    prepositional_relation = ['INB', "ATC", 'IN']

    for i in [15, 19, 23, 27, 31]:
        if list1[i] != None:
            assert list1[i] != ""
            if list1[i] not in prepositional_relation:
                return False
        else:
            break

    for i in must_be_blank:
        if list1[i] != None and list1[i] != '':
            return False
    for i in [5, 14, 18, 22, 26, 30, 34]:
        if list1[i] != None and list1[i] != "":
            if not isvariable(list1[i], "i") and not list1[9] == "=" and not i == 14:
                return False
    return True


def new_relevant_variables(list1):
    for i in noun_slots():
        if not ex(list1, i):
            break
        else:
            if list1[i] not in variable_type[3]:
                variable_type[3].append(list1[i])


def is_relevant(list1):
    abbrev_definite = abbreviations[1].get("definite")
    if list1[71] == 'relevant':
        new_relevant_variables(list1)
        return True
    if list1[9] == 'J' and list1[14] == abbrev_definite:
        return False
    for i in standard_slots():
        if list1[i] in variable_type[3]:
            return True
    return False


def check_mispellings(test_sent):
    if proof_type != 3:
        return
    global prop_name, total_sent, all_sent, attach_sent, detach_sent, prop_var, sn, abbreviations
    for k in order:
        print(k)
        prop_name = []
        total_sent = []
        all_sent = []
        attach_sent = []
        detach_sent = []
        abbreviations = [{}, {}, {}]
        prop_var = copy.deepcopy(prop_var4)
        sn = test_sent[k][-1][0] + 1

        divide_sent(test_sent[k])

        eliminate_redundant_words()
    sys.exit()


def obtain_truth_value(sent):
    sentence = tran_str(sent[1])
    add_to_total_sent("", "CLAIM " + str(sent[0]) + ": " + sentence)
    add_to_total_sent("", "")

    if sentence[7:12] == 'consi':
        return True, sentence[len("It isa consistent that "):]
    elif sentence[7:12] == 'contr':
        return False, sentence[len("It isa contradictory that "):]
    else:
        # tahir system exit
        print ("Each sentence must begin with either 'it is|a consistent that' or 'it is|a contradictory that'")


def eliminate_logical_connectives(sentence):
    return sentence.split(" and ")


def step_one(sent):
    global time_spent_reducing, all_sent
    truth_value, sentence = obtain_truth_value(sent)

    set_of_sentences = eliminate_logical_connectives(sentence)

    divide_sent(set_of_sentences)

    eliminate_redundant_words()

    replace_determinative_nouns()

    replace_synonyms()

    replace_special_synonyms()

    word_sub()

    eliminate_negative_determiners()

    transfer_negation_signs()

    aa = time.time()

    all_sent = remove_duplicates(all_sent, 0)

    all_sent = define_irregular_terms(all_sent)

    time_spent_reducing += (time.time() - aa)

    return truth_value


def is_linked_to_rare_word(word):
    global dictionary
    dict1 = {"individual" + ua: "individual"}
    linked_word = dict1.get(word)
    if word in dictionary[6]:
        temp = copy.deepcopy(dictionary[6])
        temp.remove(word)
        dictionary[6] = temp
    elif linked_word != None:
        temp = copy.deepcopy(dictionary[6])
        temp.remove(linked_word)
        dictionary[6] = temp


def divide_sent(list2):
    global sn
    sn = 0
    for str2 in list2:
        sn += 1
        str2 = str2.lower()
        if "'s" not in str2:
            str2 = str2.replace("'", "")
        str3 = name_sent(str2)
        str2 = str2.strip()
        words_in_sent = str2.split()
        sent_parts = [None] * 80
        sent_parts[0] = str2
        sent_parts[42] = str3
        sent_parts[72] = str2
        sent_parts[1] = str3
        sent_parts[2] = ""
        sent_parts[58] = sn
        for j in range(len(words_in_sent)):
            sent_parts[j + 3] = words_in_sent[j]
            is_linked_to_rare_word(words_in_sent[j])
        list4 = copy.deepcopy(sent_parts)
        detach_sent.append(sent_parts)
        add_to_total_sent(sn, str2, str3, "", "")
        all_sent.append(list4)
    sn += 1


def eliminate_redundant_words():
    # modify this if we start dealing with sentences longer than 41 words
    global all_sent
    bool1 = False
    for i in range(len(all_sent)):
        ant_sent_parts = copy.deepcopy(all_sent[i])
        rule = ""
        j = 2
        while j < 55:
            j += 1
            if all_sent[i][j] == None:
                break
            pos = dictionary[0].get(all_sent[i][j])
            if pos != None and pos[0] == 't':
                if get_words_used == 1:
                    if not isvariable(all_sent[i][j]) and all_sent[i][j] not in words_used:
                        words_used.append(all_sent[i][j])
                bool1 = True
                if rule == '':
                    rule += all_sent[i][j]
                else:
                    rule += "," + all_sent[i][j]
                del all_sent[i][j]
                j -= 1
                # this means that sentences must be shorter than 40 words
                all_sent[i].insert(40, None)
        if bool1:
            bool1 = False
            all_sent[i] = build_uncategorized_sent(all_sent[i])
            con_parts = copy.deepcopy(all_sent[i])
            prepare_att_sent_1_sent(ant_sent_parts, "RD " + rule, iff, [con_parts])

    for i, sent in enumerate(all_sent):
        sent[79] = 'do not rebuild sentence'
        sent2 = copy.deepcopy(sent)
        all_sent[i] = categorize_words(sent)
        for j in [0, 1, 2, 42, 72]:
            all_sent[i][j] = sent2[j]


def replace_determinative_nouns():
    global sn
    m = -1
    while m < len(all_sent) - 1:
        m += 1
        replacement_made = False
        while all_sent[m][45][19] != []:
            does_not_affect_decision_procedure = False
            ant_sent_parts = copy.deepcopy(all_sent[m])
            i = all_sent[m][45][19][0]
            rule = "DF " + all_sent[m][i]
            synonym = dictionary[2].get(all_sent[m][i])
            determinative = synonym[:synonym.find(" ")]
            definition = dictionary[1].get(all_sent[m][i])
            noun = synonym[synonym.find(" ") + 1:]
            determinative.strip()
            noun.strip()
            replacement_made = True
            j = 10 if i == 14 else i - 2
            all_sent[m][j] = determinative
            all_sent[m][i] = noun
            if determinative == 'every' or determinative == 'no':
                b = 15
            elif determinative == 'a':
                b = 1
            else:
                does_not_affect_decision_procedure = True
            all_sent[m][56].insert(all_sent[m][56].index(i), j)
            all_sent[m][45][19].remove(i)
            if not does_not_affect_decision_procedure:
                all_sent[m][45][b].append(j)
                all_sent[m][46].append(b)
        if replacement_made:
            all_sent[m] = build_sent_slots_known(all_sent[m])
            con_sent_parts = copy.deepcopy(all_sent[m])
            g = findinlist(rule, total_sent, 4, 0)
            if g == None:
                sn += 1
                add_to_total_sent(sn, definition, "", "", rule)
            prepare_att_sent_1_sent(ant_sent_parts, "SUB", iff, [con_sent_parts], "")


def replace_synonyms():
    global sn
    definitions_added = []
    m = -1
    while m < len(all_sent) - 1:
        m += 1
        replacement_made = False
        while all_sent[m][45][18] != []:
            ant_sent_parts = copy.deepcopy(all_sent[m])
            i = all_sent[m][45][18][0]
            synonym = dictionary[2].get(all_sent[m][i])
            assert synonym != None
            recategorize_word(synonym, m, i)
            definition = dictionary[1].get(all_sent[m][i])
            if definition not in definitions_added:
                definitions_added.append(definition)
                sn += 1
                add_to_total_sent(sn, definition, "", "", "DF " + all_sent[m][i])
            replacement_made = True
            all_sent[m][i] = synonym
            all_sent[m][45][18].remove(i)
        if replacement_made:
            all_sent[m] = build_sent_slots_known(all_sent[m])
            con_parts = copy.deepcopy(all_sent[m])
            prepare_att_sent_1_sent(ant_sent_parts, "SUB", iff, [con_parts], "")


def recategorize_word(synonym, m, i):
    # because we replace a word with a synonym we need to know it decision procedure
    # for elimination
    part_of_speech_syn = dictionary[0].get(synonym)
    pos = part_of_speech_syn[0]
    sub_pos = part_of_speech_syn[1] if len(part_of_speech_syn) > 1 else ""
    sub_sub_pos = part_of_speech_syn[1] if len(part_of_speech_syn) > 1 else ""
    b = get_used_slots(i, pos, sub_pos, pos)
    if b != 0:
        all_sent[m][45][b].append(i)
        all_sent[m][46].append(b)


def replace_special_synonyms():
    global sn
    m = -1
    while m < len(all_sent) - 1:
        m += 1
        replacement_made = False
        while all_sent[m][45][20] != []:
            ant_sent_parts = copy.deepcopy(all_sent[m])
            i = all_sent[m][45][20][0]
            rule = 'DE ' + all_sent[m][i]
            replace_special_synonyms2(m, i)
            replacement_made = True
            all_sent[m][45][20].remove(i)
        if replacement_made:
            all_sent[m] = build_sent2(all_sent[m])
            con_parts = copy.deepcopy(all_sent[m])
            prepare_att_sent_1_sent(ant_sent_parts, rule, iff, [con_parts], "")


def replace_special_synonyms2(m, i):
    dict1 = {9: 8, 15: 49, 50: 19, 51: 23, 52: 27}
    if all_sent[m][i] == 'distinct from':
        all_sent[m][i] = "is"
        all_sent[m][dict1.get(i)] = "~"


def word_sub():
    global sn
    relational_positions = [9, 15, 19, 23, 27, 31]
    m = -1
    n = len(all_sent)
    while m < n - 1:
        m += 1
        replacement_made = False
        while all_sent[m][45][0] != []:
            ant_sent_parts = copy.deepcopy(all_sent[m])
            k = all_sent[m][45][0][0]
            str2 = all_sent[m][k]
            if str2 == "not":
                all_sent[m][k] = "~"
                replacement_made = True
            elif k == 69 or k == 70:
                replacement_made = True
                str2 = str2[:-2]
                replace_word_w_variable(m, k, str2)
            elif k in relational_positions:
                relat = dictionary[3].get(str2)
                assert relat != None
                replacement_made = True
                abbreviations[0].update({str2: relat})
                all_sent[m][k] = relat
            else:
                replacement_made = True
                replace_word_w_variable(m, k, str2)
            all_sent[m][45][0].remove(k)

        if replacement_made:
            all_sent[m] = build_sent2(all_sent[m])
            con_parts = copy.deepcopy(all_sent[m])
            prepare_att_sent_1_sent(ant_sent_parts, "SUB", iff, [con_parts])


def replace_word_w_variable(m, k, str2):
    not_normally_defined = dictionary[6]
    if isvariable(str2) == False:
        if str2 in not_normally_defined:
            not_normally_defined.remove(str2)
        str3 = abbreviations[1].get(str2)
        if str3 == None:
            pos = dictionary[0].get(str2)
            if len(pos) > 1 and pos[1] == "u":
                list1 = build_sent1(variables[0], "", "=", str2)
                list1[46] = []
                list1[56] = [5, 9, 14]
                all_sent.append(list1)
                detach_sent.append(list1)
            if k == 69 or k == 70:
                all_sent[m][k] = variables[0] + "'s"
            else:
                all_sent[m][k] = variables[0]
            abbreviations[0].update({variables[0]: str2})
            abbreviations[1].update({str2: variables[0]})
            del variables[0]
        elif k == 69 or k == 70:
            all_sent[m][k] = str3 + "'s"
        else:
            all_sent[m][k] = str3


def eliminate_negative_determiners():
    # modify this if the category number of the universals change
    # modify this if we allow for two negative determiners in a sentence

    special_determinatives = ['a', 'every', 'many' + un, 'any' + un, 'many' + ud, 'few']
    for sent in all_sent:
        if sent[45] != None:
            for j in sent[45][17]:
                make_new_sentence = True
                if j == 8 and sent[10] in special_determinatives:
                    position = 10
                elif j == 49 and sent[16] in special_determinatives:
                    position = 16
                elif j == 50 and sent[20] in special_determinatives:
                    position = 20
                elif j == 51 and sent[24] in special_determinatives:
                    position = 24
                elif j == 52 and sent[28] in special_determinatives:
                    position = 28
                elif j == 47 and sent[3] in special_determinatives:
                    position = 3
                else:
                    make_new_sentence = False
                if make_new_sentence:
                    ant_sent_parts = copy.deepcopy(sent)
                    sent[j] = None
                    sent[56].remove(j)
                    if sent[position] == 'every':
                        sent[position] = 'many' + un
                        sent[45][1].append(position)
                        sent[45][15].remove(position)
                        sent[46].append(1)
                        sent[46].remove(15)
                        rule = "DE ~ every"
                    else:
                        rule = "DE ~ " + sent[position]
                        if sent[position] == 'many' + ud:
                            sent[46].remove(16)
                            sent[45][16].remove(position)
                        elif sent[position] not in ['any' + un, 'many' + un]:
                            sent[46].remove(1)
                            sent[45][1].remove(position)
                        if sent[position] == 'many' + un:
                            sent[position] = 'few'
                        else:
                            sent[position] = 'no'
                            sent[45][15].append(position)
                            sent[46].append(15)

                    sent = build_sent_slots_known(sent)
                    con_parts = copy.deepcopy(sent)
                    prepare_att_sent_1_sent(ant_sent_parts, rule, iff, [con_parts])


def transfer_negation_signs():
    m = -1
    while m < len(all_sent) - 1:
        m += 1
        if all_sent[m][15] != None and all_sent[m][60] == None and all_sent[m][8] == "~":
            ant_sent_parts = copy.deepcopy(all_sent[m])
            all_sent[m][49] = "~"
            all_sent[m][8] = None
            all_sent[m] = build_sent2(all_sent[m])
            con_parts = copy.deepcopy(all_sent[m])
            prepare_att_sent_1_sent(ant_sent_parts, "SNR", iff, [con_parts])


def quick_negation_transfer(list1):
    if list1[15] != None and list1[60] == None and list1[8] == "~":
        list1[49] = "~"
        list1[8] = None
    return list1


def is_adj_definite(list1, i):
    # modify this if more indefinite determinatives are added
    indefinite_determinatives = ['a']
    determinative_position = 10 if i == 10 else i - 1
    if list1[determinative_position] in indefinite_determinatives:
        return False
    else:
        return True


def lies_wi_subclause(word_pos, location):
    begin = location[0]
    end = location[1]
    if word_pos > begin and word_pos < end:
        return True
    else:
        return False


def relative_pronoun_lies_in_scope_univ(list1, word_pos, univ_pos, current_universal):
    bool1 = False
    absolute_univ_pos = allowable_slots().index(univ_pos)
    univ_in_subclause = lies_wi_subclause(absolute_univ_pos, list1[57])
    rel_pro_in_sub_clause = lies_wi_subclause(word_pos, list1[57])

    if not univ_in_subclause:
        if absolute_univ_pos < word_pos:
            bool1 = True
    elif univ_in_subclause and rel_pro_in_sub_clause:
        if absolute_univ_pos < word_pos:
            bool1 = True

    return bool1


def relation_lies_wi_scope_univ(list1, univ_pos, word_pos):
    absolute_univ_pos = allowable_slots().index(univ_pos)
    univ_in_subclause = lies_wi_subclause(absolute_univ_pos, list1[57])
    relat_in_subclause = lies_wi_subclause(word_pos, list1[57])
    bool1 = False

    if not univ_in_subclause:
        bool1 = True
    elif univ_in_subclause and relat_in_subclause:
        bool1 = True

    return bool1


def adjective_lies_wi_scope_of_univ(list1, i, current_universal, word_pos, univ_pos):
    # modify this is we allow sentences with more than 1 determinative

    absolute_univ_pos = allowable_slots().index(univ_pos)
    univ_in_subclause = lies_wi_subclause(absolute_univ_pos, list1[57])
    adj_in_subclause = lies_wi_subclause(word_pos, list1[57])
    is_definite = is_adj_definite(list1, i)

    bool1 = False
    if i == 13 and univ_pos == 10:
        bool1 = True
    elif i == univ_pos + 1:
        bool1 = True
    elif current_universal == 'no' and not is_definite:
        if not univ_in_subclause:
            if absolute_univ_pos < word_pos:
                bool1 = True
        else:
            if adj_in_subclause and absolute_univ_pos < word_pos:
                bool1 = True

    return bool1


def determ_lies_wi_scope_of_univ(list1, i, current_universal, word_pos, univ_pos):
    # modify this if we increase the number of indefinite determinatives
    indefinite_determinatives = ['a', 'many' + un, "any" + un, "a" + ud, 'few']
    bool1 = False
    univ_pos = allowable_slots().index(univ_pos)
    univ_in_sub_clause = lies_wi_subclause(univ_pos, list1[57])
    determ_in_sub_clause = lies_wi_subclause(word_pos, list1[57])
    if list1[i] in indefinite_determinatives and current_universal == 'no':
        if univ_in_sub_clause and determ_in_sub_clause:
            if univ_pos < word_pos:
                bool1 = True
        elif not univ_in_sub_clause and univ_pos < word_pos:
            bool1 = True
    elif list1[i] in indefinite_determinatives and current_universal == 'every':
        if not univ_in_sub_clause and determ_in_sub_clause:
            if univ_pos < word_pos:
                bool1 = True



    return bool1


def determ_lies_wi_scope_of_univ2(list1, i):
    # modify this if we increase the number of indefinite determinatives
    indefinite_determinatives = ['a', 'many' + un, "any" + un, "a" + ud, 'few']
    bool1 = False
    for j in determ_slots():
        if list1[j] in indefinite_determinatives:
            determ_position = j
            univ_pos = allowable_slots().index(i)
            univ_in_sub_clause = lies_wi_subclause(univ_pos, list1[57])
            determ_in_sub_clause = lies_wi_subclause(determ_position, list1[57])
            if list1[i] == 'no':
                if univ_in_sub_clause and determ_in_sub_clause:
                    if univ_pos < determ_position:
                        bool1 = True
                elif not univ_in_sub_clause and not determ_in_sub_clause and \
                                univ_pos < determ_position:
                    bool1 = True
                elif not univ_in_sub_clause and determ_in_sub_clause:
                    dict1 = {3: 59, 10: 60, 16: 61, 20: 62}
                    k = dict1.get(i)
                    if list1[k] != None:
                        bool1 = True
            # elif list1[i] == 'every':
            #     print ('hey')
            #     dict1 = {3: 59, 10: 60, 16: 61, 20: 62}
            #     k = dict1.get(i)
            #     if list1[k] != None:
            #         bool1 = True


    return bool1


def lies_wi_scope_of_univ_quant(list1, i):
    # modify this if you allow for more than one universal quantifier in a sentence
    # or you allow for more than one subclause
    # or you increase the number of determinatives

    bool1 = False
    if 15 in list1[46]:# 15 means there is a universal quantifier in the sentence

        adjective_positions = [4, 13, 17, 21, 25, 29, 33]
        determinative_positions = [3, 10, 16, 20, 24, 28, 32]
        relation_positions = [15, 19, 23, 27, 31]
        relative_pronoun_positions = [59, 60, 61, 62]
        univ_pos = list1[45][15][0]
        current_universal = list1[univ_pos]
        word_pos = allowable_slots().index(i)

        if i in adjective_positions:
            bool1 = adjective_lies_wi_scope_of_univ(list1, i, current_universal, word_pos, univ_pos)
        elif i in determinative_positions:
            bool1 = determ_lies_wi_scope_of_univ(list1, i, current_universal, word_pos, univ_pos)
        elif i in relation_positions:
            bool1 = relation_lies_wi_scope_univ(list1, univ_pos, word_pos)
        elif i in relative_pronoun_positions:
            bool1 = relative_pronoun_lies_in_scope_univ(list1, word_pos, univ_pos, current_universal)

    return bool1


def get_subclause_position(list1):
    begin = 0
    end = 0
    for i in [59, 60, 61, 62]:
        if list1[i] != None:
            begin = allowable_slots().index(i)

            break
    if begin != 0:
        if list1[39] == None:
            end = 55
        else:
            end = allowable_slots().index(list1[39])

    return [begin, end]


def prepare_att_sent_4_sent(ant_sent_parts, consequent, connective, rule):
    global sn
    con_parts1 = consequent[0]
    con_parts2 = consequent[1]
    con_parts3 = consequent[2]
    con_parts4 = consequent[3]

    sn += 1
    list4 = [""] * 60
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts1[68] = "121"
    con_parts1[44] = chr(950)
    con_parts2[68] = "122"
    con_parts2[44] = chr(951)
    con_parts3[68] = "123"
    con_parts3[44] = chr(952)
    con_parts4[68] = "124"
    con_parts4[44] = chr(953)
    sent_type = "e" if connective == iff else "c"

    new_equivalence = ant_sent_parts[0] + " " + connective + " (" + con_parts1[0] + \
                      " & " + con_parts2[0] + " & " + con_parts3[0] + " & " + con_parts4[0] + ")"
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " (" + con_parts1[42] + \
                    " & " + con_parts2[42] + " & " + con_parts3[42] + " & " + con_parts4[42] + ")"
    new_greek = chr(949) + " " + connective + " (" + chr(950) + " & " + chr(951) + \
                " & " + chr(952) + " & " + chr(953) + ")"
    if connective == 'e':
        ant_sent_parts[53] = 'b'
        con_parts1[53] = 'cf'
        con_parts2[53] = 'cf'
        con_parts3[53] = 'cf'
        con_parts4[53] = 'cf'
    else:
        ant_sent_parts[53] = 'a'
        con_parts1[53] = 'cq'
        con_parts2[53] = 'cq'
        con_parts3[53] = 'cq'
        con_parts4[53] = 'cq'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts1[1], con_parts1[2]], [con_parts2[1], con_parts2[2]],
                [con_parts3[1], con_parts3[2]], [con_parts4[1], con_parts4[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = ["(" + con_parts1[42] + " & " + con_parts2[42] + \
                " & " + con_parts3[42] + " & " + con_parts4[42] + ")", ""]
    list4[34] = [ant_sent_parts]
    list4[35] = [con_parts1, con_parts2, con_parts3, con_parts4]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts1[1], con_parts2[1], con_parts3[1], con_parts4[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = ["(" + con_parts1[0] + " & " + con_parts2[0] + " & " + con_parts3[0] + \
                 " & " + con_parts4[0] + ")", ""]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts1[72], con_parts1[2]], [con_parts2[72], con_parts2[2]],
                 [con_parts3[72], con_parts3[2]], [con_parts4[72], con_parts4[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule)
    attach_sent.append(list4)


def prepare_att_sent_3_sent(ant_sent_parts, connective, consequent, rule):
    # this populates the attach_sent list provided a sentence is equivalent
    # to two conjuncts

    global sn
    sn += 1
    con_parts = consequent[0]
    con_parts2 = consequent[1]
    con_parts3 = consequent[2]
    list4 = [""] * 60
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts[68] = "121"
    con_parts[44] = chr(950)
    con_parts2[68] = "122"
    con_parts2[44] = chr(951)
    con_parts3[68] = "123"
    con_parts3[44] = chr(952)
    sent_type = "e" if connective == iff else "c"
    new_equivalence = ant_sent_parts[0] + " " + connective + " (" + con_parts[0] + \
                      " & " + con_parts2[0] + " & " + con_parts3[0] + ")"
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " (" + con_parts[42] + \
                    " & " + con_parts2[42] + " & " + con_parts3[42] + ")"
    new_greek = chr(949) + " " + connective + " (" + chr(950) + " & " + chr(951) + \
                " & " + chr(952) + ")"
    if sent_type == 'e':
        ant_sent_parts[53] = 'b'
        con_parts[53] = 'cf'
        con_parts2[53] = 'cf'
        con_parts3[53] = 'cf'
    else:
        ant_sent_parts[53] = 'a'
        con_parts[53] = 'cq'
        con_parts2[53] = 'cq'
        con_parts3[53] = 'cq'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts[1], con_parts[2]], [con_parts2[1], con_parts2[2]],
                [con_parts3[1], con_parts3[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = ["(" + con_parts[42] + " & " + con_parts2[42] + " & " + con_parts3[42] + ")", ""]
    list4[34] = [copy.deepcopy(ant_sent_parts)]
    list4[35] = [copy.deepcopy(con_parts), copy.deepcopy(con_parts2), copy.deepcopy(con_parts3)]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts[1], con_parts2[1], con_parts3[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = ["(" + con_parts[0] + " & " + con_parts2[0] + " & " + con_parts3[0] + ")", ""]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts[72], con_parts[2]], [con_parts2[72], con_parts2[2]],
                 [con_parts3[72], con_parts3[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule)
    attach_sent.append(list4)


def prepare_att_sent_2_sent(ant_sent_parts, connective, consequent, rule):
    # this populates the attach_sent list provided a sentence is equivalent
    # to two conjuncts

    global sn
    sn += 1
    if rule == "RDC": connective = conditional
    con_parts = consequent[0]
    con_parts2 = consequent[1]
    list4 = [""] * 60
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts[68] = "121"
    con_parts[44] = chr(950)
    con_parts2[68] = "122"
    con_parts2[44] = chr(951)
    sent_type = "e" if connective == iff else "c"
    new_equivalence = ant_sent_parts[0] + " " + connective + " (" + con_parts[0] + " & " + con_parts2[0] + ")"
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " (" + con_parts[42] + " & " + con_parts2[42] + ")"
    new_greek = chr(949) + " " + connective + " (" + chr(950) + " & " + chr(951) + ")"
    if sent_type == 'e':
        ant_sent_parts[53] = 'b'
        con_parts[53] = 'cf'
        con_parts2[53] = 'cf'
    else:
        ant_sent_parts[53] = 'a'
        con_parts[53] = 'cq'
        con_parts2[53] = 'cq'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts[1], con_parts[2]], [con_parts2[1], con_parts2[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = ["(" + con_parts[42] + " & " + con_parts2[42] + ")", ""]
    list4[34] = [ant_sent_parts]
    list4[35] = [con_parts, con_parts2]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts[1], con_parts2[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = ["(" + con_parts[0] + " & " + con_parts2[0] + ")", ""]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts[72], con_parts[2]], [con_parts2[72], con_parts2[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule)
    attach_sent.append(list4)


def prepare_att_sent_1_sent(ant_sent_parts, rule, connective, consequent, anc1="", anc2=""):
    # this populates the attach_sent list provided a sentence is equivalent to one other sentence

    global sn
    sn += 1
    list4 = [""] * 60
    con_parts = consequent[0]
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts[68] = "12"
    con_parts[44] = chr(950)
    sent_type = "e" if connective == iff else "c"
    new_equivalence = ant_sent_parts[0] + " " + connective + " " + con_parts[0]
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " " + con_parts[42]
    new_greek = chr(949) + " " + connective + " " + chr(950)
    if connective == iff:
        ant_sent_parts[53] = 'b'
        con_parts[53] = 'f'
    else:
        ant_sent_parts[53] = 'a'
        con_parts[53] = 'q'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts[1], con_parts[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = [con_parts[1], con_parts[2]]
    list4[34] = [ant_sent_parts]
    list4[35] = [con_parts]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = [con_parts[72], con_parts[2]]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts[72], con_parts[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule, anc1, anc2)
    attach_sent.append(list4)


def noun_slots():
    return [5, 14, 18, 22, 26, 30, 34]

def determ_slots():
    return [3, 10, 16, 20, 24, 28, 32]


def ex(list1, i):
    if i > 14 and list1[i] == None:
        return False
    return True


def allowable_slots2():
    num2 = [11, 47, 3, 69, 4, 55, 5, 66, 67, 35,
            48, 59, 6, 8,
            9, 7, 48, 12, 10, 70,
            13, 14, 36, 60, 63, 40, 49, 15,
            16, 17, 18,
            61, 64, 41, 50, 19, 20, 21, 22, 62, 65, 43, 51,
            23, 24, 25, 26, 52, 27, 28,
            29, 30, 31, 32, 33, 34]

    return num2


def allowable_slots(type=0):
    if type == 0:
        num2 = [11, 47, 3, 69, 4, 55, 5, 66, 67, 35, 48, 59, 6, 8, 9, 7, 48, 12, 10, 70,
                13, 14, 36, 60, 63, 49, 15,
                16, 17, 18, 37,
                61, 64, 50, 19, 20, 21, 22, 38, 62, 65, 51,
                23, 24, 25, 26, 52, 27, 28,
                29, 30, 31, 32, 33, 34]
    elif type == 1:
        # determiners and adjectives
        num2 = [3, 4, 5, 8, 9, 10,
                13, 14, 15,
                16, 17, 18, 19, 20, 22]
        # typically the slots that appear in definitions
    elif type == 2:
        num2 = [5, 8, 9, 14, 49, 15, 18, 50, 19, 22, 51, 23, 26, 52, 27, 30, 31]

    return num2


def is_set_of_bic_wo_id(def_info):
    # if a definition is composed of a set of conditionals or biconditionals
    # and there is no identity statement within the definition then
    # we do not need to delete anything from the def_info list

    if def_info[4][0][1] == "&":
        for lst in def_info[4]:
            if len(lst[0]) == 2 and lst[1] == "":
                return False
            elif len(lst[0]) > 2:
                return True

    return False


def eliminate_conjuncts_from_definition(def_info):
    # any sentences which are not within either an iff or conditiional
    # are deleted

    # 1. if a definition has some detached conjuncts then these are eliminated
    # and a new greek definition is placed in def_info[5], the def_info
    # list will have one member
    # 2. if a definition is composed of sets of complex sentences then
    # the greek definition remains in def_info[5]
    # the other complex sentences are placed in sets of equivalences
    # and it is written on [45] that they are not yet detached and
    # hence cannot be used in the modus ponens function
    # 3. when we add to the total sent list, we do so by looping through the
    # sets of equivalences list and only adding those that say in [45]
    # 'detached'



    if is_set_of_bic_wo_id(def_info):
        for lst in def_info[4]:
            lst[0] = lst[0][1:]
        def_info = make_sets_of_equivalences(def_info)
        def_info = build_conjunction_of_biconditionals(def_info)
    elif def_info[4][0][1] == "&":
        i = 0
        while len(def_info[4][i][0]) < 3:
            if def_info[4][i][1] == "" and len(def_info[4][i][0]) == 2:
                del def_info[0][i]
                del def_info[1][i]
                del def_info[3][i]
                del def_info[4][i]
                del def_info[6][i]
            else:
                i += 1

        def_info = prepare_def_info_list(def_info)
    else:
        def_info = [def_info]

    return def_info


def prepare_def_info_list(def_info):
    connectives = [iff, conditional, idisj, xorr]
    # this removes the first number from the old numbers
    for lst in def_info[4]:
        lst[0] = lst[0][1:]

    # if the definition was composed of a set of conjunctions then the
    # the greek sentence needs to be amended, otherwise we need
    # the "do not add to total sent list
    if def_info[4][1][1] in connectives and def_info[4][2][1] in connectives:
        def_info = make_sets_of_equivalences(def_info)
        def_info = build_conjunction_of_biconditionals(def_info)
    else:
        del def_info[0][0]
        del def_info[1][0]
        del def_info[3][0]
        del def_info[4][0]
        del def_info[6][0]
        def_info[6][0] = remove_outer_paren(def_info[6][0])
        def_info[5] = def_info[6][0]
        def_info = [def_info]

    return def_info


def make_sets_of_equivalences(def_info):
    # if a definition is composed of a set of equivalences
    # then this function puts each complex sentence into a list

    num = [0, 1, 3, 4, 6]
    list2 = []
    for i in range(len(def_info[4])):
        if len(def_info[4][i][0]) == 1:
            list1 = [""] * 7
            for j in num:
                list1[j] = copy.deepcopy([def_info[j][i]])
            list2.append(list1)
        elif len(def_info[4][i][0]) > 1:
            list1 = [""] * 7
            for j in num:
                list1[j] = copy.deepcopy(def_info[j][i])
            for m in range(len(list2)):
                if list1[4][0].startswith(list2[m][4][0][0]):
                    for k in num:
                        list3 = list2[m][k]
                        list3.append(list1[k])
                        list2[m][k] = list3
                    list2[m][5] = list2[m][6][0]

    def_info = [def_info]
    for lst in list2:
        def_info.append(lst)

    return def_info


def build_conjunction_of_biconditionals(def_info):
    str1 = def_info[1][6][0]
    def_info[1][2] = 'eliminate as conjunct'
    for i in range(2, len(def_info)):
        str1 += " & " + def_info[i][5]
        def_info[i][2] = 'eliminate as conjunct'

    def_info[0][5] = str1
    def_info[0][6][0] = str1
    return def_info


def remove_i_from_45(sent):
    for i in [5, 14, 18, 22, 26, 30]:
        if sent[i] == 'i':
            sent[45][1].remove(i)
    return sent


def determine_what_is_conjunct(def_info):
    # modify this if definitions have v or xor as a main connective in their consequent
    ant_conj = 0
    con_conj = 0
    for i, num in enumerate(def_info[4]):
        if len(num[0]) == 2 and num[0][1] == '1':
            if num[1] == "&":
                ant_conj = 3
            else:
                ant_conj = 2
        elif len(num[0]) == 2 and num[0][1] == '2':
            if num[1] == "&":
                con_conj = 3
            else:
                con_conj = 2
        if ant_conj != 0 and con_conj != 0:
            return ant_conj, con_conj


def prepare_attach_sent(def_info, defin_sent, rule, r_sent_loc):
    # this populates the attach sent list

    global sn
    if r_sent_loc != []:
        sn += 1
    elif def_info[2] == 'eliminate as conjunct':
        pass
    elif findinlist(rule, total_sent, 4, 0) != None:
        sn += 2
    else:
        sn += 3

    list1 = [""] * 60
    list1[2] = sn
    greek_sent = def_info[5]
    list2 = translate_complex_sent(greek_sent, defin_sent)
    list1[4] = list2[1]
    list1[37] = list2[0]
    list1[47] = greek_sent
    ant_parts = []
    con_parts = []
    ant_variables = []
    con_variables = []
    ant_conjunction = ""
    con_conjunction = ""
    spec_conn = [iff, conditional, idisj, xorr]
    total_sent_in_attach_sent = []
    embed_att_sent = []
    isdisjunction = False
    isconjunction = False
    if def_info[4][0][1] == iff:
        list1[3] = "e"
    elif def_info[4][0][1] == conditional:
        list1[3] = "c"
    elif def_info[4][0][1] == idisj:
        list1[3] = "d"
        isdisjunction = True
    elif def_info[4][0][1] == xorr:
        list1[3] = "x"
        isdisjunction = True
    else:
        isconjunction = True

    if isdisjunction:
        list1[36] = copy.deepcopy(def_info)
        list1[44] = copy.deepcopy(defin_sent)
        list1[45] = "do not append to attach_sent list"
    elif isconjunction:
        list1[45] = "do not append to attach_sent list"
    else:
        ant_conj, con_conj = determine_what_is_conjunct(def_info)
        list1[45] = "append to attach_sent list"
        for i in range(1, len(def_info[0])):
            t_value = def_info[1][i]
            if def_info[4][i][1] == "":
                d = findposinmd(def_info[6][i], defin_sent, 44)  # d = index of def sent
                defin_sent[d] = ancestor_numbers(defin_sent[d], def_info[4][i][0], def_info)
                if defin_sent[d][53][-1] == "a" or defin_sent[d][53][-1] == "b":
                    ant_parts.append(copy.deepcopy(defin_sent[d]))
                    ant_parts[-1][68] = def_info[4][i][0]
                else:
                    con_parts.append(copy.deepcopy(defin_sent[d]))
                    con_parts[-1][68] = def_info[4][i][0]
                total_sent_in_attach_sent.append(defin_sent[d][1])
                t_value = defin_sent[d][2]
                defin_sent[d][68] = def_info[4][i][0]

            if def_info[4][i][0][1] == '1' and len(def_info[4][i][0]) == ant_conj:
                ant_variables.append([def_info[6][i], t_value])
            elif def_info[4][i][0][1] == '2' and len(def_info[4][i][0]) == con_conj:
                con_variables.append([def_info[6][i], t_value])

            if def_info[4][i][1] != "" and len(def_info[4][i][0]) == 2:
                if def_info[4][i][0][1] == '1' and ant_conjunction == "":
                    ant_conjunction = def_info[6][i]

                elif def_info[4][i][0][1] == '2' and con_conjunction == "":
                    con_conjunction = def_info[6][i]

            if def_info[4][i][1] in spec_conn:
                embed_att_sent.append(i)

        list1 = prepare_attach_sent2(ant_conjunction, con_conjunction,
                                     ant_variables, con_variables, def_info,
                                     defin_sent, embed_att_sent, rule, list1)

        list1[34] = ant_parts
        list1[35] = con_parts
        list1[38] = total_sent_in_attach_sent

    return list1


def prepare_attach_sent2(ant_conjunction, con_conjunction, ant_variables,
                         con_variables, def_info, defin_sent,
                         embed_att_sent, rule, list1):
    global sn
    embed_info = []
    list2 = translate_list_of_sentences(ant_variables, defin_sent)
    list1[0] = list2[1]
    list1[42] = list2[0]

    if ant_conjunction != "":
        list2 = translate_complex_sent(ant_conjunction, defin_sent)
        list1[40] = [list2[0], ""]
        list1[7] = [list2[1], ""]
    else:
        list1[40] = list2[0][0]
        list1[7] = list2[1][0]

    list2 = translate_list_of_sentences(con_variables, defin_sent)
    list1[1] = list2[1]
    list1[43] = list2[0]

    if con_conjunction != "":
        list2 = translate_complex_sent(con_conjunction, defin_sent)
        list1[41] = [list2[0], ""]
        list1[8] = [list2[1], ""]
    else:
        list1[41] = list2[0][0]
        list1[8] = list2[1][0]

    if embed_att_sent != []:
        for num in embed_att_sent:
            list4 = prepare_embed_att_sent(def_info, defin_sent, rule, num)
            embed_info.append(list4)
            sn -= 1

        list1[39] = embed_info

    return list1


def prepare_embed_att_sent(def_info, defin_sent, rule, num):
    # this takes those attached sentences within attached sentences and
    # prepares them to be manipulated

    def_info2 = copy.deepcopy(def_info)
    sent_num = def_info[4][num][0]
    def_info2 = delete_irrel_sent(def_info2, sent_num)
    def_info2[5] = def_info2[6][0]
    pos_of_new_num = len(sent_num) - 1
    main_conn_loc = def_info2[3][0].find(def_info2[4][0][1])
    def_info2[4][0][2] = main_conn_loc
    def_info2 = renumber_embed_sent(def_info2, pos_of_new_num)
    list1 = prepare_attach_sent(def_info2, defin_sent, rule, "EMBED")
    get_general_variables(list1)
    # note that general variables are also added in the map double definienda

    return list1


def renumber_embed_sent(def_info2, pos_of_new_num):
    for position in def_info2[4]:
        position[0] = position[0][pos_of_new_num:]

    return def_info2


def delete_irrel_sent(def_info2, sent_num):
    # this deletes sentences from the def_info list
    # which are not members of the embedded sentence
    # under investigation

    i = 0
    while i < len(def_info2[4]):

        if not def_info2[4][i][0].startswith(sent_num):
            del def_info2[0][i]
            del def_info2[1][i]
            del def_info2[3][i]
            del def_info2[4][i]
            del def_info2[6][i]
        else:
            i += 1

    return def_info2


def ancestor_numbers(list2, k, def_info):
    # this determines the number and connective of the ancestors of a
    # sent in the conditional
    list2[53] = None
    self_num = k[-1]
    if len(k) == 4:
        ggparen_num = k[0]
        gparen_num = k[:2]
        paren_num = k[:3]
        ggparen_conn = findinlist(ggparen_num, def_info[4], 0, 1)
        gparen_conn = findinlist(gparen_num, def_info[4], 0, 1)
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        ggparen_conn = convert_con_to_letter(ggparen_conn, gparen_num[-1])
        gparen_conn = convert_con_to_letter(gparen_conn, paren_num[-1])
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn + gparen_conn + ggparen_conn

    elif len(k) == 3:
        gparen_num = k[0]
        paren_num = k[:2]
        gparen_conn = findinlist(gparen_num, def_info[4], 0, 1)
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        gparen_conn = convert_con_to_letter(gparen_conn, paren_num[-1])
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn + gparen_conn

    elif len(k) == 2:
        paren_num = k[0]
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn

    elif len(k) == 5:
        gggparen_num = k[0]
        ggparen_num = k[:2]
        gparen_num = k[:3]
        paren_num = k[:4]
        gggparen_conn = findinlist(gggparen_num, def_info[4], 0, 1)
        ggparen_conn = findinlist(ggparen_num, def_info[4], 0, 1)
        gparen_conn = findinlist(gparen_num, def_info[4], 0, 1)
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        gggparen_conn = convert_con_to_letter(gggparen_conn, ggparen_num[-1])
        ggparen_conn = convert_con_to_letter(ggparen_conn, gparen_num[-1])
        gparen_conn = convert_con_to_letter(gparen_conn, paren_num[-1])
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn + gparen_conn + ggparen_conn + gggparen_conn

    elif len(k) == 6:
        print("you have not coded for attached sentences with 5 generations yet")
        sys.exit()

    if list2[53] == None:
        print("the number ancestor function is messed up")
        sys.exit()
    return list2


def convert_con_to_letter(str1, str2):
    # this converts a connective to a letter
    if str1 == iff and str2 == '1':
        return 'b'
    elif str1 == iff and str2 == '2':
        return 'f'
    elif str1 == conditional and str2 == '1':
        return 'a'
    elif str1 == conditional and str2 == '2':
        return 'q'
    elif str1 == xorr and str2 == '1':
        return 'x'
    elif str1 == xorr and str2 == '2':
        return 'y'
    elif str1 == idisj and str2 == '1':
        return 'd'
    elif str1 == idisj and str2 == '2':
        return 'g'
    elif str1 == idisj and str2 == '3':
        return 'd3'
    elif str1 == idisj and str2 == '4':
        return 'd4'
    elif str1 == idisj and str2 == '5':
        return 'd5'
    elif str1 == idisj and str2 == '6':
        return 'd6'


    elif str1 == "&":
        return 'c'
    else:
        print('the convert con to letter function is messed up')
        g = 4 / 0


def translate_list_of_sentences(to_be_converted, defin_sent):
    # this converts the 0th or 1st member of the attach_sent_list into
    # sentence variables

    abbrev_sent = copy.deepcopy(to_be_converted)
    for i in range(len(to_be_converted)):
        if os(to_be_converted[i][0]):
            d = findposinmd_alert_error(to_be_converted[i][0], defin_sent, 44)
            to_be_converted[i][0] = defin_sent[d][72]
            abbrev_sent[i][0] = defin_sent[d][1]
            defin_sent[d][74] = 'add to all_sent'
        elif not os(to_be_converted[i][0]):
            list1 = translate_complex_sent(to_be_converted[i][0], defin_sent)
            to_be_converted[i][0] = list1[0]
            abbrev_sent[i][0] = list1[1]

    return [to_be_converted, abbrev_sent]


def translate_complex_sent(to_be_translated, defin_sent):
    # this converts a conjunction in a definition into the definition
    # with the new variables

    to_translate_abbrev = to_be_translated

    for sentence in defin_sent:
        if sentence[44] in to_be_translated:
            sentence[74] = 'add to all_sent'
            to_be_translated = to_be_translated.replace(sentence[44], sentence[0])
            to_translate_abbrev = to_translate_abbrev.replace(sentence[44], sentence[42])

    return [to_be_translated, to_translate_abbrev]


def prepare_categorize_words(str2):
    # when we prepare to categorize a word, the first three slots
    # are reserved for the sentence, the sentence letter and the tvalue

    str2 = str2.strip()
    list1 = str2.split(' ')
    list2 = [None] * 80
    j = 2
    for i in range(len(list1)):
        j += 1
        list2[j] = list1[i]

    return list2


def get_empty_slots():
    return [[], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], []]


def categorize_words(list1):
    sentence_slots = [None] * 80
    relation_type = 0
    slots_used = get_empty_slots()
    places_used = []
    noun_list = ['n', 'p']
    the_is_of_group = ["I", "is" + ug, "are" + ug, "be" + ug, "was" + ug, "were" + ug, "am" + ug]
    the_is_of_adjective = ["J", "is" + ua, "be" + ua, "are" + ua, "was" + ua, "am" + ua, "were" + ua]
    spec_rel = the_is_of_adjective + the_is_of_group
    categories_used = []

    i = 2
    while list1[i + 1] != None:
        i += 1
        k = 0
        word = list1[i]

        if word == 'woman':
            bb = 8

        i, word, has_comma = determine_if_compound_word(i, list1, word)

        if get_words_used == 1:
            if word not in words_used and not isvariable(word):
                words_used.append(word)

        part_of_speech, sub_part, sub_sub_part, rest = get_part_of_speech(word, list1[79])

        if word != ' ' and word != "":
            insert_special_location = False
            if part_of_speech == "d":  # q are possessive pronouns
                if relation_type == 0:
                    k = 3
                elif relation_type == 1:
                    k = 10
                elif relation_type == 2:
                    k = 16
                elif relation_type == 3:
                    k = 20
                elif relation_type == 4:
                    k = 24
                elif relation_type == 5:
                    k = 28
                elif relation_type == 6:
                    k = 32

            elif part_of_speech == 'o' or part_of_speech == 's':
                if relation_type == 0 and sentence_slots[5] == None:
                    k = 69
                elif relation_type == 1 and sentence_slots[14] == None:
                    k = 70

            elif part_of_speech == 'a':
                if relation_type == 0:
                    k = 4
                elif relation_type == 1 and sentence_slots[9] in the_is_of_adjective:
                    k = 14
                elif relation_type == 1:
                    k = 13
                elif relation_type == 2 and sentence_slots[15] in the_is_of_adjective:
                    k = 18
                    # no decision needed here because it is defined as a concept
                elif relation_type == 2:
                    k = 17
                elif relation_type == 3 and sentence_slots[19] in the_is_of_adjective:
                    k = 22
                elif relation_type == 3:
                    k = 21
                elif relation_type == 4 and sentence_slots[23] in the_is_of_adjective:
                    k = 26
                elif relation_type == 4:
                    k = 25
                elif relation_type == 5 and sentence_slots[27] in the_is_of_adjective:
                    k = 30
                elif relation_type == 5:
                    k = 29
                elif relation_type == 6 and sentence_slots[31] in the_is_of_adjective:
                    k = 34
                elif relation_type == 6:
                    k = 33

            elif part_of_speech == 'm':
                if sentence_slots[3] == None and sentence_slots[5] == None:
                    k = 47
                elif relation_type == 0:
                    k = 8
                elif (relation_type == 1 and sentence_slots[14] == None and sentence_slots[60] == None):
                    k = 8
                    # because 'not' in this location comes after the relation we must
                    # insert to before the relation
                    insert_special_location = True
                elif relation_type == 1:
                    k = 49
                elif relation_type == 2 and sentence_slots[15] in spec_rel:
                    k = 49
                    insert_special_location = True
                elif relation_type == 2:
                    k = 50
                elif relation_type == 3 and sentence_slots[18] in spec_rel:
                    k = 50
                    insert_special_location = True
                elif relation_type == 3:
                    k = 51
                elif relation_type == 4 and sentence_slots[24] in spec_rel:
                    k = 51
                    insert_special_location = True
                elif relation_type == 4 or sentence_slots[27] in spec_rel:
                    k = 52


            elif sub_part == 't':

                if relation_type == 0 and sentence_slots[5] == None:
                    k = 5

            elif part_of_speech in noun_list:

                if relation_type == 0:
                    if sentence_slots[66] != None:
                        k = 67
                    elif sentence_slots[5] != None:
                        k = 35
                    elif sentence_slots[5] == None:
                        k = 5
                elif relation_type == 1:
                    if sentence_slots[14] == None:
                        k = 14
                    else:
                        k = 36
                elif relation_type == 2:
                    if sentence_slots[18] == None:
                        k = 18
                    else:
                        k = 37

                elif relation_type == 3:
                    if sentence_slots[22] == None:
                        k = 22
                    else:
                        k = 38
                elif relation_type == 4:
                    k = 26
                elif relation_type == 5:
                    k = 30
                elif relation_type == 6:
                    k = 34


            elif part_of_speech == 'c':
                if relation_type == 0 and sentence_slots[5] != None:
                    k = 66  # uuu

            elif part_of_speech == 'u':
                if relation_type == 0 and sentence_slots[5] != None:
                    k = 59
                elif relation_type == 1 and sentence_slots[14] != None:
                    k = 60
                elif relation_type == 2 and sentence_slots[18] != None:
                    k = 61
                elif relation_type == 3 and sentence_slots[22] != None:
                    k = 62

            elif part_of_speech == 'y':

                if sentence_slots[7] == None:
                    k = 7
                elif sentence_slots[40] == None:
                    k = 40
                elif sentence_slots[41] == None:
                    k = 41
                elif sentence_slots[43] == None:
                    k = 43

            elif part_of_speech == 'r':

                if relation_type == 0:
                    k = 9
                    relation_type = 1
                elif relation_type == 1:
                    k = 15
                    relation_type = 2
                elif relation_type == 2:
                    relation_type = 3
                    k = 19
                elif relation_type == 3:
                    relation_type = 4
                    k = 23
                elif relation_type == 4:
                    relation_type = 5
                    k = 27
                elif relation_type == 5:
                    k = 31
                    relation_type = 6

            if has_comma: sentence_slots[39] = k
            if mysql == 2 and k == 0:
                # tahir system exit
                print ("our system does not have this grammatical syntax yet")

            if k == 0 and proof_type != 3:
                print(word)
                assert k != 0

            k, sentence_slots, sub_part, part_of_speech = exceptional_parts_of_speech(word,
                                                                                      k, sentence_slots, sub_part,
                                                                                      part_of_speech)
            sentence_slots[k] = word
            if insert_special_location:
                places_used.insert(-1, k)
            else:
                places_used.append(k)
            b = get_used_slots(k, part_of_speech, sub_part, rest)

            slots_word_sub = [4, 5, 12, 13, 14, 17, 18, 22, 26, 30, 49, 50, 51, 52, 8, 47, 49, 50, 51, 52,
                              34, 35, 36, 63, 64, 65, 67, 69, 70, 9, 15, 19, 23, 27, 31, 37, 38]

            b = divide_the_i_relation(k, b, sentence_slots, the_is_of_group)
            if k in slots_word_sub and b not in [1, 14]:
                slots_used[0].append(k)
            if b != 0:
                slots_used[b].append(k)
                # line 19a
                # the categories are only used in the define_irregular_words function
                # categories about 15 are used in other functions
                if b < 17: categories_used.append(b)

    categories_used.sort()
    sentence_slots[45] = slots_used
    sentence_slots[46] = categories_used
    sentence_slots[56] = places_used
    sentence_slots[57] = get_subclause_position(sentence_slots)
    if not list1[79] == 'do not rebuild sentence':
        sentence_slots = build_sent_slots_known(sentence_slots)

    return sentence_slots


def get_used_slots(k, part_of_speech, sub_part, rest):
    # if you change the number of the slots then you must change which slot the universal is in
    # in the lies within univ quant function
    # also modify line 19a if you go above 16
    predicative_complement_positions = [14, 18, 22, 26, 29]
    sub_sub_part = rest[2] if len(rest) > 2 else ""
    fourth_letter = rest[3] if len(rest) > 3 else ""
    fifth_letter = rest[4] if len(rest) > 4 else ""

    b = 0
    if sub_sub_part == 'd':
        b = 19
    elif sub_part == 's':
        b = 18
    elif sub_part == 'z':
        b = 20
    elif (part_of_speech == 'd' or part_of_speech == 'p') \
            and sub_part != 'b' and sub_part != "i" and sub_part != 'e':  # determinative, pronouns, possessive pronouns
        b = 1
    elif part_of_speech == 'o':
        b = 3  # common name possessives
    elif part_of_speech == 's':
        b = 4  # proper name possessives
    elif k == 66:
        b = 5  # and
    elif part_of_speech == 'a' and k not in predicative_complement_positions:  # adjectives
        b = 6
    elif k in [35, 36, 37, 38]:  # CIA
        b = 7
    elif part_of_speech == 'u' and k == 59:
        b = 8
    elif part_of_speech == 'u':  # relative pronouns
        b = 9
    elif part_of_speech == 'y':  # that
        b = 10
    elif part_of_speech == 'r' and sub_part == 'd':  # AS
        b = 11
    elif part_of_speech == 'r' and k in [15, 19, 23, 27, 31] and fifth_letter != 'b':  # RDA
        # modify this if you change this then you must also change the eliminate relative pronouns function
        b = 13
    elif part_of_speech == 'n' and sub_part == 't':  # there
        b = 14
    elif part_of_speech == 'd' and sub_part == 'b':  # universals
        b = 15
    elif part_of_speech == 'd' and sub_part == 'e':
        b = 16
    elif part_of_speech == 'm':
        b = 17

    return b


def get_part_of_speech(word, str5):
    sub_part_of_speech = ""
    sub_sub_part = ""
    if isvariable(word):
        pos = 'n'
        posp = ""
        if word in variables:
            variables.remove(word)
            # if the variable stands for an adjective then its part of speech
            # is adjective
        str1 = abbreviations[0].get(word)
        if str1 != None and str5 != 'is in definition':
            posp = dictionary[0].get(str1)
            pos = posp[0]
            sub_part_of_speech = posp[1] if len(posp) > 1 else ""
            sub_sub_part = posp[2] if len(posp) > 2 else ""

    else:
        if word[-2:] == "'s":
            posp = dictionary[0].get(word[:-2])
            if len(posp) > 2 and posp[2] == 'n':
                pos = 's'
            else:
                pos = 'o'
            sub_part_of_speech = ""
            sub_sub_part = ""

        else:
            posp = dictionary[0].get(word)
            if posp == None:
                print("you misspelled " + word)
                if proof_type != 3 or mysql == 2:
                    # tahir system exit
                    g = 4 / 0
                else:
                    posp = "n"
            pos = posp[0]
            sub_part_of_speech = posp[1] if len(posp) > 1 else ""
            sub_sub_part = posp[2] if len(posp) > 2 else ""
            if word == "~":
                pos = 'm'
            elif word == ne:
                pos = 'r'
            elif word == 'not':
                pos = 'm'

    return pos, sub_part_of_speech, sub_sub_part, posp


def determine_if_compound_word(i, list1, word):
    if word == "not":
        bb = 8
    if "," in word:
        word = word.replace(",", "")
        return i, word, True
    has_comma = False
    double = dictionary[4].get(word)
    triple = dictionary[5].get(word)
    triple_word = ""

    if triple != None:
        if list1[i + 1] != None and list1[i + 2] != None and "," not in list1[i + 1]:
            if "," in list1[i + 2]:
                after_next_word = list1[i + 2].replace(",", "")
                has_comma = True
            else:
                after_next_word = list1[i + 2]
            next_word = list1[i + 1]
            triple_word = word + " " + next_word + " " + after_next_word
            if triple_word in triple:
                i += 2
                word = triple_word
            else:
                triple_word = ""
        else:
            triple_word = ""

    if triple_word == "" and double != None:
        if list1[i + 1] != None:
            if "," in list1[i + 1]:
                next_word = list1[i + 1].replace(",", "")
                has_comma = True
            else:
                next_word = list1[i + 1]
            double_word = word + " " + next_word
            if double_word in double:
                i += 1
                word = double_word

    return i, word, has_comma


def exceptional_parts_of_speech(word, k, sentence_slots, sub_part, part_of_speech):
    if word == 'doglike':
        bb = 8
    noun_pos = k + 1
    dict1 = {5: 35, 14: 36, 18: 37, 22: 38}
    if sentence_slots[noun_pos] == 'concept' + ua and part_of_speech == "a":
        k = dict1.get(noun_pos)
        part_of_speech = 'n'

    return k, sentence_slots, sub_part, part_of_speech


def divide_the_i_relation(k, b, sentence_slots, the_is_of_group):
    if k in [15, 19, 23, 27, 31]:
        dict1 = {15: 9, 19: 15, 23: 19, 27: 23, 31: 27}
        i_relation = dict1.get(k)
        if sentence_slots[i_relation] in the_is_of_group:
            dict2 = {15: 60, 19: 61, 23: 62}
            rel_pronoun_pos = dict2.get(k)
            if sentence_slots[rel_pronoun_pos] == None:
                b = 13

    return b


def build_sent_name(prop_name):
    str1 = ''
    str2 = ''
    list1 = []

    for i in range(len(prop_name)):
        if i == 24:
            bb = 8
        str3 = remove_outer_paren(prop_name[i][2])
        str3 = str3.replace("~", "")
        str1 = '(' + prop_name[i][0] + mini_e + str3 + ')'
        if len(str2) == 0 and len(str1) > 57:
            list1.append(str1)
        elif (len(str2) + len(str1)) > 57:
            list1.append(str2)
            str2 = str1
            if i + 1 == len(prop_name):
                list1.append(str2)
        elif (len(str2) + len(str1)) <= 57:
            if len(str2) == 0:
                str2 = str1
            else:
                str2 = str2 + ' ' + str1
            if i + 1 == len(prop_name):
                list1.append(str2)
    return list1


def get_right_most_connective(str1):
    split_at_space = False
    if len(str1) > 67:
        j = 67
    else:
        j = len(str1) - 1
    for i in range(j, 15, -1):
        if str1[i] in ["&", conditional, iff, xorr, idisj]:
            return i
        if i == 31:
            split_at_space = True
        if split_at_space and str1[i] == " ":
            return i


def space_sentences(str1, str2):
    b = len(str1)
    c = len(str2)
    j = 0
    second = ""
    third = ""
    if (b + c) > 70:
        location = get_right_most_connective(str1)
        first = str1[:location]
        second = str1[location:]
        if len(second) > 70:
            j += 1
            assert j != 2
            list1 = space_sentences(second, str2)
            second = list1[0]
            third = list1[1]
            str2 = ""

        spaces_needed = 65 - (len(second) + len(str2))
        space = " " * spaces_needed
        second = "     " + second + space + str2
    else:
        spaces_needed = 70 - (len(str1) + len(str2))
        space = " " * spaces_needed
        first = str1 + space + str2

    return [first, second, third]


###### tahir
def print_sent_full(test_sent, tot_prop_name, row_number):
    global result_data
    if proof_type == 0 and mysql == 0:
        return
    elif proof_type == 1:
        row_number = 1
    o = -1

    for i in order:
        # if i == 2:
        #     break
        for j in range(len(test_sent[i])):
            if j == 31:
                bb = 8
            if test_sent[i][j][0] == "" or test_sent[i][j][1] != "":
                if test_sent[i][j][3] == 'RN':
                    test_sent[i][j][3] = ""

                if test_sent[i][j][6] != "":
                    str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5]) + ',' + str(test_sent[i][j][6])
                elif test_sent[i][j][5] != "":
                    str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5])
                elif test_sent[i][j][4] != "":
                    str1 = test_sent[i][j][4]
                else:
                    str1 = ""
                if j == 0:
                    test_sent[i][j][1] = test_sent[i][j][1]
                    test_sent[i][j][4] = i
                else:
                    test_sent[i][j][4] = str1
                if proof_type == 2:
                    len_sp = 5 - len(str(test_sent[i][j][0]))
                    space = " " * len_sp
                    list3 = space_sentences(str(test_sent[i][j][0]) + space + test_sent[i][j][3] + \
                                            test_sent[i][j][1], str(test_sent[i][j][4]))
                    for str1 in list3:
                        if str1 != "":
                            print(str1)

                elif proof_type == 1:
                    w4.cell(row=row_number, column=2).value = test_sent[i][j][0]
                    w4.cell(row=row_number, column=3).value = test_sent[i][j][3] + test_sent[i][j][1]
                    w4.cell(row=row_number, column=4).value = test_sent[i][j][4]
                    if len(test_sent[i][j]) > 8:
                        if test_sent[i][j][8] == "*":
                            xls_cell = w4.cell(row=row_number, column=3)
                            xls_cell.font = xls_cell.font.copy(color='FFFF0000')



                elif mysql == 1:
                    result_data['text_' + str(row_number - 1) + '_1'] = test_sent[i][j][0]
                    result_data['text_' + str(row_number - 1) + '_2'] = test_sent[i][j][1]
                    result_data['text_' + str(row_number - 1) + '_3'] = test_sent[i][j][4]

                row_number += 1

        row_number += 1
        if proof_type == 2:
            print(" ")
        o += 1
        list1 = build_sent_name(tot_prop_name[o])
        for j in range(len(list1)):

            if proof_type == 2:
                print(list1[j])
            elif proof_type == 1:
                w4.cell(row=row_number, column=3).value = list1[j]
            elif mysql == 1:
                result_data['text_' + str(row_number) + '_2'] = list1[j]
            row_number += 1
        row_number += 1
        if proof_type == 2:
            print(" ")
        do_not_print = False
        for j in range(len(test_sent[i])):
            if j == 8:
                bb = 8
            if test_sent[i][j - 1][4] == 'ID':
                do_not_print = True
            elif test_sent[i][j][0] == "" and test_sent[i][j][1] == "":
                do_not_print = False
            if not do_not_print and test_sent[i][j][2] != "":
                if j == 0:
                    test_sent[i][j][4] == ""
                if proof_type == 2:
                    len_sp = 5 - len(str(test_sent[i][j][0]))
                    space = " " * len_sp
                    list3 = space_sentences(str(test_sent[i][j][0]) + space + test_sent[i][j][3] + \
                                            test_sent[i][j][2], str(test_sent[i][j][4]))
                    if list3[1] != "":
                        print(list3[0])
                        print(list3[1])
                    else:
                        print(list3[0])
                elif proof_type == 1:
                    w4.cell(row=row_number, column=2).value = test_sent[i][j][0]
                    w4.cell(row=row_number, column=3).value = test_sent[i][j][3] + test_sent[i][j][2]
                    w4.cell(row=row_number, column=4).value = test_sent[i][j][4]
                    if test_sent[i][j][2] == bottom or test_sent[i][j][2] == consist:
                        w4.cell(row=row_number, column=5).value = 1

                elif mysql == 1:
                    result_data['text_' + str(row_number) + '_1'] = test_sent[i][j][0]
                    result_data['text_' + str(row_number) + '_2'] = test_sent[i][j][3] + test_sent[i][j][2]
                    result_data['text_' + str(row_number - 1) + '_3'] = test_sent[i][j][4]

                row_number += 1
        row_number += 3


def determine_words_used():
    if get_words_used == 1:
        for i in range(len(words_used)):
            j = dictionary[7].get(words_used[i], 28)
            if j == 28:
                print(words_used[i])
            ws.cell(row=j, column=2).value = 1


def progress(count, total, suffix=''):
    bar_len = 60
    filled_len = int(round(bar_len * count / float(total)))

    percents = round(100.0 * count / float(total), 1)
    bar = '=' * filled_len + '-' * (bar_len - filled_len)

    sys.stdout.write('[%s] %s%s ...%s\r' % (bar, percents, '%', suffix))
    sys.stdout.flush()


def build_dict():
    global dictionary
    # tahir dictionary is built here
    ex_dict = large_dict() if get_words_used == 0 else ""

    parts_of_speech = {}
    definitions = {}
    synonyms = {}
    relations = {}
    doubles = {}
    triples = {}
    only_def_if_input = []  # words that are only defined if they appear in the input sentence
    words_to_row = {}
    almost_done = False

    i = -1
    if get_words_used == 1:
        mm = 3000
    else:
        mm = len(ex_dict)

    aa = time.time()
    while i < mm - 1:
        i += 1
        # if i % 50 == 0:
        #     print(str(i) + " " + str("{0:.7f}".format((time.time() - aa))))
        #     aa = time.time()
        if get_words_used == 1:
            if i == 0:
                i = 5
            s = ws.cell(row=i, column=1).value
            if s == 451:
                bb = 8
            pos = ws.cell(row=i, column=3).value
            word = ws.cell(row=i, column=4).value
            if word == "true*":
                word = "true"
            if word == "false*":
                word = "false"
            next_word = ws.cell(row=i, column=4).value
            if word == "" and next_word == "":
                break
        else:
            s = 0
            pos = ex_dict[i][0]
            word = ex_dict[i][1]
            if word != None:
                word = tran_str(word)

        if word == None and almost_done:
            break

        almost_done = True if word == None else False

        if pos != None and pos != "":
            if not isinstance(pos, int):
                pos = pos.strip()
            if word == 'relationship':
                bb = 7

            if isinstance(word, int): word = str(word)

            if "(" in word:
                cc = word.index("(")
                word = word[:cc - 1]

            word = word.strip()

            if get_words_used == 1:
                abbrev_relat = ws.cell(row=i, column=5).value
                defin = ws.cell(row=i, column=6).value
            else:
                abbrev_relat = ex_dict[i][2]
                defin = ex_dict[i][3]
                defin = tran_str(defin, True)
            words_to_row.update({word: s})
            if abbrev_relat != "": words_to_row.update({abbrev_relat: s})

            parts_of_speech.update({word: pos})
            parts_of_speech.update({abbrev_relat: pos})
            fir_let = pos[0]
            sec_let = pos[1] if len(pos) > 1 else ""
            thir_let = pos[2] if len(pos) > 2 else ""
            four_let = pos[3] if len(pos) > 3 else ""
            fif_let = pos[4] if len(pos) > 4 else ""

            synonyms = update_synonyms(defin, synonyms, definitions, sec_let)
            if fir_let == 'r': relations.update({word: abbrev_relat})

            if " " in word:
                m = word.count(" ")
                if m == 1:
                    word1 = copy.copy(word)
                    y = word1.find(" ")
                    word1 = word1[:y]
                    doubles.setdefault(word1, []).append(word)
                if m == 2:
                    word1 = copy.copy(word)
                    y = word1.find(" ")
                    word1 = word1[:y]
                    triples.setdefault(word1, []).append(word)

            if sec_let == 'k':
                only_def_if_input.append(word)

            if sec_let != 'a' and sec_let != 's':
                if fir_let == "r":
                    definitions.update({abbrev_relat: defin})
                else:
                    definitions.update({word: defin})

    dictionary = [parts_of_speech, definitions, synonyms, relations,
                  doubles, triples, only_def_if_input, words_to_row]


def update_synonyms(defin, synonyms, definitions, sec_let):
    if sec_let == 's':
        str6 = defin[defin.find("=") + 1:-1]
        str6 = str6.strip()
        str7 = defin[1:defin.find("=")]
        str7 = str7.strip()
        synonyms.update({str7: str6})
        definitions.update({str7: defin})

    return synonyms


def get_key(dict1, val):
    for k, v in dict1.items():
        if v == val:
            return k
    else:
        return ""


def findinlist(str1, list1, i, j):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the matching second element

    for d in range(len(list1)):
        if str1 == list1[d][i]:
            str2 = list1[d][j]
            return str2
    return None


def findposinmdlistint(i, list1, p):
    for j in range(len(list1)):
        if list1[j][p] == i:
            return j
    else:
        return -1


def findposinmd(str1, list1, p):
    # this determines the position of an element in a multidimensional list
    for i in range(len(list1)):
        if list1[i][p] == str1:
            return i
    return -1


def findposinmd_alert_error(str1, list1, p):
    # this determines the position of an element in a multidimensional list
    for i in range(len(list1)):
        if list1[i][p] == str1:
            return i
    g = 4 / 0


def isinmdlist(str1, list1, p):
    # this determines whether or not an element is in a multidimensional list
    for i in range(len(list1)):
        if list1[i][p] == str1:
            return True

    return False


def use_identity(negated_conjunction, consistent):
    if not consistent:
        return False
    identities = []
    for sent in detach_sent:
        if sent[9] == '=' and isvariable(sent[14], "i") and isvariable(sent[5]) and sent[8] != "~":
            identities.append([sent[5], sent[14], sent[58]])

    if identities != []:
        for i in range(len(identities)):
            str1 = identities[i][0]
            str2 = identities[i][1]
            anc1 = identities[i][2]
            for sent in detach_sent:
                for k in [5, 14, 18, 22]:
                    if sent[k] == str1 or sent[k] == str2:
                        ant_sent_parts = copy.deepcopy(sent)
                        con_parts = copy.deepcopy(sent)
                        if sent[k] == str1:
                            con_parts[k] = str2
                        else:
                            con_parts[k] = str1
                            con_parts = build_sent2(con_parts)
                        if con_parts[42] == "q" + l1:
                            bb = 8
                        prepare_att_sent_1_sent(ant_sent_parts, "SUB", iff, [con_parts], anc1)
                        if con_parts[5] == con_parts[14] and \
                                        con_parts[9] != "=" and con_parts[2] == "":
                            consistent = use_reflexivity1(con_parts)
                            if not consistent:
                                return consistent

        consistent, _ = detach1("do not use modus tollens", consistent, negated_conjunction)

    return consistent


def check_reflexivity(consistent):
    if not consistent:
        return consistent
    for sent in detach_sent:
        if sent[5] == sent[14] and sent[2] == "" and sent[9] != "=":
            consistent = use_reflexivity2(sent)
            break

    return consistent

def use_reflexivity1(sent):
    global sn
    anc1 = sent[58]
    anc2 = attach_sent[-1][2]
    sn += 1
    add_to_total_sent(sn, sent[72], sent[1], "", iff + "E", anc1, anc2)
    consistent = use_reflexivity2(sent)

    return consistent


def use_reflexivity2(sent):
    global sn
    new_sent = copy.deepcopy(sent)
    new_sent[8] = "~"
    new_sent = build_sent2(new_sent)
    full_conditional = sent[0] + " " + conditional + " " + new_sent[0]
    abbrev_conditional = sent[42] + " " + conditional + " " + new_sent[42]
    sn += 1
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "IRR")
    sn += 1
    consistent = add_to_total_sent_consist(sn, new_sent[72], new_sent[1], \
                                           new_sent[2], "MP", sn - 2, sn - 1, [])

    return consistent


def get_position_of_identities():
    for i in range(len(total_sent)):
        if total_sent[i][4] != "":
            return i
    g = 4 / 0


def build_list_of_abbreviations(loop_number):
    # this turns the abbreviations into a conjunction
    global sn
    position_of_identities = get_position_of_identities()
    # this loop removes duplicates in a multidimensional list
    str1 = ""
    total_attach_sent = []
    temp_detach_sent = []
    for lst in attach_sent:
        for sent in lst[38]:
            total_attach_sent.append(sent)

    for key, value in abbreviations[0].items():
        if get_words_used == 1:
            if value not in words_used:
                words_used.append(value)
        str2 = "(" + key + "=" + value + ")"
        str2p = name_sent(str2)
        if str1 == "":
            str1 = str2
            str1p = str2p
        else:
            str1 += " & " + str2
            str1p += " & " + str2p
        if str2p in total_attach_sent:
            list1 = [None] * 80
            list1[0] = str2
            list1[1] = str2p
            list1[2] = ""
            list1[5] = key
            list1[8] = ""
            list1[9] = "="
            list1[14] = value
            list1[72] = str2
            list1[42] = str2p
            temp_detach_sent.append(list1)

    if temp_detach_sent != []:
        for sent in temp_detach_sent:
            sn += 1
            sent[58] = sn
            g = findposinmd(sent[42], detach_sent, 42)
            if g != -1:
                detach_sent[g][58] = sn
            else:
                detach_sent.append(sent)
            add_to_total_sent(sn, sent[0], sent[1], "", "&E", position_of_identities + 1)

    total_sent.insert(position_of_identities, [position_of_identities - 1, str1,
                                               str1p, "", 'ID', "", "", "", "", ""])
    list2 = [""] * 9
    list2[1] = "UNTRANSLATED DEFINITIONS"
    total_sent.insert(position_of_identities, list2)
    list2 = [""] * 9
    total_sent.insert(position_of_identities, list2)
    abbreviations[3] = len([abbreviations[0].keys()])


def step_two(truth_value):
    global instan_time, sn

    negated_conjunction = []
    nonstandard_sentences = []
    loop_number = 1

    consistent, _ = detach1("do not use modus tollens", True, negated_conjunction)

    get_relevant_variables(all_sent)

    to_be_defined = copy.deepcopy(all_sent)

    while True:

        if loop_number > 4: break

        define_regular_terms(to_be_defined)

        add_necessary_conditions_for_concept(loop_number)

        build_list_of_abbreviations(loop_number)

        eliminate_attached_conjuncts(loop_number)

        consistent, _ = detach1("do not use modus tollens", consistent, negated_conjunction)

        get_nonstandard_sent(nonstandard_sentences, loop_number)

        consistent = use_identity(negated_conjunction, consistent)

        consistent = check_reflexivity(consistent)

        determine_relevance(to_be_defined)

        get_more_variable_types()

        get_object_properties(consistent, to_be_defined, loop_number)

        consistent = use_basic_lemmas(consistent)

        false_by_def = True if not consistent else False

        rearrange(nonstandard_sentences, false_by_def)

        consistent, proof_done, to_be_defined = step_three(negated_conjunction, consistent)

        if proof_done: break

        loop_number += 1

        if loop_number > 2:
            to_be_defined, proof_done = make_attached_detached(consistent, [])
            if to_be_defined == []:
                break
            else:

                print ("two loops")


    rearrange(nonstandard_sentences, false_by_def, "final")

    rename_rules()

    add_consistency_sign(consistent)

    # obtain_relevant_sentences(consistent)

    consistent = final_truth_value(consistent, truth_value)

    return consistent


def final_truth_value(consistent, truth_value):
    if truth_value == consistent:
        return True
    else:
        return False


def obtain_relevant_sentences(consistent):
    # 35, 15
    relevant_sentences = []
    if not consistent:
        relevant_sentences = [total_sent[-2][5], total_sent[-2][6]]
        i = -1
        while i < len(relevant_sentences) - 1:
            i += 1
            k = findposinmdlistint(relevant_sentences[i], total_sent, 0)
            total_sent[k][8] = "*"

            for j in [5, 6]:
                if total_sent[k][j] != "":
                    if isinstance(total_sent[k][j], str):
                        list1 = total_sent[k][j].split(",")
                        list1 = [int(m) for m in list1]
                        for m in list1:
                            if m not in relevant_sentences:
                                relevant_sentences.append(m)
                    else:
                        if total_sent[k][j] not in relevant_sentences:
                            relevant_sentences.append(total_sent[k][j])
        relevant_sentences.sort()
        rel_sent_str = [str(m) for m in relevant_sentences]
        rel_sent = " ".join(rel_sent_str)
        add_to_total_sent("", "RELEVANT SENTENCES: " + rel_sent)

def rename_rules():
    for lst in total_sent:
        lst[4] = lst[4].replace('AY', 'AX')
        lst[4] = lst[4].replace('DE', 'DF')
        lst[4] = lst[4].replace('NE', 'NC')
        lst[4] = lst[4].replace('LY', 'LE')
        lst[4] = lst[4].replace('DFF', 'DEF')
        if lst[2].startswith("INFER"):
            return

def make_attached_detached(consistent, to_be_defined):
    global sn
    if not consistent or to_be_defined != []:
        proof_done = True if not consistent else False
        return to_be_defined, proof_done
    used_sent = []
    dict1 = {}
    for lst in attach_sent:
        if lst[46] != "instantiated" and lst[50] == 'axiom of definition':
            lst[50] = ""
            for j in [34, 35]:
                for sent in lst[j]:
                    if sent[42] not in used_sent:
                        general_variable_found = 0
                        for i in noun_slots():
                            if ex(sent, i):
                                if sent[i] in variable_type[0]:
                                    if general_variable_found == 0:
                                        general_variable_found += 1
                                        new_sent = copy.deepcopy(sent)
                                        used_sent.append(sent[42])

                                    var = dict1.get(sent[i])
                                    if var == None:
                                        dict1.update({sent[i]: variables[0]})
                                        new_sent[i] = variables[0]
                                        do_not_instantiate.setdefault(sent[i], []).append(variables[0])
                                        variable_type[1].append(variables[0])
                                        variable_type[3].append(variables[0])
                                        del variables[0]
                                    else:
                                        new_sent[i] = var

                            else:
                                if general_variable_found > 0:
                                    new_sent = build_sent2(new_sent, 2)
                                    sn += 1
                                    new_sent[58] = sn
                                    new_sent[54] = None
                                    to_be_defined.append(new_sent)
                                    detach_sent.append(new_sent)
                                    add_to_total_sent(sn, new_sent[72], new_sent[1], new_sent[2], "AY DEF")
                                break

    proof_done = True if to_be_defined == [] else False

    return to_be_defined, proof_done


def eliminate_attached_conjuncts(loop_number):
    global sn
    for sent in attach_sent:
        if sent[46] == 'eliminate as conjunct':
            sent[46] = ""
            anc1 = sent[2]
            sn += 1
            sent[2] = sn
            add_to_total_sent(sn, sent[37], sent[4], "", "&E", anc1)


def match_rn_sent_to_definition2(rn_sent):
    bool1 = False
    definitions = {}
    for i, lst in enumerate(total_sent):
        if lst[1] == 'UNTRANSLATED DEFINITIONS':
            bool1 = True
        if bool1:
            definitions[lst[4][3:]] = i
            if lst[1] == '':
                break
    j = 0
    for definiendum, v in rn_sent.items():
        i = definitions.get(definiendum)
        for substitution in v:
            j += 1
            total_sent.insert(i + 1 + j, substitution)

def rearrange(nonstandard_sentences, false_by_def, type = ""):
    # if proof_type = 0 never rearrange
    #
    # if proof_type = 1 only rearrange at second point
    #
    # if proof_type = 2 rearrange at the first point, and do not rearrange at the second point if it was false by definition
    #
    # if proof_type = 2 rearrange at the first point and at the second point if it was consistent by definition at the first point
    if proof_type == 0 or (proof_type == 1 and type != 'final') or (proof_type == 2 and false_by_def and type == "final"):
         return
    global total_sent
    untran_rules = ['DF', 'LE', "AX", "NC", "NS"]
    infer_rules = [iff + "E", "MP", "MT", "EN", "&I", "&E", idisj + "E", "~~E", bottom + "I", consist + "I"]
    untranslated_sentences = []
    inferences = []
    reduce_to_standard_form = []
    rn_sent = {}
    initial_claims = []
    done_with_initial_claims = False
    total_sent2 = copy.deepcopy(total_sent)

    for i, lst in enumerate(total_sent2):
        if lst[4][:2] in untran_rules:
            untranslated_sentences.append(lst)
        elif lst[4] in infer_rules:
            done_with_initial_claims = True
            inferences.append(lst)
        elif lst[4] == "RN":
            rn_sent.setdefault(lst[3], []).append(lst)
            assert lst[3] != ""
        elif lst[4] == 'ID':
            id_sent = lst
        elif lst[4] == "" and lst[0] != "" and not done_with_initial_claims and i != 0:
            initial_claims.append(lst)
        elif lst[4] != "":
            reduce_to_standard_form.append(lst)

    match_rn_sent_to_definition(untranslated_sentences, rn_sent)
    num = 0
    new_numbers = {}
    total_sent = []
    total_sent.append(total_sent2[0])
    add_to_total_sent("","")
    for lst in initial_claims:
        num += 1
        new_numbers.update({lst[0]: num})
        lst[0] = num
        total_sent.append(lst)

    if untranslated_sentences != []:
        add_to_total_sent("", "")
        add_to_total_sent("", "UNTRANSLATED PREMISES")
        total_sent.append(id_sent)
        num += 1
        for lst in untranslated_sentences:
            num += 1
            new_numbers.update({lst[0]: num})
            lst[0] = num
            total_sent.append(lst)
    else:
        total_sent.append(id_sent)

    add_to_total_sent("", "")
    add_to_total_sent("", "REDUCE TO STANDARD FORM")

    for lst in reduce_to_standard_form:
        num += 1
        new_numbers.update({lst[0]: num})
        lst[0] = num
        total_sent.append(lst)

    add_to_total_sent("", "")
    add_to_total_sent("", "INFERENCES", "____________")

    for lst in inferences:
        num += 1
        new_numbers.update({lst[0]: num})
        lst[0] = num
        total_sent.append(lst)

    renumber_sentences(new_numbers)

    if nonstandard_sentences != []:
        add_to_total_sent("", "")
        add_to_total_sent("", "NONSTANDARD SENTENCES")
        for sent in nonstandard_sentences:
            add_to_total_sent(sent[58], sent[0])

    irrelevant_sent_found = False
    if detach_sent != []:
        for sent in detach_sent:
            if sent[71] == 'irrelevant':
                if not irrelevant_sent_found:
                    irrelevant_sent_found = True
                    add_to_total_sent("", "")
                    add_to_total_sent("", "IRRELEVANT SENTENCES")
                add_to_total_sent(sent[58], sent[0])

        standard_sent_found = False
        for sent in detach_sent:
            if sent[71] != 'irrelevant':
                if not standard_sent_found:
                    standard_sent_found = True
                    add_to_total_sent("", "")
                    add_to_total_sent("", "DETACHED SENTENCES")
                add_to_total_sent(sent[58], sent[2] + sent[72])

    if attach_sent != []:
        add_to_total_sent("", "")
        add_to_total_sent("", "HYPOTHETICAL SENTENCES")
        for sent in attach_sent:
            add_to_total_sent(sent[2], sent[37])

    print_variables()
    print_object_properties()

def renumber_sentences(new_numbers):
    # this gives attach_sent their proper number according to the new
    # numbering system as arrived at in the rearrange_total_sent function

    if new_numbers != {}:
        for i in range(len(attach_sent)):
            attach_sent[i][2] = new_numbers.get(attach_sent[i][2])
            assert attach_sent[i][2] != None

        for lst in detach_sent:
            lst[58] = new_numbers.get(lst[58])
            assert lst[58] != None

        for lst in total_sent:
            lst[5] = new_numbers.get(lst[5],"")
            lst[6] = new_numbers.get(lst[6],"")
            lst[7] = new_numbers.get(lst[7],"")


def match_rn_sent_to_definition(definitions, rn_sent):
    for definiendum, v in rn_sent.items():
        for i, definition in enumerate(definitions):
            if "." in definiendum:
                definiendum2 = definition[4]
            else:
                definiendum2 = definition[4][3:]
            if definiendum == definiendum2:
                substitutions = v
                for substitution in substitutions:
                    definitions.insert(i + 1, substitution)
                break
        else:
            g = 4 / 0

def do_sent_match(list1, list2, i):
    for j in standard_slots_w_neg():
        if j != i:
            if list1[j] != list2[j]:
                return False
    return True


def get_id_sent():
    # this function gets the list of identities
    for i in range(len(total_sent)):
        if total_sent[i][4] == "ID":
            list1 = [""] * 9
            list1[0] = total_sent[i][0]
            list1[1] = total_sent[i][1]
            return list1


def print_variables():
    # this prints out the variables within the total_sent list, just above
    # where it prints the attached sentences

    ## variable type = [general, defn, indef, same_sent]


    identities = get_id_sent()
    add_to_total_sent("", "")
    add_to_total_sent("", identities[1])

    if variable_type[0] != []:
        add_to_total_sent("", 'General Variables: ' + " ".join(variable_type[0]))

    if variable_type[1] != []:
        add_to_total_sent("", 'Indefinite Variables: ' + " ".join(variable_type[1]))

    if variable_type[2] != []:
        add_to_total_sent("", 'Constants: ' + " ".join(variable_type[2]))

    if variable_type[3] != []:
        add_to_total_sent("", 'Relevant Abbreviations: ' + " ".join(variable_type[3]))


# object_properties, abbrev, general, class, accidental properties, parts of properties
# if object is a thing then it is stated whether or not it is consequential


def get_general_variables(sent):
    # modify this if you allow for more variables beyond 14
    # modify this is you allow for more than 2 disjuncts
    ant_var = set()
    con_var = set()
    whole = []
    for i in (34, 35):
        for subsent in sent[i]:
            for j in [5, 14, 18]:
                if isvariable(subsent[j]) and i == 34:
                    if subsent[j] not in ant_var:
                        if subsent[9] == 'W' and j == 5:
                            whole.append(subsent[j])
                        ant_var.add(subsent[j])
                elif isvariable(subsent[j]) and i == 35:
                    if subsent[j] not in con_var:
                        if subsent[9] == 'W' and j == 5:
                            whole.append(subsent[j])
                        con_var.add(subsent[j])
    general_variables = ant_var.intersection(con_var)
    for var in general_variables:
        if var not in variable_type[0]:
            variable_type[0].append(var)
            variable_type[3].append(var)
        if var in variable_type[1]:
            variable_type[1].remove(var)
    if sent[3] == 'e':
        for var in ant_var:
            if var not in variable_type[2] and var not in variable_type[0] \
                    and var not in abbreviations[1].values() and var not in whole:
                # the var not in whole is a band-aid and is a temporary solution until i find something
                # better
                variable_type[0].append(var)
                variable_type[3].append(var)
                if var in variable_type[1]:
                    variable_type[1].remove(var)
        for var in con_var:
            if var not in variable_type[2] and var not in variable_type[0] \
                    and var not in abbreviations[1].values() and var not in whole:
                variable_type[0].append(var)
                variable_type[3].append(var)
                if var in variable_type[1]:
                    variable_type[1].remove(var)


def get_variable_type():
    indefinite_concept = abbreviations[1].get("indefinite")
    for abbrev in abbreviations[0].keys():
        if isvariable(abbrev):
            variable_type[2].append(abbrev)

    for sent in all_sent:
        for i in [5, 14, 18]:
            if sent[i] == 'u':
                bb = 8
            if isvariable(sent[i]) and sent[i] not in variable_type[1] \
                    and sent[i] not in variable_type[0]:
                if sent[9] == "J" and sent[14] == indefinite_concept and i == 5:
                    if sent[i] not in variable_type[1]:
                        variable_type[1].append(sent[i])
                        if sent[i] in variable_type[2]: variable_type[2].remove(sent[i])
                elif sent[i] not in variable_type[2]:
                    variable_type[2].append(sent[i])

    variable_type[2].append("i")
    print_variables(variable_type)


def rearrange_all_sent(all_sent2):
    for lst in all_sent2:
        lst[3] = lst[0]
        if lst[53] == None:
            lst[53] = 'z'
    return all_sent2


def determine_relevance(to_be_defined):
    irrelevant = []
    for var in variable_type[0]:
        if var not in variable_type[3]:
            variable_type[3].append(var)

    for lst in to_be_defined:
        if not is_relevant(lst):
            lst[71] = "irrelevant"
            irrelevant.append(lst)
            g = findposinmd(lst[42], detach_sent, 42)
            if g != -1:
                lst[58] = detach_sent[g][58]
                del detach_sent[g]

    if irrelevant != []:
        add_to_total_sent("", "")
        add_to_total_sent("", "IRRELEVANT SENTENCES")
        for lst in irrelevant:
            if lst[58] != None:
                add_to_total_sent(lst[58], lst[0])
            else:
                add_to_total_sent("", lst[0])


def star_indef_var(list1, i):
    if isinmdlist(list1[42], detach_sent, 42):
        return list1
    for j in noun_slots():
        if ex(list1, i):
            if j != i and (list1[j] in variable_type[0]):
                list1[j] = list1[j] + "*"
        else:
            break
    return list1

def is_an_exception(lst):
    if abbreviations[0].get(lst[14]) == 'thing':
        return True
    elif lst[9] == 'EX':
        return True
    elif lst[9] == 'J' and abbreviations[0].get(lst[14]) == 'extant':
        return True
    else:
        return False

def is_biconditional(lst, var, object_properties, obj_class, exception):
    if var not in variable_type[0]:
        return False
    elif lst[53][-1] != 'b' and lst[53][-1] != 'f':
        return False
    else:
        sent_type = lst[53][-1]
        object_values = object_properties.get(var)
        if sent_type == 'b':
            class_slot = 0
            prop_slot = 1
        elif sent_type == 'f':
            class_slot = 3
            prop_slot = 4
            if len(object_values) == 3:
                for k in range(3):
                    object_values.append([])
        if obj_class != "":
            object_values[class_slot].append(obj_class)
        if lst[9] == "I" and obj_class == "" and not exception:
            object_values[prop_slot].append(lst)
        elif lst[9] != "I" and not exception:
            object_values[prop_slot].append(lst)

        object_properties[var] = object_values

        return True


def get_object_properties(consistent, to_be_defined, loop_number):
    if not consistent:
        return
    global object_properties
    groups = {}
    to_be_defined2 = copy.deepcopy(to_be_defined)
    to_be_defined2 = rearrange_all_sent(to_be_defined2)
    if loop_number == 1:
        for var in variable_type[3]:
            object_properties.update({var: [[], [], []]})
    else:
        for var in variable_type[3]:
            if var not in object_properties.keys():
                object_properties.update({var: [[], [], []]})

    for lst in to_be_defined2:
        if lst[1] == 'y':
            bb = 8
        if isinmdlist(lst[42], detach_sent, 42):
            lst[53] = "detached"

        for i in noun_slots():
            if not ex(lst, i):
                break
            elif lst[i] != None and isvariable(lst[i], "i") and lst[i] in variable_type[3] \
                    and lst[71] != 'irrelevant':
                if lst[i] == 'j':
                    bb = 8
                obj_class = get_class(lst[9], lst, i)
                if obj_class == 'whole' and lst[53] == 'b':
                    groups.update({lst[i]: lst[14]})
                else:
                    if lst[i] == 'o':
                        bb = 8
                    lst2 = copy.deepcopy(lst)
                    lst2 = star_indef_var(lst2, i)
                    var = lst[i]
                    lst2[i] = alpha
                    lst2[0] = build_sent_standard(lst2)
                    lst2[i] = var
                    exception = is_an_exception(lst)
                    biconditional = is_biconditional(lst2, lst[i], object_properties, obj_class, exception)

                    if not biconditional:
                        object_values = object_properties.get(lst[i])
                        if obj_class != "":
                            if obj_class not in object_values[0]:
                                object_values[0].append(obj_class)
                                lst2[67] = obj_class
                        if lst[9] == "I" and lst[53][-1] != "q" and obj_class.endswith("-object") and not exception:
                            object_values[1].append(lst2)
                        elif lst[9] != "I" and lst[53][-1] != "q" and not exception:
                            object_values[1].append(lst2)
                        elif lst[53][-1] == "q":
                            object_values[2].append(lst2)
                        else:
                            lst2[69] = 'detached' # this is because to find out if there are category
                            # errors every sentence must be accounted for
                            object_values[2].append(lst2)
                        object_properties[lst[i]] = object_values

    object_properties = categorize_groups(groups, object_properties, to_be_defined2)

    return object_properties


def categorize_groups(groups, object_properties, to_be_defined2):
    for key in groups.keys():
        lst = object_properties.get(key)
        if 'whole' not in lst[0]:
            lst[0].append('whole')
        object_properties[key] = lst

    for group, member in groups.items():
        for lst in to_be_defined2:
            for i in noun_slots():
                if not ex(lst, i):
                    break
                elif lst[i] == member and not isinmdlist(lst[42], detach_sent, 42) and lst[9] != "W":
                    lst2 = copy.deepcopy(lst)
                    if lst[9] == "I":
                        species = abbreviations[0].get(lst[14])
                        lst2[0] = species
                    else:
                        lst2[i] = delta
                        lst2[0] = build_sent_standard(lst2)
                    object_values = object_properties.get(group)
                    object_values[1].append(lst2)
                    if object_values[0] == []:
                        object_values[0] = 'whole'
                    object_properties[group] = object_values

    return object_properties


def have_same_properties(particular_properties, general_properties, gen_var, det_var, instantiations, indef=False):
    gen_properties = copy.deepcopy(general_properties)
    part_prop = copy.deepcopy(particular_properties)
    potential_instantiations = []
    subscripts = [l1, l2, l3, l4, l5, l6, l7]
    i = -1
    while gen_properties != [] and i < len(gen_properties) - 1:
        i += 1
        gen_properties[i]
        starred = False
        if "*" in gen_properties[i][0]:
            pos = gen_properties[i][0].find("*")
            gen_properties[i][0] = gen_properties[i][0].replace("*", "")
            attached_indef_var = gen_properties[i][0][pos - 1]
            if attached_indef_var in subscripts:
                attached_indef_var = gen_properties[i][0][pos - 2: pos]
            len_gen = len(gen_properties[i][0])
            starred = True

        j = -1
        while j < len(part_prop) - 1:
            j += 1
            if starred and len(part_prop[j][0]) == len_gen and not indef:
                temp_prop = part_prop[j][0]
                detached_ind_var = temp_prop[pos - 1]
                if detached_ind_var in subscripts:
                    detached_ind_var = temp_prop[i][0][pos - 2: pos]
                det_indef_prop = object_properties.get(detached_ind_var)
                att_ind_prop = object_properties.get(attached_indef_var)
                if att_ind_prop == None:
                    bb = 8
                if att_ind_prop[0] != [] and det_indef_prop != None and \
                                detached_ind_var != attached_indef_var:
                    set1 = set(att_ind_prop[0]).difference(set(det_indef_prop[0]))
                    if len(set1) == 0:
                        match = have_same_properties(det_indef_prop[1], att_ind_prop[1], attached_indef_var,
                                                     detached_ind_var, instantiations, True)
                        if match:
                            temp_str = temp_prop[:pos - 1] + attached_indef_var + temp_prop[pos:]
                            if temp_str == gen_properties[i][0]:
                                if [attached_indef_var, detached_ind_var, [], ""] not in instantiations:
                                    potential_instantiations = [attached_indef_var, detached_ind_var, [], ""]
                                del gen_properties[i]
                                i -= 1
                                break

            elif starred and indef and gen_var != det_var:
                if len(part_prop[j][0]) == len(gen_properties[i][0]):
                    temp_str = part_prop[j][0][:pos - 1] + attached_indef_var + part_prop[j][0][pos:]
                    if temp_str == gen_properties[i][0] and part_prop[j][2] == gen_properties[i][2]:
                        del gen_properties[i]
                        i -= 1
                        break
            else:
                if part_prop[j][0] == gen_properties[i][0] and part_prop[j][2] == gen_properties[i][2]:
                    del gen_properties[i]
                    i -= 1
                    break

    if gen_properties == []:
        if potential_instantiations != []:
            instantiations.append(potential_instantiations)
            do_not_instantiate.setdefault(potential_instantiations[0], []).append(potential_instantiations[1])
        return True
    else:
        return False


def has_opp_con_prop(particular_properties, general_con_prop):
    for gen_prop in general_con_prop:
        for part_prop in particular_properties:
            if gen_prop[0] == part_prop[0] and gen_prop[2] != part_prop[2] \
                and gen_prop[69] != 'detached' and part_prop[69] != 'detached':
                return True
    return False



def instantiate():
    instantiations = []
    dict1 = {0:1, 1:4, 2:7} # the key is g and the value is index of the property in lst
    for var in variable_type[0]:
        lst = object_properties.get(var)
        no_of_sets = len(lst) // 3
        for g in range(no_of_sets):
            general_classes = set(lst[g * 3])
            for k, v in object_properties.items():
                forbidden = do_not_instantiate.get(var, [])
                if k not in variable_type[0] and k not in forbidden:
                    if var == 'u' and k == 't':
                        bb = 8
                    gen_prop_no = dict1.get(g) # this is the index of the properties
                    particular_class = set(v[0])
                    set1 = general_classes.difference(particular_class)
                    if len(set1) == 0 or lst[0] == []:
                        use_lemma_of_entity = False
                        if lst[gen_prop_no] != []:
                            match = have_same_properties(v[1], lst[gen_prop_no],
                                                         var, k, instantiations)

                        elif lst[0] != []:
                            match = True
                        else:
                            use_lemma_of_entity = True
                            match = has_opp_con_prop(v[1], lst[2])

                        if match:
                            if [var, k, [], ""] not in instantiations:
                                if use_lemma_of_entity:
                                    # the T means we have to add on to the detach sent list
                                    # that the aforesaid object is a thing
                                    instantiations.append([var, k, [], "T"])
                                    do_not_instantiate.setdefault(var, []).append(k)
                                else:
                                    instantiations.append([var, k, [], ""])
                                    do_not_instantiate.setdefault(var, []).append(k)

        employ_lemma_of_entity(instantiations)

    return instantiations


def employ_lemma_of_entity(instantiations):
    global sn
    for var, lst in object_properties.items():
        group = lst[0]
        if var == "x":
            bb = 8
        if 'thing' in group or group == []:
            if isinmdlist(var, instantiations, 1):
                for sent in detach_sent:
                    if sent[5] == var and sent[9] == 'I' and sent[14] == abbreviations[1].get('thing'):
                        break
                else:
                    new_sent = build_sent1(var, "", "I", abbreviations[1].get('thing'))
                    sn += 1
                    add_to_total_sent(sn, new_sent[0], new_sent[1], new_sent[2], "LY ENT")
                    new_sent[58] = sn
                    detach_sent.append(new_sent)

def link_gen_var_to_sent(instantiations):
    for k, instantiation in enumerate(instantiations):
        var = instantiation[0]
        next_sent = False
        for lst in attach_sent:
            for i in [34, 35]:
                if not next_sent:
                    for sent in lst[i]:
                        if not next_sent:
                            for j in noun_slots():
                                if ex(sent, j):
                                    if sent[j] == var:
                                        instantiations[k][2].append(lst[2])
                                        next_sent = True
                                        break
                                else:
                                    break
                        else:
                            break
                else:
                    break

    return instantiations


def step_three(negated_conjunction, consistent):
    global instan_used, all_sent

    if consistent and attach_sent != []:
        instan_used += 1

        instantiations = instantiate()

        if instantiations == []:
            to_be_defined, proof_done = make_attached_detached(consistent, [])
            return consistent, proof_done, to_be_defined

        use_axiom_of_definition(instantiations)

        link_gen_var_to_sent(instantiations)

        substitute_in_attach_sent(instantiations)

        print_instantiations(instantiations)

        consistent, to_be_defined = detach1("use modus tollens", consistent, negated_conjunction)

        to_be_defined, proof_done = make_attached_detached(consistent, to_be_defined)

        return consistent, proof_done, to_be_defined

    else:

        to_be_defined, proof_done = make_attached_detached(consistent, [])

    return consistent, proof_done, to_be_defined



def add_consistency_sign(consistent):
    if not consistent:
        return
    global sn
    sn += 1
    list1 = []
    bool1 = False
    conjunction = ""
    for sent in total_sent:
        if sent[2].startswith("_____"):
            bool1 = True
        if bool1 and os(sent[2]) and not sent[2].startswith("_____"):
            if [sent[3], sent[2]] not in list1:
                list1.append([sent[3], sent[2]])
                conjunction += sent[3] + sent[2] + " "
    add_to_total_sent(sn, "", conjunction)
    list1 = sorted(list1, key=operator.itemgetter(1))
    conjunction = ""
    for j in list1:
        conjunction += j[0] + j[1] + " "
    sn += 1
    add_to_total_sent(sn, "", conjunction)
    sn += 1
    add_to_total_sent(sn, "", consist, "",consist + "I", sn - 1)


def exclusive_classes():
    return ['moment', 'relationship', 'point', 'number',
            'imagination', 'concept' + un, "property" + un, 'property',
            'possible world', 'letter', 'mind', 'matter', 'sensorium']

def use_basic_lemmas(consistent):
    if not consistent:
        return consistent
    for var, value in object_properties.items():
        j = 0
        if var in variable_type[1] or var in variable_type[2]:
            groups = value[0]
            if len(groups) > 1:
                for group in groups:
                    if group in exclusive_classes():
                        g = findposinmd(group, value[1], 67)
                        b = 1
                        if g == -1:
                            g = findposinmd(group, value[2], 67)
                            b = 2
                        temp_sent = value[b][g]
                        if isinmdlist(temp_sent[42], detach_sent, 42):
                            j += 1
                        if j == 1:
                            first_sent = temp_sent
                        elif j > 1:
                            second_sent = temp_sent
                            first_sent[58] = findinlist(first_sent[42], detach_sent, 42, 58)
                            second_sent[58] = findinlist(second_sent[42], detach_sent, 42, 58)
                            consistent = add_basic_lemmas(var, first_sent, second_sent)
                            return consistent

    return consistent


def add_basic_lemmas(object, first_sent, second_sent):
    concept_thing = abbreviations[1].get("thing")
    if concept_thing == None:
        concept_thing = add_thing_to_abbreviations()
    obj_pos1 = 5 if first_sent[5] == object else 14
    obj_pos2 = 5 if second_sent[5] == object else 14
    sec_obj_pos = 5 if obj_pos2 == 14 else 14
    second_obj = second_sent[sec_obj_pos]
    first_obj_pos = 5 if obj_pos1 == 14 else 14
    first_obj = first_sent[first_obj_pos]

    build_original_lemma(obj_pos1, obj_pos2, second_sent, first_sent)

    int_thing = build_rename_sent(object, first_obj, concept_thing)

    build_renamed_lemma(obj_pos1, obj_pos2, second_sent, first_sent, object, concept_thing, int_thing)

    consistent = infer_from_lemmas(concept_thing, second_obj, int_thing, first_sent, second_sent)

    return consistent


def build_original_lemma(obj_pos1, obj_pos2, second_sent, first_sent):
    global sn
    thing_sent = "(d I e)"
    thing_concept = "(e = thing)"
    if obj_pos1 == 5 and obj_pos2 == 5:
        cond1 = "(" + "b" + " " + first_sent[9] + " " + "c" + ")"
        cond2 = "(" + "b" + " ~ " + second_sent[9] + " " + "d" + ")"
        name = "SS"
    elif obj_pos1 == 5 and obj_pos2 == 14:
        cond1 = "(" + "b" + " " + first_sent[9] + " " + "c" + ")"
        cond2 = "(" + "d" + " ~ " + second_sent[9] + " " + "b" + ")"
        name = "SO"
    elif obj_pos1 == 14 and obj_pos2 == 5:
        cond1 = "(" + "c" + " " + first_sent[9] + " " + "b" + ")"
        cond2 = "(" + "b" + " ~ " + second_sent[9] + " " + "d" + ")"
        name = "OS"
    elif obj_pos1 == 14 and obj_pos2 == 14:
        cond1 = "(" + "c" + " " + first_sent[9] + " " + "b" + ")"
        cond2 = "(" + "d" + " ~ " + second_sent[9] + " " + "b" + ")"
        name = "OO"

    original_conditional = "((" + cond1 + " & " + thing_sent + ") " + implies + " " \
                           + cond2 + ") & " + thing_concept

    add_to_total_sent("", "", "")
    add_to_total_sent("", "BASIC LEMMA USED", "")
    lemma_name = "LE." + first_sent[9] + "." + second_sent[9] + "." + name
    sn += 1
    add_to_total_sent(sn, original_conditional, "", "", lemma_name)


def build_rename_sent(key, first_obj, concept_thing):
    global sn
    int_thing = variables[0]
    del variables[0]
    sent1 = "(b" + idd + variables[0] + ")"
    sent1a = "(" + variables[0] + mini_c + key + ")"
    del variables[0]
    sent2 = "(c" + idd + variables[0] + ")"
    sent2a = "(" + variables[0] + mini_c + first_obj + ")"
    del variables[0]
    sent3 = "(d" + idd + variables[0] + ")"
    sent3a = "(" + variables[0] + mini_c + int_thing + ")"
    del variables[0]
    thing_sent = "(e" + idd + concept_thing + ")"
    list1 = [sent1, sent1a, sent2, sent2a, sent3, sent3a, thing_sent]
    rename_sent = " ".join(list1)
    sn += 1
    add_to_total_sent(sn, rename_sent, "", total_sent[-1][4], "RN")

    return int_thing


def build_renamed_lemma(obj_pos1, obj_pos2, second_sent, first_sent,
                        key, concept_thing, int_thing):
    global sn
    if obj_pos1 == 5 and obj_pos2 == 5:
        cond1 = "(" + key + " " + first_sent[9] + " " + first_sent[14] + ")"
        cond2 = "(" + key + " ~ " + second_sent[9] + " " + int_thing + ")"
    elif obj_pos1 == 5 and obj_pos2 == 14:
        cond1 = "(" + key + " " + first_sent[9] + " " + first_sent[14] + ")"
        cond2 = "(" + int_thing + " ~ " + second_sent[9] + " " + key + ")"
    elif obj_pos1 == 14 and obj_pos2 == 5:
        cond1 = "(" + first_sent[5] + " " + first_sent[9] + " " + key + ")"
        cond2 = "(" + key + " ~ " + second_sent[9] + " " + int_thing + ")"
    elif obj_pos1 == 14 and obj_pos2 == 14:
        cond1 = "(" + first_sent[5] + " " + first_sent[9] + " " + key + ")"
        cond2 = "(" + int_thing + " ~ " + second_sent[9] + " " + key + ")"
    thing_sent = "(" + int_thing + " I " + concept_thing + ")"

    cond1p = first_sent[42]
    cond2p = name_sent(cond2)
    thing_sentp = name_sent(thing_sent)
    full_conditional = "(" + cond1 + " & " + thing_sent + ") " + implies + " " + cond2
    abbrev_conditional = "(" + cond1p + " & " + thing_sentp + ") " + implies + " " + cond2p
    sn += 1
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "SUB", sn - 2, sn - 1)


def infer_from_lemmas(concept_thing, second_obj, int_thing, first_sent, second_sent):
    global sn
    thing_sent = "(" + second_obj + " I " + concept_thing + ")"
    thing_sentp = name_sent(thing_sent)
    full_conditional = second_sent[3] + " " + implies + " " + thing_sent
    abbrev_conditional = second_sent[42] + " " + implies + " " + thing_sentp
    sn += 1
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "LY ENT")

    rename_sent = "(" + int_thing + mini_c + second_obj + ")"
    sn += 1
    add_to_total_sent(sn, rename_sent, "", "", "IN", sn - 1, sn - 2)

    sec_antec = "(" + second_obj + " I " + concept_thing + ")"
    sec_antecp = name_sent(sec_antec)
    second_sent[8] = "~"
    second_sent2 = build_sent2(copy.deepcopy(second_sent), 2)
    full_conditional = "(" + first_sent[3] + " & " + sec_antec + ") " + \
                       implies + " " + second_sent2[0]
    abbrev_conditional = "(" + first_sent[42] + " & " + sec_antecp + ") " + \
                         implies + " " + second_sent2[42]
    sn += 1
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "SUB", sn - 1, sn - 3)

    sn += 1
    add_to_total_sent(sn, thing_sent, thing_sentp, "", "MP", sn - 3, second_sent[58])

    conjunction = "(" + first_sent[3] + " & " + sec_antec + ")"
    conjunctionp = "(" + first_sent[42] + " & " + sec_antecp + ")"
    assert "~" not in first_sent[42]
    sn += 1
    add_to_total_sent(sn, conjunction, conjunctionp, "", "&I", first_sent[58], sn - 1)
    sn += 1
    consistent = add_to_total_sent_consist(sn, second_sent2[72],
                                           second_sent2[1], second_sent2[2], "MP", sn - 1, sn - 3, [])

    return consistent


def add_thing_to_abbreviations():
    concept_thing = variables[0]
    del variables[0]
    abbreviations[0].update({concept_thing: 'thing'})
    abbreviations[1].update({'thing': concept_thing})
    d = findposinmd("ID", total_sent, 4)
    total_sent[d][1] += "(" + concept_thing + " = thing)"

    return concept_thing


def use_axiom_of_definition(instantiations):
    global sn
    for var_list in instantiations:
        if var_list[3] == "T":
            list1 = [None] * 80
            list1[5] = var_list[1]
            list1[9] = "I"
            list1[14] = abbreviations[1].get("thing")
            list1 = build_sent2(list1)
            if not isinmdlist(list1[1], detach_sent, 1):
                sn += 1
                list1[58] = sn
                add_to_total_sent(sn, list1[0], list1[1], "", "LE ENT")
                detach_sent.append(list1)


def substitute_in_attach_sent(instantiations):
    # this substitutes the attached variable with the detached variables

    if instantiations == []:
        return

    attach_sent2 = copy.deepcopy(attach_sent)
    for sent in attach_sent2:
        sent[25] = {}
    for sent in attach_sent:
        sent[26] = 'not new'

    for instantiation in instantiations:
        for sent_num in instantiation[2]:
            for sent in attach_sent2:
                if sent[2] == sent_num:
                    sent[46] = 'instantiated'
                    if sent_num == 43:
                        bb = 8
                    sent[26] = 'new conditional from instantiation'
                    if instantiation[0] in sent[25].keys():
                        new_attach_sent = copy.deepcopy(sent)
                        # new_attach_sent[48] = instantiation[5]
                        new_attach_sent[25][instantiation[0]] = instantiation[1]
                        attach_sent2.append(new_attach_sent)
                        break
                    else:
                        sent[25].update({instantiation[0]: instantiation[1]})
                        # if sent[48] == "":
                        #     sent[48] = instantiation[5]
                        # else:
                        #     sent[49] = instantiation[5]

    i = -1
    while i < len(attach_sent2) - 1:
        i += 1
        if attach_sent2[i][26] == None:
            del attach_sent2[i]
            i -= 1

    for sent in attach_sent2:
        for j in [34, 35, 32, 31, 30, 29]:
            if sent[j] == []:
                break
            else:
                for sub_sent in sent[j]:
                    for k in [5, 14, 18, 22]:
                        if sub_sent[k] == None and k > 14: break
                        if sub_sent[k] in sent[25].keys():
                            sub_sent[k] = sent[25].get(sub_sent[k])
                            sub_sent[74] = True

    for cond_sent in attach_sent2:
        cond_sent = make_new_attach_sent(cond_sent)
        if not isinmdlist(cond_sent[4], attach_sent, 4):
            attach_sent.append(cond_sent)
        else:
            print ('hey you')


def make_new_attach_sent(cond_sent):
    # this builds new strings within the conditional list

    prop_var_greek = cond_sent[47]
    prop_var_greek2 = prop_var_greek
    for j in [34, 35, 32, 31, 30, 29]:
        for k in range(len(cond_sent[j])):
            atom_cond_sent = cond_sent[j][k]
            if atom_cond_sent[74]:
                abs_oldp = atom_cond_sent[1]
                atom_cond_sent = build_sent2(atom_cond_sent)
                if j == 34:
                    n = 0
                    q = 42
                elif j == 35:
                    n = 1
                    q = 43
                else:
                    print("you haven't coded for this yet")
                    g = 4 / 0

                for m in range(len(cond_sent[n])):
                    if cond_sent[n][m][0] == abs_oldp:
                        cond_sent[n][m][0] = atom_cond_sent[1]
                for m in range(len(cond_sent[38])):
                    if cond_sent[38][m] == abs_oldp:
                        cond_sent[38][m] = atom_cond_sent[1]
                for m in range(len(cond_sent[q])):
                    if cond_sent[n][m][0] == atom_cond_sent[1]:
                        cond_sent[q][m][0] = atom_cond_sent[72]

                cond_sent = build_conjunction(cond_sent, q)

            prop_var_greek = prop_var_greek.replace(atom_cond_sent[44], atom_cond_sent[0])
            prop_var_greek2 = prop_var_greek2.replace(atom_cond_sent[44], atom_cond_sent[42])
    cond_sent[4] = prop_var_greek2
    cond_sent[37] = prop_var_greek

    return cond_sent


def build_conjunction(list1, q):
    if q == 42:
        n = 7
        o = 0
    else:
        n = 8
        o = 1

    if len(list1[q]) > 1:
        str1 = ""
        str1p = ""
        for i in range(len(list1[q])):
            if str1 == "":
                str1 = list1[q][i][1] + list1[q][i][0]
                str1p = list1[o][i][1] + list1[o][i][0]
            else:
                str1 += " & " + list1[q][i][1] + list1[q][i][0]
                str1p += " & " + list1[o][i][1] + list1[o][i][0]

        list1[n] = [str1p, ""]
        list1[q - 2] = [str1, ""]
    else:
        list1[n] = list1[o][0]
        list1[q - 2] = list1[q][0]

    return list1


def print_instantiations(instantiations):
    # this adds the instantiations to the total_sent list
    global sn
    rearrange = False
    if instantiations != []:
        if total_sent[-1][4] == 'AX ENT':
            rearrange = True
            list2 = [""] * 9
            total_sent.insert(-1, list2)
            list2 = [""] * 9
            list2[1] = "INSTANTIATIONS"
            total_sent.insert(-1, list2)
        else:
            add_to_total_sent("", "")
            add_to_total_sent("", "INSTANTIATIONS")

        for instantiation in instantiations:
            str1 = "(" + instantiation[0] + mini_c + instantiation[1] + ")"
            #
            #     if rearrange:
            #         total_sent.insert(-1, [instantiation[5], str1, "", "", "IN", "", "", "", ""])
            #         rearrange = False
            #     else:
            add_to_total_sent("", str1, "", "", "IN")

        for cond in attach_sent:
            g = findposinmd(cond[4], total_sent, 2)
            if g == -1:
                if cond[26] == "new conditional from instantiation":
                    sn += 1
                    anc1 = cond[2]
                    cond[2] = sn
                    num = cond[49] if cond[49] != None else ""
                    add_to_total_sent(cond[2], cond[37], cond[4], "", "SUB", anc1, cond[48], num)

        add_to_total_sent("", "")
        add_to_total_sent("", "INFERENCES FROM INSTANTIATION")


def print_object_properties():
    if object_properties == {}:
        return

    add_to_total_sent("", "")
    add_to_total_sent("", "OBJECT PROPERTIES")

    for k, v in object_properties.items():
        if len(v) > 3 and v[0] == [] and v[1] == [] and v[2] == []:
            e = 1
            f = 2
        elif len(v) > 3:
            e = 0
            f = 2
        else:
            e = 0
            f = 1

        for m in range(e, f):
            if m == 0:
                b, c, d = 0, 1, 2
            elif m == 1:
                b, c, d = 3, 4, 5

            classes = " ".join(v[b])
            if k == "j":
                bb = 8
            if classes == "":
                classes = "thing"
            str1 = k + " | " + classes
            properties = []
            if v[c] != []:
                for lst in v[c]:
                    properties.append(lst[0])
                properties2 = " ".join(properties)
                if len(properties2) > 5 and len(str1) > 40:
                    add_to_total_sent("", str1)
                    str1 = k
                str1 += " | " + properties2
            if v[d] != []:
                properties = []
                for lst in v[d]:
                    if lst[69] != 'detached':
                        properties.append(lst[2] + lst[0])
                if properties != []:
                    str1 += " [" + " ".join(properties) + "]"

            add_to_total_sent("", str1)



def get_class(relat, sent, p):
    # this determines what class or category an object belongs to
    if sent[2] == "~":
        kind = ''
    elif relat == "A" or (relat == 'T' and p == 14):
        kind = 'moment'
    elif relat == 'AB' or relat == "L" or relat == 'AB' or (relat == 'S' and p == 14):
        kind = 'point'
    elif relat == "G" or (relat == 'N' and p == 14):
        kind = 'number'
    elif relat == "M" and p == 5 or (relat == 'B' and p == 14):
        kind = 'relationship'
    elif relat == "M" and p == 14:
        kind = 'imagination'
    elif relat == "I" and p == 5:
        group = sent[14]
        kind = abbreviations[0].get(group)
        if kind == 'thing':
            kind = ""
        elif kind == None:
            kind = sent[14] + "-object"
    elif relat == "I" and p == 14:
        kind = "concept" + un
    elif relat == "H" and p == 14:
        kind = "property" + un
    elif relat == "J" and p == 14:
        kind = "property"
    elif relat == "W" and p == 5:
        kind = "whole"
    elif relat == "W" and p == 14:
        kind = ''
    elif relat == 'P' and p == 14:
        kind = 'possible world'
    elif relat == "D" and p == 14:
        kind = 'relationship'
    elif relat == 'AL':
        kind = 'letter'
    elif (relat == 'B' or relat == "D") and p == 5:
        kind = 'mind'
    elif relat == "S" and p == 5:
        kind = 'matter'
    elif relat == "O" and p == 14:
        kind = 'sensorium'
    else:
        kind = ""


    return kind


def kind_exception(str1):
    # since everything belongs to the class 'whole' or 'part' these are not
    # genuine classes
    exceptions = ['whole', 'part']
    if str1 in exceptions:
        return 'thing'
    elif str1 == None:
        return 'thing2'
    return str1
    # if str1 equals none then that means the subject belongs to an indefinite
    # concept


def get_quick_variable_type(variable, variable_type):
    # this tells us what type a certain variable is after we already know what it is

    if variable == "":
        return ""
    elif variable in variable_type[0]:
        return 'agen'
    elif variable in variable_type[2]:
        return 'definite'
    elif variable in variable_type[1]:
        return 'indefinite'


def get_nonstandard_sent(nonstandard_sentences, loop_number):
    if loop_number != 1:
        return
    i = 0
    while i < len(detach_sent):
        if detach_sent[i][42] == 'k':
            bb = 8
        if detach_sent[i][76] == None:
            detach_sent[i][76] = is_standard(detach_sent[i])
        if not detach_sent[i][76]:
            nonstandard_sentences.append(detach_sent[i])
            del detach_sent[i]
        else:
            i += 1

# never used
def is_in_md(list1, i, str1, bool1=False, k=0):
    if not bool1:
        for j in range(len(list1)):
            if list1[j][i] == str1:
                return True
        return False
    else:
        for j in range(len(list1)):
            if j != k:
                if list1[j][i] == str1:
                    return True
        return False


def determine_if_all_cond_4_detach_met(g, k):
    # this loops through all the sentences in the antecedent or the consequent
    # and determines if they all have been detached

    conditions = copy.deepcopy(attach_sent[g][k])
    del conditions[0]
    ancestors = []
    sent_type = attach_sent[g][3]
    done = False
    i = 0
    while i < len(attach_sent[g][k]) - 1 and not done:
        i += 1
        for j, sent in enumerate(detach_sent):
            temp = sent[1]

            if sent[1] == 'h' + l1:
                bb = 8
            if sent[1] == attach_sent[g][k][i][0] \
                    and sent[2] == attach_sent[g][k][i][1]:
                ancestors.append(sent[58])
                del conditions[0]
                if conditions == []:
                    output = ["", ancestors]
                    done = True
                break

            elif sent[1] == attach_sent[g][k][i][0] \
                    and sent[2] != attach_sent[g][k][i][1]:
                if (sent_type == "e" and k == 0) or k == 1:
                    done = True
                    output = ['a sentence is negated', j]
                elif sent_type == 'c' and k == 0:
                    done = True
                    output = ["", ""]
                break
        else:
            done = True
            output = ["", ""]

    return output


def detach1(str1, consistent, negated_conjunction):
    global st_log_time
    b = time.time()
    if str1 == 'do not use modus tollens':
        kind = '~MT'
    else:
        kind = 'MT'
    if not consistent:
        return False, []
    if attach_sent == []:
        return True, []
    set_of_det_sent = []
    begin_num_of_detach_sent = len(detach_sent)
    r = -1
    while consistent and r < len(detach_sent) - 1:
        r += 1
        if r == 46:
            bb = 8
        g = -1
        while consistent and g < len(attach_sent) - 1:
            g += 1
            if attach_sent[g][26] != 'not neww':
                k = -1
                while consistent and k < 1:
                    k += 1
                    sent_type = attach_sent[g][3]
                    if attach_sent[g][3] == "e" or attach_sent[g][3] == "c":
                        temp_detach_sent = detach_sent[r][1]
                        det_tvalue = detach_sent[r][2]
                        temp_attach_sent = attach_sent[g][k][0][0]
                        att_tvalue = attach_sent[g][k][0][1]

                        if temp_detach_sent == 'i':
                            bb = 8
                        if temp_detach_sent == temp_attach_sent:
                            if det_tvalue == att_tvalue and (k == 0 or (k == 1 and sent_type == "e")):
                                rule = iff + "E" if sent_type == 'e' else "MP"
                                if len(attach_sent[g][k]) == 1:

                                    # if k = 0 and sent_type = 'c' and set_of_det_sent = []
                                    # then p & (p > q)
                                    # if k = 0 and sent_type = 'e' and set_of_det_sent = []
                                    # then p & (p <> q)
                                    # if k = 1 and sent_type = 'e' and set_of_det_sent = []
                                    # then p & (q <> p)
                                    # if k = 0 and sent_type = 'c' and set_of_det_sent != []
                                    # then p & r & ((p & r) > q)
                                    # if k = 0 and sent_type = 'e' and set_of_det_sent != []
                                    # then p & r & ((p & r) <> q)
                                    # if k = 1 and sent_type = 'e' and set_of_det_sent != []
                                    # then p & r & (q <> (p & r))

                                    consistent, g, k, r = detach2(k, r, g, rule,
                                                                  set_of_det_sent, negated_conjunction)
                                    if g > len(attach_sent) - 1 or g == -1:
                                        break
                                else:
                                    output = determine_if_all_cond_4_detach_met(g, k)
                                    if output[0] == 'a sentence is negated':
                                        rule = "EN" if sent_type == 'e' else "MT"
                                        consistent, g, k, r = detach2(k, output[1], g, rule, [], negated_conjunction)
                                        if g > len(attach_sent) - 1 or g == -1:
                                            break
                                    elif output[1] != "":
                                        output[1].insert(0, detach_sent[r][58])
                                        consistent, g, k, r = detach2(k, r, g, rule, output[1], negated_conjunction)
                                        if g > len(attach_sent) - 1 or g == -1:
                                            break
                            else:
                                if k == 0:
                                    t = 27
                                else:
                                    t = 28
                                if k == 0 and sent_type == 'c':
                                    pass
                                elif kind == "MT" and attach_sent[g][t] and \
                                        ((k == 0 and sent_type == 'e') or (k == 1)):
                                    rule = "EN" if sent_type == 'e' else "MT"
                                    consistent, g, k, r = detach2(k, r, g, rule, set_of_det_sent, negated_conjunction)
                                    if g > len(attach_sent) - 1 or g == -1:
                                        break

    new_sent = get_new_detach_sent(begin_num_of_detach_sent)
    c = time.time()
    c = c - b
    st_log_time += c

    return consistent, new_sent


def detach2(k, r, g, rule, set_of_det_sent, negated_conjunction):
    global sn

    if sn == 43:
        bb = 8

    anc1 = attach_sent[g][2]
    anc2 = detach_sent[r][58]
    sn += 1
    if sn == 28:
        bb = 8
    if k == 0:
        m, h, t, s, n = 41, 43, 8, 1, 35
    else:
        m, h, t, s, n = 40, 42, 7, 0, 34

    introduce_conjunction(set_of_det_sent)

    if attach_sent[g][t][0] == 't':
        bb = 8

    consistent = build_negated_conjunction(anc1, anc2, g, h, m, negated_conjunction, rule, s, t)

    if not consistent:
        return consistent, len(attach_sent), 0, 0

    consistent = add_to_total_sent_consist(sn,
                                           attach_sent[g][m][0],
                                           attach_sent[g][t][0],
                                           attach_sent[g][m][1],
                                           rule,
                                           anc1,
                                           anc2,
                                           negated_conjunction)

    consistent, r = add_sent(consistent, g, h, n, negated_conjunction, r, rule, s, sn)

    del attach_sent[g]
    if g + 1 == len(attach_sent):
        g -= 1
    k = -1
    return consistent, g, k, r


def build_negated_conjunction(anc1, anc2, g, h, m, negated_conjunction, rule, s, t):
    consistent = True
    if rule == 'EN' or rule == 'MT':
        if attach_sent[g][m][1] == "":
            attach_sent[g][m][1] = "~"
            attach_sent[g][t][1] = "~"
        else:
            print('you need to add double negative eventuaully')
            attach_sent[g][m][1] = ""
            attach_sent[g][t][1] = ""

        if len(attach_sent[g][h]) > 1:
            negated_conjunction.append([attach_sent[g][s],
                                        attach_sent[g][m],
                                        attach_sent[g][t],
                                        []])
            consistent = check_consistency_w_neg_conj(negated_conjunction[-1], rule, anc1, anc2)
    return consistent


def add_sent(consistent, g, h, n, negated_conjunction, r, rule, s, sn):
    if rule != 'MT' and rule != "EN":
        if len(attach_sent[g][h]) == 1:
            if consistent:
                if os(attach_sent[g][s][0][0]):
                    list3 = attach_sent[g][n][0]
                    list3[58] = sn
                    is_in_detach_sent = isinmdlist(list3[42], detach_sent, 42)
                    if not is_in_detach_sent:
                        detach_sent.append(list3)
                else:
                    for lst in attach_sent[g][39]:
                        lst[2] = sn
                        if not isinmdlist(lst[4], attach_sent, 4):
                            attach_sent.append(lst)
                            r = reset_r(lst[38], r)

        else:
            consistent, r = eliminate_conjuncts(g, r, h, negated_conjunction)
    return consistent, r


def reset_r(atomic_sentences, r):
    for sent in atomic_sentences:
        det_sent_pos = findposinmd(sent, detach_sent, 1)
        if det_sent_pos != -1 and det_sent_pos < r:
            r = det_sent_pos

    return r


def introduce_conjunction(set_of_det_sent):
    global sn
    full_sent = ""
    if set_of_det_sent != []:
        for num in set_of_det_sent:
            d = findposinmdlistint(num, detach_sent, 58)
            if d == -1: g = 4 / 0
            if full_sent == "":
                full_sent = detach_sent[d][0]
                abbrev_sent = detach_sent[d][42]
                anc = str(detach_sent[d][58])
            else:
                full_sent += " & " + detach_sent[d][0]
                abbrev_sent += " & " + detach_sent[d][42]
                anc += "," + str(detach_sent[d][58])
        full_sent = "(" + full_sent + ")"
        abbrev_sent = "(" + abbrev_sent + ")"
        add_to_total_sent(sn, full_sent, abbrev_sent, "", "&I", anc)
        sn += 1


def eliminate_conjuncts(g, r, h, negated_conjunction):
    # if the detached sentences are a conjunction then this function
    # places each individual conjunct into the total_sent and detach_sent list
    global sn
    num = copy.copy(sn)
    if h == 43:
        m = 41
        k = 1
        n = 35
        c = 8
    else:
        m = 40
        k = 0
        n = 34
        c = 7
    conjunct_list = attach_sent[g][k]

    for i in range(len(conjunct_list)):
        sn += 1
        consistent = add_to_total_sent_consist(sn, attach_sent[g][h][i][0], attach_sent[g][k][i][0],
                                               attach_sent[g][h][i][1], "&E", num, "", negated_conjunction)
        if consistent:
            if os(conjunct_list[i][0]):
                d = findposinmd_alert_error(conjunct_list[i][0], attach_sent[g][n], 1)
                sent_parts = attach_sent[g][n][d]
                sent_parts[58] = sn
                is_in_detach_sent = isinmdlist(sent_parts[42], detach_sent, 42)
                if not is_in_detach_sent:
                    detach_sent.append(sent_parts)
            else:
                d = findposinmd_alert_error(conjunct_list[i][0], attach_sent[g][39], 4)
                list2 = attach_sent[g][39][d]
                list2[2] = sn
                if not isinmdlist(list2[4], attach_sent, 4):
                    attach_sent.append(list2)
                    r = reset_r(list2[38], r)
        else:
            break

    return consistent, r


def get_new_detach_sent(begin_num_of_detach_sent):
    if detach_sent == []:
        return []
    new_sent = []
    for i in range(begin_num_of_detach_sent - 1, len(detach_sent)):
        if not isinmdlist(detach_sent[i][1], all_sent, 1):
            new_sent.append(detach_sent[i])
    return new_sent


def add_to_total_sent_consist(num, str1, str2, tvalue, rule, anc1, anc2, negated_conjunction):
    list2 = [""] * 9
    list2[0] = num
    list2[1] = str1
    list2[2] = str2
    list2[3] = tvalue
    list2[4] = rule
    list2[5] = anc1
    list2[6] = anc2
    total_sent.append(list2)
    consistent = check_consistency(negated_conjunction)

    return consistent


def add_to_total_sent(num, str1, str2="", tvalue="", rule="", anc1="", anc2="", anc3=""):
    list2 = [""] * 9
    list2[0] = num
    list2[1] = str1
    list2[2] = str2
    list2[3] = tvalue
    list2[4] = rule
    list2[5] = anc1
    list2[6] = anc2
    list2[7] = anc3
    total_sent.append(list2)


def check_consistency(negated_conjunction):
    new_sent_abbr = total_sent[-1][2]
    tvalue = total_sent[-1][3]
    for i in range(len(total_sent) - 2, -1, -1):
        if total_sent[i][2] == new_sent_abbr and total_sent[i][3] != tvalue:
            build_contradiction(i)
            return False

    for lst in negated_conjunction:
        for sent in lst[0]:
            if sent[0] == new_sent_abbr and sent[1] == tvalue:
                del sent[0]
                del sent[1]
                lst[3].append(total_sent[-1][0])
                if lst[0] == []:
                    build_contradictory_conjunction(lst[3])

    return True


def check_consistency_w_neg_conj(neg_conj, rule, anc1, anc2):
    global sn
    i = -1
    while neg_conj[0] != [] or i < len(neg_conj[0]) - 1:
        i += 1
        for j in range(len(total_sent) - 1, 0, -1):
            if total_sent[j][1].startswith("INFE"):
                return
            if total_sent[j][2] == neg_conj[0][i][0] and total_sent[j][3] == neg_conj[0][i][1]:
                neg_conj[3].append(j)
                del neg_conj[0][i]
                i -= 1
                break
    if neg_conj[0] == []:
        sn += 1
        add_to_total_sent(sn, neg_conj[1][0], neg_conj[2][0], "~", rule, anc1, anc2)
        build_contradictory_conjunction(neg_conj[3])
        return False


def build_contradictory_conjunction(list1):
    global sn
    sn += 1
    list3 = tuple(str(total_sent[k][0]) for k in list1)
    anc1 = ",".join(list3)
    list2 = [total_sent[j][3] + total_sent[j][1] for j in list1]
    list2p = [total_sent[j][3] + total_sent[j][2] for j in list1]
    conjunction = "(" + " & ".join(list2) + ")"
    conjunctionp = "(" + " & ".join(list2p) + ")"
    add_to_total_sent(sn, conjunction, conjunctionp, "", "&I", anc1)
    build_contradiction(-2)


def build_contradiction(i):
    global sn
    sn += 1
    str1 = total_sent[-1][1] + " & ~" + total_sent[i][1]
    str2 = total_sent[-1][2] + " & ~" + total_sent[i][2]
    total_sent.append([sn, str1, str2, "", "&E", total_sent[-1][0], total_sent[i][0], "", ""])
    sn += 1
    total_sent.append([sn, bottom, bottom, "", bottom + "I", sn - 1, "", "", ""])


def disjunction_heirarchy(str5, d, new_disj=False):
    global prop_name
    global sn, pn

    if d > len(attach_sent) - 1:
        return
    if iff in str5 or conditional in str5:
        return

    str5 = enclose(str5)
    def_info = find_sentences(str5)
    mainc = def_info[4][0][1]
    list2 = [""] * 60
    n = 7
    if mainc == xorr:
        list2[3] = 'x'
    else:
        list2[3] = 'd'
    if attach_sent == [] or new_disj:
        list2[2] = pn
    else:
        list2[2] = attach_sent[d][2]
    list2[5] = ""
    list2[4] = def_info[0][0]  # fix this
    sentences = []

    for i in range(len(def_info[0])):
        if os(def_info[0][i]):
            siblings = []
            list3 = [None] * 9
            n += 1
            # str1 = findinlist(def_info[0][i],prop_name,1,0)
            str2 = def_info[4][i][0][:-1]
            g = findposinlist(str2, def_info[4], 0)
            parent = def_info[0][g]
            if def_info[4][g][1] == "&":
                list3[2] = 'c'
            elif def_info[4][g][1] == xorr:
                list3[2] = 'x'
            else:
                list3[2] = 'd'
            if len(str2) > 1:
                str3 = def_info[4][i][0][:-2]
                g = findposinlist(str3, def_info[4], 0)
                gparent = def_info[0][g]
            else:
                gparent = parent
            list3[1] = def_info[4][i][0]
            list3[5] = parent
            list3[6] = gparent
            list3[0] = [def_info[0][i], def_info[1][i]]
            # fix this
            b = parent.count(xorr)
            c = parent.count(idisj)
            if c > 1 or b > 1:
                list3[7] = 2
            else:
                list3[7] = 1
            sent_num = def_info[4][i][0]
            m = len(sent_num)
            for j in range(len(def_info[4])):
                if len(def_info[4][j][0]) == m and def_info[4][j][0][:-1] == str2 \
                        and j != i:
                    siblings.append([def_info[0][j], def_info[1][j]])  # fix this
            list3[4] = siblings
            list2[n] = list3
            sentences.append(list3[0][0])
            if list3[0][0] not in rel_conj:
                rel_conj.append(list3[0][0])
    list2[36] = def_info

    if attach_sent == [] or new_disj:
        list2[38] = sentences
        attach_sent.append(list2)
    else:
        list2[38] = sentences
        list2[2] = attach_sent[d][2]
        list2[37] = attach_sent[d][37]
        attach_sent[d] = list2


def proper_spacing(str1):
    str1 = str1.replace(" ", "")
    str1 = str1.replace(iff, " " + iff + " ")
    str1 = str1.replace(conditional, " " + conditional + " ")
    str1 = str1.replace(idisj, " " + idisj + " ")
    str1 = str1.replace(xorr, " " + xorr + " ")
    str1 = str1.replace("&", " & ")
    return str1


def proper_spacing2(str1):
    str1 = str1.replace(iff, " " + iff + " ")
    str1 = str1.replace(conditional, " " + conditional + " ")
    str1 = str1.replace(idisj, " " + idisj + " ")
    str1 = str1.replace(xorr, " " + xorr + " ")
    str1 = str1.replace("&", " & ")
    return str1


def bad_paren(str1):
    if str1.find("(") == -1:
        return str1
    # we first must get rid of strings of the following form ((p) & s)
    for i in range(len(str1)):
        str2 = str1[i:i + 1]
        str3 = str1[i - 1:i]
        if i > 1:
            str4 = str1[i - 2:i - 1]
        else:
            str4 = ""
        str5 = str1[i + 1:i + 2]
        if str2.islower() and str3 == "(" and str5 == ")":
            str1 = str1[:i - 1] + str1[i:i + 1] + str1[i + 2:]
        elif str2.islower() and str4 == "(" and str3 == "~" and str5 == ")":
            str1 = str1[:i - 2] + str1[i - 1:i + 1] + str1[i + 2:]

    str1 = enclose(str1)
    list1 = find_sentences(str1)
    mstr = list1[3][0]
    for i in range(1, len(list1[3])):
        if list1[4][i][1] != "":
            mc = list1[4][i][1]
            ostr = list1[3][i]
            str2 = list1[4][i][0][:-1]
            prcnt = findinlist(str2, list1[4], 0, 1)
            if mc == prcnt:
                nstr = remove_outer_paren(ostr)
                nstr = remove_outer_paren(nstr)
                mstr = mstr.replace(ostr, nstr)
    return mstr


def unenclose(str1):
    # this removes ( ) from around a sentence abbreviation
    i = -1
    if "(" not in str1:
        return str1

    while i < len(str1) - 1:
        i += 1
        str2 = str1[i:i + 1]
        str3 = str1[i - 1:i]
        str4 = str1[i + 1:i + 2]
        if str2.islower() and str3 != "~" and str4 not in subscripts:
            str1 = str1[:i - 1] + str2 + str1[i + 2:]
        elif str2.islower() and str3 == "~" and str4 not in subscripts:
            str1 = str1[:i - 2] + str3 + str2 + str1[i + 2:]
        if str2.islower() and str3 != "~" and str4 in subscripts:
            str1 = str1[:i - 1] + str2 + str4 + str1[i + 3:]
        elif str2.islower() and str3 == "~" and str4 in subscripts:
            str1 = str1[:i - 2] + str3 + str2 + str4 + str1[i + 3:]

    return str1


def new_disjunct(str1, ng, n, candd, conjt, anc1, anc2, anc3=None, anc4=None, kind=0, rule=""):
    global sn, pn
    list2 = mainconn(str1)
    if kind == 1:
        del attach_sent[n]
        consistent = new_prop(str1, ng, "&I", anc1, anc2, anc3, anc4)
        return consistent
    elif kind == 2:
        consistent = new_prop(str1, ng, "&I", anc1, anc2, anc3, anc4)
        return consistent
    else:
        if os(str1):
            del attach_sent[n]
            str1 = remove_outer_paren(str1)
            list1 = tilde_removal2(str1)
            str1 = list1[0]
            consistent = new_prop(str1, list1[1], rule + "E", anc1, anc2)
            candd.append([pn, list1[0], list1[1]])
            conjt.append([pn, list1[0], list1[1]])
            return consistent
        elif list2[0] == "&":
            del attach_sent[n]
            str1 = remove_outer_paren(str1)
            new_prop(str1, ng, rule + "E", anc1, anc2)
            g = copy.copy(pn)
            list3 = get_conjuncts(str1)
            for i in range(len(list3)):
                list4 = tilde_removal2(list3[i])
                consistent = new_prop(list4[0], list4[1], "&E", g, "")
                if no_output == False:
                    return consistent
                if list3[i].find(idisj) > -1:
                    disjunction_heirarchy(list4[0], n, True)
                else:
                    candd.append([pn, list4[0], list4[1]])
                    conjt.append([pn, list4[0], list4[1]])
            return True
        else:
            consistent = new_prop(str1, ng, idisj + "E", anc1, anc2)
            if consistent == False:
                return consistent
            if ng == "~":
                str1 = ng + str1
            else:
                str1 = remove_outer_paren(str1)
            attach_sent[n][2] = pn
            disjunction_heirarchy(str1, n, False)
            return True


def xorr_elim(n, i, parent, grandparent, whole_d, candd, anc1, anc2, conjt, kind=0):
    str9 = ""
    de_mor = False
    if kind == 0:
        for r in range(len(attach_sent[n][i][4])):
            if r != i:
                if not os(attach_sent[n][i][4][r][0]):
                    de_mor = True
                if str9 == "":
                    str9 += "~" + attach_sent[n][i][4][r][1] + attach_sent[n][i][4][r][0]
                else:
                    str9 += " & ~" + attach_sent[n][i][4][r][1] + attach_sent[n][i][4][r][0]
    else:
        grandp2 = copy.copy(grandparent)
        grandp2 = grandp2.replace(parent, "")
        for r in range(8, 38):
            if attach_sent[n][r] == "":
                break
            if attach_sent[n][r][0][0] in grandp2:
                if str9 == "":
                    str9 += "~" + attach_sent[n][r][0][1] + attach_sent[n][r][0][0]
                else:
                    str9 += " & ~" + attach_sent[n][r][0][1] + attach_sent[n][r][0][0]
    g = copy.copy(pn)
    if parent != grandparent:
        str9 = remove_outer_paren(str9)
        if grandparent == whole_d:
            mc = mainconn(str9)
            if mc[0] == '&':
                consistent = xorr_elim2(str9, candd, conjt, anc1, anc2)
                if consistent == False:
                    return consistent
            else:
                list4 = tilde_removal(str9)
                consistent = new_prop(list4[0], list4[1], xorr + "E", anc1, anc2)
                if consistent == False:
                    return consistent
        else:
            str9 = "(" + str9 + ")"
            if kind == 0:
                str9 = grandparent.replace(parent, str9)
                str9 = whole_d.replace(grandparent, str9)
                str9 = bad_paren(str9)
                consistent = new_prop(str9, "", xorr + "E", anc1, anc2)
                if str9.find("~~") > -1:
                    str9 = str9.replace("~~", "")
                    consistent = new_prop(str9, "", "~~E", pn, "")

            else:
                str9 = whole_d.replace(grandparent, str9)
                str9 = bad_paren(str9)
                consistent = new_prop(str9, "", xorr + "E", anc1, anc2)
                g = copy.copy(pn)
                if str9.find("~~") > -1:
                    str9 = str9.replace("~~", "")
                    consistent = new_prop(str9, "", "~~E", g, "")
                    if consistent == False:
                        return consistent
                disjunction_heirarchy(str9, n, True)
                del attach_sent[n]
            if de_mor:
                list1 = demorgan(all_sent, attach_sent, total_sent, detach_sent, candd, True, str9, pn)
                consistent = list1[0]
                attach_sent = list1[1]
                if consistent == False:
                    return consistent
            else:
                if str9.find(idisj) > -1 or str9.find(xorr) > -1:
                    disjunction_heirarchy(str9, n, True)
                consistent = True
    else:
        # this does not account for the case where the parent == grandparent but
        # grandparent does not == whole d
        consistent = xorr_elim2(str9, candd, conjt, anc1, anc2)
    return consistent


def xorr_elim2(str9, candd, conjt, anc1, anc2):
    str9 = bad_paren(str9)
    consistent = new_prop(str9, "", xorr + "E", anc1, anc2)
    if consistent == False:
        return False
    if str9.find("~~") > -1:
        str9 = str9.replace("~~", "")
        consistent = new_prop(str9, "", "~~E", pn, "")
        if consistent == False:
            return consistent
    list3 = get_conjuncts(str9)
    g = copy.copy(pn)
    for b in range(len(list3)):
        list4 = tilde_removal2(list3[b])
        list4[0] = remove_outer_paren(list4[0])
        consistent = new_prop(list4[0], list4[1], "&E", g, "")
        if consistent == False:
            return consistent
        if not os(list3[b]):
            if list4[1] == "~":
                list1 = demorgan(all_sent, attach_sent, total_sent, detach_sent, candd, list3[b], pn, "&E")
                consistent = list1[0]
                attach_sent = list1[1]
                if consistent == False:
                    return False
            else:
                disjunction_heirarchy(list4[0], n, True)
        else:
            candd.append([pn, list4[0], list4[1]])
            conjt.append([pn, list4[0], list4[1]])

    return True


def disjunction_elimination(candd, kind=""):
    bool1 = False
    bool2 = False
    global sn, pn
    global rel_conj

    for i in range(len(attach_sent)):
        if attach_sent[i][8] == "":
            disjunction_heirarchy(attach_sent[i][4], i)
    i = -1
    conjt = copy.deepcopy(candd)

    while i < len(conjt) - 1:
        i += 1
        if conjt[i][1] not in rel_conj:
            del conjt[i]
            i -= 1
    d = -1
    while d < len(conjt) - 1:
        d += 1
        str2 = conjt[d][2]
        conj = conjt[d][1]
        if conj == 'q':
            bb = 7
        if d == 42:
            bb = 7
        anc1 = conjt[d][0]
        n = -1
        while n < len(attach_sent) - 1:
            if bool1:
                bool1 = False
                d = -1
                break
            n += 1
            if n == 7:
                bb = 7
            i = 7
            while attach_sent != []:
                if bool2:
                    bool2 = False
                    break
                i += 1
                if conj not in attach_sent[n][38] and iff not in attach_sent[n][4] \
                        and conditional not in attach_sent[n][4]:
                    break
                else:
                    if attach_sent[n][i] == "":
                        break
                    whole_d = attach_sent[n][4]
                    anc2 = attach_sent[n][2]
                    str3 = attach_sent[n][i][0][0]
                    # 'pos or neg'
                    str4 = attach_sent[n][i][0][1]
                    # 'disjunct or conjunct
                    str5 = attach_sent[n][i][2]
                    # 'disjunct number
                    str6 = attach_sent[n][i][1]

                    if conj == str3:
                        grandparent = attach_sent[n][i][6]
                        parent = attach_sent[n][i][5]
                        parent2 = copy.copy(parent)
                        parent3 = copy.copy(parent)
                        str7 = " " + idisj + " "
                        str7a = " " + xorr + " "
                        if str2 == str4 and str5 == "d":
                            # 'if the disjuncts are not embedded within a conjunct then the disjunction
                            # is simply deleted

                            del attach_sent[n]
                            if parent != grandparent:
                                conj = str2 + conj
                                str8 = whole_d.replace(parent, conj)
                                disjunction_heirarchy(str8, n)
                            bool1 = True
                            n = -1
                            break

                        elif str2 == str4 and str5 == "x":

                            consistent = xorr_elim(n, i, parent, grandparent, whole_d, candd, anc1, anc2)
                            if not consistent:
                                return False
                            del attach_sent[n]
                            bool2 = True
                            bool1 = True
                            d = -1
                        elif str2 == str4 and str5 == "c":
                            list2 = []
                            list2.append([conj, str2])
                            anc3 = ""
                            anc4 = ""
                            list11 = attach_sent[n][i][4]
                            f = -1
                            while f < len(list11) - 1:
                                mc = mainconn(grandparent)
                                f += 1

                                for e in range(len(candd)):
                                    anc5 = candd[e][0]
                                    # since it's too hard to program, if the sibling is a disjunct then we just
                                    # ignore this

                                    if list11[f][0].find(idisj) > -1 or list11[f][0].find(xorr) > -1:
                                        break
                                    else:
                                        if candd[e][1] == list11[f][0]:
                                            if candd[e][2] == list11[f][1]:
                                                list2.append([list11[f][0], list11[f][1]])
                                                if len(list2) == 2:
                                                    anc3 = anc5
                                                elif len(list2) == 3:
                                                    anc4 = anc5
                                                del list11[f]
                                                if list11 == []:
                                                    str3 = build_sent_list2(list2)
                                                    if mc[0] == xorr:
                                                        new_prop(str3, "", "&I", anc1, anc3, anc4)
                                                        consistent = xorr_elim(n, i, parent, grandparent, whole_d,
                                                                               candd,
                                                                               anc1, anc2, conjt, 1)
                                                        if not consistent:
                                                            return False
                                                    else:
                                                        # if the conjunct is not embedded within another conjunct
                                                        # then the disjunct is simply deleted
                                                        if whole_d == grandparent:
                                                            consistent = new_disjunct(str3, "", n,
                                                                                      candd, conjt, anc1, anc3,
                                                                                      anc4, anc5, 1)

                                                        else:
                                                            str8 = whole_d.replace(grandparent, parent2)
                                                            if str8.find("(") > -1 and str8.find(idisj) > -1:
                                                                str8 = bad_paren(str8)
                                                                consistent = new_disjunct(str3, "", n,
                                                                                          candd, conjt, anc1, "",
                                                                                          anc3, anc4, 2)

                                                            consistent = new_disjunct(str8, "", n, candd, conjt, pn - 1,
                                                                                      anc2)
                                                            if not consistent:
                                                                return False
                                                    bool1 = True
                                                    bool2 = True
                                                    n = 0
                                                    d = -1
                                                    break
                                                else:
                                                    f -= 1
                                                    break

                                            elif candd[e][2] != list11[f][1]:
                                                mc = mainconn(grandparent)
                                                if mc[0] == idisj:
                                                    rule = idisj
                                                    str7 = " " + idisj + " "
                                                else:
                                                    str7 = " " + xorr + " "
                                                    rule = xorr
                                                r = grandparent.find(parent)
                                                if r > 1:
                                                    parent = str7 + parent
                                                else:
                                                    parent = parent + str7
                                                anc1 = candd[e][0]
                                                str9 = grandparent.replace(parent, "")
                                                str8 = whole_d.replace(grandparent, str9)
                                                if str8.find("(") > -1 and (
                                                                str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                                    str8 = bad_paren(str8)
                                                consistent = new_disjunct(str8, "", n, candd, conjt, anc1, anc2, None,
                                                                          None,
                                                                          0, rule)

                                                if not consistent:
                                                    return False
                                                else:
                                                    list11 = []
                                                    bool1 = True
                                                    bool2 = True
                                                    n = 0
                                                    d = -1
                                                    break

                        elif str2 != str4 and str5 == "c":
                            mc = mainconn(attach_sent[n][i][6])
                            if mc[0] == idisj:
                                str6 = str7 + parent
                                rule = idisj
                            else:
                                rule = xorr
                                str6 = str7a + parent
                            if grandparent.find(str6) > -1:
                                parent = str6
                            else:
                                parent = parent + str7
                            str9 = grandparent.replace(parent, "")
                            str8 = whole_d.replace(grandparent, str9)
                            if str8.find("(") > -1 and str8.find(idisj) > -1:
                                str8 = bad_paren(str8)
                            consistent = new_disjunct(str8, "", n, candd, conjt, anc1, anc2, None, None, 0, rule)
                            if not consistent:
                                return False
                            bool1 = True
                            n = -1
                            break

                        elif str2 != str4 and (str5 == "d" or str5 == 'x'):
                            # if the disjunct is a triple disjunct then enter below
                            if str5 == 'd':
                                rule = idisj
                            else:
                                rule = xorr
                            if attach_sent[n][i][7] > 1:
                                str6 = str4 + str3 + " " + rule + " "
                                if parent.find(str6) > -1:
                                    str5 = str6
                                else:
                                    str5 = " " + rule + " " + str4 + str3
                                str9 = parent.replace(str5, "")
                                str8 = whole_d.replace(parent, str9)
                                if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                    str8 = bad_paren(str8)
                                consistent = new_disjunct(str8, "", n, candd, conjt,
                                                          anc1, anc2, None, None, 0, rule)

                                if not consistent:
                                    return False
                                bool1 = True
                                n = -1
                                break

                            else:
                                str3 = attach_sent[n][i][4][0][0]  # ddd
                                str4 = attach_sent[n][i][4][0][1]
                                str5 = str4 + str3
                                str8 = whole_d.replace(parent, str5)
                                if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                    str8 = bad_paren(str8)
                                consistent = new_disjunct(str8, "", n, candd,
                                                          conjt, anc1, anc2, None, None, 0, rule)
                                if not consistent:
                                    return False
                                bool1 = True
                                n = -1
                                break
    return True


def use_statement_logic(kind=""):
    global st_log_time
    b = time.time()
    list1 = detach1(kind, consistent, )
    consistent = list1[0]
    attach_sent = list1[1]
    # if consistent == False:
    #     return [False, attach_sent]
    # if kind != 2:
    #     list1 = disjunction_elimination(all_sent, attach_sent, detach_sent, \
    #                                     candd, total_sent, kind)
    #     consistent = list1[0]
    #     attach_sent = list1[1]

    c = time.time()
    d = c - b
    st_log_time += d
    return consistent


def add_outer_paren(str1):
    str1 = remove_outer_paren(str1)
    return "(" + str1 + ")"


##### tahir begin
def populate_sentences():
    global result_data

    test_sent = []
    last_row_blank = False
    row_number = 1

    if mysql == 1:
        for row in w4:
            row_number += 1
            if row[0] != "":
                test_sent.append([row[0], row[1]])
                last_row_blank = False
            elif row[1] == "" and not last_row_blank:
                last_row_blank = True
            elif row[1] != "":
                last_row_blank = False
            elif row[1] == "" and last_row_blank:
                break

                # tahir I'm not really sure what this does:

                # if not first_sent:
                #     result_data['text_' + str(row_number - 2) + '_1'] = len(test_sent)

    return test_sent, row_number


def get_number_of_sent_to_prove(len_test_sent):
    global order, start, stop
    if order[2] == "#":
        order = order[3:]
        return order
    else:
        start = order[0]
        stop = order[1]
        if stop == 0: stop = len_test_sent
        return [x for x in range(start, stop)]


def calculate_time_statistics(proof_time, nonlinear):
    global instan_used, instan_time, lemmas_used

    if nonlinear == "#":
        num_of_sent = len(order)
    else:
        num_of_sent = (stop - start)

    total = time.time() - total_time

    if lemmas_used == 0: lemmas_used = 1

    if instan_used != 0:
        ee = instan_time / instan_used
    else:
        ee = 0
    print("")
    print("average " + str("{0:.4f}".format((time.time() - proof_time) / num_of_sent)))
    print("time used in statement logic " + str("{0:.4f}".format(st_log_time / num_of_sent)))
    print("time spent reducing " + str("{0:.4f}".format(time_spent_reducing / (num_of_sent))))
    print("time used in lemma function " \
          + str("{0:.5f}".format(time_spent_in_lemma_function / lemmas_used)))
    print("time used in instantiation " + str("{0:.4f}".format(ee)))
    print("time used in change variables function " + str("{0:.4f}".format(time_spent_defining / num_of_sent)))
    print("total " + str("{0:.3f}".format(total)))
    print("")


def get_result(post_data, archive_id=None, request=None, input=None):
    global ws, w4, result_data, order, propositional_constants
    global sn, total_sent, prop_name, variable_type, object_properties
    global all_sent, attach_sent, detach_sent, definite_assignments
    global prop_var, variables, stop, abbreviations, dictionary, do_not_instantiate

    ########## tahir begin
    if mysql == 1 and not input:
        archive = Archives.objects.latest('archives_date')
        test_sent, row_number = pop_sent()
    elif input:
        # "It is|a contradictory that I do not have many|n points"
        test_sent = [[0, input]]
        row_number = 1

    else:
        test_sent, row_number = pop_sent()
    dictionary = large_dict()
    not_oft_def = copy.deepcopy(dictionary[6])

    _, _, order = info()
    nonlinear = order[2]
    if mysql == 2:
        order = [0]
    else:
        order = get_number_of_sent_to_prove(len(test_sent))
    check_mispellings(test_sent)
    time_used_proving_sent = time.time()

    if mysql == 1:
        views.progressbar_send(request, 0, 100, 0, 1)
    for j, k in enumerate(order):
        if mysql == 1:
            views.progressbar_send(request, start, stop, k, 1)
        if k == 18:
            bb = 7
        st1 = time.time()
        prop_name = []
        total_sent = []
        all_sent = []
        attach_sent = []
        detach_sent = []
        definite_assignments = {}
        do_not_instantiate = {}
        object_properties = {}
        dictionary[6] = not_oft_def
        variable_type = [[], [], [], []]
        abbreviations = [{}, {}, {}, ""]
        propositional_constants = {}
        prop_var = copy.deepcopy(prop_var4)
        variables = copy.deepcopy(variables2)

        truth_value = step_one(test_sent[k])

        consistent = step_two(truth_value)

        test_sent[k] = copy.deepcopy(total_sent)
        tot_prop_name.append(prop_name)
        # progress(j+1, len(order))
        if not consistent:
            print(str(k) + " - " + str("{0:.3f}".format(time.time() - st1) + " False"))
        else:
            print(str(k) + " - " + str("{0:.3f}".format(time.time() - st1)))

    calculate_time_statistics(time_used_proving_sent, nonlinear)

    determine_words_used()

    # cProfile.run("get_result('hey')")

    print_sent_full(test_sent, tot_prop_name, row_number)

    if mysql == 1:
        views.progressbar_send(request, 0, 100, 100, 2)
        views.save_result(archive_id, result_data)
        return result_data


##############################################

########## THE CODE BEGINS HERE


if mysql == 0:
    get_result('hey')

    if proof_type == 1:
        wb4.save('/Users/kylefoley/Desktop/inference engine/temp_proof.xlsx')
    if get_words_used == 1:
        wb5.save('/Users/kylefoley/Desktop/inference engine/dictionary4.xlsx')
