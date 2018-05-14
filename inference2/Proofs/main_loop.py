import time
import sys
from openpyxl import load_workbook

try:
    from natural_language import step_one
    from general_functions import parameters
    from classes import ErrorWithCode
    from settings import *
except:
    from .natural_language import step_one
    from .general_functions import parameters
    from .classes import ErrorWithCode
    from .settings import *


# try:
#     from natural_language import step_one
#     from general_functions import parameters
#     from classes import ErrorWithCode
#     from settings import *
# except:
#     from .natural_language import step_one
#     from .general_functions import parameters
#     from .classes import ErrorWithCode
#     from .settings import *


def calculate_time_statistics(num_proved, total_time):
    total_time = time.time() - total_time
    print("")
    print("average " + str("{0:.4f}".format(total_time / num_proved)))
    print("total " + str("{0:.3f}".format(total_time)))
    print("")
    print(num_proved)


def determine_words_used(words_used):
    wb5 = load_workbook('/Users/kylefoley/Desktop/inference_engine/dictionary5.xlsx')
    w5 = wb5.worksheets[0]
    for word in words_used:
        j = dictionary.words_to_row.get(word, 28)
        if j == 28:
            print(word)
        elif j == None:
            print(word)
        else:
            try:
                w5.cell(row=j, column=2).value = 1
            except:
                pass
    wb5.save('/Users/kylefoley/Desktop/inference_engine/dictionary5.xlsx')
    words = open('words_used.pkl', 'wb')
    pickle.dump(words_used, words)
    words.close()


def print_on_error(order, dictionary, test_sent, print_type, lemmata, user):
    global words_used
    j = -1
    num_proved = 0
    correct = 0
    accurracy = []
    while j < len(order) - 1:
        j += 1
        k = order[j]

        if test_sent[k][0] != 'pass':
            num_proved += 1
            st1 = time.time()

            if k == 182:
                bb = 7
            # print (k)
            try:

                _ = step_one(dictionary, user, test_sent, lemmata, k)
                consistent, total_sent, twords_used = _
                if total_sent != 'skip':

                    test_sent[k] = json.loads(json.dumps(total_sent))
                    words_used = words_used | twords_used

                else:
                    if k in order:
                        order.remove(k)
                        j -= 1

                correct = reaction(consistent, k, print_type, st1, accurracy, correct)

            except:
                print("bug")
                accurracy.append("WRONG - bug")

                if k in order:
                    order.remove(k)
                    j -= 1


        elif print_type[0] in ["0", "4"] and k % 50 == 0:
            print(k)

    if len(test_sent) > 1:
        b = correct / len(test_sent)
        b = b * 100
        b = int(b)
        accurracy.append(str(b) + "% accuracy")
        print(accurracy[-1])

    return num_proved, accurracy


def reaction(consistent, k, print_type, st1, accurracy, correct):
    if print_type[0] in ['1', "2"]:
        if consistent:
            print('RIGHT')
            accurracy.append("RIGHT")
            correct += 1
        else:
            print('WRONG')
            accurracy.append("WRONG")


    elif print_type[0] != "4":
        if not consistent:
            if print_type[0] == "3":
                print(str(k) + " - " + str("{0:.3f}".format(time.time() - st1) + " False"))
            elif print_type[0] == "0":
                print(str(k) + " - False")
                sys.exit()


        elif print_type[0] == "3":
            print(str(k) + " - " + str("{0:.3f}".format(time.time() - st1)))
        elif print_type[0] == "1":
            print(str(k) + " - True")
    elif print_type[0] == "4" and not consistent:
        print(str(k) + " - False")

    return correct


def stop_if_error(order, dictionary, test_sent, print_type, lemmata, user):
    global words_used
    j = -1
    num_proved = 0

    while j < len(order) - 1:
        j += 1
        k = order[j]

        if test_sent[k][0] != 'pass':
            num_proved += 1
            st1 = time.time()

            if k == 182:
                bb = 7
            # print (k)
            _ = step_one(dictionary, user, test_sent, lemmata, k)
            consistent, total_sent, twords_used = _

            if total_sent != 'skip':
                test_sent[k] = json.loads(json.dumps(total_sent))
                words_used = words_used | twords_used

            reaction(consistent, k, print_type, st1)


        elif print_type[0] in ["0", "4"] and k % 50 == 0:
            print(k)

    return num_proved


def get_result(one_sent, user="", print_type="40", order=[0], get_words_used=0):
    global words_used
    total_time = time.time()

    if user == 'gs':
        user = ""
        test_sent = one_sent

    elif one_sent == 'a':
        proof_type, print_type, get_words_used, order = parameters()
        pkl_file = open(user + 'zz_claims.pkl', 'rb')
        test_sent = pickle.load(pkl_file)
        pkl_file.close()
    elif one_sent != "":
        test_sent = one_sent
    else:
        pkl_file = open(user + 'zz_claims.pkl', 'rb')
        test_sent = pickle.load(pkl_file)
        pkl_file.close()
    pkl_file = open(user + 'z_dict_words.pkl', 'rb')
    dictionary = pickle.load(pkl_file)
    pkl_file.close()
    pkl_file = open('lemmata.pkl', 'rb')
    lemmata = pickle.load(pkl_file)
    pkl_file.close()

    words_used = set()
    # print_type = "31"

    if print_type[1] == "1":
        num_proved, accurracy = print_on_error(order, dictionary, test_sent, print_type, lemmata, user)
        test_sent.append(accurracy)
    else:
        num_proved = stop_if_error(order, dictionary, test_sent, print_type, lemmata, user)

    if print_type[1] in ["0", "2", "3", "4"]:
        calculate_time_statistics(num_proved, total_time)

    if get_words_used == 1:
        determine_words_used(words_used)

    return test_sent
