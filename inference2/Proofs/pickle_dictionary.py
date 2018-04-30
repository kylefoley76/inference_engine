from openpyxl import load_workbook
import pickle
import copy
import json, operator
import sys, time
from lemmas import determine_class
from itertools import combinations
import itertools

try:
    from settings import *
    from analyze_definition import process_sentences
    from classes import get_dictionary, row_class, relata
    from general_functions import get_last_row
except:
    from .settings import *
    from .analyze_definition import process_sentences
    from .classes import get_dictionary, row_class, relata
    from .general_functions import get_last_row


#
# from settings import *
# from analyze_definition import process_sentences
# from classes import get_dictionary, row_class, relata
# from general_functions import get_last_row


def update_synonyms(rw):
    if not rw.pos[1] in ['s', "d"]: return
    definition = rw.defin
    pos = rw.pos
    str6 = definition[definition.find("=") + 1:-1]
    str6 = str6.strip()
    str7 = definition[1:definition.find("=")]
    str7 = str7.strip()
    if rw.word != str7:
        print(f"the synonym {rw.word} does not appear in its definition")
    if pos[0] == "r":
        if not isrelat(str6[0]):
            print(f"the synonym for {str7} needs to be a capital letter")
        else:
            list1 = rw.pronounce.split(",")
            list1 = [x.strip() for x in list1]
            for word in list1:
                dictionary.synonyms.update({word: str6})

    else:
        dictionary.synonyms.update({str7: str6})

    dictionary.definitions.update({str7: definition})


def cut_word(word):
    try:
        if "(" in word:
            cc = word.index("(")
            word = word[:cc - 1]
            word = word.strip()
        else:
            word = word.strip()
    except:
        word = ""

    return word


def is_a_definition(str1):
    if str1 == None:
        return False
    if "e.g." in str1:
        return False
    formal_symbols = ["&", "=", conditional, iff, xorr]
    for n in formal_symbols:
        if n in str1:
            return True
    return False


def check_ontology():
    b = len(dictionary.ontology[1].keys())
    print(f"{b} categories")
    for group in dictionary.ontology[1].keys():
        if group not in dictionary.pos.keys():
            print(f"the category: {group} has not yet been defined")

    for word, cls in dictionary.relata.items():
        if word not in dictionary.ambiguous2.keys():
            check_ontology2(cls.subject, word)
            check_ontology2(cls.object, word)
            check_ontology2(cls.object2, word)
            check_ontology2(cls.object3, word)
            check_ontology2(cls.object4, word)

    return


def check_ontology2(str1, word):
    if str1 == None: return
    pos = dictionary.pos.get(word)
    if pos[0] not in ["n", "r", "a"]: return
    if ";" in str1:
        st = str1.split(";")
        st = [x.strip() for x in st]
        for st1 in st:
            if str1 == 's being': str1 = 'sentient being'
            if str1 not in ['thing', 'postponed']:
                if dictionary.ontology[1].get(st1) == None:
                    print(f"in {word} the category {st1} is not in your ontology")
    else:
        if str1 in ['thing', 'postponed']: return
        if str1 == 's being': str1 = 'sentient being'
        if dictionary.ontology[1].get(str1) == None:
            print(f"in {word} the category {str1} is not in your ontology")


def build_ontology(i, worksheet):
    i += 2
    j = i
    forward_onto = []
    reverse_dict = {}
    cond_properties = {}
    branched_concepts = []
    branched_ontology = []
    while True:
        i += 1
        if i > j + 35: raise Exception
        rw = row_class(worksheet, i)
        branch = False
        if rw.defin == None:
            break
        elif "e.g." not in rw.defin:
            if disj in rw.defin:
                rw.defin = rw.defin.replace(disj, xorr)
            if conditional in rw.defin:
                list2 = rw.defin.split(conditional)
            elif implies in rw.defin:
                list2 = rw.defin.split(implies)
                branch = True
            else:
                list2 = rw.defin.split(iff)
            if "," in list2[1]:
                list3 = list2[1].split(",")
            else:
                list3 = list2[1].split(xorr)
            list3 = [x.strip() for x in list3]
            if list3[0] == 'male':
                bb = 8
            if branch or list2[0].strip() in branched_concepts:
                for str1 in list3:
                    if str1 not in reverse_dict.keys():
                        branched_concepts.append(str1)
                    cond_properties.update({str1: list2[0].strip()})
                branched_ontology.append([list2[0].strip(), set(list3)])

            else:
                for str1 in list3: reverse_dict.update({str1: list2[0].strip()})
                forward_onto.append([list2[0].strip(), set(list3)])

    reverse_dict2 = {}
    for k, v in reverse_dict.items():
        set1 = {v}
        new_class = v
        set1.add(k)
        while new_class != 'thing':
            new_class = reverse_dict.get(new_class)
            assert new_class != None, (f"you misspelled {v} in your ontology")
            set1.add(new_class)
        reverse_dict2.update({k: set1})

    c = set()
    c.add("thing")
    reverse_dict2.update({"thing": c})
    reverse_dict3 = {}
    for k, v in cond_properties.items():
        set1 = {v}
        new_class = v
        set1.add(k)
        while True:
            new_class = cond_properties.get(new_class)
            if new_class == None:
                break
            set1.add(new_class)
        reverse_dict3.update({k: set1})

    excl_class = {}
    excl_class = get_mutually_exclusive_classes(forward_onto, reverse_dict2, excl_class)
    excl_class2 = {}
    excl_class2 = get_mutually_exclusive_classes(branched_ontology, reverse_dict3, excl_class2)
    excl_class = {**excl_class, **excl_class2}
    get_branched_classes(reverse_dict3, reverse_dict2, excl_class)

    forward_onto = expand_forward_ontology(forward_onto)

    dictionary.ontology = [forward_onto, reverse_dict2, excl_class]

    for k in dictionary.ontology[1].keys(): dictionary.groups.update({k: k})

    return i + 1, rw


def get_branched_classes(reverse_dict3, reverse_dict2, excl_class):
    branched = []
    branched_dict = {}
    for key, value in reverse_dict3.items():
        if key not in reverse_dict2.keys():
            branched.append(key)
            for concept in value:
                if concept in reverse_dict2.keys():
                    branched_dict.update({key: concept})
                    break

    possibilities = [branched, reverse_dict2.keys()]
    possibilities = [i for i in itertools.product(*possibilities)]
    for possibility in possibilities:
        if possibility[0] == possibility[1]:
            pass
        else:
            str1 = possibility[0] + "." + possibility[1]
            str2 = possibility[1] + "." + possibility[0]
            tv = excl_class.get(str1)
            if str1 == 'male.female':
                bb = 8

            tv2 = excl_class.get(str2)

            if tv == None:
                parent = branched_dict.get(possibility[0])
                if parent == possibility[1]:
                    tv = 'necessary'
                else:
                    str3 = parent + "." + possibility[1]
                    tv = excl_class.get(str3)
                    assert tv != None
                    str4 = possibility[1] + "." + parent

                if tv == 'necessary':
                    excl_class.update({str1: 'necessary'})
                    excl_class.update({str2: 'possible'})
                elif tv == 'impossible':
                    excl_class.update({str1: 'impossible'})
                    excl_class.update({str2: 'impossible'})
                else:
                    excl_class.update({str1: 'possible'})
                    excl_class.update({str2: tv2})

    return excl_class


def expand_forward_ontology(forward_onto):
    new_dict = {}
    for lst in forward_onto[1:]:
        new_dict.update({lst[0]: list(lst[1])})
    newer_dict = {}
    for k, children in new_dict.items():
        j = 0
        while j < len(children):
            group = children[j]
            new_child = new_dict.get(group)
            if new_child != None:
                children += new_child
            j += 1
        newer_dict.update({k: children})

    return newer_dict


def in_same_set(x, k, forward_onto):
    for lst in forward_onto:
        if x in lst[1] and k in lst[1]:
            return True
    return False


def get_mutually_exclusive_classes(forward_onto, reverse_dict2, excl_class):
    combos = combinations(reverse_dict2.items(), 2)
    for combo in combos:
        pair1 = combo[1][0] + "." + combo[0][0]
        pair2 = combo[0][0] + "." + combo[1][0]
        if pair2 == 'sexless.person' or pair2 == 'person.sexless':
            bb = 8

        if in_same_set(combo[0][0], combo[1][0], forward_onto):
            excl_class.update({pair1: "impossible"})
            excl_class.update({pair2: "impossible"})
        else:
            vname = combo[0][0]
            yname = combo[1][0]
            v = combo[0][1]
            y = combo[1][1]
            if v - y != set() and y - v != set():
                sym_diff = v ^ y
                for lst in forward_onto:
                    set2 = lst[1]
                    if len(list(set2 & sym_diff)) > 1:
                        excl_class.update({pair1: "impossible"})
                        excl_class.update({pair2: "impossible"})
                        break
                else:
                    excl_class.update({pair1: "possible"})
                    excl_class.update({pair2: "possible"})

            else:
                possible = False
                if len(v & y) == len(v):
                    possible = True
                    pair1 = vname + "." + yname
                    pair2 = yname + "." + vname
                else:
                    pair2 = vname + "." + yname
                    pair1 = yname + "." + vname

                if possible:
                    excl_class.update({pair1: "possible"})
                    excl_class.update({pair2: "necessary"})
                else:
                    excl_class.update({pair1: "possible"})
                    excl_class.update({pair2: "necessary"})

    return excl_class


def get_prepositional_relations():
    str1 = third_sheet.cell(row=2, column=5).value
    str1a = third_sheet.cell(row=3, column=5).value
    str2 = third_sheet.cell(row=6, column=5).value
    str2a = third_sheet.cell(row=7, column=5).value
    str3 = third_sheet.cell(row=10, column=5).value
    str3a = third_sheet.cell(row=11, column=5).value
    prepositional = str1.split() + str1a.split()
    non_spatio_temporal = str2.split() + str2a.split()
    spatio_temporal = str3.split() + str3a.split()
    dictionary.prepositional_relations = prepositional
    dictionary.spatio_temporal_relations = spatio_temporal
    dictionary.non_spatio_temporal_relations = non_spatio_temporal


def build_dictionary(kind2):
    global ws, wsar, third_sheet, dictionary
    dictionary = get_dictionary()
    for_superscript_check = set()
    last_word = ""
    if kind2 == "":
        get_prepositional_relations()
        worksheet = ws
        last_row = get_last_row(worksheet, 4)
        print("last row " + str(last_row))

    i = 31
    while i <= last_row:

        i += 1
        if kind2 == "":
            rw = row_class(worksheet, i)
        elif kind2 == 'mysql':
            rw = fill_mysql_row(i, rw, worksheet)

        if rw.word == "feels":
            bb = 8

        if not not_blank(rw.word) and not not_blank(rw.next_word) and i > 300:
            break

        if rw.defin == 'Ontology':
            i, rw = build_ontology(i, worksheet)

        # if rw.word != "." and rw.word != last_word and rw.pos == None:
        #     print (f"{rw.word} is not defined")

        if not_blank(rw.pos):

            if not isinstance(rw.pos, int): rw.pos = rw.pos.strip()
            if isinstance(rw.word, int): rw.word = str(rw.word)
            if rw.word == "true*": rw.word = "true"
            if rw.word == "false*": rw.word = "false"
            rw.word = cut_word(rw.word)
            rw.next_word = cut_word(rw.next_word)
            for_superscript_check.add(rw.word)

            # print (rw.word)
            put_in_categories(for_superscript_check, rw)

            assign_part_of_speech(rw)

            first_relat = rw.abbrev_relat

            relata(worksheet, i, dictionary, rw)

            # universals are not defined, synonyms and determinative nouns are already done
            if rw.pos[1] not in ['a', 'b', 's', 'd'] and not_blank(rw.defin):

                rw.defin = rw.defin.strip()
                new_hprop, i = is_new_hprop2(worksheet, i, first_relat, rw)

                if rw.next_word == rw.word and "e.g." not in rw.next_defin:
                    while is_a_definition(worksheet.cell(row=i + 1, column=6).value) \
                            and rw.next_word == rw.word and not new_hprop:
                        i += 1
                        rw.defin += "| " + worksheet.cell(row=i, column=6).value.strip()
                        rw.next_word = worksheet.cell(row=i + 1, column=4).value
                        rw.next_word = cut_word(rw.next_word)
                        new_hprop, i = is_new_hprop(worksheet, i, first_relat, rw)

                if rw.pos[0] == "r":
                    dictionary.definitions.update({rw.abbrev_relat: rw.defin})

                else:
                    dictionary.definitions.update({rw.word: rw.defin})

            elif rw.pos[0] not in ['y', 'w']:
                dict1 = {"r": "r", "a": "p"}
                if rw.pos[0] == 'r':
                    dictionary.kind[rw.abbrev_relat] = dict1.get(rw.pos[0], "c")
                else:
                    dictionary.kind[rw.word] = dict1.get(rw.pos[0], "c")

            last_word = rw.word

    check_ontology()
    check_synonyms()
    check_ambiguous_words()
    check_superscripts2(for_superscript_check)

    return dictionary


def check_ambiguous_words():
    for k, v in dictionary.ambiguous2.items():
        for word in v:
            if word not in dictionary.pos.keys() and \
                    word not in dictionary.plurals.keys():
                print(f"the word {word} is not in your dictionary even though it is ambiguous")


def check_synonyms():
    for key, syn in dictionary.synonyms.items():
        pos = dictionary.pos.get(key)
        if pos[1] != 'd' and pos[0] != 'x':
            if syn not in dictionary.pos.keys():
                print(f"{syn} does not have a synonym in your dictionary")


def is_new_hprop2(worksheet, i, first_relat, rw):
    if first_relat in ["H", "W"]:
        rw.next_word = first_relat + rw.next_word
        return False, i
    next_abbrev = worksheet.cell(row=i + 1, column=5).value
    if next_abbrev in ["H", "W"]:
        return True, i
    return False, i


def is_new_hprop(worksheet, i, first_relat, rw):
    if first_relat in ["H", "W"]:
        rw.next_word = first_relat + rw.next_word
        return False, i
    next_abbrev = worksheet.cell(row=i + 1, column=5).value
    if next_abbrev in ["H", "W"]:
        i -= 1
        return True, i
    return False, i


def check_superscripts(rw):
    if rw.word[-1] in superscripts:
        letter = superscript_dict.get(rw.word[-1])
        try:
            if rw.superscript == 'arbitrary':
                pass
            elif letter != rw.superscript[0]:
                print(f"you forgot to superscript {rw.word}")
            # assert letter == rw.word[0], f"you forgot to superscript {rw.word}"
        except:
            print(f"you forgot to superscript {rw.word}")


def check_superscripts2(for_superscript_check):
    unsuperscripted_words = set()
    unsuper_dict = {}
    for word, pos in dictionary.pos.items():
        if word[-1] in superscripts and pos[0] != 't' and pos[:2] != 'nh':
            # if we use two superscripts then they both have to be the same letter
            word = word[:-2] if word[-2] == word[-1] else word[:-1]

            if word not in dictionary.ambiguous2.keys() and \
                    word + "s" not in dictionary.ambiguous2.keys():

                if pos[0] == 'r' and word[-1] != "s":
                    word2 = word + "s"

                if word not in for_superscript_check:
                    if word[-1] == "s" and dictionary.pos.get(word[:-1]) == None:

                        print(f"{word} is not unsuperscripted")
                    elif dictionary.pos.get(word2) == None:
                        print(f"{word} is not unsuperscripted")

            if pos[0] in ['n', 'r', 'a']:
                unsuperscripted_words.add(word)
                if not isrelat(word[0]):
                    unsuper_dict.update({word: pos[0]})

    unsuper_dict = sorted(unsuper_dict.items(), key=operator.itemgetter(1))

    return


def put_in_categories(for_superscript_check, rw):
    check_superscripts(rw)

    adjust_pos(rw)

    update_plurals(for_superscript_check, rw)

    place_in_decision_procedure2(rw.word, rw)

    make_doubles(rw)

    update_synonyms(rw)

    update_relations(rw)

    fix_hrelation(rw, rw.word)

    get_popular_words(rw)

    dictionary.words_to_row.update({rw.word: rw.row_num})


def update_plurals(for_superscript_check, rw):
    if rw.pos[0] == 'n':
        dictionary.plurals.update({rw.pronounce: rw.word})
        dictionary.pos.update({rw.pronounce: "nspbbb"})
    elif rw.pos[:3] == 'yas':
        ambiguous_plural(for_superscript_check, rw)


def adjust_pos(rw):
    if len(rw.pos) == 1 and rw.pos[0] not in ["y", "w"]:
        rw.pos += "zbbb"
    elif rw.pos[0] in ['y', 'w']:
        rw.pos += 'abbb'
    else:
        rw.pos += "bbbb"


def update_relations(rw):
    if rw.pos[0] == 'r' and rw.pos[1] not in ['s', 'a'] and not not_blank(rw.abbrev_relat):
        print(f"you forgot to give {word} an abbreviation or youre calling a noun a relation")

    if not_blank(rw.abbrev_relat):
        if rw.pos[0] == 'r' and rw.abbrev_relat in dictionary.definitions.keys():
            raise Exception("the relation " + rw.abbrev_relat + " has two definitions")

        place_in_decision_procedure2(rw.abbrev_relat, rw)
        if rw.abbrev_relat not in dictionary.words_to_row.keys():
            dictionary.words_to_row.update({rw.abbrev_relat: rw.row_num})
        words_to_relation(rw)
        if len(rw.abbrev_relat) == 4 and rw.abbrev_relat[-1] == "P": dictionary.past_participles.append(rw.abbrev_relat)


def ambiguous_plural(for_superscript_check, rw):
    list1 = rw.pronounce.split(",")
    if len(list1) != 2:
        raise Exception(f" {rw.word} must have two words in the alternatives colum")
    list1 = [x.strip() for x in list1]
    dictionary.plurals.update({list1[1]: list1[0]})
    for_superscript_check.add(list1[1])


def words_to_relation(rw):
    if rw.pronounce != None:
        list1 = rw.pronounce.split(",")
        for word in list1:
            word = word.strip()
            dictionary.rel_abbrev.update({word: rw.abbrev_relat})
    else:
        dictionary.rel_abbrev.update({rw.word: rw.abbrev_relat})


def fix_hrelation(rw, word):
    if rw.abbrev_relat == 'H' and rw.pos[0] != 'r':
        if rw.pos[2] != "x":
            if dictionary.pos.get(rw.word) == None:
                print(f"your defining the H word before {word}")
        rw.word = "H" + word
    elif rw.abbrev_relat == 'W' and rw.pos[0] != 'r':
        if dictionary.pos.get(rw.word) == None:
            print(f"your defining the H word before {word}")
        rw.word = "W" + word


def assign_part_of_speech(rw):
    word = rw.word
    pos = rw.pos
    if word == 'experience':
        bb = 8

    if "(for calculation)" in word:
        list1 = []

    elif rw.pronounce != None and rw.pronounce != 'variable':
        if pos[0] in ['w', 'r', 'y']:
            list1 = rw.pronounce.split(",")
            list1 = [x.strip() for x in list1]
        elif pos[:2] in ['nh', 'nw']:
            list1 = [word]
        else:
            if rw.pronounce != word:
                if "," in word:
                    print(f" {word} has a comma in pronoun")
                list1 = [word, rw.pronounce]

            else:
                list1 = [word]

    else:
        list1 = [word]

    for word1 in list1:
        hey = dictionary.pos.get(word1)
        if hey != None:
            print(f"{rw.word} is defined twice")

        dictionary.pos.update({word1: pos})

    if pos[0] == 'r' and pos[1] != 's' and \
            rw.word != rw.abbrev_relat:
        hey = dictionary.pos.get(rw.abbrev_relat)
        if hey != None:
            print(f"{rw.word} is defined twice")

        dictionary.pos.update({rw.abbrev_relat: pos})


def get_popular_words(rw):
    if rw.pos[0] == 'r' and rw.pos[1] != 's':
        if rw.popular == 1:
            dictionary.popular.append([rw.word, rw.abbrev_relat, rw.pos[0], 0])
    elif rw.pos[1] != "s" and rw.popular == 1:
        dictionary.popular.append([rw.word, rw.word, rw.pos[0], 0])

    # if rw.pos[1] == 'a':
    #     print (f"{rw.word} has no definition")

    return


def make_doubles(rw):
    for i in range(2):
        if i == 0:
            word = rw.pronounce
        elif rw.pos[0] != 'r' and i == 1:
            word = rw.word
        else:
            break
        if word != None:
            m = word.count(" ")
            if m == 0: return
            if "," in word:
                list1 = word.split(",")
            else:
                list1 = [word]
            for tword in list1:
                tword = tword.strip()
                m = tword.count(" ")
                if m == 0:
                    pass
                elif m == 1:
                    dictionary.doubles.add(tword)
                elif m == 2:
                    dictionary.triples.add(tword)
                elif m == 3:
                    dictionary.quadruples.add(tword)
                elif m == 4:
                    raise Exception(f'{tword} is a quintuple')
                else:
                    raise Exception(f'{tword} is a sextuple')


def place_in_decision_procedure2(word, rw):
    part_of_speech = rw.pos[0]
    sub_part = rw.pos[1]
    if part_of_speech == 'x': return
    if sub_part in ['h', 'j', 'w']: return
    category = None
    if sub_part == 'd':  # determinative nouns
        category = 19
    elif word in ['distinct from', 'different']:  # special synonyms, distinct from
        category = 20
    elif sub_part == 's':  # synonyms
        category = 18
    elif word == "i":
        category = .5
    elif part_of_speech == 'd' and sub_part not in ['b', 'i', 'e', 'd']:
        category = 1  # determinative, possessive pronouns
    elif part_of_speech == 'p':
        category = 2
    elif word == 'and' + uc:
        category = 5  # and
    elif part_of_speech == 'u':  # relative pronouns
        category = 9
    elif word in ['AS', 'as']:  # AS
        category = 11
    elif word in dictionary.spatio_temporal_relations:
        category = 13
    elif word in dictionary.non_spatio_temporal_relations:
        category = 13.5
    elif word == 'there':  # there
        category = 14
    elif part_of_speech == 'd' and sub_part == 'b':  # universals
        category = 15
    elif part_of_speech == 'd' and sub_part == 'e':  # many
        category = 16
    elif part_of_speech == 'm':  # negative determiners
        category = .4

    if category != None: dictionary.decision_procedure.update({word: category})


def unembed(def_info, definiendum, red_kind):
    abbrev = []
    ogreek = def_info[5]
    type3 = False
    if red_kind == 6:
        type3 = True

    for sent, greek_sent in zip(def_info[3], def_info[6]):
        if mini_e in sent and len(greek_sent) == 1:
            abbrev.append(greek_sent)

    new_conjunct = " & ".join(abbrev)
    new_conjunct = " & " + new_conjunct

    e = 0
    special = False
    for lst1, lst2, lst3, lst4, lst6 in zip(def_info[1], def_info[2], def_info[3], def_info[4], def_info[6]):
        e += 1

        if e > 1:
            parent = ".".join(lst1[:-1])
            parent_pos = def_info[2].index(parent)
            parent_type = def_info[4][parent_pos][1]

            if parent_type == iff and parent in ['1', '1.1', '1.2']:
                if lst1[-1] == "2":

                    ogreek_definiens = lst6
                    if "&" not in ogreek_definiens:
                        special = True

                    break

    for gsent in abbrev: ogreek = ogreek.replace(" & " + gsent, "")

    if special:
        ngreek_definiens = "(" + ogreek_definiens + new_conjunct + ")"

    elif type3:

        ngreek_definiens = ogreek_definiens[:-2] + new_conjunct + "))"
    else:
        ngreek_definiens = ogreek_definiens[:-1] + new_conjunct + ")"
    ogreek = ogreek.replace(ogreek_definiens, ngreek_definiens)

    for greek_sent, english in zip(def_info[6], def_info[3]):
        ogreek = ogreek.replace(greek_sent, english)

    ogreek = ogreek.strip()
    ogreek = remove_outer_paren(ogreek)

    return find_sentences(ogreek, definiendum), ogreek


def print_new_definitions():
    change_made = False
    for word, def_info in dictionary.bad_paren.items():
        print("parentheses altered for " + word)
        definition = def_info[7]
        print(definition)
        for sent in def_info[3]:
            print(sent)
        print("")
        for lst in def_info[8]:
            print(f"""
            {lst[0]} was replaced with {lst[1]}
            """)

        bool1 = input("do you accept this alteration? 1 for yes, 0 for no: ")
        if bool1 == "1":
            change_made = True
            row_num = dictionary.words_to_row.get(word)
            if "|" in definition:
                list1 = definition.split("|")
                for i, part in enumerate(list1):
                    ws.cell(row=row_num + i, column=6).value = part
            else:
                ws.cell(row=row_num, column=6).value = definition

    if dictionary.bad_paren != {} and change_made:
        wb4.save('/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/dictionary5.xlsx')

    return


def print_categories(kind4):
    if kind4 == 1:
        for word, kind in dictionary.kind.items():
            if kind == 'p':
                print(word)


def print_lists(dictionary):
    bb = 8
    for word in dictionary.rel_abbrev.items():
        bb = 8


def reduce_definitions(dictionary):
    j = 0
    s = 0
    kind3 = 2
    kind4 = 4

    for definiendum, definition in dictionary.definitions.items():

        pos = dictionary.pos.get(definiendum)

        if definiendum != "0":

            bb = 8

            # print (definiendum)

            if pos[1] not in ["s", "d", "b"]:
                j += 1
                definition = dictionary.definitions.get(definiendum)
                if kind3 == 1:
                    try:
                        dictionary = process_sentences(definition, definiendum, dictionary, [])
                    except:
                        print(f"bug found in {definiendum}")
                else:
                    dictionary = process_sentences(definition, definiendum, dictionary, [])

            elif pos[1] in ['s', 'd']:
                s += 1

    print_categories(kind4)

    print(str(j) + " definitions")
    print(str(s) + " synonyms")
    print(str(len(dictionary.popular)) + " popular words")
    return dictionary


def print_period_definitions():
    row_num = 1
    ws6 = wb4.worksheets[5]

    for k, v in dictionary.period_definitions.items():
        ws6.cell(row=row_num, column=1).value = k
        ws6.cell(row=row_num, column=2).value = v
        row_num += 1
    wb4.save('/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/dictionary5.xlsx')


def save_dictionary(dictionary, user=""):
    for k, v in dictionary.categorized_sent.items():
        lst2 = [x.sentences for x in v]
        with open(user + "json_dict/" + k + ".json", "w") as fp:
            json.dump(lst2, fp)
        for x in v: x.sentences = None


def start_pickle(user="", kind2=""):
    global dictionary, wb4, ws, third_sheet, kind

    if user == "":

        if kind2 != "":
            kind = kind2

        arguments = sys.argv
        if len(arguments) > 1:
            kind = int(arguments[1])
        else:
            kind = 6
    # kind = 10
    if kind in [6, 7, 8]:
        wb4 = load_workbook('dictionary5.xlsx')
        ws = wb4.worksheets[0]
        third_sheet = wb4.worksheets[2]

    if kind == 6:
        print('dictionary built and reduced')
        dictionary = build_dictionary(kind2)
        dictionary = reduce_definitions(dictionary)
        print_new_definitions()
        save_dictionary(dictionary, user)
        # dictionary = determine_class(user, "small", dictionary)

    elif kind == 8:
        dictionary = build_dictionary(kind2)
        save_dictionary(dictionary, user)

    elif kind in [9, 10]:
        pkl_file = open(user + 'z_dict_words.pkl', 'rb')
        dictionary = pickle.load(pkl_file)
        pkl_file.close()
        if kind == 9:
            dictionary = reduce_definitions(dictionary)
            print_new_definitions()
            save_dictionary(dictionary, user)
        else:
            print_lists(dictionary)



    elif kind == 7:
        wb4 = load_workbook('/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/dictionary5.xlsx')
        ws = wb4.worksheets[0]
        dictionary = build_dictionary()
        for word in dictionary.easy_embed:
            definition = dictionary[1].get(word)
            definition = adjust_definition(word, definition, dictionary)
            dictionary[1][word] = definition

        print_new_definitions()
        wb4.save('/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/dictionary5.xlsx')

    kind3 = 2
    if kind3 == 1:
        print_period_definitions()

    if user == "":
        dictionary2 = open(user + 'z_dict_words.pkl', 'wb')
        pickle.dump(dictionary, dictionary2)
        dictionary2.close()
    else:
        return dictionary


start_pickle()
