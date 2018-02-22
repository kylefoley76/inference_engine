
import shutil, os, re, copy, sys
import pickle
from openpyxl import load_workbook
import collections
from itertools import chain
import datetime
from general_functions import get_key



# to add a new file to the list go to the differences file and put the name of the file
# 3 times into that list


kind2 = ""
kind3 = ""

try:
    arguments = sys.argv
    if len(arguments) > 1:
        kind = arguments[1]
        if len(arguments) > 2:
            kind2 = arguments[2]

        if kind == "commands":
            print ("""
           vo: count lines of very old file   
            rc: recounts lines using the text file
           and also eliminates old replacements
            up1: updates files to 1
            up2: updates files from very old file
            up3: updates files from the oldest file
            wl: writes the number of lines on modular stability
            wlvo: writes the number of lines on 6 hour line count
            me: puts a message on excel file
            ex: puts in the number of hours spent on excel and terminal
           """)
        elif kind == 'rules':
            print ("""
            local word replacements must be preceded by #lo!
            global word replacements must be preceded by #go!
            lines not counted must end with #nc!
            function renames must end with #fo! followed by the function's old name
            extracted methods end with #ex! followed by the function from which extracted
            functions modeled on a template must end with #tm followed by the function's old name 
            """)
            sys.exit()

        elif kind == 'wl':
            kind = "1"
            kind2 = "wl"

        elif kind == 'wlvo':
            kind = 'vo'
            kind2 = 'wlvo'



    else:
        kind = "1"
except:
    kind = "1"


if kind == 'commands':
    sys.exit()




new_files = ["put_words_in_slots.py",
             "main_loop.py", "standard_order.py", "general_functions.py",
             "pickle_dictionary.py", "pickle_claims.py", "uninstantiable_definitions.py",
         "change_abbreviations.py", "search_for_instantiation.py", "use_lemmas.py",
         "analyze_sentence.py", "natural_language.py", "settings.py",
             "prepare_for_print.py", "analyze_definition.py", "classes.py",
             "lemmas.py", "begin_code.py", ]


old_files = ["pw",
             "nc", "so", "gl", "pd", "pc", "ud", "ca", "sf", "ul", "as", "nl",
             "se", "pp", "ad", "cl", "le", "bc",]


# begin_file = "/Users/kylefoley/PycharmProjects/inference_engine2/inference2/differences/"

begin_file = "/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Proofs/"
begin_file_diff = "/Users/kylefoley/Desktop/inference_engine/differences/"
text_file = begin_file_diff + "new_lines.txt"
file_mistake = "/Users/kylefoley/Desktop/inference_engine/differences_mistakes/"
wb5 = load_workbook(begin_file + "modular_stability.xlsx")

if kind == 'rc':
    f = open(text_file, "r")
    j = 0
    for line in f:
        line = line.replace(" ", "")
        if not line.startswith("#"):
            j += 1
    print (j)
    f.close()
    sys.exit()



def delete_files():
    if kind3 in ['up1', 'up2', 'up3']:

        for docu, old_file in zip(new_files, old_files):

            proof_file = begin_file + docu
            mistake_file = file_mistake + old_file + ".py"

            if kind3 == "up1":
                new_file = begin_file_diff + old_file + "1.py"
                shutil.copy2(proof_file, new_file)
                shutil.copy2(proof_file, mistake_file)
            elif kind3 == "up2":
                new_file = begin_file_diff + old_file + "2.py"
                new_file2 = begin_file_diff + old_file + "1.py"
                shutil.copy2(proof_file, new_file)
                shutil.copy2(proof_file, mistake_file)
                shutil.copy2(proof_file, new_file2)
                output = open(begin_file_diff + 'extracted_methods.pkl', 'wb')
                pickle.dump({}, output)
                output.close()
                output = open(begin_file_diff + 'replacements.pkl', 'wb')
                pickle.dump({}, output)
                output.close()

            elif kind3 == "up3":
                new_file = begin_file_diff + old_file + "3.py"
                new_file2 = begin_file_diff + old_file + "1.py"
                shutil.copy2(proof_file, new_file)
                shutil.copy2(proof_file, mistake_file)
                shutil.copy2(proof_file, new_file2)

            all_new_lines = [line for line in open(new_file)]
            use_for_replace_deletion = []
            replacement_found = False
            for i, line in enumerate(all_new_lines):
                if "#!" in line:
                    use_for_replace_deletion.append(line.replace("#!", ""))
                    replacement_found = True
                elif not re.search(r'#..!', line):
                    use_for_replace_deletion.append(line)
                else:
                    replacement_found = True

            if kind3 == 'up1' and replacement_found:
                os.unlink(proof_file)
                temp_docu = open(proof_file, "w")
                for line in use_for_replace_deletion: temp_docu.write(line)
                temp_docu.close()

        sys.exit()


def get_current_cell():
    row_number = 2
    sheet1 = wb5.worksheets[0]
    now = datetime.datetime.now()
    month = now.month
    day = now.day
    excel_date = sheet1.cell(row=row_number, column=1).value
    excel_day = excel_date.day
    excel_month = excel_date.month
    monthDict = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
                 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
    month = monthDict.get(month)
    the_date = str(day) + "-" + month

    while True:

        if excel_date == None:
            old_amount = sheet1.cell(row=row_number - 1, column=2).value
            break

        elif (excel_month == now.month and excel_day < now.day) or excel_month < now.month:
            row_number += 1
            excel_date = sheet1.cell(row=row_number, column=1).value
            if isinstance(excel_date, str):
                list2 = excel_date.split("-")
                excel_day = int(list2[0])
                excel_month = list2[1]
                excel_month = get_key(monthDict, excel_month)

            else:
                try:
                    excel_day = excel_date.day
                    excel_month = excel_date.month
                except:
                    pass

        elif (excel_month == now.month or excel_day == now.day):
            old_amount = sheet1.cell(row=row_number, column=2).value
            break

    return row_number, old_amount, the_date



def calc_modular_stability():
    now = datetime.datetime.now()
    month = now.month
    day = now.day
    monthDict = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
                 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
    if day < 10:
        day = "0" + str(day)
    else:
        day = str(day)


    sheet1 = wb5.worksheets[0]
    sheet3 = wb5.worksheets[2]
    module_stats = {}
    module_stats2 = []
    for k, v in modular_stability.items():
        num_func = len(v)
        size = 0
        for func in v:
            size += len(new_functions2.get(func, []))
        module_stats.update({k: [num_func, size]})
        module_stats2.append([k, num_func, size])

    total_lines = 0
    for x in module_stats2: total_lines += x[2]
    total_functions = 0
    for x in module_stats2: total_functions += x[1]
    print ("total_lines: " + str(total_lines))
    print ("total_functions: " + str(total_functions))
    print ("total modules: " + str(len(module_stats)))

    month = monthDict.get(month)


    row_number, old_amount, the_date = get_current_cell()
    sheet1.cell(row=row_number, column=1).value = str(day) + "-" + month
    sheet1.cell(row=row_number, column=2).value = total_lines
    sheet1.cell(row=row_number, column=3).value = total_functions
    sheet1.cell(row=row_number, column=4).value = len(module_stats)

    new_gross_lines = total_lines - old_amount
    print ("new gross lines: " + str(new_gross_lines))
    print_modular_stability(wb5, module_stats, row_number, the_date)

    if kind2 in ['wl', 'wlvo']:

        wb5.save(begin_file + "modular_stability.xlsx")


    return


def print_modular_stability(wb5, module_stats, row_date, the_date):
    global p
    sheet2 = wb5.worksheets[1]
    current_modules = []
    sheet2.cell(row=row_date, column=1).value = the_date
    i = 2
    for col in range(len(modular_stability)):
        mod = sheet2.cell(row=1, column=i).value

        if mod == None:
            break
        lst = module_stats.get(mod)

        sheet2.cell(row=row_date, column=i).value = lst[1]
        sheet2.cell(row=row_date, column=i + 1).value = lst[0]

        current_modules.append(mod)
        i += 2

    new_mod = [x for x, y in modular_stability.items() if x not in current_modules]

    if new_mod != []:
        e = 0
        for col2 in range(len(new_mod)):
            sheet2.cell(row=1, column=i).value = new_mod[e]
            sheet2.cell(row=2, column=i).value = "lines"
            sheet2.cell(row=2, column=i + 1).value = "functions"
            sheet2.cell(row=row_date - 1, column=i).value = "birth"
            i += 2
            e += 1

    if kind2 == "wl":
        sheet3 = wb5.worksheets[2]

        row_number = 1
        while True:
            row_number += 1
            num = sheet3.cell(row=row_number, column=2).value
            if num == None:
                sheet3.cell(row=row_number, column=2).value = p
                sheet3.cell(row=row_number, column=1).value = the_date
                end = row_number - 5
                orow_number = row_number
                avg = p
                while row_number != end:
                    row_number -= 1
                    avg += sheet3.cell(row=row_number, column=2).value
                avg = avg // 6
                sheet3.cell(row=orow_number, column=3).value = avg
                print ('average: ' + str(avg))
                if sheet3.cell(row=row_number-1, column=4).value == 0:
                    print ('you must update 6 hours now')
                break



    elif kind2 == 'wlvo':
        sheet3 = wb5.worksheets[2]
        sheet4 = wb5.worksheets[3]
        row_number = 1
        while True:
            row_number += 1
            num = sheet4.cell(row=row_number, column=2).value
            avg = 0
            if num == None:
                sheet4.cell(row=row_number, column=3).value = p
                sheet4.cell(row=row_number, column=1).value = the_date
                break

        row_number2 = 1
        while True:
            row_number2 += 1
            num = sheet3.cell(row=row_number2, column=2).value
            if num == None:
                sheet3.cell(row=row_number2-1, column=4).value = 0
                no_delete = 0
                row_number2 -= 1
                end = row_number2 - 6
                while row_number2 != end:
                    no_delete += sheet3.cell(row=row_number2, column=2).value
                    row_number2 -= 1

                if p > no_delete:

                    p = no_delete
                    avg2 = 1
                    sheet4.cell(row=row_number, column=3).value = p
                else:
                    avg2 = p / no_delete
                    avg2 = str("{0:.2f}".format(avg2))


                sheet4.cell(row=row_number, column=2).value = no_delete
                sheet4.cell(row=row_number, column=4).value = avg2
                print ('6 hour average: ' + str(avg2))

                end = row_number - 5
                orow_number = row_number
                avg = p // 6
                while row_number != end:
                    row_number -= 1
                    avg += int(sheet4.cell(row=row_number, column=5).value)

                avg = avg // 6

                sheet4.cell(row=orow_number, column=5).value = p // 6
                sheet4.cell(row=orow_number, column=6).value = avg
                print ('moving average: ' + str(avg))


                break
    return







def cut_hashtag(line):
    if "#" in line and "!" not in line:
        return line[:line.index("#")]
    else:
        return line


def make_single_lines(all_lines):
    new_list = []
    m = 0
    while m < len(all_lines):
        if all_lines[m].endswith(",\n"):
            current_line = all_lines[m][:-1]
            m += 1
            while all_lines[m].endswith(",\n"):
                current_line += all_lines[m][:-1]
                m += 1
            current_line += all_lines[m]
            current_line = cut_hashtag(current_line)
            new_list.append(current_line)
        else:
            all_lines[m] = cut_hashtag(all_lines[m])
            new_list.append(all_lines[m])
        m += 1

    return new_list




def weed_out_same_functions(new_functions2, orig_new_functions2, old_functions2,
                            orig_old_functions2):
    brand_new_lines = []
    j = 0
    for k, v in new_functions2.items():
        old_counterpart = old_functions2.get(k)
        if k.startswith("rearrange"):
            bb = 8


        if old_counterpart == None:

            brand_new_lines.append(k)
            temp_list = list(orig_new_functions2.values())[j]
            # brand_new_lines += list(orig_new_functions2.values())[j]
            brand_new_lines += v

        elif old_counterpart != v:

            new_functions.update({k:v})
            v_orig = orig_new_functions2.get(k)
            v_old_orig = orig_old_functions2.get(k)
            orig_new_functions.update({k:v_orig})
            orig_old_functions.update({k:v_old_orig})
            old_functions.update({k: old_counterpart})
        j += 1


    return brand_new_lines



def is_an_exception(list1, line):
    exceptions = ["return\n", "break\n", "else:\n", "pass\n", "try:\n",
                  "except:\n"]

    line = line.replace(" ", "")

    if line in exceptions:
        return True
    if line.startswith("#"):
        return True
    if not re.search(r'\S', line):
        return True
    if line in ["", '']:
        return True

    if "#!" in line:
        return True
    if line.endswith("#nc\n"):
        return True
    if line.startswith("bb=8"):
        if "if" in list1[-1]:
            del list1[-1]
            if list1 != []:
                del list1[-1]
        return True
    if line == "\n":
        return True


    return False


def check_replacements(line, file_abbreviation, subroutine):


    current_replacements = replacements.get(subroutine)
    if current_replacements == None:
        current_replacements = replacements.get(file_abbreviation)


    if current_replacements != None:
        for k, v in current_replacements.items():
            if k in line:

                list1 = re.split("[, ()=.]+", line)

                for i, word in enumerate(list1):
                    if word == k:
                        line = line.replace(k, v)

                        return line
    return line
#
# """
# local word replacements must be preceded by #lo
# global word replacements must be preceded by #go
# lines not counted must end with #nc
# function renames must end with #fo followed by the function's old name
# extracted methods end with #ex followed by the function from which extracted
# functions modeled on a template must end with #tm followed by the function's old na
# """

def change_functions(all_new_lines, file_abbreviation):
    new_word = ""
    for i, line in enumerate(all_new_lines):
        edit_line = line.replace(" ", "")
        try:
            b = edit_line.index("!") + 1
        except:
            b = 0

        if edit_line.startswith("#fo!"):
            new_word, old_word = edit_line[b:].split(",")
            new_word = new_word.strip()
            old_word = old_word.strip()

        if line.startswith("def ") and new_word != "":

            subroutine = line[line.index(" "):line.index("(")]
            subroutine = subroutine.strip()

            if subroutine == new_word:

                line = line.replace(new_word, old_word)

                all_new_lines[i] = line


            if file_abbreviation not in file_has_replacements:
                file_has_replacements.append(file_abbreviation)
            new_word = ""

    return all_new_lines







def get_replacements(all_new_lines, file_abbreviation):

    for line in all_new_lines:
        replacement_found = True
        edit_line = line.replace(" ","")
        edit_line = edit_line.replace("\n","")
        try:
            b = edit_line.index("!") + 1
        except:
            b = 0

        if "last_res" in line:
            bb = 8
        if line.startswith("def "):
            subroutine = line[line.index(" "):line.index("(")]
            subroutine = subroutine.strip()

        if len(edit_line) > 3 and edit_line[3] == "!" and "," in edit_line:
            new_word, old_word = edit_line[b:].split(",")
            new_word = new_word.strip()
            old_word = old_word.strip()


            if edit_line.startswith("#lo!"):


                replacements.setdefault(subroutine, {}).update({new_word: old_word})

            elif edit_line.startswith("#go!"):

                replacements.setdefault(file_abbreviation, {}).update({new_word: old_word})



            elif edit_line.startswith("#ex!"):


                extracted_methods.setdefault(old_word, []).append(new_word)

            else:
                replacement_found = False

            if replacement_found:
                if file_abbreviation not in file_has_replacements:
                    file_has_replacements.append(file_abbreviation)



    return


def get_lines(list1, orig_functions2, functions2, file_abbreviation, new=False):
    dupl_on = False
    lines = []
    original_lines = []
    put_on_beginning = []
    if file_abbreviation in file_has_replacements:
        check_replace = True
    else:
        check_replace = False

    
    subroutine = file_abbreviation
    global_found = False
    end = len(list1)

    for line in reversed(list1):

        end -= 1
        if line.startswith("def ") or line.startswith("class "):
            list12 = copy.deepcopy(list1)
            for num in range(end + 1, len(list12)):
                line = list12[num]
                line = line.replace("\n", " ")
                if global_found:
                    list1[num] = "#"
                    put_on_beginning.append(line)

                if line.startswith("defput_def"):
                    bb = 8

                if re.search(r'\w', line[0]) and not global_found:
                    global_found = True
                    put_on_beginning.append(line)
                    list1[num] = "#"

            for line in put_on_beginning: list1.insert(0, line)
            break

    for i, line in enumerate(list1):

        edit_line = line.replace(" ", "")

        if new:

            if line.startswith("#dupl"):
                dupl_on = True
            elif line.startswith("#edupl"):
                dupl_on = False

        if not dupl_on:

            if edit_line.startswith("defput_def"):
                bb = 8

            if check_replace:
                line = check_replacements(line, file_abbreviation, subroutine)

            if line.startswith("def ") or line.startswith("class "):
                if new:
                    modular_stability.setdefault(file_abbreviation[:-3], []).append(subroutine)
                functions2.update({subroutine: copy.deepcopy(lines)})
                orig_functions2.update({subroutine: copy.deepcopy(original_lines)})
                lines = []
                original_lines = []
                if "(" not in line:
                    subroutine = line[line.index(" "):line.index(":")]
                else:
                    subroutine = line[line.index(" "):line.index("(")]

                subroutine = subroutine.strip()
                if subroutine == 'categorize_words2':
                    bb = 8



            edit_line = line.replace(" ", "")
            lines.append(edit_line)
            original_lines.append(line)

    functions2.update({subroutine: copy.deepcopy(lines)})
    orig_functions2.update({subroutine: copy.deepcopy(original_lines)})

    return




def get_total_lines():
    global complete_new_code


    for file_abbreviation, old_file in zip(new_files, old_files):
        new_file = begin_file + file_abbreviation

        # print (proof)
        if file_abbreviation.startswith('search'):
            bb = 8

        if kind == "1":
            old_file = begin_file_diff + old_file + "1.py"
        elif kind == "vo":
            old_file = begin_file_diff + old_file + "2.py"
        elif kind == "vvo":
            old_file = begin_file_diff + old_file + "3.py"

        all_new_lines = [line for line in open(new_file)]
        all_old_lines = [line for line in open(old_file)]
        list14 = []
        for line in all_new_lines:
            if not is_an_exception(all_new_lines, line):
                list14.append(line)
        all_new_lines = list14
        list15 = []
        for line in all_old_lines:
            if not is_an_exception(all_old_lines, line):
                list15.append(line)
        all_old_lines = list15


        all_new_lines = make_single_lines(all_new_lines)
        all_old_lines = make_single_lines(all_old_lines)

        all_new_lines = change_functions(all_new_lines, file_abbreviation)

        get_replacements(all_new_lines, file_abbreviation)

        get_lines(all_new_lines, orig_new_functions2, new_functions2, file_abbreviation, True)
        get_lines(all_old_lines, orig_old_functions2, old_functions2, file_abbreviation, False)

    restore_extracts(orig_new_functions2, new_functions2, orig_old_functions2, old_functions2)

    brand_new_lines = weed_out_same_functions(new_functions2, orig_new_functions2,
                            old_functions2, orig_old_functions2)

    complete_new_code = brand_new_lines


def restore_extracts(orig_new_functions2, new_functions2, orig_old_functions2, old_functions2):
    to_be_deleted = []
    for old, new_func in extracted_methods.items():
        for new in new_func:
            lines = new_functions2.get(new)
            assert lines != None
            orig_lines = orig_new_functions2.get(new)
            old_lines = new_functions2.get(old)
            orig_old_lines = orig_new_functions2.get(old)
            lines += old_lines
            orig_lines += orig_old_lines
            to_be_deleted.append(new)

    for lst in to_be_deleted: del new_functions2[lst]

    return


def compare_two_lists(new_lines, old_lines, original_lines):
    new_num = 0
    old_num = 0
    temp_list = []
    if old_lines != []:
        while new_num < len(new_lines):
            new_line = new_lines[new_num]
            old_num = len(old_lines) -1 if old_num > len(old_lines) -1 else old_num
            old_line = old_lines[old_num]
            original_line = original_lines[new_num]
            if new_line != old_line:

                if new_num in [108, 112]:
                    bb = 8

                # print (new_num)
                # print (old_num)
                try:
                    old_num = old_lines.index(new_line)
                except:
                    complete_new_code.append(original_line)
                    temp_list.append(original_line)

            new_num += 1
            old_num += 1
    else:
        new_num += len(new_lines)

    return


def determine_new_lines():
    global q

    for new, old, original in zip(new_functions, old_functions, orig_new_functions):
        q += 1
        new_lines = new[1]
        old_lines = old[1]
        original_lines = original[1]
        compare_two_lists(new_lines, old_lines, original_lines)

    return



if kind == 'me':
    str1 = input("message: ")
    sheet3 = wb5.worksheets[2]
    j = 2
    while sheet3.cell(row=j, column=2).value != None:
        j += 1
    sheet3.cell(row=j-1, column=6).value = str1
    wb5.save(begin_file + "modular_stability.xlsx")
    sys.exit()
elif kind == 'ex':
    kind2 = int(arguments[2])
    kind2 = kind2 / 60
    minutes = str("{0:.2f}".format(kind2))

    sheet3 = wb5.worksheets[2]
    j = 2
    while sheet3.cell(row=j, column=2).value != None:
        j += 1
    sheet3.cell(row=j-1, column=5).value = minutes
    wb5.save(begin_file + "modular_stability.xlsx")
    sys.exit()



if kind in ["up1", "up2"]:
    kind3 = kind
    delete_files()

q = 0
p = 0
new_functions2 = {}
orig_new_functions2 = {}
old_functions2 = {}
orig_old_functions2 = {}
new_functions = {}
orig_new_functions = {}
old_functions = {}
orig_old_functions = {}
modular_stability = {}
complete_new_code = []
replacements = {}
extracted_methods = {}
orig_extracted_methods = {}
global_replacements = {}
template_methods = {}
orig_template_methods = {}
file_has_replacements = []



if kind == 'vo':

    pkl_file = open(begin_file_diff + 'extracted_methods.pkl', 'rb')
    extracted_methods = pickle.load(pkl_file)
    pkl_file.close()

    pkl_file = open(begin_file_diff + 'replacements.pkl', 'rb')
    replacements = pickle.load(pkl_file)
    pkl_file.close()




get_total_lines()



orig_new_functions = sorted(orig_new_functions.items())
new_functions = sorted(new_functions.items())
orig_old_functions = sorted(orig_old_functions.items())
old_functions = sorted(old_functions.items())

determine_new_lines()


p = len(complete_new_code)


calc_modular_stability()




os.unlink(text_file)
f = open(text_file, "w")
for line in complete_new_code: f.write(line)
f.close()
print (p)


output = open(begin_file_diff + 'extracted_methods.pkl', 'wb')
pickle.dump(extracted_methods, output)
output.close()
output = open(begin_file_diff + 'replacements.pkl', 'wb')
pickle.dump(replacements, output)
output.close()


if kind2 == 'wl': kind3 = 'up1'
if kind2 == 'wlvo': kind3 = 'up2'

delete_files()







# chain1 = collections.ChainMap(new_functions, old_functions, orig_new_functions)

# for k, new_lines in new_functions.items():
#     old_lines = old_functions.get(k)
#     original_lines = orig_new_functions.get(k)
#     compare_two_lists(new_lines, old_lines, original_lines)


