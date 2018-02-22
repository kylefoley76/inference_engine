
import datetime, sys
from general_functions import get_key
from openpyxl import load_workbook



excel_file = "/Users/kylefoley/Desktop/inference_engine/exercise.xlsx"

wb5 = load_workbook(excel_file)

arguments = sys.argv

now = datetime.datetime.now()
month = now.month
day = now.day

arg1 = arguments[1]

if arg1 == 'commands':
    print("""
    pu = pushups
    su = situps
    jg = jog
    di = distance
    cu = curls
    sh = shoulders
    st = standing up
    if none of these arguments are chosen
    it will go into girl name
    second argument for gl
    new = new
    bi = binge
    """)
    sys.exit()



row_number = 2
sheet1 = wb5.worksheets[0]
excel_date = sheet1.cell(row=row_number, column=1).value
excel_day = excel_date.day
excel_month = excel_date.month



if arg1 in ['pu', 'su', 'jg', 'di', 'cu', 'sh', "st", "wt"]:
    while True:
        if (excel_month == now.month and excel_day == now.day):
            break
        row_number += 1
        excel_date = sheet1.cell(row=row_number, column=1).value
        excel_day = excel_date.day
        excel_month = excel_date.month



    dict1 = {"pu": 4, "di": 3, "jg": 2, "su": 5, "cu": 6, "sh": 7, "st": 8, "wt":9}
    col = dict1.get(arg1)
    old_num = sheet1.cell(row=row_number, column=col).value
    if arg1 not in ["jg", "di", "wt"]:
        new_num = int(arguments[2])
        if old_num != None:
            new_num += old_num
    else:
        new_num = arguments[2]
    sheet1.cell(row=row_number, column=col).value = new_num


else:
    new_girl = arguments[1]
    new_girl = new_girl.replace("_", " ")
    girls = []
    row_number = 1
    found = False
    try:
        new = arguments[2]
    except:
        new = ""

    sheet2 = wb5.worksheets[1]
    while True:
        girl = sheet2.cell(row=row_number, column=1).value
        if girl == new_girl:
            old_num = sheet2.cell(row=row_number, column=2).value
            old_num += 1
            sheet2.cell(row=row_number, column=2).value = old_num
            found = True
            break

        if girl == None:
            if new == "":
                print ('you mispelled the girls name')
            break
        girls.append(girl)
        row_number += 1

    if new != "new" and not found:
        first_letter = new_girl[0]
        for girl in girls:
            if girl[0] == first_letter:
                print (girl)
    else:
        sheet2.cell(row=row_number, column=1).value = new_girl
        sheet2.cell(row=row_number, column=2).value = 1




wb5.save(excel_file)

