import time
import sys
from natural_language import step_one
from general_functions import parameters
from classes import ErrorWithCode
from settings import *
from openpyxl import load_workbook

def calculate_time_statistics(num_proved, total_time):
    total_time = time.time() - total_time
    print("")
    print("average " + str("{0:.4f}".format(total_time / num_proved)))
    print("total " + str("{0:.3f}".format(total_time)))
    print("")
    print (num_proved)


def determine_words_used(words_used):
    wb5 = load_workbook('/Users/kylefoley/Desktop/inference_engine/dictionary5.xlsx')
    w5 = wb5.worksheets[0]
    for word in words_used:
        j = dictionary.words_to_row.get(word, 28)
        if j == 28:
            print(word)
        elif j == None:
            print (word)
        else:
            try:
                w5.cell(row=j, column=2).value = 1
            except:
                pass
    wb5.save('/Users/kylefoley/Desktop/inference_engine/dictionary5.xlsx')
    words = open('words_used.pkl', 'wb')
    pickle.dump(words_used, words)
    words.close()


def get_result(one_sent, user = "", print_type=4, order=[0], get_words_used=0):
    total_time = time.time()

    if one_sent == 'a':
        proof_type, print_type, get_words_used, order = parameters()
        pkl_file = open(user + 'zz_claims.pkl', 'rb')
        test_sent = pickle.load(pkl_file)
        pkl_file.close()
    elif one_sent != "":
        test_sent = [[one_sent]]
    else:
        pkl_file = open(user + 'zz_claims.pkl', 'rb')
        test_sent = pickle.load(pkl_file)
        pkl_file.close()
        pkl_file = open(user + 'z_dict_words.pkl', 'rb')
        dictionary = pickle.load(pkl_file)
        pkl_file.close()

    words_used = set()

    j = -1
    num_proved = 0
    while j < len(order) - 1:
        j += 1
        k = order[j]

        if test_sent[k][0] != 'pass':
            num_proved += 1
            st1 = time.time()

            if k == 182:
                bb = 7
            try:
                _ = step_one(dictionary, user, test_sent[k])
                consistent, total_sent, twords_used = _
                if total_sent != 'skip':

                    test_sent[k] = json.loads(json.dumps(total_sent))
                    words_used = words_used | twords_used

                else:
                    order.remove(k)
                    j -= 1

                if print_type != 4:
                    if not consistent:
                        if print_type == 3:
                            print(str(k) + " - " + str("{0:.3f}".format(time.time() - st1) + " False"))
                        elif print_type == 0:
                            print (str(k) + " - False")
                            sys.exit()
                    elif print_type == 3:
                        print(str(k) + " - " + str("{0:.3f}".format(time.time() - st1)))
                elif print_type == 4 and not consistent:
                    print (str(k) + " - False")


            except ErrorWithCode:
                if print_type != 4:
                    print (str(k) + " - infinite loop")
                order.remove(k)
                j -= 1


        elif print_type in [0, 4] and k % 50 == 0:
            print (k)



    if print_type in [0, 2, 3, 4]:
        if print_type == 0:
            print ("success")
        calculate_time_statistics(num_proved, total_time)

    if get_words_used == 1:
        determine_words_used(words_used)

    return test_sent






