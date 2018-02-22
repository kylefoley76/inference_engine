
from openpyxl import load_workbook
import operator, copy
import types, sys
from ast import literal_eval

kind = 1
if kind == 1:

    from begin_code import get_result
    m = types.ModuleType("m","begin_code.py")
    sys.modules['m'] = m
    sys.modules['m']
    code = compile("get_result('a')", "m.py", "exec")
    exec in m.__dict__
else:
    from pickle_dictionary import start_pickle

    m = types.ModuleType("m", "pickle_dictionary.py")
    sys.modules['m'] = m
    sys.modules['m']
    code = compile("start_pickle()", "m.py", "exec")
    exec in m.__dict__






#
# begin_file = "/Users/kylefoley/PycharmProjects/inference_engine2/inference2/Ancient/intermed_code.py"
# import importlib.machinery
# bb = 8
# loader = importlib.machinery.SourceFileLoader('report', begin_file)
# handle = loader.load_module('report')
#
# handle.getresult_old()



import cProfile, pstats, profile
import re





list1 = ["new_code","pickle_dictionary","pickle_claims","general_functions",
        "settings", "change_abbreviations", "standard_order", "start_and_stop", "use_lemmas",
         "analyze_sentence", "uninstantiable_definitions", 'natural_language',
         "analyze_definition", "prepare_for_print",
         "put_words_in_slots", "z_dict_words", "zz_claims", "search_for_instantiation"]

print ("start")


cProfile.run(code, 'data_stats')
p = pstats.Stats('data_stats')
p.strip_dirs().sort_stats(-1).print_stats()
p.sort_stats('name')
primitive_calls = p.prim_calls
all_time = str("{0:.1f}".format(p.total_tt))
num_inputs = 1
time_per_input = p.total_tt/num_inputs
time_per_input = str("{0:.3f}".format(time_per_input))
calls_per_input = primitive_calls // num_inputs

calls_per_input = str("{:,}".format(calls_per_input))
primitive_calls = str("{:,}".format(primitive_calls))


output = []
built_ins = []
j = 0
for k, v in p.stats.items():
    j += 1
    list2 = k[2].split(",")
    list3 = list1[0].split(".")
    module_location = k[0].replace(".py", "")
    function_name = list2[0].replace(".py","")

    if "hypothetical" in k[2]:
        bb = 8

    total_time = int(v[3] * 1_000_000)
    function_calls = v[0]
    prim_num_of_calls = v[1]
    average = total_time//function_calls
    percentage = int((v[3]/ p.total_tt) * 100)
    if k[2] == "change_abbrev":
        bb = 8



    # if average > 100:
    #     average = str(int(average))
    # else:
    # average = str("{0:.1f}".format(total_time/function_calls))

    if module_location in list1:
        output.append([module_location, function_name, prim_num_of_calls, total_time, average, percentage])
    else:
        built_ins.append([function_name, prim_num_of_calls, total_time, average, percentage])

_deepcopy = []

for k, v in p.stats.items():
    if 'deepcopy' in k and _deepcopy == []:
        for y, z in v[4].items():
            module_location_dc = y[0].replace(".py", "")
            line_number = y[1]
            total_time2 = int(z[3] * 1_000_000)
            prim_num_of_calls2 = z[1]
            average = total_time2 // prim_num_of_calls2
            # average = str("{0:.0f}".format(average))
            # total_time2 = str("{0:.0f}".format(total_time2))
            percentage = int((z[3] / p.total_tt) * 100)

            if module_location_dc in list1:

                _deepcopy.append([module_location_dc, line_number, prim_num_of_calls2, total_time2, average, percentage])





by_module = copy.deepcopy(output)
output = sorted(output, key=operator.itemgetter(3), reverse=True)
by_module = sorted(by_module, key=operator.itemgetter(0,3), reverse=True)
built_ins = sorted(built_ins, key=operator.itemgetter(2), reverse=True)


total_time = str("{:,}".format(total_time))
prim_num_of_calls = str("{:,}".format(prim_num_of_calls))


for lst, lst2 in zip(output, by_module):
    lst[4] = str("{:,}".format(lst[4]))
    lst2[4] = str("{:,}".format(lst2[4]))

    lst[3] = str("{:,}".format(lst[3]))
    lst2[3] = str("{:,}".format(lst2[3]))

for lst in built_ins:
    lst[3] = str("{:,}".format(lst[3]))
    lst[2] = str("{:,}".format(lst[2]))

for lst in _deepcopy:
    lst[3] = str("{:,}".format(lst[3]))
    lst[4] = str("{:,}".format(lst[4]))
    lst[2] = str("{:,}".format(lst[2]))


# output.sort(reverse = True)

wb4 = load_workbook('/Users/kylefoley/Desktop/inference_engine/code_stats.xlsx')
w4 = wb4.worksheets[0]

w4.cell(row=1+3, column=3).value = all_time
w4.cell(row=2+3, column=3).value = primitive_calls
w4.cell(row=3+3, column=3).value = time_per_input
w4.cell(row=4+3, column=3).value = calls_per_input
w4.cell(row=1+3, column=2).value = "time spent: "
w4.cell(row=2+3, column=2).value = "primitive calls:"
w4.cell(row=3+3, column=2).value = "time per input"
w4.cell(row=4+3, column=2).value = "calls per input"

w4.cell(row=1, column=2).value = "module name"
w4.cell(row=1, column=3).value = "function name"
w4.cell(row=1, column=4).value = "primitive calls"
w4.cell(row=1, column=5).value = "total time"
w4.cell(row=1, column=6).value = "average"
w4.cell(row=1, column=7).value = "percentage"




for row_number, lst in enumerate(output):
    w4.cell(row=row_number + 11, column=2).value = lst[0]
    w4.cell(row=row_number + 11, column=3).value = lst[1]
    w4.cell(row=row_number + 11, column=4).value = lst[2]
    w4.cell(row=row_number + 11, column=5).value = lst[3]
    w4.cell(row=row_number + 11, column=6).value = lst[4]
    w4.cell(row=row_number + 11, column=7).value = lst[5]

b = row_number + 11
b += 5

for row_number, lst in enumerate(by_module):
    w4.cell(row=row_number + b, column=2).value = lst[0]
    w4.cell(row=row_number + b, column=3).value = lst[1]
    w4.cell(row=row_number + b, column=4).value = lst[2]
    w4.cell(row=row_number + b, column=5).value = lst[3]
    w4.cell(row=row_number + b, column=6).value = lst[4]
    w4.cell(row=row_number + b, column=7).value = lst[5]

b += row_number
b += 5

for row_number, lst in enumerate(built_ins):
    w4.cell(row=row_number + b, column=3).value = lst[0]
    w4.cell(row=row_number + b, column=4).value = lst[1]
    w4.cell(row=row_number + b, column=5).value = lst[2]
    w4.cell(row=row_number + b, column=6).value = lst[3]
    w4.cell(row=row_number + b, column=7).value = lst[4]

b += row_number
b += 5

for row_number, lst in enumerate(_deepcopy):

    w4.cell(row=row_number + b, column=3).value = lst[0]
    w4.cell(row=row_number + b, column=4).value = lst[1]
    w4.cell(row=row_number + b, column=5).value = lst[2]
    w4.cell(row=row_number + b, column=6).value = lst[3]
    w4.cell(row=row_number + b, column=7).value = lst[4]
    w4.cell(row=row_number + b, column=8).value = lst[5]


wb4.save('/Users/kylefoley/Desktop/inference_engine/code_stats.xlsx')



