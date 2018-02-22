
from general_functions import *
from settings import *
from openpyxl import load_workbook
from analyze_sentence import find_sentences, period_elimination

variables = [chr(122 - t) for t in range(25)]
variables.remove("i")
variables.remove("l")
variables3 = [chr(122 - t) + l1 for t in range(26)]
variables3.remove("l" + l1)
variables4 = [chr(122 - t) + l2 for t in range(26)]
variables4.remove("l" + l2)
variables5 = [chr(122 - t) + l3 for t in range(26)]
variables5.remove("l" + l3)
variables = variables + variables3 + variables4 + variables5


variables9 = [chr(98 + t) for t in range(25)]
variables9.remove("i")
variables9.remove("l")
variables8 = [chr(97 + t) + l1 for t in range(26)]
variables8.remove("l" + l1)
variables7 = [chr(97 + t) + l2 for t in range(26)]
variables7.remove("l" + l2)
variables9 = variables9 + variables8 + variables7



greek = [chr(945 + t) for t in range(40)]
dict1 = {}


print ("""
type 1: planned instantiations | abbreviations | definitions to be changed | last variable

type 2: rearrange variables in definition. put definiendum to the left of /

type 3: place planned instantiations before | and the sentence to be instantiated after |

type 4: properly spaces definition and changes variables, can also split definition up with |

type 5: adjust definition on excel sheet

type 6: eliminates periods

type 7: eliminates periods
""")


# type = '2'
# total_sent = "(f=seventy five percent of the time) & (((p E e T d) & (k SUT d)) → (q OC f T k)) ≡ (((p E e T d) & (k A d)) → (((r W s) ≡ ((q P s) & (e H s) & (s I z))) & ((t W u) ≡ ((q ~ P s) & (e H s) & (s I z)) & (r N x) & (t N y) & ((x / (x + y) = f))))) & (z=offspring)"
# type = input("enter type: ")
# total_sent = input("enter sentence: ")
# total_sent = "((b DJ c d E f) ≡ |(((n W u) ≡(((o I s) & (t H o) & (u P o)) →(c.d ~ P o))) & | (t H f) & (f I s) & | ((p W q) ≡ ((q I s) & (f H q) & (m.¬u.¬j P q))) & | ((r W v) ≡ ((v I s) & (f H v) & (j.¬u.¬c P v))) & | (p EQ r)) & | (m ⇿ b CS e) & (h ⇿ e CS c) & (j ⇿ b CS f) & (k ⇿ f CS d) & (s=offspring)"



def replace_letters(new_sent, dict1, type, variables6 = []):
    i = -1
    star_used = False
    while i < len(new_sent) - 1:
        i += 1
        letter = new_sent[i]
        if letter == 'y':
            bb = 8
        j = 0
        if i < len(new_sent) - 1:
            if new_sent[i + 1] in subscripts:
                letter += new_sent[i + 1]
                j = 1
        first_half = new_sent[:i]
        second_half = new_sent[i + 1 + j:]
        if letter == "*":
            star_used = True
            k = new_sent.find(")", i, len(new_sent))
            i = k + 2
        elif (letter in dict1.keys() and new_sent[i + 1 + j] == "="):
            new_var = dict1.get(letter)
            new_sent = first_half + new_var + second_half
            k = new_sent.find(")", i, len(new_sent))
            i = k + 2
        elif letter in dict1.keys():
            new_var = dict1.get(letter)
            new_sent = first_half + new_var + second_half
            i += j
        elif letter.islower() and new_sent[i + 1 + j] == "=" and type in ["1","2","4"]:
            k = new_sent.find(")", i, len(new_sent))
            i = k + 2
            if variables6 == []:
                new_sent = first_half + variables9[0] + second_half
                dict1.update({letter: variables9[0]})
                del variables9[0]
            else:
                new_sent = first_half + variables6[0] + second_half
                dict1.update({letter: variables6[0]})
                del variables6[0]
        elif letter[0].islower() and type in ["1", "2", "4"]:
            if variables6 == []:
                new_sent = first_half + variables9[0] + second_half
                dict1.update({letter: variables9[0]})
                del variables9[0]
            else:
                new_sent = first_half + variables6[0] + second_half
                dict1.update({letter: variables6[0]})
                del variables6[0]
    if star_used:
        new_sent = new_sent.replace("*","")

    return new_sent


def no_space_around_sym(sent, sym):
    if sym + " " in sent:
        sent = sent.replace(sym + " ", sym)
    if " " + sym in sent:
        sent = sent.replace(" " + sym, sym)
    if sym in sent:
        sent = sent.replace(sym, " " + sym + " ")
    return sent

def adjust_definition(word, definition):
    pkl_file = open('z_dict_words.pkl', 'rb')
    dictionary = pickle.load(pkl_file)
    pkl_file.close()

    pos = dictionary[0].get(word)
    sent = definition
    if sent.count("(") == 1:
        sent = sent.replace("= ", "=")
        sent = sent.replace(" =", "=")
        sent = sent.replace("=", " = ")
    else:
        variables6 = copy.copy(variables9)
        def_info = find_sentences(sent)
        def_info[5] = no_space_around_sym(def_info[5], mini_e)

        for i in range(len(def_info[3])):
            sent = def_info[3][i]
            if def_info[4][i][1] == "":
                try:
                    if word in sent:
                        if (pos[0] == "r" and "=" in sent) or (pos[0] != 'r' and "=" not in sent):
                            sent = sent.replace(word, "%")
                            sent = "*" + sent
                except:

                    pass

                if "=" in sent:
                    sent = sent.replace("= ", "=")
                    sent = sent.replace(" =", "=")
                else:
                    sent = sent.replace(" ", "")
                    # sent = no_space_around_sym(sent, mini_e)
                    sent_list = list(sent)
                    j = 0
                    while j < len(sent_list) - 2:
                        j += 1
                        if sent_list[j + 1] not in [")", "."]:
                            if sent_list[j] not in ["(", ")", neg, "."]:
                                if sent_list[j].isupper():
                                    if not sent_list[j + 1].isupper():
                                        sent_list.insert(j + 1, " ")
                                        j += 1
                                else:
                                    sent_list.insert(j + 1, " ")
                                    j += 1
                    sent = "".join(sent_list)
                def_info[5] = def_info[5].replace(def_info[6][i], sent)
        def_info[5] = replace_letters(def_info[5], {}, "4", variables6)
        def_info[5] = def_info[5].replace("%", word + " ")
        def_info[5] = def_info[5].replace(" )", ")")
        def_info[5] = def_info[5].replace("  ", " ")
        sent = def_info[5]

    return sent


def change_variables(total_sent, type):
    if type == "1":
        # to use type 1, we need the string before the first | to be the planned instantiations
        # the string between the first | and second | are the abbreivations
        # the string after | needs to be the definitions that you want to fix
        # the last variable needs to be the last variable used in the definition
        last_variable = input("enter last variable: ")
        # last_variable = "b"
        try:
            if last_variable[1] == "1":
                last_variable = last_variable[0] + l1
        except:
            pass

        while True:
            if variables9[0] != last_variable:
                variables9.remove(variables9[0])
            else:
                variables9.remove(variables9[0])
                break


        #total_sent = "(p⇒u) (b≍b) | (b=smith) & (c=jones) & (d=job) & (e=coin) & (f=10) & (g=pocket) & (o=man) & (s=important) & (t=true) & (w=partially caused) & (b₁=relation) | ((b KN p) ≡ ((b B p) & (p J c) & ((d W e) ≡ ((b B k) & (k J c) & (e I o))) & ((n W q) ≡ ((b B m) & (r J c) & (q I o))) & (d J u) & (n J v))) & (k ⇿ (e p H f) & (f I g) & (f I j) & (j J w)) & (m ⇿ (q p H s) & (s I g) & (s ~ I t) & (t J w)) & (r ⇿ (q p H s) & (s I g) & (s I t) & (t J w)) & (u=many) & (v=few) & (g=relation) & (o=relationship) & (w=important) & (c=true)"
        # total_sent = "* (c=extant) & (d=mind) & (e=real world) & (f=abstract) & (v=level two fundamental object) |  (d ~ J c E e) & (f J c E e) ⊥ (f=abstract) & ((b J f) ≡ ((b I z) ⊻ (b I y) ⊻ (b I x) ⊻ (b I g) ⊻ (b I h) ⊻ (b I j) ⊻ (b I k) ⊻ (b I m) ⊻ (b I w))) & (z=moment) & (y=number) & (x=point) & (g=property) & (h=relationship) & (j=whole) & (k=part) & (m=relation) & (w=symbol) ((d ~ J c U f) → (v ~ J c U f)) | ((z ~ J c U d) → (q ~ J c U d)) & (c=extant) & (q=abstract) & (z=level two fundamental object)"
        # total_sent = "* (c=extant) & (d=mind) & (e=real world) & (f=abstract) |  (d ~ J c E e) & (f J c E e) ⊥ | (c=abstract) & ((b J c) ≡ ((b I d) ⊻ (b I e) ⊻ (b I f) ⊻ (b I g) ⊻ (b I h) ⊻ (b I j) ⊻ (b I k) ⊻ (b I m) ⊻ (b I n))) & (d=moment) & (e=number) & (f=point) & (g=property) & (h=relationship) & (j=whole) & (k=part) & (m=relation) & (n=symbol)"
        # total_sent = "* (g=real) & (t=possible world) |  (p ⇿ n J b) (q ⇿ i D p) (p E c T f DP q T e) & (f SUT e) | (((b D p T c) & (p ~= q) & (d SUT c)) # (r P e T d)) & (p ⇿ r P e T d) & (q ⇿ (r M f T d) & (b W f T d))"
        # total_sent = " | (b=henry) & (c⇿this isᵍ aʳ realᵍ barn) & (d⇿thatᵈ isᵍ aʳ realᵍ barn) & (e=fact) & (f=degree) & (g=true) & (h=false) & (j=greatestᵈ) & (m=SMD) | ((b KN c) ≡ ((b B c) & (c J d) &(((e W f) ≡ (f SMD c g)) & ((h W j) ≡ (j SMD c k)) & (g G k) & ((m W n) ≡ ((e W n) & (b B n) & (n J d))) & ((o W p) ≡ ((e W p) & (b IG p))) & (m N q) & (o N r) & (s * q / (q + r)) & ((t W u) ≡ ((h W u) & (b B u) & (u J d))) & ((v W w) ≡ ((h W w) & (b IG w))) & (t N x) & (v N y) & (z * x / (x + y))) → (s GR z)) & (((b₁ SMD c c₁) & (c₁ GRA d₁ e₁)) → ((b B b₁) & (b₁ J d)))) & (d₁=degree) & (d=true) & (e₁=SMD)"

        first_pipe = total_sent.find("|")
        second_pipe = total_sent.find("|", first_pipe + 1)
        planned_instantiations = total_sent[:first_pipe]
        abbrev_sent = total_sent[first_pipe:second_pipe]
        new_sent = total_sent[second_pipe + 1:]
        new_sent = new_sent.strip()
        abbreviations = {}

        i = -1
        while i < len(abbrev_sent) - 1:
            i += 1
            if i == 82:
                bb = 8
            letter = abbrev_sent[i]
            if letter == "=":
                for j in range(i + 1, len(abbrev_sent)):
                    if abbrev_sent[j] == ")":
                        word = abbrev_sent[i + 1:j]
                        word.strip()
                        if abbrev_sent[i - 1] in subscripts:
                            var = abbrev_sent[i-2:i]
                        else:
                            var = abbrev_sent[i-1]
                        assert var[0].islower()
                        abbreviations.update({word: var})
                        i = j
                        break


        planned_instantiations += "*     "
        trans_sent = ""
        temp_abbrev = {}

        i = 0
        while i < len(planned_instantiations) -1:
            i += 1
            if planned_instantiations[i].islower():
                if planned_instantiations[i+1] in subscripts:
                    str1 = planned_instantiations[i:i+2]
                    if planned_instantiations[i+4] in subscripts:
                        str2 = planned_instantiations[i+3:i+5]
                        i += 4
                    else:
                        str2 = planned_instantiations[i + 3]
                        i += 3
                else:
                    str1 = planned_instantiations[i]
                    if planned_instantiations[i+3] in subscripts:
                        str2 = planned_instantiations[i+2:i+4]
                        i += 3
                    else:
                        str2 = planned_instantiations[i + 2]
                        i += 2
                dict1.update({str1:str2})


        i = -1
        while i < len(new_sent) - 1:
            i += 1
            if new_sent[i] == "=" and new_sent[i - 1] != "~":
                j = i
                letter = new_sent[i - 1]
                if letter in subscripts:
                    letter = new_sent[i-2:i]
                assert letter[0].islower()
                bool1 = True
                while bool1:
                    j += 1
                    if new_sent[j] == ")":
                        word = new_sent[i + 1:j]
                        word.strip()
                        temp_abbrev.update({word: letter})
                        i = j
                        break

        for k, v in temp_abbrev.items():
            if k in abbreviations.keys():
                new_var = abbreviations.get(k)
                dict1.update({v: new_var})

            else:
                if v not in variables9:
                    dict1.update({v: variables9[0]})
                    del variables9[0]
                else:
                    variables9.remove(v)

        i = 1
        while i < len(new_sent) - 1:
            i += 1

            if letter.islower() and new_sent[i - 1] != "=":
                if letter not in variables9 and letter not in dict1:
                    new_letter = variables9[0]
                    del variables9[0]
                    dict1.update({letter: new_letter})

            elif letter.islower() and new_sent[i - 1] == "=":
                k = new_sent.find(")", i, len(new_sent))
                i = k + 2

        new_sent = replace_letters(new_sent, dict1, type)

        for k, v in dict1.items():
            if k != v:
                trans_sent += "(" + k + idd + v + ") "


        print (trans_sent + "\n" + new_sent)

    elif type == "2":
        #this is for rearranging the variables in a definition
        list1 = total_sent.split("/")
        word = list1[0].strip()
        total_sent = list1[1].strip()
        if total_sent.count("(") > total_sent.count(")"):
            print ("more open paren than closed")
        elif total_sent.count("(") < total_sent.count(")"):
            print ("more closed parent than open")
        else:
            total_sent = replace_letters(total_sent, dict1, type)
            total_sent = adjust_definition(word, total_sent)

        print (" ")
        list1 = total_sent.split("|")
        for x in list1:
            x = x.strip()
            print (x)




    elif type == "3":


        planned_instantiations = total_sent[:total_sent.find("|")]
        new_sent = total_sent[total_sent.find("|"):]
        i = 0
        while i < len(planned_instantiations) - 1:
            i += 1
            if planned_instantiations[i].islower():
                if planned_instantiations[i + 1] in subscripts:
                    str1 = planned_instantiations[i:i + 2]
                    if planned_instantiations[i + 4] in subscripts:
                        str2 = planned_instantiations[i + 3:i + 5]
                        i += 4
                    else:
                        str2 = planned_instantiations[i + 3]
                        i += 3
                else:
                    str1 = planned_instantiations[i]
                    if planned_instantiations[i + 3] in subscripts:
                        str2 = planned_instantiations[i + 2:i + 4]
                        i += 3
                    else:
                        str2 = planned_instantiations[i + 2]
                        i += 2
                dict1.update({str1: str2})

        total_sent = replace_letters(new_sent, dict1, type)
        print (total_sent)


# definition = "((b ABF c) ≡ ((b S d T e.f) & (b H g T e.f) & (g.h I j) & (g N k) & (c ~ S d T e) & (c S d T f) & (f SUT e) & (m SUT f) & ((n I o) → (c ~ S n T m)) & (b S p T m) & (b H h T m) & (h N o) & (o G k))) & (j=energy) & (o=point)"
# type = "7"
type = input("enter type: ")

if type in ['1','2','3']:
    total_sent = input("sentence: ")
    change_variables(total_sent, type)
elif type == '5':
    row_num = input("row number: ")
    wb5 = load_workbook('/Users/kylefoley/Desktop/inference engine/dictionary5.xlsx')
    ws = wb5.worksheets[0]
    word = ws.cell(row=int(row_num), column=4).value
    definition = ws.cell(row=int(row_num), column=6).value
    definition = adjust_definition(word, definition)
    print ("\n" + definition + "\n")

elif type == '4':
    word = input("word: ")
    definition = input("definition: ")
    definition = adjust_definition(word, definition)
    print ("\n")
    if "|" in definition:
        list1 = definition.split("|")
        for x in list1:
            x = x.strip()
            print (x)
        print ("\n")
    else:
        print ("\n" + definition + "\n")

elif type == '6':
    definition = input("definition: ")
    def_info = find_sentences(definition)
    def_info, definition = period_elimination([def_info], definition)
    print ("\n" + definition + "\n")

elif type == '7':
    definition = input("definition: ")
    variables10 = [chr(98 + t) for t in range(25)]
    variables10.remove("i")
    variables10.remove("l")
    variables11 = [chr(97 + t) + l1 for t in range(26)]
    variables11.remove("l" + l1)
    i = 24
    dict1 = {}
    for var in variables11:
        i += 1
        dict1.update({var: str(i)})
    i = 0
    for var in variables10:
        i += 1
        dict1.update({var: str(i)})

    def_info = find_sentences(definition)
    def_info, definition = period_elimination([def_info], definition)
    for i, sent in enumerate(def_info[0][3]):
        if def_info[0][4][i][1] == "":
            if "=" not in sent:
                sent2 = sent
                for var, num in dict1.items():
                    if var in sent2:
                        sent2 = sent2.replace(var, num)
                def_info[0][5] = def_info[0][5].replace(def_info[0][6][i], sent2)
            else:
                def_info[0][5] = def_info[0][5].replace(def_info[0][6][i], sent)
    print ("\n" + def_info[0][5] + "\n")





