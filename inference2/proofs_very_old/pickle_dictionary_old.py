from general_functions_old import *
from settings_old import *
from openpyxl import load_workbook
import pickle
import json
import sys
import time
from standard_order_old import order_sentence
from analyze_sentence_old import find_sentences, eliminate_conjuncts_from_definition
from analyze_definition_old import build_hypo_list, build_greek_num_dict2, get_sent_kind,\
    get_sets_of_conditions, embedded_hypothetical_sentences, get_disjuncts


variables = [chr(122 - t) for t in range(25)]
variables.remove("i")
variables.remove("l")
variables3 = [chr(122 - t) + l1 for t in range(26)]
variables3.remove("l" + l1)
variables4 = [chr(122 - t) + l2 for t in range(26)]
variables4.remove("l" + l2)
variables5 = [chr(122 - t) + l3 for t in range(26)]
variables5.remove("l" + l3)
variables = variables + variables3 + variables4 + variables5

btime = time.time()

def update_synonyms(definition):

    str6 = definition[definition.find("=") + 1:-1]
    str6 = str6.strip()
    str7 = definition[1:definition.find("=")]
    str7 = str7.strip()
    dictionary[2].update({str7: str6})
    dictionary[1].update({str7: definition})


def cut_word(word):
    try:
        if "(" in word:
            cc = word.index("(")
            word = word[:cc - 1]
            word = word.strip()
        else:
            word = word.strip()
    except:
        word = ""

    return word


def is_a_definition(str1):
    if str1 == None:
        return False
    if "e.g." in str1:
        return False
    formal_symbols = ["&", "=", conditional, iff, xorr]
    for n in formal_symbols:
        if n in str1:
            return True
    return False


def get_prepositional_relations():
    str1 = third_sheet.cell(row=2, column=5).value
    str1a = third_sheet.cell(row=3, column=5).value
    str2 = third_sheet.cell(row=6, column=5).value
    str2a = third_sheet.cell(row=7, column=5).value
    str3 = third_sheet.cell(row=10, column=5).value
    str3a = third_sheet.cell(row=11, column=5).value
    prepositional = str1.split() + str1a.split()
    non_spatio_temporal = str2.split() + str2a.split()
    spatio_temporal = str3.split() + str3a.split()

    dictionary[8] = prepositional
    dictionary[13] = spatio_temporal
    dictionary[14] = non_spatio_temporal


def build_dictionary():
    global ws, wsar, third_sheet, dictionary
    dictionary = [{}, {}, {}, {}, {}, {}, [], {}, [], {}, {}, [], [], [], [], [], {}, []]

    #	0	parts of speech
    #	1	definitions
    #	2	synonyms
    #	3	abbreviated relations
    #	4	doubles
    #	5	triples
    #	6	individuals
    #	7	words to row
    #	8	prepositional relations
    #	9	categorized sent
    #	10	word decision procedure
    #	11	list of easy embeds
    #	12	embed type {}
    #	13	spatio temporal relations
    #	14	non spatio temporal relations
    #	15	concepts
    #   16  number of conjunctions
    #   17  past participles

    worksheet = ws

    get_prepositional_relations()


    i = 25
    while i < 3000:

        i += 1
        row_num = worksheet.cell(row=i, column=1).value
        pos = worksheet.cell(row=i, column=3).value
        word = worksheet.cell(row=i, column=4).value
        next_word = worksheet.cell(row=i + 1, column=4).value
        abbrev_relat = worksheet.cell(row=i, column=5).value
        defin = worksheet.cell(row=i, column=6).value
        next_defin = worksheet.cell(row=i + 1, column=6).value
        edisj = worksheet.cell(row=i, column=13).value
        embed = worksheet.cell(row=i, column=7).value
        easy_embed = worksheet.cell(row=i, column=8).value

        if word == 'spies on' or abbrev_relat == 'OFG':
            bb = 8

        if not not_blank(word) and not not_blank(next_word) and i > 300:
            break

        if not_blank(pos):
            if easy_embed in [1,6]: embed = False
            if embed == 'done': embed = False
            if not isinstance(pos, int): pos = pos.strip()
            if isinstance(word, int): word = str(word)
            if word == "true*": word = "true"
            if word == "false*": word = "false"
            word = cut_word(word)
            next_word = cut_word(next_word)

            fir_let, sec_let = put_in_categories(abbrev_relat, pos, row_num, word, defin)

            # universals are not defined, synonyms and determinative nouns are already done
            if sec_let not in ['a', 'b', 's', 'd'] and not embed and not_blank(defin):

                defin = defin.strip()

                if next_word == word and "e.g." not in next_defin:
                    while is_a_definition(worksheet.cell(row=i + 1, column=6).value) \
                    and next_word == word:
                        i += 1
                        defin += "| " + worksheet.cell(row=i, column=6).value.strip()
                        next_word = worksheet.cell(row=i + 1, column=4).value
                        next_word = cut_word(next_word)


                if fir_let == "r":
                    dictionary[1].update({abbrev_relat: defin})
                    if easy_embed in [1,6] and fir_let == 'd':
                        dictionary[11].append(abbrev_relat)
                        dictionary[12].update({abbrev_relat: easy_embed})


                else:
                    if easy_embed in [1,6] and fir_let == 'd':
                        dictionary[11].append(word)
                        # embed_type.update({word: easy_embed})
                    dictionary[1].update({word: defin})





def put_in_categories(abbrev_relat, pos, row_num, word, defin):
    if len(pos) == 1: pos += "z"
    dictionary[0].update({word: pos})
    fir_let = pos[0]
    sec_let = pos[1]
    thir_let = pos[2] if len(pos) > 2 else ""
    four_let = pos[3] if len(pos) > 3 else ""
    fif_let = pos[4] if len(pos) > 4 else ""
    place_in_decision_procedure2(word, fir_let, sec_let, dictionary[10])
    if " " in word: make_doubles(word)
    dictionary[7].update({word: row_num})
    if pos[1] in ['s', "d"]: update_synonyms(defin)
    if not_blank(abbrev_relat):
        place_in_decision_procedure2(abbrev_relat, fir_let, sec_let, dictionary[10])
        if abbrev_relat not in dictionary[7].keys():
            dictionary[7].update({abbrev_relat: row_num})
        if fir_let == 'r': dictionary[0].update({abbrev_relat: pos})
        if fir_let == 'r': dictionary[3].update({word: abbrev_relat})
        if len(abbrev_relat) == 4 and abbrev_relat[-1] == "P":dictionary[17].append(abbrev_relat)
    if fir_let == 'r' and sec_let not in ['s','a'] and not not_blank(abbrev_relat):
        raise Exception (f"you forgot to give {word} an abbreviation")

    return fir_let, sec_let

def make_doubles(word):
    m = word.count(" ")
    if m == 1:
        word1 = copy.copy(word)
        y = word1.find(" ")
        word1 = word1[:y]
        dictionary[4].setdefault(word1, []).append(word)
    if m == 2:
        word1 = copy.copy(word)
        y = word1.find(" ")
        word1 = word1[:y]
        dictionary[5].setdefault(word1, []).append(word)


def place_in_decision_procedure2(word, part_of_speech, sub_part, categories):
    category = None
    if sub_part == 'd':  # determinative nouns
        category = 19
    elif word in ['distinct from', 'different']:  # special synonyms, distinct from
        category = 20
    elif sub_part == 's':  # synonyms
        category = 18
    elif word == "i":
        category = .5
    elif part_of_speech == 'd' and sub_part not in ['b', 'i', 'e', 'd']:
        category = 1  # determinative, possessive pronouns
    elif part_of_speech == 'p':
        category = 2
    elif word == 'and' + uc:
        category = 5  # and
    elif part_of_speech == 'u':  # relative pronouns
        category = 9
    elif word in ['AS', 'as']:  # AS
        category = 11
    elif word in dictionary[13]:
        category = 13
    elif word in dictionary[14]:
        category = 13.5

    elif word == 'there':  # there
        category = 14
    elif part_of_speech == 'd' and sub_part == 'b':  # universals
        category = 15
    elif part_of_speech == 'd' and sub_part == 'e':  # many
        category = 16
    elif part_of_speech == 'm':  # negative determiners
        category = .4

    if category != None: categories.update({word: category})





def get_abbreviations_from_definition(def_info):
    # this function picks out that variables in the id sentences of the
    # definition

    constants = {}
    constant_map = {}
    temp_propositional_constants = {}

    list3 = []
    for i in range(len(def_info[3])):
        if one_sentence(def_info[3][i]) and mini_e in def_info[3][i]:
            list3.append(def_info[3][i])
        elif one_sentence(def_info[3][i]) and "=" in def_info[3][i]:
            str1 = def_info[3][i]
            g = str1.find("=")
            var = str1[1:g]
            word = str1[g + 1:-1]
            if isvariable(var):
                if not isvariable(word):
                    constants.update({var: word})

    return constants




#############################
##########  the following are for the determiners

def get_determiner_info(pos, definiendum, sentences):
    if definiendum == 'he':
        bb = 8

    instance_names = []
    instance_position = []
    concept_names = []
    concept_positions = []
    tvalue = []
    rsent_positions = []
    m = 11
    while sentences[0][m] != None:
        if sentences[0][m][13] == "R":
            rsent_positions.append(m)
            tvalue.append(sentences[0][m][3])
        if sentences[0][m][13] == "I":
            instance_names.append(sentences[0][m][10])
            instance_position.append([m, 10])
            concept_names.append(sentences[0][m][14])
            concept_positions.append([m, 14])
        m += 1
    sentences[0][30] = instance_names
    sentences[0][31] = instance_position
    sentences[0][32] = concept_names
    sentences[0][33] = concept_positions
    sentences[0][34] = tvalue
    sentences[0][35] = rsent_positions
    sentences[0][45] = []


def is_a_concept(sentences):
    n = 0 if len(sentences) == 1 else 1
    m = 10
    while sentences[n][m] != None:
        if sentences[n][m][6] in ['1.1', '2.1', '3.1']:
            if sentences[n][m][13] == 'I':
                dictionary[15].append(sentences[n][m][4])
            elif sentences[n][m][13] == "=":
                dictionary[6].append(sentences[n][m][4])
            break
        m += 1
    return


def renumber_sentences(definition, definiendum):

    def_info = find_sentences(definition)

    constants = get_abbreviations_from_definition(def_info)

    def_info = eliminate_conjuncts_from_definition(def_info)

    sentences = build_hypo_list(def_info, definiendum, constants)

    build_greek_num_dict2(sentences, def_info)

    definition, ordered = order_sentence(def_info, definition, sentences)

    assert not ordered

    return def_info, sentences, definition


def unembed(def_info, definiendum, red_kind):
    abbrev = []
    ogreek = def_info[5]
    type3 = False
    if red_kind == 6:
        type3 = True

    for sent, greek_sent in zip(def_info[3], def_info[6]):
        if mini_e in sent and len(greek_sent) == 1:
            abbrev.append(greek_sent)





    new_conjunct = " & ".join(abbrev)
    new_conjunct = " & " + new_conjunct

    e = 0
    special = False
    for lst1, lst2, lst3, lst4, lst6 in zip(def_info[1], def_info[2], def_info[3], def_info[4], def_info[6]):
        e += 1

        if e > 1:
            parent = ".".join(lst1[:-1])
            parent_pos = def_info[2].index(parent)
            parent_type = def_info[4][parent_pos][1]

            if parent_type == iff and parent in ['1', '1.1', '1.2']:
                if lst1[-1] == "2":



                    ogreek_definiens = lst6
                    if "&" not in ogreek_definiens:
                        special = True

                    break

    for gsent in abbrev: ogreek = ogreek.replace(" & " + gsent, "")



    if special:
        ngreek_definiens = "(" + ogreek_definiens + new_conjunct + ")"

    elif type3:

        ngreek_definiens = ogreek_definiens[:-2] + new_conjunct + "))"
    else:
        ngreek_definiens = ogreek_definiens[:-1] + new_conjunct + ")"
    ogreek = ogreek.replace(ogreek_definiens, ngreek_definiens)




    for greek_sent, english in zip(def_info[6], def_info[3]):
        ogreek = ogreek.replace(greek_sent, english)

    ogreek = ogreek.strip()
    ogreek = remove_outer_paren(ogreek)


    return find_sentences(ogreek, definiendum), ogreek


def print_new_definitions():
    for word in dictionary[11]:
        definition = dictionary[1].get(word)
        row_num = dictionary[7].get(word)
        if "|" in definition:
            list1 = definition.split("|")
            for i, part in enumerate(list1):
                ws.cell(row=row_num + i, column=6).value = part


        else:

            ws.cell(row=row_num, column=6).value = definition
        ws.cell(row=row_num, column=8).value = "done"
    return


def reduce_definitions():
    debug = False#
    z = 0

    for definiendum, definition in dictionary[1].items():

        pos = dictionary[0].get(definiendum)

        if definiendum == "natural" + ur:
            bb = 8

            if pos[1] not in ["s", "d"]:
                bb = 8

                definition = dictionary[1].get(definiendum)

                definition = definition.replace("|", "")

                try:

                    def_info = find_sentences(definition, definiendum)



                    # if definiendum in dictionary[11]:
                    #
                    #     embed_type = dictionary[12].get(definiendum)
                    #
                    #     def_info, definition = unembed(def_info, definiendum, embed_type)
                    #
                    #     dictionary[1][definiendum] = definition



                    constants = get_abbreviations_from_definition(def_info)

                    def_info = eliminate_conjuncts_from_definition(def_info)

                    n = len(def_info) - 1 if len(def_info) > 1 else len(def_info)

                    dictionary[16].update({definiendum: n})

                    sentences = build_hypo_list(def_info, definiendum, constants)

                    is_a_concept(sentences)

                    sent_kind = get_sent_kind(def_info)

                    if pos[0] not in ["d", "p"]:

                        definition, ordered = order_sentence(def_info, definition, sentences)

                        if ordered:
                            z += 1
                            def_info, sentences, definition = renumber_sentences(definition, definiendum)

                        def_info = get_sets_of_conditions(def_info, definiendum, sentences, sent_kind)

                        get_disjuncts(def_info, sentences)


                    else:

                        get_determiner_info(pos, definiendum, sentences)

                    sentences[0][80] = definition

                    dictionary[9].update({definiendum: sentences})

                    embedded_hypothetical_sentences(def_info, sentences, dictionary, definiendum)


                except ErrorWithCode as e:

                    print(e.code)
    print (z)


kind = 6


if kind == 5:
    print("dictionary built")
    reduce_definitions()

    output = open('z_dict_words_old.pkl', 'wb')
    pickle.dump(dictionary, output)
    output.close()
    # with open("z_dictionary.json", "w") as fp:
    #     json.dump(dictionary, fp)


elif kind == 6:
    wb4 = load_workbook('/Users/kylefoley/Desktop/inference_engine/dictionary5.xlsx')
    ws = wb4.worksheets[0]
    third_sheet = wb4.worksheets[2]
    print ('dictionary built')
    build_dictionary()
    reduce_definitions()
    print (time.time() - btime)
    output = open('z_dict_words_old.pkl', 'wb')
    pickle.dump(dictionary, output)
    output.close()





elif kind == 7:
    wb4 = load_workbook('/Users/kylefoley/Desktop/inference_engine/dictionary5.xlsx')
    ws = wb4.worksheets[0]
    dictionary = build_dictionary()
    for word in dictionary[11]:

        definition = dictionary[1].get(word)
        definition = adjust_definition(word, definition, dictionary)
        dictionary[1][word] = definition

    print_new_definitions()
    wb4.save('/Users/kylefoley/Desktop/inference_engine/dictionary5.xlsx')
























