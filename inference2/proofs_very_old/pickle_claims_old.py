
from openpyxl import load_workbook
import pickle
# from start_and_stop import info2
import copy

# proof_type, order, print_type = info2()
#
# if proof_type == 2:
#     wb = load_workbook('/Users/kylefoley/Desktop/inference_engine/artificial_tests.xlsm')
#     ws = wb.worksheets[0]
#     test_sent = []
#     list1 = []
#
#     for row in ws:
#         if row[1].value != None:
#             b = 0
#             list1.append(row[1].value)
#         else:
#             b += 1
#             if list1 != []:
#                 list2 = list1
#                 test_sent.append(copy.deepcopy(list2))
#             list1 = []
#             if b == 2:
#                 list1 = []
#                 break
#     if list1 != []:
#         test_sent.append(copy.deepcopy(list1))
#     output = open('zz_claims_artificial.pkl', 'wb')
#     pickle.dump(test_sent, output)
#     output.close()




wb = load_workbook('/Users/kylefoley/Desktop/inference_engine/inference engine new.xlsx')
ws = wb.worksheets[0]

test_sent = []
num = 0
last_string = 0
set_of_sentences = []
for row in ws:

    str1 = row[2].value

    if str1 != None and "*" not in str1:
        if row[3].value == 0: row[2].value = 'pass'

        if row[2].value != None and row[2].value != "":
            set_of_sentences.append(row[2].value)
        last_string = row[2].value
        if row[0].value == None:
            break
    elif set_of_sentences != []:
        test_sent.append(copy.deepcopy(set_of_sentences))
        set_of_sentences = []



output = open('zz_claims_old.pkl', 'wb')
pickle.dump(test_sent, output)
output.close()
