import json
import os
from django.core.serializers.json import DjangoJSONEncoder
from django.core.urlresolvers import reverse
from django.shortcuts import render, redirect
from wsgiref.util import FileWrapper

# from django.core.servers.basehttp import FileWrapper
from django.http import HttpResponse
from django.conf import settings
import time

from .models import Output, InstructionFile, Algorithm, Profile, Define3Notes, Settings, TestedDictionary
import importlib
from inference2.models import Input

from .models import Define3, Archives
import openpyxl
from openpyxl.cell import get_column_letter
from django.contrib import messages
from .main_loop import get_result

DEFAULT_ROWS = 40000


def save_result(archive_id, post_data):
    Output.objects.all().delete()
    archive = Archives.objects.get(pk=archive_id)
    Rows = []
    data_found = False

    rows_settings = Settings.objects.first()
    rows_to_show = rows_settings.rows_to_show if rows_settings else (len(post_data) - 1)
    if rows_to_show > (len(post_data) - 1):
        rows_to_show = len(post_data) - 1
    for idx in range(min(rows_to_show, len(post_data))):
        if post_data.get("text_" + str(idx) + "_2", '') or post_data.get("text_" + str(idx) + "_3", ''):
            data_found = True
        if not data_found:
            continue
        c1 = post_data.get("text_" + str(idx) + "_1", '')
        c2 = post_data.get("text_" + str(idx) + "_2", '')
        c3 = post_data.get("text_" + str(idx) + "_3", '')
        if type(c1) == type([]):
            c1 = c1[0]
        if type(c2) == type([]):
            c2 = c2[0]
        if type(c3) == type([]):
            c3 = c3[0]
        R = Output(col1=c1,
                   col2=c2,
                   col3=c3,
                   archives=archive,
                   )
        Rows.append(R)
    Output.objects.bulk_create(Rows)


def current_archive():
    archive = Archives.objects.latest('archives_date')
    return archive


def id_file_pdf(inst_file):
    if inst_file and inst_file.file_extension == InstructionFile.PDF:
        return True
    return False


def make_file_path(inst_file):
    if (inst_file):
        return '/' + str(inst_file.data)
    return ''


def index(request, archive=None):
    ins_file = InstructionFile.objects.filter(file_type='0').order_by('-id').first()
    download_dict_file = InstructionFile.objects.filter(file_type='1').order_by('-id').first()
    rules_in_brief_file = InstructionFile.objects.filter(file_type='2').order_by('-id').first()
    arguments = InstructionFile.objects.filter(file_type='3').order_by('-id').first()

    is_pdf_file = id_file_pdf(ins_file)
    is_dict_pdf_file = id_file_pdf(download_dict_file)
    is_rules_in_bried_pdf_file = id_file_pdf(rules_in_brief_file)
    is_arguments_pdf_file = id_file_pdf(arguments)

    ins_file = make_file_path(ins_file)
    download_dict_file = make_file_path(download_dict_file)
    rules_in_brief_file = make_file_path(rules_in_brief_file)
    arguments_file = make_file_path(arguments)

    progressbar_send(request, 1, 100, 1)
    url_path = ''
    archive_date = ''
    if not archive:
        archive = current_archive()
        url_path = '/'
    else:
        url_path = '/archives/{}/'.format(archive.id)
        archive_date = archive.archives_date
    input = Input.objects.filter(archives_id=archive.id)
    result = {}
    output = []
    show_column = False
    if request.method == 'POST':
        show_column = True
        Output.objects.all().delete()
        post_data = request.POST.copy()
        prove_algorithm = importlib.import_module('.' + archive.algorithm.split('.py')[0], package='inference2.Proofs')
        prove_dictionary = importlib.import_module('.' + archive.dictionary.split('.py')[0],
                                                   package='inference2.Proofs')
        post_data = prove_algorithm.get_result(
            request.POST.copy(), archive.id, request, prove_dict=prove_dictionary)
        print(post_data)
        if post_data:
            post_data["type"] = "prove"
            result = json.dumps(post_data, cls=DjangoJSONEncoder)

            save_result(archive.id, post_data)
        output = Output.objects.all()

    algo = Algorithm.objects.all().order_by('id')

    template_args = {'result': result, 'input': input,
                     'url_path': url_path, 'archive_date': archive_date,
                     'output': output, 'ins_file': ins_file, 'download_dict_file': download_dict_file,
                     'download_dict_pdf': is_dict_pdf_file,
                     'rules_in_brief_file': rules_in_brief_file,
                     'argument_file': arguments_file,
                     'is_rules_in_bried_pdf_file': is_rules_in_bried_pdf_file,
                     'is_arguments_pdf_file': is_arguments_pdf_file,
                     'archive': archive, 'show_column': show_column, 'algo': algo[0].name if algo else archive,
                     'notes': algo[0].notes if algo else '', 'pdf': is_pdf_file
                     }
    return render(request, "inference2/index.html", template_args)


def try_input(request, archive=None):
    output = []
    template_args = {}
    template_args['success'] = 'Right'
    if not archive:
        archive = current_archive()
        url_path = '/'
    if request.method == 'POST':
        try:
            # input = "It is|a contradictory that I do not have many|n points"
            input = request.POST.get('try_input')
            Output.objects.all().delete()
            # prove_algorithm = importlib.import_module('.' + archive.test_machine.split('.py')[0],
            #                                           package='inference2.Proofs')
            result_string = get_result(input)
            # result_string = prove_algorithm.get_result_from_views(
            #     request.POST.copy(), archive.id, request, input)
            print_output(result_string)

            template_args['result'] = result_string
            print(post_data)
            if post_data:
                post_data["type"] = "prove"
                result = json.dumps(post_data, cls=DjangoJSONEncoder)

                save_result(archive.id, post_data)
            output = Output.objects.all()
        except Exception as e:
            messages.error(request, str(e))
            template_args['success'] = 'Wrong'
            template_args['result'] = 'Wrong'

    algo = Algorithm.objects.all().order_by('id')
    template_args['notes'] = algo[0].try_input_notes if algo else ''
    template_args['url_path'] = url_path
    template_args['output'] = output
    template_args['archive'] = archive

    return render(request, "inference2/try_input.html", template_args)


def print_output(test_sent):

    row_number = 1

    for i in order:
        for j in range(len(test_sent[i])):
            rule = ""
            if test_sent[i][j][4] != "":
                rule = test_sent[i][j][4] + " "
                if test_sent[i][j][5] != "":
                    rule += str(test_sent[i][j][5])

                    if test_sent[i][j][6] != "":
                        rule += "," + str(test_sent[i][j][6])

            result_data['text_' + str(row_number - 1) + '_1'] = test_sent[i][j][0]
            result_data['text_' + str(row_number - 1) + '_2'] = test_sent[i][j][3] + test_sent[i][j][1]
            result_data['text_' + str(row_number - 1) + '_3'] = rule
                
            row_number += 1

    return result_data

def progress(count, total, suffix=''):
    bar_len = 60
    filled_len = int(round(bar_len * count / float(total)))

    percents = round(100.0 * count / float(total), 1)
    bar = '=' * filled_len + '-' * (bar_len - filled_len)

    sys.stdout.write('[%s] %s%s ...%s\r' % (bar, percents, '%', suffix))
    sys.stdout.flush()
    ERROR_MESSAGES = ["Each sentence must begin with either 'it is|a consistent that' or 'it is|a contradictory that",
                      "our system does not have this grammatical syntax yet", "you misspelled", ]

    # if print_type in [8, 9]:
    #     views.progressbar_send(request, 0, 100, 100, 2)
    #     views.save_result(archive_id, result_data)
    #     if print_type == 9:
    #         return result_data, result
    #     return result_data
    
# 
# def get_result_from_views(post_data, archive_id=None, request=None, input=None, prove_dict=None):
#     if print_type == 9:
#         try:
#             return get_result(post_data, archive_id, request, input, prove_dict)
#         except Exception as e:
#             if str(e) in ERROR_MESSAGES or str(e).startswith(ERROR_MESSAGES[2]):
#                 raise (e)
#             else:
#                 raise Exception("Our Fault Not Yours")
#     else:
#         return get_result(post_data, archive_id, request, input, prove_dict)


def export_xlsx(request, archives_id=None):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"
    only_output = request.GET.get('only_output', None)
    only_input = request.GET.get('only_input', None)

    queryset = Output.objects.filter(archives_id=int(archives_id))
    input_queryset = Input.objects.filter(archives_id=int(archives_id))
    row_num = 0

    columns = [
        (u"ID", 15),
        (u"Title", 70),
        (u"Description", 70),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(
            col_num + 1)].width = columns[col_num][1]
    if not only_output:
        for obj in input_queryset:
            row_num += 1
            row = [
                obj.col1,
                obj.col2,
                obj.col3,
            ]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                c.style.alignment.wrap_text = True

    if not only_input:
        for obj in queryset:
            row_num += 1
            row = [
                obj.col1,
                obj.col2,
                obj.col3,
            ]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                c.style.alignment.wrap_text = True

    wb.save(response)
    return response


def stream_response_generator(request):
    for x in range(1, 11):
        yield "%s\n" % x  # Returns a chunk of the response to the browser
        request.session['idx'] = x
        request.session.modified = True

        request.session.save()
        time.sleep(1)


def prove(request, archive=None):
    progressbar_send(request, 1, 100, 1)
    if not archive:
        archive = current_archive()
    result = {}
    if request.method == 'POST':
        post_data = request.POST.copy()
        prove_algorithm = importlib.import_module(
            '.' + archive.algorithm, package='inference2.Proofs')
        post_data = prove_algorithm.get_result(
            request.POST.copy(), archive.id, request)
        result = json.dumps(post_data, cls=DjangoJSONEncoder)

    return result


def dictionary(request, archive=None):
    from inference2.Proofs.dictionary_new import large_dict
    outputs = Define3.objects.all()
    notes = Define3Notes.objects.all().order_by('id')
    return render(request, "inference2/dict.html",
                  {'result': large_dict, 'url_path': '/', 'output': outputs,
                   'notes': notes[0].notes if notes else '', })


def tested_dict(request, archive=None):
    from inference2.Proofs.dictionary_new import large_dict
    outputs = TestedDictionary.objects.all()
    notes = Define3Notes.objects.all().order_by('id')
    return render(request, "inference2/dict.html",
                  {'result': large_dict, 'url_path': '/', 'output': outputs,
                   'notes': notes[1].notes if notes.count() > 1 else '', 'tested_dict': True})


def tested_dictionary(request, archive=None):
    from inference2.Proofs.dictionary_tested import large_test_dict
    return render(request, "inference2/tested_dict.html", {'result': large_test_dict, 'url_path': '/'})


def download_files(request):
    ins_files = InstructionFile.objects.filter(
        file_type='1').order_by('-id')
    return render(request, "inference2/files.html", {'ins_files': ins_files})


# def download_files_in_brief(request):
#     ins_files = InstructionFile.objects.filter(
#         file_type='2').order_by('-id')
#     return render(request, "inference2/files_in_brief.html", {'ins_files': ins_files})


def archives(request):
    dict = Archives.objects.all()
    return render(request, "inference2/archives.html", {'result': dict})


def assign_archives(request, num=-1, type=None):
    if num == -1:
        return
    archive = Archives.objects.get(pk=num)
    if type:
        return globals()[type](request, archive)
    else:
        return index(request, archive)


def manual(request):
    filename = os.path.join(settings.MANUAL_PATH)
    wrapper = FileWrapper(file(filename))
    response = HttpResponse(
        wrapper, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Length'] = os.path.getsize(filename)
    response['Content-Disposition'] = 'attachment; filename=' + \
                                      os.path.basename(filename)
    return response


def getdict(request, archive=None):
    if not archive:
        archive = current_archive()

    filename = os.path.join(settings.DICT_DIRS, archive.algorithm + ".csv")
    if not os.path.exists(filename):
        return HttpResponse("No CSV found for Algorithm %s" % archive.algorithm)
    wrapper = FileWrapper(file(filename))
    response = HttpResponse(
        wrapper, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Length'] = os.path.getsize(filename)
    response['Content-Disposition'] = 'attachment; filename=' + \
                                      os.path.basename(filename)
    return response


def progress(request):
    contxt = {"K": request.session['idx']}
    if request.session.get('status', 0) == 2:
        progressbar_send(request, 1, 100, 1)
        print("prog123 %s" % contxt)
    return HttpResponse(json.dumps(contxt), content_type="application/json")


def progressbar_send(request, strt, stp, k, status=0):
    if request is not None:
        request.session.modified = True
        request.session['strt'] = 0
        request.session['stp'] = 100
        request.session['idx'] = [0, 100, k]
        request.session['status'] = status
        request.session.modified = True
        request.session.save()


def author(request):
    try:
        profile = Profile.objects.latest('id')
    except Profile.DoesNotExist:
        profile = None
    return render(request, "inference2/author.html", {'profile': profile})


def clear(request):
    return redirect(reverse('index'))


def clear_output(request):
    Output.objects.all().delete()
    return redirect(reverse('try_input'))
